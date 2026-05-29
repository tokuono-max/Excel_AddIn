# -*- coding: utf-8 -*-
"""ODN-164 / ODN375 実シナリオ JSON と手元テストファイルでの結合確認（任意）。"""
from __future__ import annotations

import json
import os
import sys
from pathlib import Path
from typing import Any

import pytest

_root = Path(__file__).resolve().parents[1]
if str(_root) not in sys.path:
    sys.path.insert(0, str(_root))

from svc.svc_data_agg import compute_batch_table_rows  # noqa: E402

_TEST_DATA_ROOT = Path(
    os.environ.get("DATA_AGG_TEST_DATA_ROOT", r"C:\Project\データ集約テストファイル")
)
_ODN164_JSON = _TEST_DATA_ROOT / "ODN164" / "ODN-164出荷履歴試験用.json"
_ODN375_JSON = _TEST_DATA_ROOT / "ODN375_ALL_1NCM90024.json"
_ODN375_DEFAULT_XLSX = (
    _TEST_DATA_ROOT / "出荷履歴" / "ODN375_A0512M100000.xlsx"
)


def _odn375_xlsx_path() -> str:
    xlsx = os.environ.get("DATA_AGG_ODN375_XLSX", "").strip()
    if xlsx:
        return xlsx
    if _ODN375_DEFAULT_XLSX.is_file():
        return str(_ODN375_DEFAULT_XLSX)
    return ""


def _odn375_sheet_cells(xlsx: str) -> dict[str, Any]:
    """ユニット実装チェックの結合・主キー参照セル（シナリオ JSON と同じ）。"""
    import openpyxl  # noqa: E402

    wb = openpyxl.load_workbook(xlsx, read_only=True, data_only=True)
    try:
        ws = wb["ユニット実装チェック"]
        return {
            "AS30": ws["AS30"].value,
            "AS31": ws["AS31"].value,
            "AV30": ws["AV30"].value,
            "AV31": ws["AV31"].value,
        }
    finally:
        wb.close()


def _odn164_xlsx_files() -> list[str]:
    d = _TEST_DATA_ROOT / "ODN164"
    if not d.is_dir():
        return []
    return [str(p) for p in sorted(d.glob("*.xlsx")) if not p.name.startswith("~")]


def _odn164_link_and_hist_paths() -> tuple[str, str]:
    """紐づけ履歴 xlsx と光特性履歴 xlsx のパス（ファイル名に依存）。"""
    link_f = ""
    hist_f = ""
    for f in _odn164_xlsx_files():
        name = Path(f).name
        if "紐" in name or "づ" in name:
            link_f = f
        elif "光" in name or "特性" in name:
            hist_f = f
    return link_f, hist_f


def _odn164_link_first_row(link_xlsx: str) -> dict[str, Any]:
    """紐付け履歴シート先頭データ行（C5/P5/J5。シナリオの join/link 起点）。"""
    import openpyxl  # noqa: E402

    wb = openpyxl.load_workbook(link_xlsx, read_only=True, data_only=True)
    try:
        ws = wb["紐付け履歴"]
        return {
            "pt": ws["C5"].value,
            "mac": ws["P5"].value,
            "seq": ws["J5"].value,
        }
    finally:
        wb.close()


def _load_json(path: Path) -> dict[str, Any]:
    with open(path, encoding="utf-8") as fp:
        return json.load(fp)


@pytest.mark.skipif(
    not _ODN164_JSON.is_file() or len(_odn164_xlsx_files()) < 2,
    reason="ODN-164 シナリオまたは xlsx が手元に無い",
)
def test_odn164_scenario_row_count_and_join_columns() -> None:
    """光特性行のみ出力（≒3597）。PT・製番は MAC 一致分に付与。7195 行化しない。"""
    data = _load_json(_ODN164_JSON)
    headers, rows, _, join_events = compute_batch_table_rows(
        data, _odn164_xlsx_files(), max_primary_rows=5000, max_table_rows=5000
    )
    ix_dev = headers.index("機器番号")
    ix_pt = headers.index("PT番号")
    ix_seq = headers.index("製番")
    ix_mac = headers.index("MACアドレス")

    assert join_events == 0
    assert 3500 <= len(rows) <= 3700
    assert all(r[ix_dev] not in (None, "") for r in rows)
    assert all(r[ix_mac] not in (None, "") for r in rows)

    pt_filled = sum(1 for r in rows if r[ix_pt] not in (None, ""))
    seq_filled = sum(1 for r in rows if r[ix_seq] not in (None, ""))
    assert pt_filled > 0
    assert pt_filled == seq_filled
    assert pt_filled <= len(rows)


@pytest.mark.skipif(
    not _ODN164_JSON.is_file() or len(_odn164_xlsx_files()) < 2,
    reason="ODN-164 シナリオまたは xlsx が手元に無い",
)
def test_odn164_pt_seq_join_on_matching_mac() -> None:
    """
    紐づけ先頭行の MAC と一致する光特性行へ PT番号・製番が横断結合で付く（セル値で検証）。
    """
    link_f, hist_f = _odn164_link_and_hist_paths()
    if not link_f or not hist_f:
        pytest.skip("光特性・紐づけの xlsx が見つからない")

    sample = _odn164_link_first_row(link_f)
    mac = sample.get("mac")
    if mac in (None, ""):
        pytest.skip("紐づけ P5（MAC）が空")

    data = _load_json(_ODN164_JSON)
    headers, rows, _, _ = compute_batch_table_rows(
        data, [hist_f, link_f], max_table_rows=5000
    )
    ix_mac = headers.index("MACアドレス")
    ix_pt = headers.index("PT番号")
    ix_seq = headers.index("製番")

    mac_s = str(mac).strip()
    matched = [r for r in rows if str(r[ix_mac] or "").strip() == mac_s]
    assert len(matched) >= 1, "MAC 一致の光特性行が無い"
    row = matched[0]
    assert row[ix_pt] == sample.get("pt"), "PT番号は紐づけ C5 と一致"
    assert row[ix_seq] == sample.get("seq"), "製番は紐づけ J5 と一致"


@pytest.mark.skipif(
    not _ODN164_JSON.is_file() or len(_odn164_xlsx_files()) < 2,
    reason="ODN-164 シナリオまたは xlsx が手元に無い",
)
def test_odn164_master_preview_stepped_pt_item_has_join_data() -> None:
    """マスタプレビュー（preview_master）で PT 項目到達時も paths が残り PT が空にならない。"""
    from svc.data_agg_master_preview import scenario_for_stepped_preview

    data = _load_json(_ODN164_JSON)
    items = data.get("items") or []
    mi_pt = next(
        (i for i, it in enumerate(items) if str(it.get("name") or "") == "PT番号"),
        -1,
    )
    if mi_pt < 0:
        pytest.skip("PT番号 項目なし")
    act = [0]
    scen = scenario_for_stepped_preview(
        data,
        mi_idx=mi_pt,
        master_step_idx=1,
        active_slot_indices=act,
    )
    fps = _odn164_xlsx_files()
    headers, rows, _, _ = compute_batch_table_rows(
        scen,
        fps,
        max_primary_rows=500,
        max_table_rows=500,
        probe_caller="test_master_preview",
    )
    ix_pt = headers.index("PT番号")
    pt_filled = sum(1 for r in rows if r[ix_pt] not in (None, ""))
    assert len(rows) > 0
    assert pt_filled > 0


@pytest.mark.skipif(
    not _ODN375_JSON.is_file(),
    reason="ODN375 シナリオ JSON が手元に無い",
)
def test_odn375_scenario_runs_without_error() -> None:
    """ODN375 JSON が読める xlsx 1 件と組み合わせて例外なく完走する（データは環境依存）。"""
    data = _load_json(_ODN375_JSON)
    xlsx = _odn375_xlsx_path()
    if not xlsx:
        pytest.skip("DATA_AGG_ODN375_XLSX 未設定かつ既定 xlsx が無い")
    fp = Path(xlsx)
    if not fp.is_file():
        pytest.skip("ODN375 xlsx が存在しない")
    headers, rows, _, _ = compute_batch_table_rows(data, [str(fp)], max_table_rows=50)
    assert isinstance(headers, list)
    assert isinstance(rows, list)


@pytest.mark.skipif(
    not _ODN375_JSON.is_file(),
    reason="ODN375 シナリオ JSON が手元に無い",
)
def test_odn375_mac_loc_rmt_join_on_matching_device() -> None:
    """
    機器番号＝AS30 の行: MAC 空、MAC LOC/RMT＝AV30/AV31（結合キー＋空連携の仕様）。
    既定: 出荷履歴/ODN375_A0512M100000.xlsx（DATA_AGG_ODN375_XLSX で上書き可）。
    """
    xlsx = _odn375_xlsx_path()
    if not xlsx or not Path(xlsx).is_file():
        pytest.skip("ODN375 xlsx が無い")

    cells = _odn375_sheet_cells(xlsx)
    join_dev = cells.get("AS30")
    if join_dev in (None, ""):
        pytest.skip("AS30 が空のため結合検証不可")

    data = _load_json(_ODN375_JSON)
    headers, rows, _, _ = compute_batch_table_rows(
        data, [xlsx], max_table_rows=500
    )
    ix_dev = headers.index("機器番号")
    ix_mac = headers.index("MAC")
    ix_loc = headers.index("MAC LOC")
    ix_rmt = headers.index("MAC RMT")

    matched = [r for r in rows if r[ix_dev] == join_dev]
    assert len(matched) >= 1, "AS30 と一致する機器番号行が無い"
    row = matched[0]
    assert row[ix_mac] in (None, ""), "一致行の MAC は空クリア"
    assert row[ix_loc] == cells.get("AV30"), "MAC LOC は AV30 の主キー値"
    if cells.get("AV31") not in (None, ""):
        assert row[ix_rmt] == cells.get("AV31"), "MAC RMT は AV31 の主キー値"

    # 結合対象外の行に MAC LOC が付かないこと（先頭ユニット行など）
    others = [r for r in rows if r[ix_dev] not in (None, "") and r[ix_dev] != join_dev]
    assert others, "比較用の非一致行が無い"
    assert all(r[ix_loc] in (None, "") for r in others[:5])
