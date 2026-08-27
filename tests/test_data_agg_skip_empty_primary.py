# -*- coding: utf-8 -*-
"""N件 + skip_empty_primary の空主キー反復スキップ。"""
from __future__ import annotations

import sys
from pathlib import Path

_root = Path(__file__).resolve().parents[1]
if str(_root) not in sys.path:
    sys.path.insert(0, str(_root))

from openpyxl import Workbook  # noqa: E402

from svc.svc_data_agg_extract import extract_item_bundle, xlsx_workbook_scope  # noqa: E402


def _make_xlsx(tmp_path: Path) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = "S"
    # A1=A, A2=空, A3=B  / B1=L1, B2=L2, B3=L3  / C1=J1, C2=J2, C3=J3
    ws["A1"] = "A"
    ws["A2"] = None
    ws["A3"] = "B"
    ws["B1"] = "L1"
    ws["B2"] = "L2"
    ws["B3"] = "L3"
    ws["C1"] = "J1"
    ws["C2"] = "J2"
    ws["C3"] = "J3"
    fp = tmp_path / "skip_empty.xlsx"
    wb.save(fp)
    wb.close()
    return fp


def _item(*, skip_empty: bool) -> dict:
    return {
        "name": "PK",
        "sources": [
            {
                "type": "cell",
                "sheet_name": "S",
                "cell_ref": "A1",
                "row_offset": 1,
                "col_offset": 0,
                "repeat_direction": "vertical",
                "repeat_until_empty": False,
                "repeat_max": 3,
                "skip_empty_primary": skip_empty,
                "ui_scenario_source_v1": {
                    "link_defs": [
                        {
                            "item": "連携",
                            "mode": "セル座標",
                            "cell": "B1",
                            "row": 1,
                            "col": 0,
                        }
                    ],
                    "join_defs": [
                        {
                            "item": "結合",
                            "mode": "セル座標",
                            "cell": "C1",
                            "row": 1,
                            "col": 0,
                        }
                    ],
                },
            }
        ],
    }


def test_skip_empty_primary_drops_blank_with_link_join(tmp_path: Path) -> None:
    fp = _make_xlsx(tmp_path)
    with xlsx_workbook_scope():
        b = extract_item_bundle(str(fp), _item(skip_empty=True), item_id="i1")
    prim = [str(x).lstrip("'") for x in (b["primary_values"] or [])]
    assert prim == ["A", "B"]
    links = [str(x).lstrip("'") for x in (b["link_values"]["連携"] or [])]
    joins = [str(x).lstrip("'") for x in (b["join_values"]["結合"] or [])]
    assert links == ["L1", "L3"]
    assert joins == ["J1", "J3"]
    ictx = b.get("iteration_contexts") or []
    assert [c.get("rule_iter_index") for c in ictx] == [0, 2]
    assert [c.get("iter_index") for c in ictx] == [0, 1]
    # 本番の行割当は file_path+iter_index。圧縮後は連番に揃える。
    lctx = (b.get("link_contexts") or {}).get("連携") or []
    jctx = (b.get("join_contexts") or {}).get("結合") or []
    assert [c.get("iter_index") for c in lctx] == [0, 1]
    assert [c.get("iter_index") for c in jctx] == [0, 1]


def test_skip_empty_primary_assign_by_context_fills_all_rows(tmp_path: Path) -> None:
    """空スキップ後も _assign_series_to_rows_by_context で全主値行に連携が載る。"""
    from svc.svc_data_agg import _assign_series_to_rows_by_context

    fp = _make_xlsx(tmp_path)
    with xlsx_workbook_scope():
        b = extract_item_bundle(str(fp), _item(skip_empty=True), item_id="i1")
    prim = list(b["primary_values"] or [])
    rows = [
        {
            "PK": v,
            "__file_path": str(fp),
            "__iter_index": int(i),
        }
        for i, v in enumerate(prim)
    ]
    _assign_series_to_rows_by_context(
        rows,
        "連携",
        list((b.get("link_values") or {}).get("連携") or []),
        list((b.get("link_contexts") or {}).get("連携") or []),
        str(fp),
        write_mode="overwrite",
    )
    assert [str(r.get("連携")).lstrip("'") for r in rows] == ["L1", "L3"]


def test_skip_empty_primary_off_keeps_legacy_empty_drop(tmp_path: Path) -> None:
    """フラグOFF: 空主キーは主値から除外される（従来）。"""
    fp = _make_xlsx(tmp_path)
    with xlsx_workbook_scope():
        b = extract_item_bundle(str(fp), _item(skip_empty=False), item_id="i1")
    prim = [str(x).lstrip("'") for x in (b["primary_values"] or [])]
    assert prim == ["A", "B"]


def test_skip_empty_multi_sheet_oneshot_link_context_iter_offset(tmp_path: Path) -> None:
    """含む＋複数シートの一括抽出で、連携コンテキスト iter_index がシート連結オフセットになる。"""
    from openpyxl import Workbook

    from svc.svc_data_agg import _assign_series_to_rows_by_context

    fp = tmp_path / "multi_oneshot.xlsx"
    wb = Workbook()
    wb.active.title = "Other"
    for sh, pfx in (("R_1", "T"), ("R_2", "M")):
        s = wb.create_sheet(sh)
        rows = [
            (f"{pfx}S1", f"{pfx}L1"),
            (None, f"{pfx}L2"),
            (f"{pfx}S3", f"{pfx}L3"),
        ]
        for i, (a, b) in enumerate(rows, start=1):
            s[f"A{i}"] = a
            s[f"B{i}"] = b
    wb.save(fp)
    wb.close()

    item = {
        "name": "出荷番号",
        "sources": [
            {
                "type": "cell",
                "sheet_name": "R_",
                "cell_ref": "A1",
                "row_offset": 1,
                "col_offset": 0,
                "repeat_direction": "vertical",
                "repeat_until_empty": False,
                "repeat_max": 3,
                "skip_empty_primary": True,
                "ui_scenario_source_v1": {
                    "sheet_rule": "含む",
                    "link_defs": [
                        {
                            "item": "連携",
                            "mode": "セル座標",
                            "cell": "B1",
                            "row": 1,
                            "col": 0,
                        }
                    ],
                },
            }
        ],
    }
    with xlsx_workbook_scope():
        b = extract_item_bundle(str(fp), item, item_id="i1")
    prim = [str(x).lstrip("'") for x in (b["primary_values"] or [])]
    links = [str(x).lstrip("'") for x in (b["link_values"]["連携"] or [])]
    lctx = [c.get("iter_index") for c in (b["link_contexts"]["連携"] or [])]
    assert prim == ["TS1", "TS3", "MS1", "MS3"]
    assert links == ["TL1", "TL3", "ML1", "ML3"]
    assert lctx == [0, 1, 2, 3]

    rows = [
        {"出荷番号": v, "__file_path": str(fp), "__iter_index": int(i)}
        for i, v in enumerate(b["primary_values"] or [])
    ]
    _assign_series_to_rows_by_context(
        rows,
        "連携",
        list(b["link_values"]["連携"]),
        list(b["link_contexts"]["連携"]),
        str(fp),
        write_mode="overwrite",
    )
    assert [str(r.get("連携")).lstrip("'") for r in rows] == ["TL1", "TL3", "ML1", "ML3"]


def test_skip_empty_primary_phased_multi_sheet_keeps_link_join_aligned(
    tmp_path: Path,
) -> None:
    """含む＋複数シートの段階抽出でも、空スキップ後の連携/結合が元オフセットを保つ。"""
    from openpyxl import Workbook

    from svc.svc_data_agg_debug_run import scenario_debug_phase_result

    fp = tmp_path / "multi_skip.xlsx"
    wb = Workbook()
    wb.active.title = "Other"
    wb.active["A1"] = "X"
    s1 = wb.create_sheet("R_1")
    s1["A1"] = "A"
    s1["A2"] = None
    s1["A3"] = "B"
    s1["B1"] = "L1"
    s1["B2"] = "L2"
    s1["B3"] = "L3"
    s1["C1"] = "J1"
    s1["C2"] = "J2"
    s1["C3"] = "J3"
    s2 = wb.create_sheet("R_2")
    s2["A1"] = "C"
    s2["A2"] = ""
    s2["A3"] = "D"
    s2["B1"] = "L4"
    s2["B2"] = "L5"
    s2["B3"] = "L6"
    s2["C1"] = "J4"
    s2["C2"] = "J5"
    s2["C3"] = "J6"
    wb.save(fp)
    wb.close()

    item = {
        "id": "i1",
        "name": "PK",
        "sources": [
            {
                "type": "cell",
                "sheet_name": "R_",
                "cell_ref": "A1",
                "row_offset": 1,
                "col_offset": 0,
                "repeat_direction": "vertical",
                "repeat_until_empty": False,
                "repeat_max": 3,
                "skip_empty_primary": True,
                "ui_scenario_source_v1": {
                    "sheet_rule": "含む",
                    "ext_checked": [".xlsx"],
                    "link_defs": [
                        {
                            "item": "連携",
                            "mode": "セル座標",
                            "cell": "B1",
                            "row": 1,
                            "col": 0,
                        }
                    ],
                    "join_defs": [
                        {
                            "item": "結合",
                            "mode": "セル座標",
                            "cell": "C1",
                            "row": 1,
                            "col": 0,
                        }
                    ],
                },
            }
        ],
    }
    paths = [str(fp)]
    cache: dict = {}
    _s2, cols2, _e2, _t2 = scenario_debug_phase_result(item, paths, 2, 200, cache)
    _s3, cols3, _e3, _t3 = scenario_debug_phase_result(item, paths, 3, 200, cache)
    _s4, cols4, _e4, _t4 = scenario_debug_phase_result(item, paths, 4, 200, cache)

    def _body(v: object) -> str:
        s = str(v).lstrip("'")
        # UI 表示: #1[連携] value
        if "] " in s:
            s = s.split("] ", 1)[-1].lstrip("'")
        return s

    assert [_body(x) for x in cols2] == ["A", "B", "C", "D"]
    assert [_body(x) for x in cols3] == ["L1", "L3", "L4", "L6"]
    assert [_body(x) for x in cols4] == ["J1", "J3", "J4", "J6"]
    b = cache[str(fp)]
    assert [str(x).lstrip("'") for x in (b.get("link_values") or {}).get("連携", [])] == [
        "L1",
        "L3",
        "L4",
        "L6",
    ]
    assert [str(x).lstrip("'") for x in (b.get("join_values") or {}).get("結合", [])] == [
        "J1",
        "J3",
        "J4",
        "J6",
    ]


def test_skip_primary_match_token_not_blank(tmp_path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "S"
    ws["A1"] = "A"
    ws["A2"] = "-"
    ws["A3"] = "B"
    fp = tmp_path / "skip_token.xlsx"
    wb.save(fp)
    wb.close()
    item = {
        "name": "PK",
        "sources": [
            {
                "type": "cell",
                "sheet_name": "S",
                "cell_ref": "A1",
                "row_offset": 1,
                "col_offset": 0,
                "repeat_direction": "vertical",
                "repeat_until_empty": False,
                "repeat_max": 3,
                "skip_empty_primary": True,
                "skip_primary_match": "-",
            }
        ],
    }
    with xlsx_workbook_scope():
        b = extract_item_bundle(str(fp), item, item_id="i1")
    prim = [str(x).lstrip("'") for x in (b["primary_values"] or [])]
    assert prim == ["A", "B"]


def test_until_last_keeps_middle_blank_without_skip(tmp_path: Path) -> None:
    fp = _make_xlsx(tmp_path)
    item = {
        "name": "PK",
        "sources": [
            {
                "type": "cell",
                "sheet_name": "S",
                "cell_ref": "A1",
                "row_offset": 1,
                "col_offset": 0,
                "repeat_direction": "vertical",
                "repeat_until_empty": False,
                "repeat_until_last": True,
                "repeat_max": None,
                "skip_empty_primary": False,
                "ui_scenario_source_v1": {
                    "link_defs": [
                        {
                            "item": "連携",
                            "mode": "セル座標",
                            "cell": "B1",
                            "row": 1,
                            "col": 0,
                        }
                    ],
                },
            }
        ],
    }
    with xlsx_workbook_scope():
        b = extract_item_bundle(str(fp), item, item_id="i1")
    prim = b["primary_values"] or []
    assert len(prim) == 3
    assert str(prim[0]).lstrip("'") == "A"
    assert prim[1] is None or str(prim[1]).strip() == ""
    assert str(prim[2]).lstrip("'") == "B"
    links = (b.get("link_values") or {}).get("連携") or []
    assert [str(x).lstrip("'") for x in links] == ["L1", "L2", "L3"]


def test_until_empty_skip_token_keeps_blank_as_stop(tmp_path: Path) -> None:
    """空白までは空欄で停止。スキップ文字のみ落とす（空欄トークンは無効）。"""
    wb = Workbook()
    ws = wb.active
    ws.title = "S"
    ws["A1"] = "A"
    ws["A2"] = "-"
    ws["A3"] = "B"
    ws["A4"] = None
    ws["A5"] = "C"
    fp = tmp_path / "until_empty_skip.xlsx"
    wb.save(fp)
    wb.close()
    item = {
        "name": "PK",
        "sources": [
            {
                "type": "cell",
                "sheet_name": "S",
                "cell_ref": "A1",
                "row_offset": 1,
                "col_offset": 0,
                "repeat_direction": "vertical",
                "repeat_until_empty": True,
                "repeat_until_last": False,
                "skip_empty_primary": True,
                "skip_primary_match": ",-",  # 空欄トークンは空白まででは無効
            }
        ],
    }
    with xlsx_workbook_scope():
        b = extract_item_bundle(str(fp), item, item_id="i1")
    prim = [str(x).lstrip("'") for x in (b["primary_values"] or [])]
    assert prim == ["A", "B"]


def test_align_skips_copy_when_same_length_and_iters() -> None:
    from svc.svc_data_agg_extract import _align_one_rule_series

    vals = ["a", "b", "c"]
    ctxs = [{"file_path": "f", "iter_index": i} for i in range(3)]
    nv, nc = _align_one_rule_series(vals, ctxs, 3, file_path="f")
    assert nv is vals
    assert nc is ctxs


def test_align_places_short_series_on_primary_iters() -> None:
    from svc.svc_data_agg_extract import _align_one_rule_series

    vals = ["card0", "card1"]
    ctxs = [
        {"file_path": "f", "iter_index": 1},
        {"file_path": "f", "iter_index": 2},
    ]
    nv, nc = _align_one_rule_series(vals, ctxs, 3, file_path="f")
    assert nv is not vals
    assert ["" if x is None else str(x) for x in nv] == ["", "card0", "card1"]
    assert [c.get("iter_index") for c in nc] == [0, 1, 2]


def test_scenario_debug_skip_n1_hides_row(tmp_path: Path) -> None:
    """N件=1 でスキップ対象なら、シナリオデバッグは空欄行を出さず主値 0 件。"""
    from svc.svc_data_agg_debug_run import scenario_debug_phase_result

    wb = Workbook()
    ws = wb.active
    ws.title = "S"
    ws["A1"] = None
    fp = tmp_path / "skip_n1.xlsx"
    wb.save(fp)
    wb.close()

    item = {
        "id": "i1",
        "name": "PK",
        "sources": [
            {
                "type": "cell",
                "sheet_name": "S",
                "cell_ref": "A1",
                "row_offset": 1,
                "col_offset": 0,
                "repeat_direction": "vertical",
                "repeat_until_empty": False,
                "repeat_max": 1,
                "skip_empty_primary": True,
                "ui_scenario_source_v1": {
                    "ext_checked": [".xlsx"],
                    "link_defs": [],
                    "join_defs": [],
                },
            }
        ],
    }
    paths = [str(fp)]
    cache: dict = {}
    summary, cols, _e, _t = scenario_debug_phase_result(item, paths, 2, 200, cache)
    assert summary[2] == "0"
    assert cols == ["（主値 0 件）"]
    assert (cache.get(str(fp)) or {}).get("primary_values") == []


def test_scenario_debug_phase2_primary_only_then_prefetch_links(tmp_path: Path) -> None:
    """フェーズ2は主キーのみ。追記後のフェーズ3はキャッシュを再利用する。"""
    from svc.svc_data_agg_debug_run import (
        fill_scenario_link_join_after_primary,
        scenario_debug_phase_result,
    )

    fp = _make_xlsx(tmp_path)
    item = _item(skip_empty=True)
    item["id"] = "i1"
    item["sources"][0]["ui_scenario_source_v1"]["ext_checked"] = [".xlsx"]
    paths = [str(fp)]
    cache: dict = {}
    _s2, cols2, _e2, _t2 = scenario_debug_phase_result(item, paths, 2, 200, cache)
    b2 = cache[str(fp)]
    assert b2.get("_dbg_primary_only") is True
    assert b2.get("_dbg_full_extract") is not True
    assert not (b2.get("link_values") or {}).get("連携")
    assert [_body_dbg(x) for x in cols2] == ["A", "B"]
    fill_scenario_link_join_after_primary(item, paths, cache, "i1")
    b3pre = cache[str(fp)]
    assert b3pre.get("_dbg_full_extract") is True
    assert [_txt_skip(x) for x in (b3pre.get("link_values") or {}).get("連携") or []] == [
        "L1",
        "L3",
    ]
    id3 = id(b3pre)
    _s3, cols3, _e3, _t3 = scenario_debug_phase_result(item, paths, 3, 200, cache)
    assert cache[str(fp)] is b3pre or id(cache[str(fp)]) == id3
    assert [_body_dbg(x) for x in cols3] == ["L1", "L3"]


def test_scenario_debug_skipped_file_hides_link(tmp_path: Path) -> None:
    """主キーが全スキップのファイルは、連携フェーズにも値を出さない。"""
    from svc.svc_data_agg_debug_run import scenario_debug_phase_result

    fp1 = tmp_path / "skip.xlsx"
    wb1 = Workbook()
    ws1 = wb1.active
    ws1.title = "S"
    ws1["A1"] = None
    ws1["B1"] = "HIDE"
    wb1.save(fp1)
    wb1.close()
    fp2 = tmp_path / "keep.xlsx"
    wb2 = Workbook()
    ws2 = wb2.active
    ws2.title = "S"
    ws2["A1"] = "P"
    ws2["B1"] = "Lkeep"
    wb2.save(fp2)
    wb2.close()
    item = {
        "id": "i1",
        "name": "PK",
        "sources": [
            {
                "type": "cell",
                "sheet_name": "S",
                "cell_ref": "A1",
                "row_offset": 1,
                "col_offset": 0,
                "repeat_direction": "vertical",
                "repeat_until_empty": False,
                "repeat_max": 1,
                "skip_empty_primary": True,
                "ui_scenario_source_v1": {
                    "ext_checked": [".xlsx"],
                    "link_defs": [
                        {
                            "item": "連携",
                            "mode": "セル座標",
                            "cell": "B1",
                            "row": 1,
                            "col": 0,
                        }
                    ],
                },
            }
        ],
    }
    paths = [str(fp1), str(fp2)]
    cache: dict = {}
    _s2, cols2, _e2, _t2 = scenario_debug_phase_result(item, paths, 2, 200, cache)
    _s3, cols3, _e3, _t3 = scenario_debug_phase_result(item, paths, 3, 200, cache)
    assert [_body_dbg(x) for x in cols2] == ["P"]
    assert [_body_dbg(x) for x in cols3] == ["Lkeep"]


def test_flatten_hides_link_when_primary_skipped() -> None:
    """主キー 0 件のファイルは連携を出さない。途中スキップ後の長さは主キーに合わせる。"""
    from svc.svc_data_agg_debug_run import _flatten_map_values_by_defs

    cache = {
        "empty.xlsx": {
            "primary_values": [],
            "link_values": {"連携": ["HIDE1", "HIDE2"]},
        },
        "kept.xlsx": {
            "primary_values": ["A", "B"],
            "link_values": {"連携": ["L1", "L3", "EXTRA"]},
        },
    }
    cols = _flatten_map_values_by_defs(
        cache, "link_values", [{"item": "連携"}], 50, "PK"
    )
    assert [_body_dbg(x) for x in cols] == ["L1", "L3"]


def test_scenario_debug_stops_at_display_row_cap(tmp_path: Path) -> None:
    """表示上限に達したら以降のファイルを開かない。"""
    from svc.svc_data_agg_debug_run import scenario_debug_phase_result

    paths: list[str] = []
    for i in range(4):
        fp = tmp_path / ("f%d.xlsx" % i)
        wb = Workbook()
        ws = wb.active
        ws.title = "S"
        ws["A1"] = "P%d_1" % i
        ws["A2"] = "P%d_2" % i
        wb.save(fp)
        wb.close()
        paths.append(str(fp))
    item = {
        "id": "i1",
        "name": "PK",
        "sources": [
            {
                "type": "cell",
                "sheet_name": "S",
                "cell_ref": "A1",
                "row_offset": 1,
                "col_offset": 0,
                "repeat_direction": "vertical",
                "repeat_until_empty": False,
                "repeat_max": 2,
                "skip_empty_primary": True,
                "ui_scenario_source_v1": {
                    "ext_checked": [".xlsx"],
                    "link_defs": [],
                },
            }
        ],
    }
    cache: dict = {}
    _s, cols, _e, _t = scenario_debug_phase_result(item, paths, 2, 3, cache)
    bodies = [_body_dbg(x) for x in cols if not str(x).startswith("…")]
    assert bodies == ["P0_1", "P0_2", "P1_1"]
    assert paths[0] in cache
    assert paths[1] in cache
    assert paths[2] not in cache
    assert paths[3] not in cache
    assert any(str(x).startswith("…") for x in cols)


def test_append_rule_series_picks_by_rule_iter_holes(tmp_path: Path) -> None:
    """rule_iter が 0,2 でも列一括読取で L1/L3 を拾う。"""
    from svc.svc_data_agg_extract import _append_rule_series_to_bundle, xlsx_workbook_scope

    fp = _make_xlsx(tmp_path)
    src = {
        "type": "cell",
        "sheet_name": "S",
        "cell_ref": "A1",
    }
    rule = {"item": "連携", "mode": "セル座標", "cell": "B1", "row": 1, "col": 0}
    bundle: dict = {"link_values": {}, "link_contexts": {}}
    ictx = [
        {"file_path": str(fp), "iter_index": 0, "rule_iter_index": 0, "base_cell": "A1"},
        {"file_path": str(fp), "iter_index": 1, "rule_iter_index": 2, "base_cell": "A3"},
    ]
    with xlsx_workbook_scope():
        _append_rule_series_to_bundle(
            bundle=bundle,
            values_key="link_values",
            contexts_key="link_contexts",
            target="連携",
            file_path=str(fp),
            src=src,
            rule=rule,
            iter_contexts=ictx,
            n_src=2,
        )
    assert [_txt_skip(x) for x in bundle["link_values"]["連携"]] == ["L1", "L3"]


def _txt_skip(v: object) -> str:
    if v is None:
        return ""
    return str(v).lstrip("'")


def _body_dbg(v: object) -> str:
    s = str(v).lstrip("'")
    if "] " in s:
        s = s.split("] ", 1)[-1].lstrip("'")
    return s


def test_batch_all_skipped_primary_emits_no_blank_row(tmp_path: Path) -> None:
    """
    主キーが全スキップで primary_values=[] のとき、本番一括は空行を出さない。
    （旧: `or [None]` で余白1行が残るバグの回帰）
    """
    from svc.svc_data_agg import compute_batch_table_rows

    fp_empty = tmp_path / "ODN-623_empty.xlsx"
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "チェックシート"
    ws["D8"] = None
    wb.save(fp_empty)
    wb.close()

    fp_ok = tmp_path / "ODN-623_ok.xlsx"
    wb2 = Workbook()
    ws2 = wb2.active
    assert ws2 is not None
    ws2.title = "チェックシート"
    ws2["D8"] = "PT123"
    wb2.save(fp_ok)
    wb2.close()

    data = {
        "id": "odn623_blank_singleton",
        "items": [
            {
                "id": "item_eq",
                "name": "機器番号",
                "write_mode": "append",
                "sources": [
                    {
                        "type": "cell",
                        "sheet_name": "チェックシート",
                        "cell_ref": "D8",
                        "row_offset": 0,
                        "col_offset": 0,
                        "repeat_direction": "vertical",
                        "repeat_until_empty": False,
                        "repeat_max": 1,
                        "skip_empty_primary": True,
                        "skip_primary_match": ",-,ー",
                        "ui_scenario_source_v1": {
                            "file_pattern": "ODN-623",
                            "file_name_rule": "含む",
                            "sheet_rule": "含む",
                            "link_defs": [
                                {
                                    "cell": "FIXED",
                                    "mode": "固定値",
                                    "item": "製品コード",
                                    "row": 0,
                                    "col": 0,
                                }
                            ],
                            "join_defs": [],
                        },
                    }
                ],
            },
            {
                "id": "item_pc",
                "name": "製品コード",
                "write_mode": "fill_in",
                "sources": [],
            },
        ],
        "match_keys": [],
        "excel_options": {"output_target": "active_sheet", "write_mode": "append"},
    }
    headers, rows, _, _ = compute_batch_table_rows(
        data,
        [str(fp_empty), str(fp_ok)],
        max_table_rows=50,
        probe_caller="excel_batch_submit",
    )
    assert headers[:2] == ["機器番号", "製品コード"]
    assert len(rows) == 1
    eq = str(rows[0][0]).lstrip("'") if rows[0][0] is not None else ""
    assert eq == "PT123"
    # 全列空の余白行が混ざらない
    assert not any(
        all((c is None or str(c).strip() in ("", "'")) for c in r) for r in rows
    )


def test_batch_no_matching_sheet_emits_no_blank_row(tmp_path: Path) -> None:
    """ファイル名は条件に合うがシート不一致 → 空バンドルでも余白行を出さない。"""
    from svc.svc_data_agg import compute_batch_table_rows

    fp = tmp_path / "【other】no_sheet.xlsx"
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "別シート"
    ws["A1"] = "x"
    wb.save(fp)
    wb.close()

    data = {
        "id": "odn623_no_sheet",
        "items": [
            {
                "id": "item_eq",
                "name": "機器番号",
                "write_mode": "append",
                "sources": [
                    {
                        "type": "cell",
                        "sheet_name": "チェックシート",
                        "cell_ref": "D8",
                        "row_offset": 0,
                        "col_offset": 0,
                        "repeat_direction": "vertical",
                        "repeat_until_empty": False,
                        "repeat_max": 1,
                        "skip_empty_primary": True,
                        "skip_primary_match": ",-,ー",
                        "ui_scenario_source_v1": {
                            "file_pattern": "【",
                            "file_name_rule": "含む",
                            "sheet_rule": "含む",
                            "link_defs": [],
                            "join_defs": [],
                        },
                    }
                ],
            }
        ],
        "match_keys": [],
        "excel_options": {"output_target": "active_sheet", "write_mode": "append"},
    }
    _headers, rows, _, _ = compute_batch_table_rows(
        data,
        [str(fp)],
        max_table_rows=50,
        probe_caller="excel_batch_submit",
    )
    assert rows == []
