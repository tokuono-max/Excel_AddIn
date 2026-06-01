# -*- coding: utf-8 -*-
"""本番一括の安定性回帰（速度・UI 表示は変更しない）。"""
# pyright: reportPrivateUsage=false
from __future__ import annotations

import copy
import sys
import time
from pathlib import Path
from typing import Any

import pytest
from openpyxl import Workbook

_root = Path(__file__).resolve().parents[1]
if str(_root) not in sys.path:
    sys.path.insert(0, str(_root))

from svc.data_agg_cancel import DataAggCancelled  # noqa: E402
from svc.svc_data_agg import (  # noqa: E402
    _apply_join_key_search_link_write,
    _apply_join_key_search_write,
    _batch_file_extract_and_merge,
    _log_compute_batch_result_invariants,
    compute_batch_table_rows,
)


def _active_worksheet(wb: Any) -> Any:
    ws = wb.active
    assert ws is not None
    return ws


def _join_link_host_item() -> dict[str, Any]:
    return {
        "sources": [
            {
                "type": "cell",
                "sheet_name": "S",
                "cell_ref": "B1",
                "ui_scenario_source_v1": {
                    "join_defs": [{"item": "機器番号", "cell": "A1", "row": 0, "col": 0}],
                    "link_defs": [
                        {
                            "item": "MAC",
                            "cell": "",
                            "mode": "固定値",
                            "row": 0,
                            "col": 0,
                        }
                    ],
                },
            }
        ],
    }


def test_join_write_unified_matches_two_pass() -> None:
    """write+link 1 パス（header_set）と 2 パス呼び出しの pool 結果が一致する。"""
    pool_base = [
        {"機器番号": "PT1", "MAC": "keep", "M": "", "__file_path": "f.xlsx", "__iter_index": 0},
        {"機器番号": "PT1", "MAC": "keep", "M": "", "__file_path": "f.xlsx", "__iter_index": 1},
        {"機器番号": "PT2", "MAC": "other", "M": "", "__file_path": "f.xlsx", "__iter_index": 0},
    ]
    item = _join_link_host_item()
    bundle = {
        "primary_values": ["HOST_VAL"],
        "join_values": {"機器番号": ["PT1"]},
        "link_values": {"MAC": [""]},
    }
    hdrs = {"機器番号", "MAC", "M"}

    pool_two = copy.deepcopy(pool_base)
    _apply_join_key_search_write(pool_two, item, "M", bundle, "overwrite")
    _apply_join_key_search_link_write(pool_two, item, bundle, "overwrite", hdrs)

    pool_one = copy.deepcopy(pool_base)
    _apply_join_key_search_write(
        pool_one, item, "M", bundle, "overwrite", header_set=hdrs
    )

    assert pool_one == pool_two


def _cross_join_mini_scenario(tmp_path: Path) -> tuple[dict[str, Any], list[str]]:
    anchor = tmp_path / "光特性履歴_test.xlsx"
    wb = Workbook()
    ws = _active_worksheet(wb)
    ws.title = "ﾃﾞｰﾀ"
    ws["C7"] = "DEV1"
    ws["M7"] = "MAC-A"
    wb.save(anchor)

    join_f = tmp_path / "紐づけ履歴_test.xlsx"
    wb2 = Workbook()
    ws2 = _active_worksheet(wb2)
    ws2.title = "紐付け履歴"
    ws2["C5"] = "PT-001"
    ws2["P5"] = "MAC-A"
    ws2["J5"] = "SEQ-1"
    wb2.save(join_f)

    data: dict[str, Any] = {
        "id": "cross_join_stability",
        "items": [
            {
                "id": "i_dev",
                "name": "機器番号",
                "write_mode": "append",
                "sources": [
                    {
                        "type": "cell",
                        "sheet_name": "ﾃﾞｰﾀ",
                        "cell_ref": "C7",
                        "ui_scenario_source_v1": {
                            "file_pattern": "光特性",
                            "link_defs": [
                                {
                                    "item": "MACアドレス",
                                    "cell": "M7",
                                    "mode": "セル座標",
                                    "row": 0,
                                    "col": 0,
                                }
                            ],
                        },
                        "repeat_direction": "vertical",
                        "repeat_max": 1,
                    }
                ],
            },
            {"id": "i_mac", "name": "MACアドレス", "write_mode": "fill_in", "sources": []},
            {
                "id": "i_pt",
                "name": "PT番号",
                "write_mode": "overwrite",
                "sources": [
                    {
                        "type": "cell",
                        "sheet_name": "紐付け履歴",
                        "cell_ref": "C5",
                        "ui_scenario_source_v1": {
                            "file_pattern": "紐づけ",
                            "join_defs": [
                                {"item": "MACアドレス", "cell": "P5", "row": 0, "col": 0}
                            ],
                            "link_defs": [
                                {
                                    "item": "製番",
                                    "cell": "J5",
                                    "mode": "セル座標",
                                    "row": 0,
                                    "col": 0,
                                }
                            ],
                        },
                        "repeat_direction": "vertical",
                        "repeat_max": 1,
                    }
                ],
            },
            {"id": "i_seq", "name": "製番", "write_mode": "fill_in", "sources": []},
        ],
        "match_keys": [],
    }
    return data, [str(anchor), str(join_f)]


def test_compute_batch_parallel_matches_sequential(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    """並列 extract と逐次 extract で最終表が一致する。"""
    data, paths = _cross_join_mini_scenario(tmp_path)
    monkeypatch.setenv("DATA_AGG_FILE_PARALLEL_WORKERS", "0")
    h_seq, rows_seq, _, je_seq = compute_batch_table_rows(
        data, paths, max_primary_rows=10, max_table_rows=10
    )
    monkeypatch.setenv("DATA_AGG_FILE_PARALLEL_WORKERS", "2")
    h_par, rows_par, _, je_par = compute_batch_table_rows(
        data, paths, max_primary_rows=10, max_table_rows=10
    )
    assert h_seq == h_par
    assert rows_seq == rows_par
    assert je_seq == je_par


def test_file_pattern_skip_join_bundle_index(tmp_path: Path) -> None:
    """file_pattern 不一致項目をスキップしても join 用 bundle 位置がずれない。"""
    data, paths = _cross_join_mini_scenario(tmp_path)
    data["items"].insert(
        2,
        {
            "id": "i_noise",
            "name": "ノイズ列",
            "write_mode": "fill_in",
            "sources": [
                {
                    "type": "cell",
                    "sheet_name": "S",
                    "cell_ref": "Z1",
                    "ui_scenario_source_v1": {"file_pattern": "存在しないパターン"},
                }
            ],
        },
    )
    headers, rows, _, _ = compute_batch_table_rows(
        data, paths, max_primary_rows=10, max_table_rows=10
    )
    ix_pt = headers.index("PT番号")
    assert len(rows) == 1
    assert rows[0][ix_pt] == "PT-001"


def test_cancel_during_parallel_extract(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    """並列 extract 中の協調キャンセルが DataAggCancelled で終了する。"""
    data, paths = _cross_join_mini_scenario(tmp_path)
    paths = paths * 3
    monkeypatch.setenv("DATA_AGG_FILE_PARALLEL_WORKERS", "2")

    orig = _batch_file_extract_and_merge
    calls = [0]

    def _slow_extract(*args: Any, **kwargs: Any) -> Any:
        cc = kwargs.get("cancel_check")
        if cc is not None:
            cc(force=True)
        time.sleep(0.015)
        return orig(*args, **kwargs)

    monkeypatch.setattr(
        "svc.svc_data_agg._batch_file_extract_and_merge",
        _slow_extract,
    )

    n_poll = [0]

    def _cancel_check(*, force: bool = False) -> None:
        n_poll[0] += 1
        if n_poll[0] >= 4:
            raise DataAggCancelled()

    with pytest.raises(DataAggCancelled):
        compute_batch_table_rows(
            data,
            paths,
            max_primary_rows=10,
            max_table_rows=10,
            cancel_check=_cancel_check,
        )


def test_log_compute_batch_invariants_warns_empty_table(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    """pool のみ存在し table が空のとき診断 warning を出す（結果は変えない）。"""
    warnings: list[str] = []

    def _capture(msg: str, *args: Any) -> None:
        warnings.append(msg % args if args else msg)

    monkeypatch.setattr("svc.svc_data_agg._agg_diag.warning", _capture)
    _log_compute_batch_result_invariants(
        scenario_id="t",
        n_files=2,
        table_rows=[],
        join_search_global_pool=[{"x": 1}],
        use_join_search_merge=True,
        preview_master_mode=False,
        max_table_rows=None,
        parallel_expected=0,
        parallel_got=0,
    )
    assert any("invariant table_empty pool_nonempty" in w for w in warnings)
