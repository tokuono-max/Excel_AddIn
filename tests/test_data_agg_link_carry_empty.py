# -*- coding: utf-8 -*-
"""連携キー carry_empty（空欄は前回値を保持）の抽出・要約。"""
from __future__ import annotations

import sys
from pathlib import Path

_root = Path(__file__).resolve().parents[1]
if str(_root) not in sys.path:
    sys.path.insert(0, str(_root))

from openpyxl import Workbook  # noqa: E402

from svc.data_agg_cell_coordinate_summary import (  # noqa: E402
    _rule_def_extra_suffix,
    cell_coordinate_setting_lines,
)
from svc.data_agg_scenario_export import _fmt_link_line  # noqa: E402
from svc.svc_data_agg_extract import (  # noqa: E402
    _apply_carry_empty_link_values,
    _is_blank_link_carry_value,
    extract_item_bundle,
    link_def_wants_carry_empty,
    xlsx_workbook_scope,
)


def _txt(v: object) -> str:
    return str(v).lstrip("'") if v is not None else ""


def _link_item(
    *,
    carry: bool,
    skip_empty: bool = False,
    skip_carry_seed: bool = False,
    sheet_rule: str | None = None,
) -> dict:
    ui: dict = {
        "link_defs": [
            {
                "item": "連携",
                "mode": "セル座標",
                "cell": "B1",
                "row": 1,
                "col": 0,
                "carry_empty": carry,
            }
        ],
        "join_defs": [
            {
                "item": "結合",
                "mode": "セル座標",
                "cell": "C1",
                "row": 1,
                "col": 0,
            }
        ],
    }
    if sheet_rule:
        ui["sheet_rule"] = sheet_rule
    src: dict = {
        "type": "cell",
        "sheet_name": "S" if sheet_rule is None else "R_",
        "cell_ref": "A1",
        "row_offset": 1,
        "col_offset": 0,
        "repeat_direction": "vertical",
        "repeat_until_empty": False,
        "repeat_max": 5 if sheet_rule is None else 3,
        "skip_empty_primary": skip_empty,
        "skip_carry_seed": bool(skip_empty and skip_carry_seed),
        "ui_scenario_source_v1": ui,
    }
    return {"name": "PK", "sources": [src]}


def test_link_def_wants_carry_empty_truthy() -> None:
    assert link_def_wants_carry_empty({"carry_empty": True}) is True
    assert link_def_wants_carry_empty({"carry_empty": 1}) is True
    assert link_def_wants_carry_empty({"carry_empty": "true"}) is True
    assert link_def_wants_carry_empty({"carry_empty": False}) is False
    assert link_def_wants_carry_empty({}) is False
    assert link_def_wants_carry_empty(None) is False


def test_source_wants_skip_carry_seed() -> None:
    from svc.svc_data_agg_extract import source_wants_skip_carry_seed

    assert source_wants_skip_carry_seed({"skip_empty_primary": True, "skip_carry_seed": True})
    assert not source_wants_skip_carry_seed({"skip_empty_primary": False, "skip_carry_seed": True})
    assert not source_wants_skip_carry_seed({"skip_empty_primary": True, "skip_carry_seed": False})
    assert not source_wants_skip_carry_seed({})
    assert not source_wants_skip_carry_seed(None)


def test_is_blank_link_carry_value() -> None:
    assert _is_blank_link_carry_value(None) is True
    assert _is_blank_link_carry_value("") is True
    assert _is_blank_link_carry_value("  ") is True
    assert _is_blank_link_carry_value("'") is True
    assert _is_blank_link_carry_value("'  ") is True
    assert _is_blank_link_carry_value("A") is False
    assert _is_blank_link_carry_value("'A") is False


def test_apply_carry_empty_resets_per_source_span() -> None:
    """ソース区間ごとに前回値をリセットする（シートまたぎ相当）。"""
    bundle = {
        "primary_values": ["p"] * 6,
        "link_values": {"連携": ["A", "", "", "B", "", ""]},
        "_cell_source_spans": {0: (0, 3), 1: (3, 3)},
    }
    src = {
        "type": "cell",
        "ui_scenario_source_v1": {
            "link_defs": [{"item": "連携", "carry_empty": True, "mode": "セル座標"}]
        },
    }
    _apply_carry_empty_link_values(bundle, [src, dict(src)])
    assert bundle["link_values"]["連携"] == ["A", "A", "A", "B", "B", "B"]


def test_apply_carry_empty_leading_blank_stays() -> None:
    bundle = {
        "primary_values": ["p"] * 4,
        "link_values": {"連携": ["", "", "A", ""]},
        "_cell_source_spans": {0: (0, 4)},
    }
    src = {
        "type": "cell",
        "ui_scenario_source_v1": {
            "link_defs": [{"item": "連携", "carry_empty": True}]
        },
    }
    _apply_carry_empty_link_values(bundle, [src])
    assert bundle["link_values"]["連携"] == ["", "", "A", "A"]


def test_same_sheet_fills_blank_with_previous(tmp_path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "S"
    prim = ["P1", "P2", "P3", "P4", "P5"]
    links = ["A", None, "", "B", None]
    joins = ["J1", None, "", "J4", None]
    for i, (p, ln, jn) in enumerate(zip(prim, links, joins), start=1):
        ws[f"A{i}"] = p
        ws[f"B{i}"] = ln
        ws[f"C{i}"] = jn
    fp = tmp_path / "carry.xlsx"
    wb.save(fp)
    wb.close()
    with xlsx_workbook_scope():
        b = extract_item_bundle(str(fp), _link_item(carry=True), item_id="i1")
    assert [_txt(x) for x in (b["link_values"]["連携"] or [])] == [
        "A",
        "A",
        "A",
        "B",
        "B",
    ]
    assert [_txt(x) for x in (b["join_values"]["結合"] or [])] == [
        "J1",
        "",
        "",
        "J4",
        "",
    ]


def test_carry_empty_off_keeps_blanks(tmp_path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "S"
    for i, (p, ln) in enumerate([("P1", "A"), ("P2", None), ("P3", "")], start=1):
        ws[f"A{i}"] = p
        ws[f"B{i}"] = ln
        ws[f"C{i}"] = "J"
    fp = tmp_path / "carry_off.xlsx"
    wb.save(fp)
    wb.close()
    item = _link_item(carry=False)
    item["sources"][0]["repeat_max"] = 3
    with xlsx_workbook_scope():
        b = extract_item_bundle(str(fp), item, item_id="i1")
    assert [_txt(x) for x in (b["link_values"]["連携"] or [])] == ["A", "", ""]


def test_multi_sheet_does_not_carry_across_sheets(tmp_path: Path) -> None:
    fp = tmp_path / "carry_multi.xlsx"
    wb = Workbook()
    wb.active.title = "Other"
    s1 = wb.create_sheet("R_1")
    s1["A1"] = "P1"
    s1["A2"] = "P2"
    s1["A3"] = "P3"
    s1["B1"] = "A"
    s1["B2"] = None
    s1["B3"] = ""
    s2 = wb.create_sheet("R_2")
    s2["A1"] = "Q1"
    s2["A2"] = "Q2"
    s2["A3"] = "Q3"
    s2["B1"] = None
    s2["B2"] = "X"
    s2["B3"] = ""
    wb.save(fp)
    wb.close()
    with xlsx_workbook_scope():
        b = extract_item_bundle(
            str(fp), _link_item(carry=True, sheet_rule="含む"), item_id="i1"
        )
    assert [_txt(x) for x in (b["link_values"]["連携"] or [])] == [
        "A",
        "A",
        "A",
        "",
        "X",
        "X",
    ]


def test_skip_empty_primary_then_carry_does_not_use_dropped_seed(
    tmp_path: Path,
) -> None:
    """主キースキップ後に前埋め。落ちた行の連携値は種にしない（skip_carry_seed OFF）。"""
    wb = Workbook()
    ws = wb.active
    ws.title = "S"
    # 主: P1, 空(skip), P3  / 連携: L1, L2(落ちる), 空
    ws["A1"] = "P1"
    ws["A2"] = None
    ws["A3"] = "P3"
    ws["B1"] = "L1"
    ws["B2"] = "L2"
    ws["B3"] = None
    ws["C1"] = "J1"
    ws["C2"] = "J2"
    ws["C3"] = "J3"
    fp = tmp_path / "carry_skip.xlsx"
    wb.save(fp)
    wb.close()
    item = _link_item(carry=True, skip_empty=True)
    item["sources"][0]["repeat_max"] = 3
    with xlsx_workbook_scope():
        b = extract_item_bundle(str(fp), item, item_id="i1")
    assert [_txt(x) for x in (b["primary_values"] or [])] == ["P1", "P3"]
    assert [_txt(x) for x in (b["link_values"]["連携"] or [])] == ["L1", "L1"]


def test_skip_carry_seed_on_uses_dropped_row_link(tmp_path: Path) -> None:
    """skip_carry_seed ON: スキップ行の非空連携値を次行の前置種にする。"""
    wb = Workbook()
    ws = wb.active
    ws.title = "S"
    ws["A1"] = "P1"
    ws["A2"] = None
    ws["A3"] = "P3"
    ws["B1"] = "L1"
    ws["B2"] = "L2"
    ws["B3"] = None
    ws["C1"] = "J1"
    ws["C2"] = "J2"
    ws["C3"] = "J3"
    fp = tmp_path / "carry_skip_seed_on.xlsx"
    wb.save(fp)
    wb.close()
    item = _link_item(carry=True, skip_empty=True, skip_carry_seed=True)
    item["sources"][0]["repeat_max"] = 3
    with xlsx_workbook_scope():
        b = extract_item_bundle(str(fp), item, item_id="i1")
    assert [_txt(x) for x in (b["primary_values"] or [])] == ["P1", "P3"]
    assert [_txt(x) for x in (b["link_values"]["連携"] or [])] == ["L1", "L2"]


def test_skip_carry_seed_on_blank_skip_keeps_previous_last(tmp_path: Path) -> None:
    """skip_carry_seed ON でもスキップ行の連携が空なら last を上書きしない。"""
    wb = Workbook()
    ws = wb.active
    ws.title = "S"
    ws["A1"] = "P1"
    ws["A2"] = None
    ws["A3"] = "P3"
    ws["B1"] = "L1"
    ws["B2"] = None
    ws["B3"] = None
    ws["C1"] = "J1"
    ws["C2"] = "J2"
    ws["C3"] = "J3"
    fp = tmp_path / "carry_skip_seed_blank.xlsx"
    wb.save(fp)
    wb.close()
    item = _link_item(carry=True, skip_empty=True, skip_carry_seed=True)
    item["sources"][0]["repeat_max"] = 3
    with xlsx_workbook_scope():
        b = extract_item_bundle(str(fp), item, item_id="i1")
    assert [_txt(x) for x in (b["primary_values"] or [])] == ["P1", "P3"]
    assert [_txt(x) for x in (b["link_values"]["連携"] or [])] == ["L1", "L1"]


def test_scenario_debug_skip_carry_seed_on(tmp_path: Path) -> None:
    """シナリオ段階実行でも skip_carry_seed ON ならスキップ行の連携を種にする。"""
    from svc.svc_data_agg_debug_run import scenario_debug_phase_result

    wb = Workbook()
    ws = wb.active
    ws.title = "S"
    ws["A1"] = "P1"
    ws["A2"] = None
    ws["A3"] = "P3"
    ws["B1"] = "L1"
    ws["B2"] = "SEED"
    ws["B3"] = None
    fp = tmp_path / "carry_skip_seed_dbg.xlsx"
    wb.save(fp)
    wb.close()
    item = _link_item(carry=True, skip_empty=True, skip_carry_seed=True)
    item["id"] = "i1"
    item["sources"][0]["repeat_max"] = 3
    item["sources"][0]["ui_scenario_source_v1"]["ext_checked"] = [".xlsx"]
    cache: dict = {}
    paths = [str(fp)]
    _s2, cols2, _e2, _t2 = scenario_debug_phase_result(item, paths, 2, 200, cache)
    _s3, cols3, _e3, _t3 = scenario_debug_phase_result(item, paths, 3, 200, cache)

    def _body(v: object) -> str:
        s = str(v).lstrip("'")
        if "] " in s:
            s = s.split("] ", 1)[-1].lstrip("'")
        return s

    assert [_body(x) for x in cols2] == ["P1", "P3"]
    assert [_body(x) for x in cols3] == ["L1", "SEED"]
    b = cache[str(fp)]
    assert b.get("_dbg_full_extract") is True
    assert [_txt(x) for x in (b.get("link_values") or {}).get("連携", [])] == ["L1", "SEED"]


def test_debug_link_phase_applies_carry_per_sheet(tmp_path: Path) -> None:
    """シナリオデバッグの連携フェーズでもシート内前埋めし、シートまたぎはしない。"""
    from svc.svc_data_agg_debug_run import scenario_debug_phase_result

    fp = tmp_path / "carry_debug.xlsx"
    wb = Workbook()
    wb.active.title = "Other"
    s1 = wb.create_sheet("R_1")
    s1["A1"] = "P1"
    s1["A2"] = "P2"
    s1["B1"] = "A"
    s1["B2"] = None
    s2 = wb.create_sheet("R_2")
    s2["A1"] = "Q1"
    s2["A2"] = "Q2"
    s2["B1"] = None
    s2["B2"] = "X"
    wb.save(fp)
    wb.close()
    item = _link_item(carry=True, sheet_rule="含む")
    item["id"] = "i1"
    item["sources"][0]["repeat_max"] = 2
    cache: dict = {}
    paths = [str(fp)]
    _s2, cols2, _e2, _t2 = scenario_debug_phase_result(item, paths, 2, 200, cache)
    _s3, cols3, _e3, _t3 = scenario_debug_phase_result(item, paths, 3, 200, cache)

    def _body(v: object) -> str:
        s = str(v).lstrip("'")
        if "] " in s:
            s = s.split("] ", 1)[-1].lstrip("'")
        return s

    assert [_body(x) for x in cols2] == ["P1", "P2", "Q1", "Q2"]
    assert [_body(x) for x in cols3] == ["A", "A", "", "X"]


def test_summary_and_export_show_carry_flag() -> None:
    ld = {
        "cell": "B1",
        "row": 1,
        "col": 0,
        "item": "連携",
        "carry_empty": True,
    }
    assert "前置保持" in _rule_def_extra_suffix(ld)
    lines = cell_coordinate_setting_lines(
        {},
        {"link_defs": [ld], "join_defs": []},
        {},
    )
    assert any("前置保持" in x and "連携" in x for x in lines)
    exported = _fmt_link_line({}, 0, ld, is_join=False)
    assert "前置保持" in exported
    assert "前置保持" not in _rule_def_extra_suffix({"item": "連携"})
