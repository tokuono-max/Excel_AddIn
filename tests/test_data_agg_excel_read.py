# -*- coding: utf-8 -*-
"""Excel 抽出読取: 文字列保全と日付/小数書式の寄せ。"""
from __future__ import annotations

from datetime import datetime

from svc.data_agg_excel_read import extract_read_scalar
from svc.data_agg_value_post import postprocess_cell_primary, postprocess_link_rule_value


def test_text_zero_padding_never_reinterpreted() -> None:
    assert extract_read_scalar("00123") == "00123"
    assert extract_read_scalar("00123", "00000") == "00123"
    assert extract_read_scalar("00123", "0") == "00123"
    s = postprocess_cell_primary("00123", {"cell_checks": ["トリム"]})
    assert "00123" in s


def test_numeric_zero_pad_format_is_not_applied() -> None:
    """0 パディングは書式では付けない前提。数値+00000 でもゼロ埋めしない。"""
    v = extract_read_scalar(123, "00000")
    assert v == 123
    out = postprocess_link_rule_value(v, {"checks": ["トリム"]})
    assert "00123" not in out
    assert "123" in out


def test_general_float_keeps_short_decimal() -> None:
    v = extract_read_scalar(2020.4, "General")
    assert v == 2020.4
    s = postprocess_link_rule_value(v, {"checks": ["トリム"]})
    assert "2020.4" in s
    assert "03999" not in s


def test_fixed_decimal_format_0_00() -> None:
    assert extract_read_scalar(2020.4, "0.00") == "2020.40"


def test_datetime_to_ymd() -> None:
    assert extract_read_scalar(datetime(2022, 5, 27, 0, 0, 0)) == "2022/05/27"
    assert extract_read_scalar(datetime(2022, 5, 27, 15, 30)) == "2022/05/27 15:30"


def test_date_number_format_converts_serial() -> None:
    s = extract_read_scalar(44708.0, "yyyy/mm/dd")
    assert s == "2022/05/27"


def test_plain_number_not_treated_as_date() -> None:
    assert extract_read_scalar(44708.0, "General") == 44708.0
    assert extract_read_scalar(123, None) == 123


def test_xlsx_text_padding_and_general_float(tmp_path) -> None:
    from openpyxl import Workbook

    from svc.svc_data_agg_extract import extract_item_values, xlsx_workbook_scope

    p = tmp_path / "pad.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "S"
    ws["A1"] = "00123"
    ws["A2"] = 123
    ws["A2"].number_format = "00000"
    ws["A3"] = 2020.4
    wb.save(p)
    wb.close()
    item = {
        "sources": [
            {
                "type": "cell",
                "sheet_name": "S",
                "cell_ref": "A1",
                "row_offset": 1,
                "repeat_direction": "vertical",
                "repeat_until_empty": False,
                "repeat_until_last": True,
            }
        ]
    }
    with xlsx_workbook_scope():
        vals = extract_item_values(str(p), item)
    got = [str(x)[1:] if str(x).startswith("'") else str(x) for x in vals]
    assert got[0] == "00123"
    assert got[1] == "123"
    assert got[2] == "2020.4"
