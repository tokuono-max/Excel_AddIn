# -*- coding: utf-8 -*-
"""連携一括読取が row_step を無視しないことの回帰。"""
from __future__ import annotations

import csv
import sys
from pathlib import Path

_root = Path(__file__).resolve().parents[1]
if str(_root) not in sys.path:
    sys.path.insert(0, str(_root))

from openpyxl import Workbook  # noqa: E402

from svc.svc_data_agg_extract import (  # noqa: E402
    extract_item_bundle,
    xlsx_workbook_scope,
    _extract_cell_rules_series_fast_map,
)


def _strip_excel_text(v: object) -> str:
    s = "" if v is None else str(v)
    return s[1:] if s.startswith("'") else s


def _make_three_row_xlsx(tmp_path: Path, *, n_units: int = 8) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = "管理表"
    for i in range(n_units):
        r = 6 + i * 3
        ws.cell(row=r, column=3, value="WJ%03d" % i)  # C
        ws.cell(row=r, column=4, value="PT%03d" % i)  # D
        ws.cell(row=r, column=5, value="AX%03d" % i)  # E
        ws.cell(row=r, column=8, value="QR1_%03d" % i)  # H  QR_1
        ws.cell(row=r + 1, column=8, value=2020.4)  # H  QR_2
        ws.cell(row=r + 2, column=8, value="WJ%03d" % i)  # H  QR_3
    fp = tmp_path / "three_row.xlsm"
    # .xlsx で十分（OpenXML）。拡張子は openxml 判定できればよい。
    fp = tmp_path / "three_row.xlsx"
    wb.save(fp)
    wb.close()
    return fp


def _item_like_oinet() -> dict:
    return {
        "name": "製造番号",
        "sources": [
            {
                "type": "cell",
                "sheet_name": "管理表",
                "cell_ref": "C6",
                "row_offset": 3,
                "col_offset": 0,
                "repeat_direction": "vertical",
                "repeat_until_empty": False,
                "repeat_until_last": True,
                "repeat_max": None,
                "ui_scenario_source_v1": {
                    "link_defs": [
                        {"item": "PT番号", "mode": "セル座標", "cell": "D6", "row": 3, "col": 0},
                        {"item": "AX番号", "mode": "セル座標", "cell": "E6", "row": 3, "col": 0},
                        {"item": "QR_1", "mode": "セル座標", "cell": "H6", "row": 3, "col": 0},
                        {"item": "QR_2", "mode": "セル座標", "cell": "H7", "row": 3, "col": 0},
                        {"item": "QR_3", "mode": "セル座標", "cell": "H8", "row": 3, "col": 0},
                    ],
                    "join_defs": [],
                },
            }
        ],
    }


def test_link_fast_map_respects_row_step_three(tmp_path: Path) -> None:
    fp = _make_three_row_xlsx(tmp_path, n_units=8)
    with xlsx_workbook_scope():
        b = extract_item_bundle(str(fp), _item_like_oinet(), item_id="item_1")
    prim = [_strip_excel_text(x) for x in (b["primary_values"] or [])]
    assert prim == ["WJ%03d" % i for i in range(8)]
    lv = b.get("link_values") or {}
    pt = [_strip_excel_text(x) for x in (lv.get("PT番号") or [])]
    ax = [_strip_excel_text(x) for x in (lv.get("AX番号") or [])]
    q1 = [_strip_excel_text(x) for x in (lv.get("QR_1") or [])]
    q2 = [_strip_excel_text(x) for x in (lv.get("QR_2") or [])]
    q3 = [_strip_excel_text(x) for x in (lv.get("QR_3") or [])]
    assert pt == ["PT%03d" % i for i in range(8)]
    assert ax == ["AX%03d" % i for i in range(8)]
    assert q1 == ["QR1_%03d" % i for i in range(8)]
    assert q2 == ["2020.4"] * 8
    assert q3 == ["WJ%03d" % i for i in range(8)]


def test_fast_map_direct_step3_xlsx(tmp_path: Path) -> None:
    fp = _make_three_row_xlsx(tmp_path, n_units=8)
    src = {"sheet_name": "管理表", "cell_ref": "C6"}
    rules = [
        {"item": "PT番号", "mode": "セル座標", "cell": "D6", "row": 3, "col": 0},
        {"item": "AX番号", "mode": "セル座標", "cell": "E6", "row": 3, "col": 0},
        {"item": "QR_1", "mode": "セル座標", "cell": "H6", "row": 3, "col": 0},
        {"item": "QR_2", "mode": "セル座標", "cell": "H7", "row": 3, "col": 0},
        {"item": "QR_3", "mode": "セル座標", "cell": "H8", "row": 3, "col": 0},
    ]
    with xlsx_workbook_scope():
        got = _extract_cell_rules_series_fast_map(str(fp), src, rules, n_src=8)
    assert isinstance(got, dict)
    # 同一開始行 D6/E6/H6 のみ一括。H7/H8 は個別経路。
    assert 0 in got and 1 in got and 2 in got
    assert 3 not in got and 4 not in got
    pt = [_strip_excel_text(x) for x in got[0]]
    q1 = [_strip_excel_text(x) for x in got[2]]
    assert pt == ["PT%03d" % i for i in range(8)]
    assert q1 == ["QR1_%03d" % i for i in range(8)]


def test_fast_map_csv_step3(tmp_path: Path) -> None:
    fp = tmp_path / "step3.csv"
    rows = [[""] * 4 for _ in range(30)]
    for i in range(8):
        r = 5 + i * 3  # 0-based row 5 = Excel 6
        rows[r][0] = "A%d" % i
        rows[r][1] = "B%d" % i
        rows[r][2] = "C%d" % i
        rows[r + 1][0] = "skip"
    with fp.open("w", encoding="utf-8", newline="") as f:
        csv.writer(f).writerows(rows)
    src = {"sheet_name": "", "cell_ref": "A6"}
    rules = [
        {"mode": "セル座標", "cell": "A6", "row": 3, "col": 0},
        {"mode": "セル座標", "cell": "B6", "row": 3, "col": 0},
        {"mode": "セル座標", "cell": "C6", "row": 3, "col": 0},
    ]
    with xlsx_workbook_scope():
        got = _extract_cell_rules_series_fast_map(str(fp), src, rules, n_src=8)
    assert got is not None
    a = [_strip_excel_text(x) for x in got[0]]
    assert a == ["A%d" % i for i in range(8)]
    assert "skip" not in a
