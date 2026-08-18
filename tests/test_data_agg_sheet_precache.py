# -*- coding: utf-8 -*-
"""Phase B: シート事前 materialize と行列経路の反復読取。"""
from __future__ import annotations

from pathlib import Path

import pytest

from svc import svc_data_agg_extract as ex


@pytest.fixture
def vertical_repeat_xlsx(tmp_path: Path) -> Path:
    openpyxl = pytest.importorskip("openpyxl")
    p = tmp_path / "repeat.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data"
    for i in range(1, 6):
        ws.cell(row=i, column=1, value=f"v{i}")
    wb.save(p)
    wb.close()
    return p


def test_precache_collects_heavy_repeat_sheets() -> None:
    items = [
        {
            "sources": [
                {
                    "type": "cell",
                    "sheet_name": "Data",
                    "cell_ref": "A1",
                    "repeat_direction": "vertical",
                    "row_offset": 1,
                    "repeat_max": 100,
                }
            ]
        }
    ]
    sheets = ex._collect_xlsx_sheets_for_precache(items)
    assert "Data" in sheets
    light = [
        {
            "sources": [
                {
                    "type": "cell",
                    "repeat_direction": "vertical",
                    "repeat_max": 10,
                }
            ]
        }
    ]
    # セル参照シートは反復が小さくても precache 対象（複数項目の都度読取を1回化）
    assert ex._collect_xlsx_sheets_for_precache(light) == {None}


def test_vertical_repeat_via_matrix_in_scope(vertical_repeat_xlsx: Path) -> None:
    import openpyxl

    p = vertical_repeat_xlsx
    wb = openpyxl.load_workbook(p, read_only=True, data_only=True)
    try:
        ws, _ = ex._resolve_readonly_worksheet(wb, "Data")
        mat = ex._materialize_readonly_sheet_matrix(ws)
        got = ex._read_repeated_series_from_matrix(
            mat,
            base_col=0,
            base_row=0,
            row_step=1,
            col_step=0,
            limit=5,
            repeat_until_empty=False,
        )
        assert got == ["v1", "v2", "v3", "v4", "v5"]
    finally:
        wb.close()

    item = {
        "id": "col_a",
        "sources": [
            {
                "type": "cell",
                "sheet_name": "Data",
                "cell_ref": "A1",
                "repeat_direction": "vertical",
                "row_offset": 1,
                "repeat_max": 5,
                "repeat_until_empty": False,
            }
        ],
    }
    with ex.xlsx_workbook_scope():
        ex.precache_xlsx_workbook_sheets_for_items(p, [item])
        vals = ex.extract_item_values(p, item)
    got = [str(x)[1:] if str(x).startswith("'") else str(x) for x in vals]
    assert got == ["v1", "v2", "v3", "v4", "v5"]
