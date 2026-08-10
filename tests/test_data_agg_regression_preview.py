# -*- coding: utf-8 -*-
from __future__ import annotations

import sys
from pathlib import Path

_root = Path(__file__).resolve().parents[1]
if str(_root) not in sys.path:
    sys.path.insert(0, str(_root))

from svc import svc_data_agg_extract as extract_mod  # noqa: E402
from svc.svc_data_agg import filter_file_paths_for_master_preview  # noqa: E402


def test_stepped_preview_future_name_extract_empty_sources_path_col_hint() -> None:
    """未到達の名前取得はソース空。照合列は path_col_hint と compute 補完に任せる（セル取得と対称）。"""
    from svc.data_agg_master_preview import scenario_for_stepped_preview  # noqa: E402
    from svc.svc_data_agg import resolve_path_column_for_merge  # noqa: E402

    base = {
        "items": [
            {"name": "製番", "sources": []},
            {
                "name": "品名",
                "sources": [{"type": "cell", "sheet_name": "S", "cell_ref": "A1"}],
            },
            {
                "name": "出荷番号",
                "sources": [
                    {
                        "type": "name_extract",
                        "ui_scenario_source_v1": {"path_item": "品名"},
                        "source_type": "file_name",
                        "search_condition": "include",
                        "search_text": "ODN",
                        "delimiter": "_",
                        "part_index": 2,
                    }
                ],
            },
        ]
    }
    headers = [it["name"] for it in base["items"]]
    out = scenario_for_stepped_preview(
        base,
        mi_idx=1,
        master_step_idx=0,
        active_slot_indices=[0],
    )
    items2 = out["items"]
    assert len(items2[2].get("sources") or []) == 0
    assert out["__debug_diag"].get("path_col_hint") == "品名"
    assert resolve_path_column_for_merge(items2, headers) == ""


def test_master_preview_mi_idx_skips_future_name_extract_column(tmp_path) -> None:
    """品名ステップでは出荷番号列を埋めず、出荷番号項目まで進んだときだけ名前取得を反映する。"""
    from openpyxl import Workbook  # noqa: WPS433

    from svc.data_agg_master_preview import scenario_for_stepped_preview  # noqa: E402
    from svc.svc_data_agg import compute_batch_table_rows  # noqa: E402

    p = tmp_path / "ODN_seg2_seg3_tail.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "S"
    ws["A1"] = "PN"
    wb.save(p)
    fp = str(p)

    base = {
        "items": [
            {"name": "製番", "sources": []},
            {
                "name": "品名",
                "sources": [{"type": "cell", "sheet_name": "S", "cell_ref": "A1"}],
            },
            {
                "name": "出荷番号",
                "sources": [
                    {
                        "type": "name_extract",
                        "ui_scenario_source_v1": {"path_item": "品名"},
                        "source_type": "file_name",
                        "search_condition": "include",
                        "search_text": "ODN",
                        "delimiter": "_",
                        "part_index": 2,
                    }
                ],
            },
        ]
    }
    scen_品名 = scenario_for_stepped_preview(
        base,
        mi_idx=1,
        master_step_idx=1,
        active_slot_indices=[0],
    )
    headers, rows, _, _ = compute_batch_table_rows(
        scen_品名,
        [fp],
        max_primary_rows=50,
        max_table_rows=50,
    )
    ix_ship = headers.index("出荷番号")
    assert rows
    assert all(r[ix_ship] in (None, "") for r in rows)

    scen_ship = scenario_for_stepped_preview(
        base,
        mi_idx=2,
        master_step_idx=1,
        active_slot_indices=[0],
    )
    headers2, rows2, _, _ = compute_batch_table_rows(
        scen_ship,
        [fp],
        max_primary_rows=50,
        max_table_rows=50,
    )
    ix2 = headers2.index("出荷番号")
    assert rows2
    assert rows2[0][ix2] not in (None, "")


def test_stepped_preview_path_col_hint_when_current_strips_only_name_extract() -> None:
    """最終項目が名前取得のみで step0 のときも path_col_hint で照合列を補えること。"""
    from svc.data_agg_master_preview import scenario_for_stepped_preview  # noqa: E402
    from svc.svc_data_agg import resolve_path_column_for_merge  # noqa: E402

    base = {
        "items": [
            {"name": "品名", "sources": [{"type": "cell", "sheet_name": "S", "cell_ref": "A1"}]},
            {
                "name": "出荷番号",
                "sources": [
                    {
                        "type": "name_extract",
                        "ui_scenario_source_v1": {"path_item": "品名"},
                        "source_type": "file_name",
                        "search_text": "X",
                        "delimiter": "_",
                        "part_index": 1,
                    }
                ],
            },
        ]
    }
    headers = [it["name"] for it in base["items"]]
    out = scenario_for_stepped_preview(
        base,
        mi_idx=1,
        master_step_idx=0,
        active_slot_indices=[0],
    )
    assert out["__debug_diag"].get("path_col_hint") == "品名"
    assert resolve_path_column_for_merge(out["items"], headers) == ""


def test_stepped_preview_current_item_no_name_extract_until_picked() -> None:
    """現在項目は active でピックしたソースのみ（未ピックの path name_extract は付けない）。"""
    from svc.data_agg_master_preview import scenario_for_stepped_preview  # noqa: E402

    base = {
        "items": [
            {"name": "品名", "sources": [{"type": "cell", "sheet_name": "S", "cell_ref": "A1"}]},
            {
                "name": "出荷番号",
                "sources": [
                    {
                        "type": "name_extract",
                        "ui_scenario_source_v1": {"path_item": "品名"},
                        "source_type": "file_name",
                        "search_text": "X",
                        "delimiter": "_",
                        "part_index": 1,
                    }
                ],
            },
        ]
    }
    out = scenario_for_stepped_preview(
        base,
        mi_idx=1,
        master_step_idx=0,
        active_slot_indices=[0],
    )
    assert (out["items"][1].get("sources") or []) == []


def test_extract_item_values_repeat_caps_by_max_primary_rows_csv(monkeypatch) -> None:
    calls: list[str] = []

    def _fake_extract_cell(file_path, sheet_name=None, cell_ref="A1"):
        calls.append(str(cell_ref))
        return "v"

    monkeypatch.setattr(extract_mod, "extract_cell", _fake_extract_cell)
    item = {
        "sources": [
            {
                "type": "cell",
                "sheet_name": "S",
                "cell_ref": "A1",
                "repeat_direction": "vertical",
                "row_offset": 1,
                "col_offset": 0,
                "repeat_until_empty": False,
                "repeat_max": None,
            }
        ]
    }
    vals = extract_mod.extract_item_values("dummy.csv", item, max_primary_rows=4)
    assert [str(x).lstrip("'") for x in vals] == ["v", "v", "v", "v"]
    # 上限判定のため終端覗き読みが1回入る場合がある
    assert len(calls) == 5


def test_extract_item_values_repeat_xlsx_opens_workbook_once(monkeypatch) -> None:
    class _FakeWorkbook:
        def __init__(self) -> None:
            self.closed = False
            self.sheetnames = ["S"]
            self.active = object()

        def close(self) -> None:
            self.closed = True

    wb = _FakeWorkbook()
    load_calls: list[str] = []
    read_calls: list[str] = []

    def _fake_load_workbook(path, read_only=True, data_only=True):
        load_calls.append(str(path))
        return wb

    def _fake_load_readonly(path):
        load_calls.append(str(path))
        return wb

    def _fake_read_from_wb(_wb, _sheet_name, cell_ref, **_kwargs):
        read_calls.append(str(cell_ref))
        return "x"

    import openpyxl  # noqa: WPS433

    monkeypatch.setattr(openpyxl, "load_workbook", _fake_load_workbook)
    monkeypatch.setattr(extract_mod, "_load_workbook_readonly", _fake_load_readonly)
    monkeypatch.setattr(extract_mod, "_xlsx_cell_value_open_workbook", _fake_read_from_wb)
    item = {
        "sources": [
            {
                "type": "cell",
                "sheet_name": "S",
                "cell_ref": "B2",
                "repeat_direction": "vertical",
                "row_offset": 1,
                "col_offset": 0,
                "repeat_until_empty": False,
                "repeat_max": 9999,
            }
        ]
    }
    vals = extract_mod.extract_item_values("dummy.xlsx", item, max_primary_rows=3)
    assert [str(x).lstrip("'") for x in vals] == ["x", "x", "x"]
    assert len(load_calls) == 1
    # 上限判定のため終端覗き読みが1回入る場合がある
    assert len(read_calls) == 4
    assert wb.closed is True


def test_extract_item_values_cell_source_spans_per_source(monkeypatch) -> None:
    def _fake_extract_cell(file_path, sheet_name=None, cell_ref="A1"):
        return str(cell_ref)

    monkeypatch.setattr(extract_mod, "extract_cell", _fake_extract_cell)
    item = {
        "sources": [
            {"type": "cell", "sheet_name": "S", "cell_ref": "A1"},
            {
                "type": "cell",
                "sheet_name": "S",
                "cell_ref": "B1",
                "repeat_direction": "vertical",
                "row_offset": 1,
                "col_offset": 0,
                "repeat_until_empty": False,
                "repeat_max": 2,
            },
        ]
    }
    spans: dict[int, tuple[int, int]] = {}
    extract_mod.extract_item_values(
        "dummy.csv", item, cell_source_spans_out=spans
    )
    assert spans.get(0) == (0, 1)
    assert spans.get(1) == (1, 2)


def test_extract_item_values_empty_sources_returns_blank_not_filename() -> None:
    """sources が空の項目はファイル名を主値にしない（一括出力で空欄にする）。"""
    vals = extract_mod.extract_item_values(
        r"C:\data\ODN375_A0512M000001.xlsx", {"name": "群番/副番", "sources": []}
    )
    assert vals == [""]


def test_compute_batch_empty_sources_column_not_filename(tmp_path) -> None:
    """compute_batch でソース未設定列に入力ファイル名が載らないこと。"""
    from openpyxl import Workbook  # noqa: WPS433

    from svc.svc_data_agg import compute_batch_table_rows  # noqa: E402

    p = tmp_path / "ODN375_A0512M000001.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "S"
    ws["A1"] = "PN123"
    wb.save(p)
    fp = str(p)

    data = {
        "version": 1,
        "items": [
            {
                "name": "製番",
                "sources": [
                    {"type": "cell", "sheet_name": "S", "cell_ref": "A1"},
                ],
            },
            {"name": "群番/副番", "sources": []},
        ],
        "match_keys": [],
    }
    headers, rows, _, _ = compute_batch_table_rows(
        data,
        [fp],
        max_primary_rows=50,
        max_table_rows=50,
    )
    ix_blank = headers.index("群番/副番")
    assert rows
    assert all(r[ix_blank] in (None, "") for r in rows)
    ix_filled = headers.index("製番")
    assert rows[0][ix_filled] is not None
    assert str(rows[0][ix_filled]).lstrip("'") == "PN123"


def test_filter_file_paths_for_master_preview_single_pattern() -> None:
    files = [
        r"C:\data\A_card.xlsx",
        r"C:\data\A_main.xlsx",
        r"C:\data\B_card.xlsx",
    ]
    items = [
        {
            "sources": [
                {
                    "type": "cell",
                    "ui_scenario_source_v1": {"file_pattern": "A_", "file_name_rule": "含む"},
                }
            ]
        },
    ]
    out = filter_file_paths_for_master_preview(files, items)
    assert out == [r"C:\data\A_card.xlsx", r"C:\data\A_main.xlsx"]


def test_filter_file_paths_for_master_preview_multi_pattern_or_union() -> None:
    """光特性×紐づけのように file_pattern が異なるときは OR（和集合）。"""
    files = [
        r"C:\data\光特性履歴.xlsx",
        r"C:\data\紐づけ履歴.xlsx",
        r"C:\data\other.xlsx",
    ]
    items = [
        {
            "sources": [
                {
                    "type": "cell",
                    "ui_scenario_source_v1": {
                        "file_pattern": "光特性",
                        "file_name_rule": "含む",
                    },
                }
            ]
        },
        {
            "sources": [
                {
                    "type": "cell",
                    "ui_scenario_source_v1": {
                        "file_pattern": "紐づけ",
                        "file_name_rule": "含む",
                    },
                }
            ]
        },
    ]
    out = filter_file_paths_for_master_preview(files, items)
    assert out == [
        r"C:\data\光特性履歴.xlsx",
        r"C:\data\紐づけ履歴.xlsx",
    ]
