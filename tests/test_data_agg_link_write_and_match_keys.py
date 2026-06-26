# -*- coding: utf-8 -*-
"""連携代入の書込みモード・照合キー経路での連携列保持の回帰。"""
# pyright: reportPrivateUsage=false
from __future__ import annotations

import sys
from pathlib import Path
from typing import Any

from openpyxl import Workbook

_root = Path(__file__).resolve().parents[1]
if str(_root) not in sys.path:
    sys.path.insert(0, str(_root))

from svc.svc_data_agg import (  # noqa: E402
    _apply_join_key_search_link_write,
    _apply_join_key_search_write,
    _assign_series_to_rows_by_context,
    compute_batch_table_rows,
)
from core.core_join_compare import join_compare_display_key  # noqa: E402


def _disp(val: Any) -> str:
    """table_rows セル値を表示本体で比較する。"""
    return join_compare_display_key(val)


def _active_worksheet(wb: Any) -> Any:
    ws = wb.active
    assert ws is not None
    return ws


def test_assign_series_link_respects_overwrite_and_empty_string() -> None:
    """連携代入は write_mode に従い、固定値の空文字も new として反映する。"""
    rows = [
        {"__file_path": "f.xlsx", "__iter_index": 0, "MAC": "old"},
    ]
    _assign_series_to_rows_by_context(
        rows,
        "MAC",
        [""],
        [{"file_path": "f.xlsx", "iter_index": 0}],
        "f.xlsx",
        write_mode="overwrite",
    )
    assert rows[0]["MAC"] == ""

    rows2 = [
        {"__file_path": "f.xlsx", "__iter_index": 0, "MAC": "old"},
    ]
    _assign_series_to_rows_by_context(
        rows2,
        "MAC",
        ["NEW"],
        [{"file_path": "f.xlsx", "iter_index": 0}],
        "f.xlsx",
        write_mode="fill_in",
    )
    assert rows2[0]["MAC"] == "old"

    rows3 = [
        {"__file_path": "f.xlsx", "__iter_index": 0, "MAC": ""},
    ]
    _assign_series_to_rows_by_context(
        rows3,
        "MAC",
        ["Z"],
        [{"file_path": "f.xlsx", "iter_index": 0}],
        "f.xlsx",
        write_mode="fill_in",
    )
    assert rows3[0]["MAC"] == "Z"


def test_match_keys_path_carries_linked_column_to_table(tmp_path: Path) -> None:
    """照合キーありでも連携先列が最終 table_rows に載る。"""
    p = tmp_path / "link_match.xlsx"
    wb = Workbook()
    ws = _active_worksheet(wb)
    ws.title = "S"
    ws["A1"] = "K1"
    ws["B1"] = "M1"
    wb.save(p)
    fp = str(p)

    data: dict[str, Any] = {
        "id": "sc_link_match",
        "items": [
            {
                "id": "i_k",
                "name": "K",
                "write_mode": "fill_in",
                "sources": [{"type": "cell", "sheet_name": "S", "cell_ref": "A1"}],
            },
            {
                "id": "i_m",
                "name": "M",
                "write_mode": "overwrite",
                "sources": [
                    {
                        "type": "cell",
                        "sheet_name": "S",
                        "cell_ref": "B1",
                        "ui_scenario_source_v1": {
                            # 結合キーが無いと file/iter 以外で行が横結合されず連携列が別行に残るため、
                            # 同一行照合用に最小の join_defs を付与する。
                            "join_defs": [
                                {"item": "K", "cell": "A1", "row": 0, "col": 0},
                            ],
                            "link_defs": [
                                {
                                    "item": "MAC",
                                    "cell": "DEL",
                                    "mode": "固定値",
                                    "row": 0,
                                    "col": 0,
                                }
                            ],
                        },
                    }
                ],
            },
            {"id": "i_mac", "name": "MAC", "write_mode": "fill_in", "sources": []},
        ],
        "match_keys": ["i_k"],
    }
    headers, rows, _, _ = compute_batch_table_rows(
        data,
        [fp],
        max_primary_rows=50,
        max_table_rows=50,
    )
    ix_mac = headers.index("MAC")
    assert rows
    assert _disp(rows[0][ix_mac]) == "DEL"


def _join_link_host_item() -> dict[str, Any]:
    return {
        "sources": [
            {
                "type": "cell",
                "sheet_name": "S",
                "cell_ref": "B1",
                "ui_scenario_source_v1": {
                    "join_defs": [{"item": "機器番号", "cell": "A1", "row": 0, "col": 0}],
                    "link_defs": [
                        {
                            "item": "MAC",
                            "cell": "",
                            "mode": "固定値",
                            "row": 0,
                            "col": 0,
                        }
                    ],
                },
            }
        ],
    }


def test_join_search_link_empty_on_matched_rows_overwrite() -> None:
    """結合一致行へ空連携を overwrite で載せる（G2/G4）。"""
    pool = [
        {"機器番号": "PT1", "MAC": "keep", "M": "", "__file_path": "f.xlsx", "__iter_index": 0},
        {"機器番号": "PT1", "MAC": "keep", "M": "", "__file_path": "f.xlsx", "__iter_index": 1},
        {"機器番号": "PT2", "MAC": "other", "M": "", "__file_path": "f.xlsx", "__iter_index": 0},
    ]
    item = _join_link_host_item()
    bundle = {
        "primary_values": ["HOST_VAL"],
        "join_values": {"機器番号": ["PT1"]},
        "link_values": {"MAC": [""]},
    }
    hdrs = {"機器番号", "MAC", "M"}
    _apply_join_key_search_write(pool, item, "M", bundle, "overwrite")
    _apply_join_key_search_link_write(pool, item, bundle, "overwrite", hdrs)
    assert pool[0]["M"] == "HOST_VAL"
    assert pool[1]["M"] == "HOST_VAL"
    assert pool[0]["MAC"] == ""
    assert pool[1]["MAC"] == ""
    assert pool[2]["MAC"] == "other"


def test_join_search_link_fill_in_keeps_existing_mac() -> None:
    """fill_in では一致行でも既存 MAC を空連携で潰さない。"""
    pool = [
        {"機器番号": "PT1", "MAC": "existing", "M": "", "__file_path": "f.xlsx", "__iter_index": 0},
    ]
    item = _join_link_host_item()
    bundle = {
        "primary_values": ["H"],
        "join_values": {"機器番号": ["PT1"]},
        "link_values": {"MAC": [""]},
    }
    _apply_join_key_search_link_write(
        pool, item, bundle, "fill_in", {"機器番号", "MAC", "M"}
    )
    assert pool[0]["MAC"] == "existing"


def test_join_search_link_skipped_without_link_defs() -> None:
    pool = [{"機器番号": "PT1", "MAC": "x", "__file_path": "f.xlsx", "__iter_index": 0}]
    item = {
        "sources": [
            {
                "type": "cell",
                "ui_scenario_source_v1": {
                    "join_defs": [{"item": "機器番号", "cell": "A1"}],
                    "link_defs": [],
                },
            }
        ],
    }
    bundle = {
        "primary_values": ["H"],
        "join_values": {"機器番号": ["PT1"]},
        "link_values": {"MAC": [""]},
    }
    _apply_join_key_search_link_write(pool, item, bundle, "overwrite", {"機器番号", "MAC"})
    assert pool[0]["MAC"] == "x"


def test_cross_file_join_writes_to_anchor_row_only(tmp_path: Path) -> None:
    """光特性行へ PT・製番を横断結合で反映し、紐づけ専用行は出力しない。"""
    anchor = tmp_path / "光特性履歴_test.xlsx"
    wb = Workbook()
    ws = _active_worksheet(wb)
    ws.title = "ﾃﾞｰﾀ"
    ws["C7"] = "DEV1"
    ws["M7"] = "MAC-A"
    wb.save(anchor)

    join_f = tmp_path / "紐づけ履歴_test.xlsx"
    wb2 = Workbook()
    ws2 = _active_worksheet(wb2)
    ws2.title = "紐付け履歴"
    ws2["C5"] = "PT-001"
    ws2["P5"] = "MAC-A"
    ws2["J5"] = "SEQ-1"
    wb2.save(join_f)

    data: dict[str, Any] = {
        "id": "cross_join",
        "items": [
            {
                "id": "i_dev",
                "name": "機器番号",
                "write_mode": "append",
                "sources": [
                    {
                        "type": "cell",
                        "sheet_name": "ﾃﾞｰﾀ",
                        "cell_ref": "C7",
                        "ui_scenario_source_v1": {
                            "file_pattern": "光特性",
                            "link_defs": [
                                {
                                    "item": "MACアドレス",
                                    "cell": "M7",
                                    "mode": "セル座標",
                                    "row": 0,
                                    "col": 0,
                                }
                            ],
                        },
                        "repeat_direction": "vertical",
                        "repeat_max": 1,
                    }
                ],
            },
            {"id": "i_mac", "name": "MACアドレス", "write_mode": "fill_in", "sources": []},
            {
                "id": "i_pt",
                "name": "PT番号",
                "write_mode": "overwrite",
                "sources": [
                    {
                        "type": "cell",
                        "sheet_name": "紐付け履歴",
                        "cell_ref": "C5",
                        "ui_scenario_source_v1": {
                            "file_pattern": "紐づけ",
                            "join_defs": [
                                {"item": "MACアドレス", "cell": "P5", "row": 0, "col": 0}
                            ],
                            "link_defs": [
                                {
                                    "item": "製番",
                                    "cell": "J5",
                                    "mode": "セル座標",
                                    "row": 0,
                                    "col": 0,
                                }
                            ],
                        },
                        "repeat_direction": "vertical",
                        "repeat_max": 1,
                    }
                ],
            },
            {"id": "i_seq", "name": "製番", "write_mode": "fill_in", "sources": []},
        ],
        "match_keys": [],
    }
    headers, rows, _, _ = compute_batch_table_rows(
        data, [str(anchor), str(join_f)], max_primary_rows=10, max_table_rows=10
    )
    ix_dev = headers.index("機器番号")
    ix_pt = headers.index("PT番号")
    ix_seq = headers.index("製番")
    assert len(rows) == 1
    assert _disp(rows[0][ix_dev]) == "DEV1"
    assert _disp(rows[0][ix_pt]) == "PT-001"
    assert _disp(rows[0][ix_seq]) == "SEQ-1"


def test_cross_file_join_writes_all_mac_matches_ignore_iter(tmp_path: Path) -> None:
    """横断結合: 光特性の iter と紐づけの k が異なっても MAC 一致行へ PT を書く。"""
    anchor = tmp_path / "光特性履歴_multi.xlsx"
    wb = Workbook()
    ws = _active_worksheet(wb)
    ws.title = "ﾃﾞｰﾀ"
    ws["C7"] = "DEV-A"
    ws["M7"] = "MAC-X"
    ws["C8"] = "DEV-B"
    ws["M8"] = "MAC-Y"
    wb.save(anchor)

    join_f = tmp_path / "紐づけ履歴_multi.xlsx"
    wb2 = Workbook()
    ws2 = _active_worksheet(wb2)
    ws2.title = "紐付け履歴"
    ws2["C5"] = "PT-X"
    ws2["P5"] = "MAC-X"
    ws2["J5"] = "SEQ-X"
    ws2["C6"] = "PT-Y"
    ws2["P6"] = "MAC-Y"
    ws2["J6"] = "SEQ-Y"
    wb2.save(join_f)

    data: dict[str, Any] = {
        "id": "cross_join_multi",
        "items": [
            {
                "id": "i_dev",
                "name": "機器番号",
                "write_mode": "append",
                "sources": [
                    {
                        "type": "cell",
                        "sheet_name": "ﾃﾞｰﾀ",
                        "cell_ref": "C7",
                        "ui_scenario_source_v1": {
                            "file_pattern": "光特性",
                            "link_defs": [
                                {
                                    "item": "MACアドレス",
                                    "cell": "M7",
                                    "mode": "セル座標",
                                    "row": 1,
                                    "col": 0,
                                }
                            ],
                        },
                        "row_offset": 1,
                        "col_offset": 0,
                        "repeat_direction": "vertical",
                        "repeat_max": 2,
                    }
                ],
            },
            {"id": "i_mac", "name": "MACアドレス", "write_mode": "fill_in", "sources": []},
            {
                "id": "i_pt",
                "name": "PT番号",
                "write_mode": "overwrite",
                "sources": [
                    {
                        "type": "cell",
                        "sheet_name": "紐付け履歴",
                        "cell_ref": "C5",
                        "ui_scenario_source_v1": {
                            "file_pattern": "紐づけ",
                            "join_defs": [
                                {"item": "MACアドレス", "cell": "P5", "row": 1, "col": 0}
                            ],
                            "link_defs": [
                                {
                                    "item": "製番",
                                    "cell": "J5",
                                    "mode": "セル座標",
                                    "row": 1,
                                    "col": 0,
                                }
                            ],
                        },
                        "row_offset": 1,
                        "col_offset": 0,
                        "repeat_direction": "vertical",
                        "repeat_max": 2,
                    }
                ],
            },
            {"id": "i_seq", "name": "製番", "write_mode": "fill_in", "sources": []},
        ],
        "match_keys": [],
    }
    headers, rows, _, _ = compute_batch_table_rows(
        data, [str(anchor), str(join_f)], max_primary_rows=10, max_table_rows=10
    )
    ix_dev = headers.index("機器番号")
    ix_pt = headers.index("PT番号")
    ix_seq = headers.index("製番")
    assert len(rows) == 2
    by_dev = {_disp(r[ix_dev]): r for r in rows}
    assert _disp(by_dev["DEV-A"][ix_pt]) == "PT-X"
    assert _disp(by_dev["DEV-A"][ix_seq]) == "SEQ-X"
    assert _disp(by_dev["DEV-B"][ix_pt]) == "PT-Y"
    assert _disp(by_dev["DEV-B"][ix_seq]) == "SEQ-Y"


def test_cross_file_chained_join_uses_accumulated_emit_row_values(tmp_path: Path) -> None:
    """
    ②で光特性行へ入れた QR を ③が照合し、Excel 出力行にダミーQR が載る。
    旧実装（紐づけ行のみ索引）では出力行のダミーQR が空のままになる。
    """
    anchor = tmp_path / "光特性履歴_chain.xlsx"
    wb = Workbook()
    ws = _active_worksheet(wb)
    ws.title = "ﾃﾞｰﾀ"
    ws["C7"] = "DEV1"
    ws["M7"] = "MAC-A"
    wb.save(anchor)

    join_f = tmp_path / "紐づけ履歴_chain.xlsx"
    wb2 = Workbook()
    ws2 = _active_worksheet(wb2)
    ws2.title = "紐付け履歴"
    ws2["C5"] = "PT-001"
    ws2["P5"] = "MAC-A"
    ws2["J5"] = "SEQ-1"
    ws2["Q5"] = "QR-MATCH"
    wb2.save(join_f)

    pack_f = tmp_path / "梱包出荷履歴_chain.xlsx"
    wb3 = Workbook()
    ws3 = _active_worksheet(wb3)
    ws3["H4"] = "QR-MATCH"
    wb3.save(pack_f)

    data: dict[str, Any] = {
        "id": "cross_chain",
        "items": [
            {
                "id": "i_dev",
                "name": "機器番号",
                "write_mode": "append",
                "sources": [
                    {
                        "type": "cell",
                        "sheet_name": "ﾃﾞｰﾀ",
                        "cell_ref": "C7",
                        "ui_scenario_source_v1": {
                            "file_pattern": "光特性",
                            "link_defs": [
                                {
                                    "item": "MACアドレス",
                                    "cell": "M7",
                                    "mode": "セル座標",
                                    "row": 0,
                                    "col": 0,
                                }
                            ],
                        },
                        "repeat_direction": "vertical",
                        "repeat_max": 1,
                    }
                ],
            },
            {"id": "i_mac", "name": "MACアドレス", "write_mode": "fill_in", "sources": []},
            {"id": "i_qr", "name": "QR装置銘板", "write_mode": "fill_in", "sources": []},
            {
                "id": "i_pt",
                "name": "PT番号",
                "write_mode": "overwrite",
                "sources": [
                    {
                        "type": "cell",
                        "sheet_name": "紐付け履歴",
                        "cell_ref": "C5",
                        "ui_scenario_source_v1": {
                            "file_pattern": "紐づけ",
                            "join_defs": [
                                {"item": "MACアドレス", "cell": "P5", "row": 0, "col": 0}
                            ],
                            "link_defs": [
                                {
                                    "item": "QR装置銘板",
                                    "cell": "Q5",
                                    "mode": "セル座標",
                                    "row": 0,
                                    "col": 0,
                                }
                            ],
                        },
                        "repeat_direction": "vertical",
                        "repeat_max": 1,
                    }
                ],
            },
            {
                "id": "i_dummy",
                "name": "ダミーQR機器番号",
                "write_mode": "fill_in",
                "sources": [
                    {
                        "type": "cell",
                        "sheet_name": "",
                        "cell_ref": "H4",
                        "ui_scenario_source_v1": {
                            "file_pattern": "梱包",
                            "join_defs": [
                                {"item": "QR装置銘板", "cell": "H4", "row": 0, "col": 0}
                            ],
                        },
                        "repeat_direction": "vertical",
                        "repeat_max": 1,
                    }
                ],
            },
        ],
        "match_keys": [],
    }
    headers, rows, _, _ = compute_batch_table_rows(
        data,
        [str(anchor), str(join_f), str(pack_f)],
        max_primary_rows=10,
        max_table_rows=10,
    )
    ix_dev = headers.index("機器番号")
    ix_qr = headers.index("QR装置銘板")
    ix_dummy = headers.index("ダミーQR機器番号")
    assert len(rows) == 1
    assert _disp(rows[0][ix_dev]) == "DEV1"
    assert _disp(rows[0][ix_qr]) == "QR-MATCH"
    assert _disp(rows[0][ix_dummy]) == "QR-MATCH"


def test_paired_join_respects_iter_index(tmp_path: Path) -> None:
    """n_prim==n_join ではスライス k が __iter_index==k の行だけに書く。"""
    p = tmp_path / "paired.xlsx"
    wb = Workbook()
    ws = _active_worksheet(wb)
    ws.title = "S"
    ws["A1"] = "K0"
    ws["A2"] = "K1"
    ws["B1"] = "V0"
    ws["B2"] = "V1"
    wb.save(p)
    fp = str(p)

    data: dict[str, Any] = {
        "id": "paired",
        "items": [
            {
                "id": "i_k",
                "name": "K",
                "write_mode": "fill_in",
                "sources": [
                    {
                        "type": "cell",
                        "sheet_name": "S",
                        "cell_ref": "A1",
                        "row_offset": 1,
                        "col_offset": 0,
                        "repeat_direction": "vertical",
                        "repeat_max": 2,
                    }
                ],
            },
            {
                "id": "i_v",
                "name": "V",
                "write_mode": "overwrite",
                "sources": [
                    {
                        "type": "cell",
                        "sheet_name": "S",
                        "cell_ref": "B1",
                        "row_offset": 1,
                        "col_offset": 0,
                        "ui_scenario_source_v1": {
                            "join_defs": [{"item": "K", "cell": "A1", "row": 1, "col": 0}],
                        },
                        "repeat_direction": "vertical",
                        "repeat_max": 2,
                    }
                ],
            },
        ],
        "match_keys": [],
    }
    headers, rows, _, _ = compute_batch_table_rows(
        data, [fp], max_primary_rows=10, max_table_rows=10
    )
    ix_k = headers.index("K")
    ix_v = headers.index("V")
    assert len(rows) == 2
    assert _disp(rows[0][ix_k]) == "K0"
    assert _disp(rows[0][ix_v]) == "V0"
    assert _disp(rows[1][ix_k]) == "K1"
    assert _disp(rows[1][ix_v]) == "V1"


def test_same_file_single_row_host_join_as30(tmp_path: Path) -> None:
    """ODN375 型: 同一 xlsx 内で AS30(row=0) と 機器番号を照合し MAC LOC を1行ホストへ。"""
    p = tmp_path / "ODN375_sample.xlsx"
    wb = Workbook()
    ws = _active_worksheet(wb)
    ws.title = "ユニット実装チェック"
    ws["P14"] = "UNIT-A"
    ws["P16"] = "DEV-001"
    ws["AS30"] = "DEV-001"
    ws["AV30"] = "AA:BB:CC:DD:EE:FF"
    wb.save(p)

    data: dict[str, Any] = {
        "id": "odn375_like",
        "items": [
            {
                "id": "i_hin",
                "name": "品名",
                "write_mode": "append",
                "sources": [
                    {
                        "type": "cell",
                        "sheet_name": "ユニット実装チェック",
                        "cell_ref": "P14",
                        "ui_scenario_source_v1": {
                            "file_pattern": "ODN375",
                            "link_defs": [
                                {
                                    "item": "機器番号",
                                    "cell": "P16",
                                    "mode": "セル座標",
                                    "row": 0,
                                    "col": 0,
                                }
                            ],
                        },
                        "repeat_direction": "vertical",
                        "repeat_max": 1,
                    }
                ],
            },
            {"id": "i_dev", "name": "機器番号", "write_mode": "append", "sources": []},
            {"id": "i_mac", "name": "MAC", "write_mode": "append", "sources": []},
            {
                "id": "i_loc",
                "name": "MAC LOC",
                "write_mode": "overwrite",
                "sources": [
                    {
                        "type": "cell",
                        "sheet_name": "ユニット実装チェック",
                        "cell_ref": "AV30",
                        "ui_scenario_source_v1": {
                            "file_pattern": "ODN375",
                            "link_defs": [
                                {
                                    "item": "MAC",
                                    "cell": "",
                                    "mode": "固定値",
                                    "row": 0,
                                    "col": 0,
                                }
                            ],
                            "join_defs": [
                                {"item": "機器番号", "cell": "AS30", "row": 0, "col": 0}
                            ],
                        },
                        "repeat_direction": "vertical",
                        "repeat_max": 1,
                    }
                ],
            },
        ],
        "match_keys": [],
    }
    headers, rows, _, _ = compute_batch_table_rows(
        data, [str(p)], max_primary_rows=10, max_table_rows=10
    )
    ix_dev = headers.index("機器番号")
    ix_loc = headers.index("MAC LOC")
    assert len(rows) == 1
    assert _disp(rows[0][ix_dev]) == "DEV-001"
    assert _disp(rows[0][ix_loc]) == "AA:BB:CC:DD:EE:FF"
