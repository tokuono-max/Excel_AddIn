# -*- coding: utf-8 -*-
"""skip_hidden_rows: 非表示行を主キー走査から除外。"""
from __future__ import annotations

import sys
from pathlib import Path

_root = Path(__file__).resolve().parents[1]
if str(_root) not in sys.path:
    sys.path.insert(0, str(_root))

from openpyxl import Workbook  # noqa: E402

from svc.data_agg_row_visibility import (  # noqa: E402
    get_hidden_excel_rows,
    source_wants_skip_hidden_rows,
)
from svc.svc_data_agg_extract import extract_item_bundle, xlsx_workbook_scope  # noqa: E402


def _item(*, skip_hidden: bool = False, skip_empty: bool = False) -> dict:
    return {
        "id": "i1",
        "name": "機器番号",
        "sources": [
            {
                "type": "cell",
                "sheet_name": "S",
                "cell_ref": "A1",
                "row_offset": 1,
                "col_offset": 0,
                "repeat_until_empty": False,
                "repeat_until_last": True,
                "repeat_direction": "vertical",
                "skip_empty_primary": skip_empty,
                "skip_primary_match": ",",
                "skip_hidden_rows": skip_hidden,
                "ui_scenario_source_v1": {
                    "link_defs": [
                        {
                            "item": "局名",
                            "mode": "セル座標",
                            "cell": "B1",
                            "row": 1,
                            "col": 0,
                            "carry_empty": False,
                        }
                    ]
                },
            }
        ],
    }


def _txt(v: object) -> str:
    s = "" if v is None else str(v)
    if s.startswith("'"):
        s = s[1:]
    return s.strip()


def test_source_wants_skip_hidden_rows() -> None:
    assert not source_wants_skip_hidden_rows({})
    assert not source_wants_skip_hidden_rows({"skip_hidden_rows": False})
    assert source_wants_skip_hidden_rows({"skip_hidden_rows": True})
    assert source_wants_skip_hidden_rows({"skip_hidden_rows": "true"})


def test_get_hidden_excel_rows_xlsx(tmp_path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "S"
    ws["A1"] = "P1"
    ws["A2"] = "P2"
    ws["A3"] = "P3"
    ws.row_dimensions[2].hidden = True
    fp = tmp_path / "hid.xlsx"
    wb.save(fp)
    wb.close()
    hidden = get_hidden_excel_rows(fp, "S")
    assert 1 in hidden  # 0-based row 1 = Excel row 2
    assert 0 not in hidden
    assert 2 not in hidden


def test_skip_hidden_rows_excludes_primary_and_aligns_link(tmp_path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "S"
    ws["A1"] = "P1"
    ws["B1"] = "L1"
    ws["A2"] = "P2"
    ws["B2"] = "L2"
    ws["A3"] = "P3"
    ws["B3"] = "L3"
    ws.row_dimensions[2].hidden = True
    fp = tmp_path / "skip_hid.xlsx"
    wb.save(fp)
    wb.close()
    with xlsx_workbook_scope():
        b_off = extract_item_bundle(str(fp), _item(skip_hidden=False), item_id="i1")
        b_on = extract_item_bundle(str(fp), _item(skip_hidden=True), item_id="i1")
    assert [_txt(x) for x in (b_off["primary_values"] or [])] == ["P1", "P2", "P3"]
    assert [_txt(x) for x in (b_on["primary_values"] or [])] == ["P1", "P3"]
    assert [_txt(x) for x in (b_on["link_values"]["局名"] or [])] == ["L1", "L3"]


def test_skip_hidden_off_keeps_all(tmp_path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "S"
    ws["A1"] = "P1"
    ws["A2"] = "P2"
    ws.row_dimensions[2].hidden = True
    fp = tmp_path / "skip_hid_off.xlsx"
    wb.save(fp)
    wb.close()
    with xlsx_workbook_scope():
        b = extract_item_bundle(str(fp), _item(skip_hidden=False), item_id="i1")
    assert [_txt(x) for x in (b["primary_values"] or [])] == ["P1", "P2"]
