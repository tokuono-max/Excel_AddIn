# -*- coding: utf-8 -*-
from __future__ import annotations

import sys
from pathlib import Path

import pytest
from openpyxl import Workbook
from core.core_join_compare import join_compare_display_key

_root = Path(__file__).resolve().parents[1]
if str(_root) not in sys.path:
    sys.path.insert(0, str(_root))

from svc.svc_data_agg import (  # noqa: E402
    _master_preview_join_item_effective,
    _master_preview_per_file_pool_cap,
    compute_batch_table_rows,
    filter_file_paths_for_master_preview,
    reorder_paths_for_master_preview_join_priority,
)
from svc.data_agg_master_preview import table_rows_to_join_search_seed_pool  # noqa: E402
from tests.test_data_agg_batch_stability import (  # noqa: E402
    _active_worksheet,
    _cross_join_mini_scenario,
)


def test_reorder_paths_interleaves_side_and_host(tmp_path: Path) -> None:
    """PT番号ホスト: side（光特性）と host（紐づけ）をラウンドロビンで交互に並べる。"""
    data, paths = _cross_join_mini_scenario(tmp_path)
    anchor, join_f = paths[0], paths[1]
    shuffled = [join_f, join_f, anchor, join_f]
    headers = [str(it.get("name") or "") for it in data["items"]]
    dd: dict = {"mi_idx": 2}
    ordered = reorder_paths_for_master_preview_join_priority(
        shuffled, data["items"], headers, dd
    )
    assert ordered[0] == anchor
    assert ordered[1] == join_f
    assert dd.get("master_preview_priority_files")
    assert dd.get("master_preview_join_side_patterns")
    assert dd.get("master_preview_join_host_patterns")


def test_master_preview_per_file_pool_cap_is_read_limit() -> None:
    data, _paths = _cross_join_mini_scenario(Path("."))
    host = data["items"][2]
    headers = [str(it.get("name") or "") for it in data["items"]]
    cap = _master_preview_per_file_pool_cap(200, host, data["items"], headers)
    assert cap == 200


def test_master_preview_join_priority_fills_link_columns(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    """横断 join プレビューで PT番号・製番が結合結果に載る。"""
    data, paths = _cross_join_mini_scenario(tmp_path)
    anchor, join_f = paths[0], paths[1]
    data["__debug_diag"] = {
        "enabled": True,
        "source": "ui_data_agg_debug.master_preview",
        "mi_idx": 2,
        "join_search_skip_seed": True,
    }
    paths_ordered = [join_f, anchor]
    monkeypatch.setenv("DATA_AGG_FILE_PARALLEL_WORKERS", "0")
    monkeypatch.setenv("DATA_AGG_MASTER_PARALLEL_EXTRACT", "0")
    headers, rows, _ev, _je = compute_batch_table_rows(
        data,
        paths_ordered,
        max_primary_rows=4,
        max_table_rows=4,
        probe_caller="test_join_priority",
    )
    idx_pt = headers.index("PT番号")
    idx_seq = headers.index("製番")
    pt_vals = [r[idx_pt] for r in rows if r[idx_pt]]
    seq_vals = [r[idx_seq] for r in rows if r[idx_seq]]
    assert pt_vals, "PT番号が空（join プレビューが効いていない）"
    assert seq_vals, "製番が空（link allowlist 経由の取得が効いていない）"


def test_master_preview_per_file_cap_allows_second_file_type(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    """1 ファイルが cap を独占しない（side と host の両方がプールに入る）。"""
    data, paths = _cross_join_mini_scenario(tmp_path)
    anchor = Path(paths[0])
    wb = Workbook()
    ws = _active_worksheet(wb)
    ws.title = "ﾃﾞｰﾀ"
    for i in range(8):
        ws.cell(row=7 + i, column=3, value="DEV%s" % i)
        ws.cell(row=7 + i, column=13, value="MAC-A")
    wb.save(anchor)
    for src in data["items"][0]["sources"]:
        src["repeat_max"] = 8
    data["__debug_diag"] = {
        "enabled": True,
        "source": "ui_data_agg_debug.master_preview",
        "mi_idx": 2,
        "join_search_skip_seed": True,
    }
    monkeypatch.setenv("DATA_AGG_FILE_PARALLEL_WORKERS", "0")
    monkeypatch.setenv("DATA_AGG_MASTER_PARALLEL_EXTRACT", "0")
    _h, rows, _ev, _je = compute_batch_table_rows(
        data,
        paths,
        max_primary_rows=4,
        max_table_rows=4,
        probe_caller="test_per_file_cap",
    )
    headers = list(_h)
    dd = data.get("__debug_diag") or {}
    idx_pt = headers.index("PT番号")
    idx_seq = headers.index("製番")
    pt_vals = [r[idx_pt] for r in rows if r[idx_pt]]
    seq_vals = [r[idx_seq] for r in rows if r[idx_seq]]
    assert pt_vals, "PT番号が空（side と host の両方がプールに入っていない）"
    assert seq_vals, "製番が空"
    assert len(rows) <= 4
    assert len(rows) >= 1
    assert int(dd.get("master_preview_pool_row_cap") or 0) == 8


def test_master_preview_interleave_reads_host_after_side_cap(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    """光特性が複数あっても host（紐づけ）がプールに入り join が走る。"""
    data, paths = _cross_join_mini_scenario(tmp_path)
    anchor, join_f = paths[0], paths[1]
    anchor2 = tmp_path / "光特性履歴_test2.xlsx"
    anchor3 = tmp_path / "光特性履歴_test3.xlsx"
    for extra in (anchor2, anchor3):
        extra.write_bytes(Path(anchor).read_bytes())
    wb = Workbook()
    ws = _active_worksheet(wb)
    ws.title = "ﾃﾞｰﾀ"
    for i in range(50):
        ws.cell(row=7 + i, column=3, value="DEV%s" % i)
        ws.cell(row=7 + i, column=13, value="MAC-A")
    wb.save(anchor)
    for src in data["items"][0]["sources"]:
        src["repeat_max"] = 50
    data["__debug_diag"] = {
        "enabled": True,
        "source": "ui_data_agg_debug.master_preview",
        "mi_idx": 2,
        "join_search_skip_seed": True,
    }
    paths_many_side = [str(anchor), str(anchor2), str(anchor3), str(join_f)]
    monkeypatch.setenv("DATA_AGG_FILE_PARALLEL_WORKERS", "0")
    monkeypatch.setenv("DATA_AGG_MASTER_PARALLEL_EXTRACT", "0")
    headers, rows, _ev, _je = compute_batch_table_rows(
        data,
        paths_many_side,
        max_primary_rows=200,
        max_table_rows=100,
        probe_caller="test_interleave_host",
    )
    dd = data.get("__debug_diag") or {}
    idx_pt = headers.index("PT番号")
    idx_seq = headers.index("製番")
    pt_vals = [r[idx_pt] for r in rows if r[idx_pt]]
    seq_vals = [r[idx_seq] for r in rows if r[idx_seq]]
    assert pt_vals, "PT番号が空（host がプールに入っていない）"
    assert seq_vals, "製番が空（link が効いていない）"
    assert len(rows) >= 1
    pool_cap = int(dd.get("master_preview_pool_row_cap") or 0)
    assert pool_cap == 800


def test_master_preview_anchor_rows_keep_prior_row_identity(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    data, paths = _cross_join_mini_scenario(tmp_path)
    headers = [str(it.get("name") or "") for it in data["items"]]
    seed_rows = [["DEV1", "MAC-A", None, None]]
    seed_pool = table_rows_to_join_search_seed_pool(headers, seed_rows)
    data["__debug_diag"] = {
        "enabled": True,
        "source": "ui_data_agg_debug.master_preview",
        "mi_idx": 2,
        "join_search_skip_seed": False,
        "join_search_seed_pool": seed_pool,
        "preview_anchor_row_keys": [[seed_pool[0]["__file_path"], seed_pool[0]["__iter_index"]]],
    }
    monkeypatch.setenv("DATA_AGG_FILE_PARALLEL_WORKERS", "0")
    monkeypatch.setenv("DATA_AGG_MASTER_PARALLEL_EXTRACT", "0")
    out_headers, out_rows, _ev, _je = compute_batch_table_rows(
        data,
        paths,
        max_primary_rows=10,
        max_table_rows=10,
        probe_caller="test_anchor_rows",
    )
    idx_dev = out_headers.index("機器番号")
    idx_pt = out_headers.index("PT番号")
    idx_seq = out_headers.index("製番")
    assert len(out_rows) == 1
    assert out_rows[0][idx_dev] == "DEV1"
    assert join_compare_display_key(out_rows[0][idx_pt]) == "PT-001"
    assert join_compare_display_key(out_rows[0][idx_seq]) == "SEQ-1"


def test_reorder_paths_uses_topology_when_prior_sources_cleared(
    tmp_path: Path,
) -> None:
    """carry-forward で前項目 sources が空でも topology で横断 join 判定できる。"""
    import copy

    data, paths = _cross_join_mini_scenario(tmp_path)
    stepped_items = copy.deepcopy(data["items"])
    for j, it in enumerate(stepped_items):
        if j < 2:
            it["sources"] = []
    headers = [str(it.get("name") or "") for it in data["items"]]
    dd: dict = {
        "mi_idx": 2,
        "preview_join_topology_items": copy.deepcopy(data["items"]),
    }
    ordered = reorder_paths_for_master_preview_join_priority(
        paths, stepped_items, headers, dd
    )
    assert ordered[0] == paths[0]
    assert dd.get("master_preview_join_side_patterns")
    assert dd.get("master_preview_join_host_patterns")


def test_carry_forward_topology_enables_cross_join_pt(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    """mpv UI 相当: carry-forward + topology で PT が table_rows に載る。"""
    import copy

    from svc.data_agg_master_preview import (  # noqa: WPS433
        MASTER_PREVIEW_DIAG_SOURCE,
        scenario_for_stepped_preview,
        table_rows_to_join_search_seed_pool,
    )
    from svc.svc_data_agg import master_preview_extract_item_allowlist  # noqa: WPS433

    data, paths = _cross_join_mini_scenario(tmp_path)
    headers = [str(it.get("name") or "") for it in data["items"]]
    prior_rows = [["DEV1", "MAC-A", None, None]]
    scen = scenario_for_stepped_preview(
        data,
        mi_idx=2,
        master_step_idx=1,
        active_slot_indices=[0],
        carry_forward_completed_items=True,
    )
    dd = scen["__debug_diag"]
    dd["preview_join_topology_items"] = copy.deepcopy(data["items"])
    dd["source"] = MASTER_PREVIEW_DIAG_SOURCE
    dd["master_preview_join_read_full_files"] = True
    allow_ix = master_preview_extract_item_allowlist(data, mi_idx=2)
    if allow_ix is not None:
        dd["preview_extract_item_allowlist"] = list(allow_ix)
    dd["join_search_seed_pool"] = table_rows_to_join_search_seed_pool(
        headers, prior_rows
    )
    monkeypatch.setenv("DATA_AGG_FILE_PARALLEL_WORKERS", "0")
    monkeypatch.setenv("DATA_AGG_MASTER_PARALLEL_EXTRACT", "0")
    out_headers, out_rows, _ev, _je = compute_batch_table_rows(
        scen,
        paths,
        max_primary_rows=10,
        max_table_rows=10,
        probe_caller="test_carry_forward_topology",
    )
    idx_pt = out_headers.index("PT番号")
    pt_vals = [r[idx_pt] for r in out_rows if r[idx_pt] not in (None, "")]
    assert pt_vals, "carry-forward + topology でも PT が空"
    assert join_compare_display_key(pt_vals[0]) == "PT-001"


def test_record_join_host_patterns_for_non_cross_chain() -> None:
    from svc.svc_data_agg import _master_preview_record_join_host_patterns_only

    host = {
        "sources": [
            {
                "type": "cell",
                "ui_scenario_source_v1": {"file_pattern": "梱包"},
            }
        ]
    }
    dd: dict = {}
    _master_preview_record_join_host_patterns_only(dd, host=host, topo=[host])
    assert dd.get("master_preview_join_host_patterns") == ["梱包"]


def test_seed_pool_prefers_anchor_file_path_over_synthetic() -> None:
    rows = table_rows_to_join_search_seed_pool(
        ["機器番号", "MACアドレス"],
        [["DEV1", "MAC-A"]],
        anchor_file_path=r"C:\data\光特性履歴.xlsx",
    )
    assert rows[0]["__file_path"] == r"C:\data\光特性履歴.xlsx"
    assert not str(rows[0]["__file_path"]).startswith("mpv_table_seed://")


def test_seed_pool_stacked_join_uses_row_file_paths() -> None:
    fps = [r"C:\a\file1.xlsx", r"C:\a\file2.xlsx", r"C:\a\file3.xlsx"]
    rows = table_rows_to_join_search_seed_pool(
        ["品名", "機器番号"],
        [["A", "1"], ["B", "2"]],
        anchor_file_path=r"C:\a\anchor.xlsx",
        row_file_paths=fps,
        stacked_join=True,
    )
    assert rows[0]["__file_path"] == fps[0]
    assert rows[1]["__file_path"] == fps[1]
    assert rows[0]["__file_path"] != r"C:\a\anchor.xlsx"


def test_table_row_file_paths_groups_by_device_id() -> None:
    from svc.data_agg_master_preview import table_row_file_paths_for_stacked_seed  # noqa: WPS433

    headers = ["品名", "機器番号", "実装装置番号"]
    rows = [
        ["ユニット", "PT1", "'DEV-A"],
        ["IMX-VCH", "PT1 '1", "'DEV-A"],
        ["ユニット", "PT2", "'DEV-B"],
    ]
    scan = [r"C:\a\sub1.xlsx", r"C:\a\sub2.xlsx"]
    fps = table_row_file_paths_for_stacked_seed(headers, rows, scan_paths=scan)
    assert fps == [scan[0], scan[0], scan[1]]


def test_table_row_file_paths_prefers_stored_iteration_paths() -> None:
    from svc.data_agg_master_preview import table_row_file_paths_for_stacked_seed  # noqa: WPS433

    headers = ["品名", "機器番号"]
    rows = [["A", "1"], ["B", "2"]]
    stored = [r"C:\real\file1.xlsx", r"C:\real\file1.xlsx"]
    fps = table_row_file_paths_for_stacked_seed(
        headers,
        rows,
        scan_paths=[r"C:\wrong\file1.xlsx", r"C:\wrong\file2.xlsx"],
        stored_row_paths=stored,
    )
    assert fps == stored


def test_patch_stacked_join_pool_skips_multi_distinct_device_ids() -> None:
    from svc.svc_data_agg import _patch_stacked_join_pool_row_join_targets  # noqa: WPS433

    pool = [
        {
            "__file_path": r"C:\data\sub1.xlsx",
            "機器番号": "'PT4400243777",
            "__iter_index": 0,
        },
        {
            "__file_path": r"C:\data\sub1.xlsx",
            "機器番号": "'PT4400243777 '1",
            "__iter_index": 1,
        },
        {
            "__file_path": r"C:\data\sub2.xlsx",
            "機器番号": "OTHER",
            "__iter_index": 0,
        },
    ]
    _patch_stacked_join_pool_row_join_targets(
        pool,
        file_path=r"C:\data\sub1.xlsx",
        join_defs=[{"item": "機器番号", "cell": "AS30"}],
        bundle={"join_values": {"機器番号": ["'PT4400243777"]}},
    )
    assert pool[0]["機器番号"] == "'PT4400243777"
    assert pool[1]["機器番号"] == "'PT4400243777 '1"
    assert pool[2]["機器番号"] == "OTHER"


def test_patch_stacked_join_pool_patches_single_wrong_row() -> None:
    from svc.svc_data_agg import _patch_stacked_join_pool_row_join_targets  # noqa: WPS433

    pool = [
        {
            "__file_path": r"C:\data\sub1.xlsx",
            "機器番号": "WRONG-1",
            "__iter_index": 0,
        },
        {
            "__file_path": r"C:\data\sub2.xlsx",
            "機器番号": "OTHER",
            "__iter_index": 0,
        },
    ]
    _patch_stacked_join_pool_row_join_targets(
        pool,
        file_path=r"C:\data\sub1.xlsx",
        join_defs=[{"item": "機器番号", "cell": "AS30"}],
        bundle={"join_values": {"機器番号": ["PT4400243777"]}},
    )
    assert pool[0]["機器番号"] == "PT4400243777"
    assert pool[1]["機器番号"] == "OTHER"


def test_patch_stacked_join_pool_skips_when_multiple_rows_share_host_file() -> None:
    """同一 __file_path の seed が複数あるとき比較列（機器番号）を上書きしない。"""
    from svc.svc_data_agg import _patch_stacked_join_pool_row_join_targets  # noqa: WPS433

    pool = [
        {
            "__file_path": r"C:\data\sub1.xlsx",
            "機器番号": "'PT4300312992",
            "__iter_index": 0,
        },
        {
            "__file_path": r"C:\data\sub1.xlsx",
            "機器番号": "'PT4300312993",
            "__iter_index": 1,
        },
    ]
    _patch_stacked_join_pool_row_join_targets(
        pool,
        file_path=r"C:\data\sub1.xlsx",
        join_defs=[{"item": "機器番号", "cell": "AS30"}],
        bundle={"join_values": {"機器番号": ["'PT4290015938"]}},
    )
    assert pool[0]["機器番号"] == "'PT4300312992"
    assert pool[1]["機器番号"] == "'PT4300312993"


def test_patch_stacked_join_pool_skips_same_host_single_distinct_value() -> None:
    """複数行が同一機器番号でも、複数行なら patch しない（一括同一化を防ぐ）。"""
    from svc.svc_data_agg import _patch_stacked_join_pool_row_join_targets  # noqa: WPS433

    pool = [
        {
            "__file_path": r"C:\data\sub1.xlsx",
            "機器番号": "OLD",
            "__iter_index": 0,
        },
        {
            "__file_path": r"C:\data\sub1.xlsx",
            "機器番号": "OLD",
            "__iter_index": 1,
        },
    ]
    _patch_stacked_join_pool_row_join_targets(
        pool,
        file_path=r"C:\data\sub1.xlsx",
        join_defs=[{"item": "機器番号", "cell": "AS30"}],
        bundle={"join_values": {"機器番号": ["NEW-PT"]}},
    )
    assert pool[0]["機器番号"] == "OLD"
    assert pool[1]["機器番号"] == "OLD"


def test_stacked_join_search_pool_not_filtered_by_host_file() -> None:
    from svc.svc_data_agg import _join_search_pool_scope  # noqa: WPS433

    pool = [
        {"__file_path": r"C:\host\a.xlsx", "機器番号": "1", "__iter_index": 0},
        {"__file_path": r"C:\host\b.xlsx", "機器番号": "2", "__iter_index": 1},
    ]
    scoped = _join_search_pool_scope(
        pool,
        r"C:\host\b.xlsx",
        False,
        stacked_join=True,
    )
    assert len(scoped) == 2
    narrow = _join_search_pool_scope(pool, r"C:\host\b.xlsx", False)
    assert len(narrow) == 1


def test_narrow_join_matched_rows_stacked_join_filters_by_host_file() -> None:
    from svc.svc_data_agg import _narrow_join_matched_rows_for_write  # noqa: WPS433

    rows = [
        {
            "__file_path": r"C:\host\a.xlsx",
            "機器番号": "PT-A",
            "MAC LOC": "",
            "__iter_index": 0,
        },
        {
            "__file_path": r"C:\host\b.xlsx",
            "機器番号": "PT-A",
            "MAC LOC": "",
            "__iter_index": 1,
        },
    ]
    narrowed = _narrow_join_matched_rows_for_write(
        rows,
        0,
        1,
        1,
        stacked_join=True,
        host_file_path=r"C:\host\b.xlsx",
    )
    assert len(narrowed) == 1
    assert narrowed[0]["__file_path"] == r"C:\host\b.xlsx"
    value_only = _narrow_join_matched_rows_for_write(
        rows,
        0,
        1,
        1,
        stacked_join=True,
        host_file_path=r"C:\host\b.xlsx",
        stacked_join_value_match_only=True,
    )
    assert len(value_only) == 2


def test_stacked_join_cell_write_uses_value_match_not_host_path() -> None:
    """table_rows seed のセル結合: __file_path 不一致でも join 比較列一致で書込み。"""
    from svc.data_agg_master_preview import table_rows_to_join_search_seed_pool  # noqa: WPS433
    from svc.svc_data_agg import _apply_join_key_search_write  # noqa: WPS433

    headers = ["品名", "機器番号", "MAC LOC"]
    host_a = r"C:\data\host_a.xlsm"
    host_b = r"C:\data\host_b.xlsm"
    pool = table_rows_to_join_search_seed_pool(
        headers,
        [["ユニット", "PT-A", None]],
        row_file_paths=[host_a],
        stacked_join=True,
    )
    item = {
        "name": "MAC LOC",
        "write_mode": "overwrite",
        "sources": [
            {
                "type": "cell",
                "ui_scenario_source_v1": {
                    "join_defs": [{"item": "機器番号", "cell": "AS30", "row": 0, "col": 0}],
                },
            }
        ],
    }
    bundle = {
        "primary_values": ["D8:4A:87:FF:B2:14"],
        "join_values": {"機器番号": ["PT-A"]},
    }
    _apply_join_key_search_write(
        pool,
        item,
        "MAC LOC",
        bundle,
        "overwrite",
        header_set=set(headers),
        search_pool=pool,
        stacked_join=True,
        host_file_path=host_b,
        stacked_join_value_match_only=True,
    )
    from core.core_join_compare import join_compare_display_key  # noqa: WPS433

    assert join_compare_display_key(pool[0].get("MAC LOC")) == "D8:4A:87:FF:B2:14"


def test_stacked_join_write_requires_matching_host_file_path_without_value_only() -> None:
    """積み上げ join（path 絞り込みあり）: __file_path がホストと一致しないと書込みされない。"""
    from svc.data_agg_master_preview import table_rows_to_join_search_seed_pool  # noqa: WPS433
    from svc.svc_data_agg import _apply_join_key_search_write  # noqa: WPS433

    headers = ["品名", "機器番号", "MAC LOC"]
    host_a = r"C:\data\host_a.xlsm"
    host_b = r"C:\data\host_b.xlsm"
    pool = table_rows_to_join_search_seed_pool(
        headers,
        [["ユニット", "PT-A", None]],
        row_file_paths=[host_a],
        stacked_join=True,
    )
    item = {
        "name": "MAC LOC",
        "write_mode": "overwrite",
        "sources": [
            {
                "type": "cell",
                "ui_scenario_source_v1": {
                    "join_defs": [{"item": "機器番号", "cell": "AS30", "row": 0, "col": 0}],
                },
            }
        ],
    }
    bundle = {
        "primary_values": ["D8:4A:87:FF:B2:14"],
        "join_values": {"機器番号": ["PT-A"]},
    }
    _apply_join_key_search_write(
        pool,
        item,
        "MAC LOC",
        bundle,
        "overwrite",
        header_set=set(headers),
        search_pool=pool,
        stacked_join=True,
        host_file_path=host_b,
        stacked_join_value_match_only=False,
    )
    assert pool[0].get("MAC LOC") in (None, "")

    pool[0]["MAC LOC"] = None
    _apply_join_key_search_write(
        pool,
        item,
        "MAC LOC",
        bundle,
        "overwrite",
        header_set=set(headers),
        search_pool=pool,
        stacked_join=True,
        host_file_path=host_a,
        stacked_join_value_match_only=False,
    )
    from core.core_join_compare import join_compare_display_key  # noqa: WPS433

    assert join_compare_display_key(pool[0].get("MAC LOC")) == "D8:4A:87:FF:B2:14"


def test_row_file_paths_real_count_prefers_iteration_paths() -> None:
    from svc.data_agg_master_preview import (  # noqa: WPS433
        is_synthetic_mpv_row_file_path,
        row_file_paths_real_count,
    )

    assert is_synthetic_mpv_row_file_path("mpv_table_seed://0")
    real = [r"C:\a\f1.xlsm", r"C:\a\f2.xlsm"]
    synth = ["mpv_table_seed://0", "mpv_table_seed://1"]
    assert row_file_paths_real_count(real) == 2
    assert row_file_paths_real_count(synth) == 0
    assert row_file_paths_real_count(real) > row_file_paths_real_count(synth)


def test_cross_detection_uses_topology_not_empty_stepped_sources(
    tmp_path: Path,
) -> None:
    import copy

    from svc.svc_data_agg import (  # noqa: WPS433
        _join_host_needs_cross_file_pool,
    )

    data, _paths = _cross_join_mini_scenario(tmp_path)
    headers = [str(it.get("name") or "") for it in data["items"]]
    stepped = copy.deepcopy(data["items"])
    for it in stepped:
        it["sources"] = []
    host_topo = data["items"][2]
    host_stepped = stepped[2]
    assert not _join_host_needs_cross_file_pool(host_stepped, stepped, headers)
    assert _join_host_needs_cross_file_pool(host_topo, data["items"], headers)


def test_carry_forward_step0_topology_cross_join_pt(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    """step0（n_pick=0）でも topology により横断 join が成立する。"""
    import copy

    from svc.data_agg_master_preview import (  # noqa: WPS433
        MASTER_PREVIEW_DIAG_SOURCE,
        scenario_for_stepped_preview,
        table_rows_to_join_search_seed_pool,
    )
    from svc.svc_data_agg import master_preview_extract_item_allowlist  # noqa: WPS433

    data, paths = _cross_join_mini_scenario(tmp_path)
    headers = [str(it.get("name") or "") for it in data["items"]]
    anchor, join_f = paths[0], paths[1]
    prior_rows = [["DEV1", "MAC-A", None, None]]
    scen = scenario_for_stepped_preview(
        data,
        mi_idx=2,
        master_step_idx=0,
        active_slot_indices=[0],
        carry_forward_completed_items=True,
    )
    dd = scen["__debug_diag"]
    dd["preview_join_topology_items"] = copy.deepcopy(data["items"])
    dd["source"] = MASTER_PREVIEW_DIAG_SOURCE
    dd["master_preview_join_read_full_files"] = True
    allow_ix = master_preview_extract_item_allowlist(data, mi_idx=2)
    if allow_ix is not None:
        dd["preview_extract_item_allowlist"] = list(allow_ix)
    dd["join_search_seed_pool"] = table_rows_to_join_search_seed_pool(
        headers,
        prior_rows,
        anchor_file_path=anchor,
    )
    monkeypatch.setenv("DATA_AGG_FILE_PARALLEL_WORKERS", "0")
    monkeypatch.setenv("DATA_AGG_MASTER_PARALLEL_EXTRACT", "0")
    out_headers, out_rows, _ev, _je = compute_batch_table_rows(
        scen,
        [join_f],
        max_primary_rows=10,
        max_table_rows=10,
        probe_caller="test_step0_topology",
    )
    idx_pt = out_headers.index("PT番号")
    pt_vals = [r[idx_pt] for r in out_rows if r[idx_pt] not in (None, "")]
    assert pt_vals, "step0 + topology でも PT が空"
    assert join_compare_display_key(pt_vals[0]) == "PT-001"


def test_master_preview_join_item_effective_restores_defs_at_step0(
    tmp_path: Path,
) -> None:
    """step0 で stepped sources が空でも join_defs を topology から復元する。"""
    import copy

    from svc.data_agg_master_preview import (  # noqa: WPS433
        scenario_for_stepped_preview,
    )
    from svc.svc_data_agg import (  # noqa: WPS433
        _item_join_defs_list,
        _master_preview_join_item_effective,
    )

    data, _paths = _cross_join_mini_scenario(tmp_path)
    scen = scenario_for_stepped_preview(
        data,
        mi_idx=2,
        master_step_idx=0,
        active_slot_indices=[0],
        carry_forward_completed_items=True,
    )
    items = list(scen.get("items") or [])
    dd = scen.get("__debug_diag") or {}
    dd["preview_join_topology_items"] = copy.deepcopy(data["items"])
    stepped_host = items[2]
    assert not _item_join_defs_list(stepped_host)
    eff = _master_preview_join_item_effective(
        items,
        2,
        dd,
        preview_master_mode=True,
    )
    assert _item_join_defs_list(eff)
    assert _item_join_defs_list(eff) == _item_join_defs_list(data["items"][2])


def test_reorder_paths_when_stepped_host_has_no_join_defs(
    tmp_path: Path,
) -> None:
    """step0 の結合ホスト（sources 空）でも topology で path 優先が効く。"""
    import copy

    from svc.data_agg_master_preview import scenario_for_stepped_preview  # noqa: WPS433

    data, paths = _cross_join_mini_scenario(tmp_path)
    scen = scenario_for_stepped_preview(
        data,
        mi_idx=2,
        master_step_idx=0,
        active_slot_indices=[0],
        carry_forward_completed_items=True,
    )
    items = list(scen.get("items") or [])
    headers = [str(it.get("name") or "") for it in data["items"]]
    dd = scen.get("__debug_diag") or {}
    dd["mi_idx"] = 2
    dd["preview_join_topology_items"] = copy.deepcopy(data["items"])
    ordered = reorder_paths_for_master_preview_join_priority(
        [paths[1], paths[0]], items, headers, dd
    )
    assert ordered[0] == paths[0]
    assert dd.get("master_preview_join_host_patterns")


def test_filter_master_preview_topology_fallback_when_stepped_empty(
    tmp_path: Path,
) -> None:
    """carry-forward で stepped sources が全空でも topology で path を復元する。"""
    import copy

    from svc.data_agg_master_preview import MASTER_PREVIEW_DIAG_SOURCE  # noqa: WPS433

    data, paths = _cross_join_mini_scenario(tmp_path)
    stepped = copy.deepcopy(data["items"])
    for it in stepped:
        it["sources"] = []
    files = [str(paths[0]), str(paths[1]), r"C:\data\other.xlsx"]
    dd = {
        "source": MASTER_PREVIEW_DIAG_SOURCE,
        "mi_idx": 2,
        "preview_join_topology_items": copy.deepcopy(data["items"]),
        "preview_extract_item_allowlist": [0, 1, 2],
    }
    out = filter_file_paths_for_master_preview(files, stepped, dd)
    assert str(paths[0]) in out
    assert str(paths[1]) in out
    assert r"C:\data\other.xlsx" not in out


def test_filter_master_preview_cross_join_prefers_topology_over_single_stepped(
    tmp_path: Path,
) -> None:
    """横断 join: stepped がホスト 1 パターンだけでも topology で side を含める。"""
    import copy

    from svc.data_agg_master_preview import MASTER_PREVIEW_DIAG_SOURCE  # noqa: WPS433

    data, paths = _cross_join_mini_scenario(tmp_path)
    stepped = copy.deepcopy(data["items"])
    for j, it in enumerate(stepped):
        if j < 2:
            it["sources"] = []
    files = [str(paths[0]), str(paths[1])]
    dd = {
        "source": MASTER_PREVIEW_DIAG_SOURCE,
        "mi_idx": 2,
        "preview_join_topology_items": copy.deepcopy(data["items"]),
        "preview_extract_item_allowlist": [0, 1, 2],
    }
    out = filter_file_paths_for_master_preview(files, stepped, dd)
    assert len(out) == 2
    assert set(out) == {str(paths[0]), str(paths[1])}


def test_master_preview_pool_cap_applies_with_join_full_read(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    """join full read 有効時もプール行 cap が効く（巨大 pool で join 列が空になるのを防ぐ）。"""
    import copy

    from svc.data_agg_master_preview import (  # noqa: WPS433
        MASTER_PREVIEW_DIAG_SOURCE,
        scenario_for_stepped_preview,
        table_rows_to_join_search_seed_pool,
    )

    data, paths = _cross_join_mini_scenario(tmp_path)
    headers = [str(it.get("name") or "") for it in data["items"]]
    scen = scenario_for_stepped_preview(
        data,
        mi_idx=2,
        master_step_idx=1,
        active_slot_indices=[0],
        carry_forward_completed_items=True,
    )
    dd = scen["__debug_diag"]
    dd["source"] = MASTER_PREVIEW_DIAG_SOURCE
    dd["preview_join_topology_items"] = copy.deepcopy(data["items"])
    dd["master_preview_join_read_full_files"] = True
    dd["join_search_seed_pool"] = table_rows_to_join_search_seed_pool(
        headers, [["DEV1", "MAC-A", None, None]]
    )
    monkeypatch.setenv("DATA_AGG_FILE_PARALLEL_WORKERS", "0")
    monkeypatch.setenv("DATA_AGG_MASTER_PARALLEL_EXTRACT", "0")
    _h, _rows, _ev, _je = compute_batch_table_rows(
        scen,
        paths,
        max_primary_rows=4,
        max_table_rows=4,
        probe_caller="test_pool_cap_join_full_read",
    )
    cap = int(dd.get("master_preview_pool_row_cap") or 0)
    assert cap > 0
    pool_rows = int(dd.get("master_preview_pool_rows") or 0)
    assert pool_rows <= cap


def test_filter_master_preview_stacked_join_host_only(
    tmp_path: Path,
) -> None:
    """積み上げ join: path filter は当ステップのホストファイルのみ。"""
    import copy

    from svc.data_agg_master_preview import MASTER_PREVIEW_DIAG_SOURCE  # noqa: WPS433

    data, paths = _cross_join_mini_scenario(tmp_path)
    stepped = copy.deepcopy(data["items"])
    for j, it in enumerate(stepped):
        if j < 2:
            it["sources"] = []
    files = [str(paths[0]), str(paths[1])]
    dd = {
        "source": MASTER_PREVIEW_DIAG_SOURCE,
        "mi_idx": 2,
        "preview_join_topology_items": copy.deepcopy(data["items"]),
        "master_preview_stacked_join": True,
        "join_search_seed_from_table_rows": True,
        "join_search_seed_pool": [{"機器番号": "DEV1"}],
    }
    out = filter_file_paths_for_master_preview(files, stepped, dd)
    assert out == [str(paths[1])]


def test_stacked_join_cross_join_pt_host_only(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    """積み上げ join: 前項目表示行 + ホストのみ読込で PT・製番が埋まる。"""
    import copy

    from svc.data_agg_master_preview import (  # noqa: WPS433
        MASTER_PREVIEW_DIAG_SOURCE,
        scenario_for_stepped_preview,
        table_rows_to_join_search_seed_pool,
    )

    data, paths = _cross_join_mini_scenario(tmp_path)
    anchor, join_f = paths[0], paths[1]
    headers = [str(it.get("name") or "") for it in data["items"]]
    prior_rows = [["DEV1", "MAC-A", None, None]]
    scen = scenario_for_stepped_preview(
        data,
        mi_idx=2,
        master_step_idx=1,
        active_slot_indices=[0],
        carry_forward_completed_items=True,
    )
    dd = scen["__debug_diag"]
    dd["source"] = MASTER_PREVIEW_DIAG_SOURCE
    dd["mi_idx"] = 2
    dd["preview_join_topology_items"] = copy.deepcopy(data["items"])
    seed_pool = table_rows_to_join_search_seed_pool(headers, prior_rows)
    dd["join_search_seed_pool"] = seed_pool
    dd["join_search_seed_from_table_rows"] = True
    dd["master_preview_stacked_join"] = True
    dd["preview_extract_item_allowlist"] = [2]
    dd["master_preview_join_read_full_files"] = False
    dd["preview_anchor_row_keys"] = [
        [str(seed_pool[0]["__file_path"]), int(seed_pool[0]["__iter_index"])]
    ]
    monkeypatch.setenv("DATA_AGG_FILE_PARALLEL_WORKERS", "0")
    monkeypatch.setenv("DATA_AGG_MASTER_PARALLEL_EXTRACT", "0")
    out_headers, out_rows, _ev, _je = compute_batch_table_rows(
        scen,
        [anchor, join_f],
        max_primary_rows=10,
        max_table_rows=10,
        probe_caller="test_stacked_join_pt",
    )
    assert int(dd.get("master_preview_stats_files_read") or 0) <= 1
    idx_pt = out_headers.index("PT番号")
    idx_seq = out_headers.index("製番")
    pt_vals = [r[idx_pt] for r in out_rows if r[idx_pt] not in (None, "")]
    seq_vals = [r[idx_seq] for r in out_rows if r[idx_seq] not in (None, "")]
    assert pt_vals, "積み上げ join で PT が空"
    assert seq_vals, "積み上げ join で製番が空"
    assert join_compare_display_key(pt_vals[0]) == "PT-001"


def test_apply_master_preview_join_max_files_caps_join_item(
    tmp_path: Path,
) -> None:
    """結合項目のみ filter/reorder 後の paths を MASTER_DEBUG_JOIN_MAX_FILES 相当で打切る。"""
    from svc.data_agg_master_preview_perf import (  # noqa: WPS433
        apply_master_preview_join_max_files,
        master_preview_join_max_files_cap,
        master_preview_should_apply_join_file_cap,
    )

    data, paths = _cross_join_mini_scenario(tmp_path)
    paths_many = [paths[0]] * 15 + [paths[1]] * 15
    dd: dict = {
        "source": "ui_data_agg_debug.master_preview",
        "mi_idx": 2,
        "master_preview_join_max_files": 20,
    }
    assert master_preview_join_max_files_cap(dd) == 20
    assert master_preview_should_apply_join_file_cap(data["items"], 2)
    assert not master_preview_should_apply_join_file_cap(data["items"], 0)
    capped = apply_master_preview_join_max_files(paths_many, data["items"], dd)
    assert len(capped) == 20
    assert dd["master_preview_join_file_cap_hit"] is True
    assert dd["master_preview_join_files_detected"] == 30
    assert dd["master_preview_join_files_read"] == 20


def test_apply_master_preview_join_max_files_skips_non_join_item(
    tmp_path: Path,
) -> None:
    data, paths = _cross_join_mini_scenario(tmp_path)
    paths_many = paths * 15
    dd: dict = {
        "source": "ui_data_agg_debug.master_preview",
        "mi_idx": 0,
        "master_preview_join_max_files": 20,
    }
    from svc.data_agg_master_preview_perf import apply_master_preview_join_max_files  # noqa: WPS433

    capped = apply_master_preview_join_max_files(paths_many, data["items"], dd)
    assert len(capped) == len(paths_many)
    assert "master_preview_join_file_cap_hit" not in dd


def test_master_preview_join_file_cap_limits_compute_files(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    """compute_batch 経路でも結合項目の読込ファイル数が上限以下になる。"""
    data, paths = _cross_join_mini_scenario(tmp_path)
    paths_many = [paths[0]] * 10 + [paths[1]] * 10
    data["__debug_diag"] = {
        "enabled": True,
        "source": "ui_data_agg_debug.master_preview",
        "mi_idx": 2,
        "master_preview_join_max_files": 1,
        "join_search_skip_seed": True,
    }
    monkeypatch.setenv("DATA_AGG_FILE_PARALLEL_WORKERS", "0")
    monkeypatch.setenv("DATA_AGG_MASTER_PARALLEL_EXTRACT", "0")
    compute_batch_table_rows(
        data,
        paths_many,
        max_primary_rows=4,
        max_table_rows=4,
        probe_caller="test_join_file_cap",
    )
    dd = data.get("__debug_diag") or {}
    assert dd.get("master_preview_join_file_cap_hit") is True
    assert int(dd.get("master_preview_join_files_read") or 0) == 1
    assert int(dd.get("master_preview_stats_files_read") or 0) <= 1


def test_stacked_seed_join_targets_fill_ratio() -> None:
    from svc.data_agg_master_preview_perf import (  # noqa: WPS433
        master_preview_stacked_seed_join_targets_fill_ratio,
        master_preview_stacked_seed_usable,
    )

    headers = ["品名", "機器番号", "MAC LOC"]
    good = [["A", "D1", None], ["B", "D2", None]]
    bad = [["A", None, None], ["B", None, None]]
    assert master_preview_stacked_seed_join_targets_fill_ratio(
        good, headers, ["機器番号"]
    ) == 1.0
    assert master_preview_stacked_seed_join_targets_fill_ratio(
        bad, headers, ["機器番号"]
    ) == 0.0
    assert master_preview_stacked_seed_usable(good, headers, ["機器番号"])
    assert not master_preview_stacked_seed_usable(bad, headers, ["機器番号"])


def test_stacked_join_odn375_like_same_file(tmp_path: Path) -> None:
    """ODN375 型: 品名表示行 seed + 積み上げ join で MAC LOC が埋まる。"""
    from openpyxl import Workbook

    from svc.data_agg_master_preview import (  # noqa: WPS433
        MASTER_PREVIEW_DIAG_SOURCE,
        scenario_for_stepped_preview,
        table_rows_to_join_search_seed_pool,
    )

    p = tmp_path / "ODN375_sample.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "ユニット実装チェック"
    ws["P14"] = "UNIT-A"
    ws["P16"] = "DEV-001"
    ws["AS30"] = "DEV-001"
    ws["AV30"] = "AA:BB:CC:DD:EE:FF"
    wb.save(p)

    data: dict = {
        "items": [
            {
                "name": "品名",
                "write_mode": "append",
                "sources": [
                    {
                        "type": "cell",
                        "sheet_name": "ユニット実装チェック",
                        "cell_ref": "P14",
                        "ui_scenario_source_v1": {
                            "file_pattern": "ODN375",
                            "link_defs": [
                                {
                                    "item": "機器番号",
                                    "cell": "P16",
                                    "mode": "セル座標",
                                    "row": 0,
                                    "col": 0,
                                }
                            ],
                        },
                        "repeat_direction": "vertical",
                        "repeat_max": 1,
                    }
                ],
            },
            {"name": "機器番号", "write_mode": "append", "sources": []},
            {"name": "MAC", "write_mode": "append", "sources": []},
            {
                "name": "MAC LOC",
                "write_mode": "overwrite",
                "sources": [
                    {
                        "type": "cell",
                        "sheet_name": "ユニット実装チェック",
                        "cell_ref": "AV30",
                        "ui_scenario_source_v1": {
                            "file_pattern": "ODN375",
                            "link_defs": [
                                {
                                    "item": "MAC",
                                    "cell": "",
                                    "mode": "固定値",
                                    "row": 0,
                                    "col": 0,
                                }
                            ],
                            "join_defs": [
                                {"item": "機器番号", "cell": "AS30", "row": 0, "col": 0}
                            ],
                        },
                        "repeat_direction": "vertical",
                        "repeat_max": 1,
                    }
                ],
            },
        ]
    }
    headers = [str(it.get("name") or "") for it in data["items"]]
    scen0 = scenario_for_stepped_preview(
        data, mi_idx=0, master_step_idx=1, active_slot_indices=[0]
    )
    _, prior_rows, _, _ = compute_batch_table_rows(
        scen0, [str(p)], max_primary_rows=10, max_table_rows=10
    )
    assert prior_rows
    ix_dev = headers.index("機器番号")
    assert prior_rows[0][ix_dev] not in (None, "")

    scen = scenario_for_stepped_preview(
        data,
        mi_idx=3,
        master_step_idx=1,
        active_slot_indices=[0],
        carry_forward_completed_items=True,
    )
    dd = scen["__debug_diag"]
    dd["source"] = MASTER_PREVIEW_DIAG_SOURCE
    dd["mi_idx"] = 3
    dd["preview_join_topology_items"] = data["items"]
    seed = table_rows_to_join_search_seed_pool(
        headers,
        prior_rows,
        row_file_paths=[str(p)],
        stacked_join=True,
    )
    dd["join_search_seed_pool"] = seed
    dd["join_search_seed_from_table_rows"] = True
    dd["master_preview_stacked_join"] = True
    dd["preview_extract_item_allowlist"] = [3]
    dd["master_preview_join_read_full_files"] = False
    _, rows, _, _ = compute_batch_table_rows(
        scen, [str(p)], max_primary_rows=10, max_table_rows=10, probe_caller="test_odn375"
    )
    ix_loc = headers.index("MAC LOC")
    ix_mac = headers.index("MAC")
    assert rows[0][ix_loc] not in (None, "")
    assert rows[0][ix_mac] in (None, "")


def test_stacked_join_seed_join_matches_as30_without_patch(
    tmp_path: Path,
) -> None:
    """積み上げ join: 前段 seed の機器番号を書き換えず AS30 と join_compare で結合する。"""
    from openpyxl import Workbook

    from svc.data_agg_master_preview import (  # noqa: WPS433
        MASTER_PREVIEW_DIAG_SOURCE,
        scenario_for_stepped_preview,
        table_rows_to_join_search_seed_pool,
    )

    p = tmp_path / "ODN375_pt_match.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "ユニット実装チェック"
    ws["P14"] = "UNIT-A"
    ws["P16"] = "PT4400243759"
    ws["AS30"] = "PT4400243759"
    ws["AV30"] = "D8:4A:87:FF:C3:58"
    wb.save(p)

    data: dict = {
        "items": [
            {
                "name": "品名",
                "write_mode": "append",
                "sources": [
                    {
                        "type": "cell",
                        "sheet_name": "ユニット実装チェック",
                        "cell_ref": "P14",
                        "ui_scenario_source_v1": {
                            "file_pattern": "ODN375",
                            "link_defs": [
                                {
                                    "item": "機器番号",
                                    "cell": "P16",
                                    "mode": "セル座標",
                                    "row": 0,
                                    "col": 0,
                                }
                            ],
                        },
                        "repeat_direction": "vertical",
                        "repeat_max": 1,
                    }
                ],
            },
            {"name": "機器番号", "write_mode": "append", "sources": []},
            {"name": "MAC", "write_mode": "append", "sources": []},
            {
                "name": "MAC LOC",
                "write_mode": "overwrite",
                "sources": [
                    {
                        "type": "cell",
                        "sheet_name": "ユニット実装チェック",
                        "cell_ref": "AV30",
                        "ui_scenario_source_v1": {
                            "file_pattern": "ODN375",
                            "link_defs": [
                                {
                                    "item": "MAC",
                                    "cell": "",
                                    "mode": "固定値",
                                    "row": 0,
                                    "col": 0,
                                }
                            ],
                            "join_defs": [
                                {"item": "機器番号", "cell": "AS30", "row": 0, "col": 0}
                            ],
                        },
                        "repeat_direction": "vertical",
                        "repeat_max": 1,
                    }
                ],
            },
        ]
    }
    headers = [str(it.get("name") or "") for it in data["items"]]
    prior_rows = [["UNIT-A", "PT4400243759", None, None]]
    scen = scenario_for_stepped_preview(
        data,
        mi_idx=3,
        master_step_idx=1,
        active_slot_indices=[0],
        carry_forward_completed_items=True,
    )
    dd = scen["__debug_diag"]
    dd["source"] = MASTER_PREVIEW_DIAG_SOURCE
    dd["mi_idx"] = 3
    dd["preview_join_topology_items"] = data["items"]
    seed = table_rows_to_join_search_seed_pool(
        headers,
        prior_rows,
        row_file_paths=[str(p)],
        stacked_join=True,
    )
    dd["join_search_seed_pool"] = seed
    dd["join_search_seed_from_table_rows"] = True
    dd["master_preview_stacked_join"] = True
    dd["preview_extract_item_allowlist"] = [3]
    dd["master_preview_join_read_full_files"] = False
    _, rows, _, _ = compute_batch_table_rows(
        scen,
        [str(p)],
        max_primary_rows=10,
        max_table_rows=10,
        probe_caller="test_stacked_seed_patch",
    )
    ix_loc = headers.index("MAC LOC")
    from core.core_join_compare import join_compare_display_key  # noqa: WPS433

    assert join_compare_display_key(rows[0][ix_loc]) == "D8:4A:87:FF:C3:58"
    ix_dev = headers.index("機器番号")
    assert join_compare_display_key(rows[0][ix_dev]) == "PT4400243759"


def test_stacked_join_mac_loc_preserves_distinct_device_ids() -> None:
    """積み上げ join: 行ごとに異なる機器番号を AS30 一括上書きせず、一致行のみ MAC LOC。"""
    from svc.data_agg_master_preview import table_rows_to_join_search_seed_pool  # noqa: WPS433
    from svc.svc_data_agg import (  # noqa: WPS433
        _apply_join_key_search_write,
        _patch_stacked_join_pool_row_join_targets,
    )

    headers = ["品名", "機器番号", "MAC", "MAC LOC"]
    prior_rows = [
        ["ユニット", "'PT4400243770", None, None],
        ["IMX-VCH", "'PT4400243777", None, None],
        ["IMX-REP", "'PT4400243777 '1", None, None],
    ]
    fp = r"C:\data\sub1.xlsx"
    pool = table_rows_to_join_search_seed_pool(
        headers,
        prior_rows,
        row_file_paths=[fp, fp, fp],
        stacked_join=True,
    )
    item = {
        "name": "MAC LOC",
        "write_mode": "overwrite",
        "sources": [
            {
                "type": "cell",
                "ui_scenario_source_v1": {
                    "join_defs": [{"item": "機器番号", "cell": "AS30", "row": 0, "col": 0}],
                },
            }
        ],
    }
    bundle = {
        "primary_values": ["D8:4A:87:FF:C3:57"],
        "join_values": {"機器番号": ["'PT4400243777"]},
    }
    _patch_stacked_join_pool_row_join_targets(
        pool,
        file_path=fp,
        join_defs=[{"item": "機器番号", "cell": "AS30"}],
        bundle=bundle,
    )
    _apply_join_key_search_write(
        pool,
        item,
        "MAC LOC",
        bundle,
        "overwrite",
        header_set=set(headers),
        search_pool=pool,
    )
    from core.core_join_compare import join_compare_display_key  # noqa: WPS433

    assert join_compare_display_key(pool[0]["機器番号"]) == "PT4400243770"
    assert join_compare_display_key(pool[1]["機器番号"]) == "PT4400243777"
    assert join_compare_display_key(pool[2]["機器番号"]) == "PT4400243777 '1"
    assert join_compare_display_key(pool[0].get("MAC LOC")) == ""
    assert join_compare_display_key(pool[1].get("MAC LOC")) == "D8:4A:87:FF:C3:57"
    assert join_compare_display_key(pool[2].get("MAC LOC")) == ""
