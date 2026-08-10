# -*- coding: utf-8 -*-
"""path_item 付き名前取得が余分行を作らない回帰。"""
from __future__ import annotations

import sys
from pathlib import Path
from typing import Any

from openpyxl import Workbook

_root = Path(__file__).resolve().parents[1]
if str(_root) not in sys.path:
    sys.path.insert(0, str(_root))

from svc.svc_data_agg import (  # noqa: E402
    _name_extract_item_emits_own_rows,
    compute_batch_table_rows,
)
from svc.svc_data_agg_extract import extract_item_values, xlsx_workbook_scope  # noqa: E402


def _ws(wb: Any) -> Any:
    ws = wb.active
    assert ws is not None
    return ws


def test_name_extract_search_miss_does_not_pad_none(tmp_path: Path) -> None:
    fp = tmp_path / "938B_unit.xlsx"
    wb = Workbook()
    wb.save(fp)
    wb.close()
    item = {
        "name": "装置タイプ",
        "sources": [
            {
                "type": "name_extract",
                "source_type": "file_name",
                "search_condition": "exclude",
                "search_text": "B",
                "ui_scenario_source_v1": {"extract_mode": "fixed", "path_item": "機器番号"},
                "length_value": "ODN-938",
            },
            {
                "type": "name_extract",
                "source_type": "file_name",
                "search_condition": "include",
                "search_text": "B",
                "ui_scenario_source_v1": {"extract_mode": "fixed", "path_item": "機器番号"},
                "length_value": "ODN-938B",
            },
        ],
    }
    with xlsx_workbook_scope():
        vals = extract_item_values(str(fp), item, item_id="t")
    assert [str(x).lstrip("'") for x in vals] == ["ODN-938B"]


def test_name_extract_with_path_item_does_not_emit_own_rows() -> None:
    it = {
        "sources": [
            {
                "type": "name_extract",
                "ui_scenario_source_v1": {"path_item": "機器番号"},
            }
        ]
    }
    assert _name_extract_item_emits_own_rows(it) is False
    it2 = {
        "sources": [
            {
                "type": "name_extract",
                "ui_scenario_source_v1": {"path_item": ""},
            }
        ]
    }
    assert _name_extract_item_emits_own_rows(it2) is True


def test_batch_no_orphan_device_type_only_rows(tmp_path: Path) -> None:
    """
    座標取得（機器番号）＋ path_item 付き名前取得（B含む→ODN-938B）で、
    装置タイプだけの余分行が table に出ないこと。
    """
    fp = tmp_path / "hist_938B_card.xlsx"
    wb = Workbook()
    ws = _ws(wb)
    ws.title = "S"
    ws["A1"] = "EQ1"
    ws["A2"] = "EQ2"
    wb.save(fp)
    wb.close()

    data = {
        "id": "odn938_orphan",
        "items": [
            {
                "id": "item_eq",
                "name": "機器番号",
                "write_mode": "append",
                "sources": [
                    {
                        "type": "cell",
                        "sheet_name": "S",
                        "cell_ref": "A1",
                        "row_offset": 1,
                        "col_offset": 0,
                        "repeat_direction": "vertical",
                        "repeat_until_empty": False,
                        "repeat_max": 2,
                        "ui_scenario_source_v1": {
                            "file_pattern": "",
                            "file_name_rule": "含む",
                            "sheet_rule": "左端シート",
                        },
                    }
                ],
            },
            {
                "id": "item_type",
                "name": "装置タイプ",
                "write_mode": "overwrite",
                "sources": [
                    {
                        "type": "name_extract",
                        "source_type": "file_name",
                        "search_condition": "exclude",
                        "search_text": "B",
                        "length_value": "ODN-938",
                        "ui_scenario_source_v1": {
                            "extract_mode": "fixed",
                            "path_item": "機器番号",
                        },
                    },
                    {
                        "type": "name_extract",
                        "source_type": "file_name",
                        "search_condition": "include",
                        "search_text": "B",
                        "length_value": "ODN-938B",
                        "ui_scenario_source_v1": {
                            "extract_mode": "fixed",
                            "path_item": "機器番号",
                        },
                    },
                ],
            },
        ],
        "match_keys": [],
        "excel_options": {"output_target": "active_sheet", "write_mode": "append"},
    }
    headers, rows, _, _ = compute_batch_table_rows(
        data,
        [str(fp)],
        max_table_rows=50,
        probe_caller="excel_batch_submit",
    )
    assert headers[:2] == ["機器番号", "装置タイプ"]
    assert len(rows) == 2
    for row in rows:
        eq = str(row[0]).lstrip("'") if row[0] is not None else ""
        typ = str(row[1]).lstrip("'") if row[1] is not None else ""
        assert eq in ("EQ1", "EQ2")
        assert typ == "ODN-938B"
    # 装置タイプだけ埋まった行がない
    assert not any(
        (str(r[0]).strip() if r[0] is not None else "") == ""
        and (str(r[1]).lstrip("'") if r[1] is not None else "") == "ODN-938B"
        for r in rows
    )
