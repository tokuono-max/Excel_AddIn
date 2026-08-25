# -*- coding: utf-8 -*-
"""同一項目の複数ソース（ユニット行＋カード行）で、カード専用連携がユニット行に載らないこと。"""
from __future__ import annotations

import sys
from pathlib import Path

_root = Path(__file__).resolve().parents[1]
if str(_root) not in sys.path:
    sys.path.insert(0, str(_root))

from openpyxl import Workbook  # noqa: E402

from svc.svc_data_agg import _assign_series_to_rows_by_context  # noqa: E402
from svc.svc_data_agg_extract import extract_item_bundle, xlsx_workbook_scope  # noqa: E402


def _txt(v: object) -> str:
    if v is None:
        return ""
    return str(v).lstrip("'")


def _unit_card_item(*, skip_empty: bool, sheet_name: str = "S", sheet_rule: str | None = None) -> dict:
    ui_unit: dict = {
        "link_defs": [
            {
                "item": "品名",
                "mode": "固定値",
                "cell": "ユニット",
                "row": 0,
                "col": 0,
            },
            {
                "item": "PT",
                "mode": "セル座標",
                "cell": "B1",
                "row": 0,
                "col": 0,
            },
        ],
    }
    ui_card: dict = {
        "link_defs": [
            {
                "item": "品名",
                "mode": "セル座標",
                "cell": "B2",
                "row": 1,
                "col": 0,
            },
            {
                "item": "SYS",
                "mode": "セル座標",
                "cell": "C2",
                "row": 1,
                "col": 0,
                "carry_empty": True,
            },
        ],
        "join_defs": [
            {
                "item": "結合",
                "mode": "セル座標",
                "cell": "D2",
                "row": 1,
                "col": 0,
            }
        ],
    }
    if sheet_rule:
        ui_unit["sheet_rule"] = sheet_rule
        ui_card["sheet_rule"] = sheet_rule
    return {
        "name": "機器番号",
        "sources": [
            {
                "type": "cell",
                "sheet_name": sheet_name,
                "cell_ref": "A1",
                "row_offset": 0,
                "col_offset": 0,
                "repeat_direction": "vertical",
                "repeat_until_empty": False,
                "repeat_max": 1,
                "skip_empty_primary": skip_empty,
                "skip_primary_match": ",-",
                "ui_scenario_source_v1": ui_unit,
            },
            {
                "type": "cell",
                "sheet_name": sheet_name,
                "cell_ref": "A2",
                "row_offset": 1,
                "col_offset": 0,
                "repeat_direction": "vertical",
                "repeat_until_empty": False,
                "repeat_max": 3,
                "skip_empty_primary": skip_empty,
                "skip_primary_match": ",-",
                "ui_scenario_source_v1": ui_card,
            },
        ],
    }


def _assign_col(b: dict, fp: str, col: str) -> list[str]:
    rows = [
        {
            "機器番号": v,
            "__file_path": fp,
            "__iter_index": int(i),
        }
        for i, v in enumerate(b.get("primary_values") or [])
    ]
    _assign_series_to_rows_by_context(
        rows,
        col,
        list((b.get("link_values") or {}).get(col) or []),
        list((b.get("link_contexts") or {}).get(col) or []),
        fp,
        write_mode="append",
    )
    return [_txt(r.get(col)) for r in rows]


def test_card_only_link_stays_off_unit_row_when_skip_drops_blank(tmp_path: Path) -> None:
    """カード主キーの空行スキップ後も、SYS / 結合はユニット行に載らない。"""
    wb = Workbook()
    ws = wb.active
    ws.title = "S"
    ws["A1"] = "UNIT"
    ws["B1"] = "UNIT_PT"
    ws["A2"] = "C1"
    ws["B2"] = "CARD1"
    ws["C2"] = "SYS1"
    ws["D2"] = "J1"
    ws["A3"] = None
    ws["B3"] = "SKIPPED"
    ws["C3"] = "SYS_SKIP"
    ws["D3"] = "J_SKIP"
    ws["A4"] = "C2"
    ws["B4"] = "CARD2"
    ws["C4"] = "SYS2"
    ws["D4"] = "J2"
    fp = tmp_path / "unit_card_skip.xlsx"
    wb.save(fp)
    wb.close()
    item = _unit_card_item(skip_empty=True)
    with xlsx_workbook_scope():
        b = extract_item_bundle(str(fp), item, item_id="i1")
    prim = [_txt(x) for x in (b.get("primary_values") or [])]
    assert prim == ["UNIT", "C1", "C2"]
    assert _assign_col(b, str(fp), "品名") == ["ユニット", "CARD1", "CARD2"]
    assert _assign_col(b, str(fp), "PT") == ["UNIT_PT", "", ""]
    assert _assign_col(b, str(fp), "SYS") == ["", "SYS1", "SYS2"]
    joins = [_txt(x) for x in (b.get("join_values") or {}).get("結合") or []]
    assert joins == ["", "J1", "J2"]
    jctx = (b.get("join_contexts") or {}).get("結合") or []
    assert [c.get("iter_index") for c in jctx] == [0, 1, 2]


def test_card_only_link_stays_off_unit_row_multi_sheet(tmp_path: Path) -> None:
    """含む＋2シート連結でも、カード専用 SYS は各シートのユニット行に載らない。"""
    fp = tmp_path / "unit_card_sheets.xlsx"
    wb = Workbook()
    wb.active.title = "Other"
    for sh, pfx in (("R_1", "T"), ("R_2", "M")):
        s = wb.create_sheet(sh)
        s["A1"] = f"{pfx}U"
        s["B1"] = f"{pfx}PT"
        s["A2"] = f"{pfx}C"
        s["B2"] = f"{pfx}CARD"
        s["C2"] = f"{pfx}SYS"
        s["D2"] = f"{pfx}J"
        s["A3"] = None
        s["A4"] = None
    wb.save(fp)
    wb.close()
    item = _unit_card_item(skip_empty=False, sheet_name="R_", sheet_rule="含む")
    item["sources"][1]["repeat_max"] = 1
    with xlsx_workbook_scope():
        b = extract_item_bundle(str(fp), item, item_id="i1")
    prim = [_txt(x) for x in (b.get("primary_values") or [])]
    assert prim == ["TU", "TC", "MU", "MC"]
    assert _assign_col(b, str(fp), "品名") == ["ユニット", "TCARD", "ユニット", "MCARD"]
    assert _assign_col(b, str(fp), "SYS") == ["", "TSYS", "", "MSYS"]
    joins = [_txt(x) for x in (b.get("join_values") or {}).get("結合") or []]
    assert joins == ["", "TJ", "", "MJ"]


def test_card_sys_carry_empty_does_not_fill_unit_row(tmp_path: Path) -> None:
    """カード区間の carry_empty はユニット行へ伝播しない。"""
    wb = Workbook()
    ws = wb.active
    ws.title = "S"
    ws["A1"] = "UNIT"
    ws["B1"] = "UNIT_PT"
    ws["A2"] = "C1"
    ws["B2"] = "CARD1"
    ws["C2"] = "SYS1"
    ws["A3"] = "C2"
    ws["B3"] = "CARD2"
    ws["C3"] = None
    fp = tmp_path / "unit_card_carry.xlsx"
    wb.save(fp)
    wb.close()
    item = _unit_card_item(skip_empty=False)
    item["sources"][1]["repeat_max"] = 2
    with xlsx_workbook_scope():
        b = extract_item_bundle(str(fp), item, item_id="i1")
    assert _assign_col(b, str(fp), "SYS") == ["", "SYS1", "SYS1"]
    assert _assign_col(b, str(fp), "品名") == ["ユニット", "CARD1", "CARD2"]
