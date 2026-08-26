# -*- coding: utf-8 -*-
"""
Python: 3.12+
Module: svc/svc_data_agg_debug_run.py
Purpose:
  データ集約デバッグ UI 向けの抽出指標・§10.8 形式ログ行・一括ドライラン。
  本番ロジックは svc_data_agg.compute_batch_table_rows / svc_data_agg_extract に委譲する。
  fill_bundles_for_scenario_phase に任意の progress_hook(done,total) があり、デバッグ UI の長時間ステップ向け。
  フェーズ2は主キーのみ（連続実行は一括）。空主キーの連携は結果に出さない。
  表示上限（SCENARIO_DEBUG_VALUE_ROWS）に達した主キー件数で以降のファイルを開かない。
  ファイル単位で xlsx_workbook_scope を張り、連携・結合フェーズのセル読取で load_workbook を再利用する。
  _name_extract_association_matches も同様（名前取得デバッグの照合カウント）。
"""
from __future__ import annotations

import importlib
import json
from pathlib import Path
from collections.abc import Callable
from typing import Any

from svc.data_agg_sheet_resolve import (
    SHEET_MISS_LABEL,
    classify_sheet_rule,
    resolve_all_sheet_names_by_rule,
)
from svc.data_agg_source_ui import source_ui_block

# 名前取得デバッグ・フェーズ2 の既定文言（MAIN.DEBUG.NAME_EXTRACT_DEBUG で上書き可）
_NE_DBG_DEFAULTS: dict[str, str] = {
    "TIP_FILE": "ファイル名",
    "TIP_SCAN_ROOT": "基準フォルダ",
    "TIP_ITER": "反復",
    "TIP_PK": "PK",
    "TIP_PATH_ITEM_READ": "path_item 読取",
    "TIP_PATH_ITEM_LABEL": "参照ラベル",
    "TIP_COMPARE": "比較",
    "TIP_EXPECTED": "期待",
    "TIP_ACTUAL": "実際",
    "COMPARE_OK": "OK",
    "COMPARE_NG": "NG",
    "COMPARE_SKIP_NO_PATH_ITEM": "比較なし（path_item 未設定）",
    "COMPARE_SKIP_EMPTY_CELL": "比較なし（セル値なし）",
    "PHASE2_MORE_OMITTED_FMT": "他 %d 件は省略（上限 %d 件）",
    "PHASE2_EMPTY": "（展開行なし）",
    "PHASE0_EMPTY": "（検索条件に一致するファイルがありません）",
    "PHASE1_EMPTY": "（主値なし）",
}

NAME_EXTRACT_PHASE2_DISPLAY_MAX = 50

# ユーザー向けデバッグ表示ではフルパス・正規化パスを出さない（方針 B）
_NE_PATH_USER_HIDDEN = "（パスは表示しません）"

_POLARS_MODULE: Any | None = None
_POLARS_CHECKED = False


def _get_polars() -> Any | None:
    global _POLARS_MODULE, _POLARS_CHECKED
    if _POLARS_CHECKED:
        return _POLARS_MODULE
    try:
        _POLARS_MODULE = importlib.import_module("polars")
    except Exception:
        _POLARS_MODULE = None
    _POLARS_CHECKED = True
    return _POLARS_MODULE


def _ne_debug_labels(overlay: dict[str, Any] | None) -> dict[str, str]:
    o = overlay if isinstance(overlay, dict) else {}
    out = dict(_NE_DBG_DEFAULTS)
    for k, v in o.items():
        if v is not None and str(v).strip():
            out[str(k)] = str(v).strip()
    return out


def _debug_path_display_pair(file_path: str, scan_root: str | None) -> tuple[str, str]:
    """表示用文字列（相対優先）とフルパス（ツールチップ用）。"""
    try:
        p = Path(file_path).resolve()
        full = str(p)
    except OSError:
        p = Path(file_path)
        full = str(p)
    root_s = (scan_root or "").strip()
    if not root_s:
        return full, full
    try:
        root = Path(root_s).resolve()
        rel = p.relative_to(root)
        return rel.as_posix(), full
    except (OSError, ValueError):
        return full, full


def _lines_tip(pairs: list[tuple[str, str]]) -> str:
    lines: list[str] = []
    for a, b in pairs:
        lines.append(a)
        lines.append(b)
    return "\n".join(lines)


def _cap_list_capped(lst: list[str], cap: int) -> list[str]:
    if len(lst) <= cap:
        return list(lst)
    out = list(lst[:cap])
    out.append("…（以降省略・上限%d件）" % cap)
    return out


def debug_event_lines_from_sheet_rows(event_log_rows: list[list[Any]]) -> list[str]:
    """§10.8 シート列（時刻・理由コード・シナリオID・参照パス・詳細）をログ用 1 行テキストに。"""
    lines: list[str] = []
    for row in event_log_rows:
        if len(row) >= 5:
            lines.append(
                "EVENT\t%s\t%s\t%s\t%s\t%s" % (row[0], row[1], row[2], row[3], row[4])
            )
        else:
            lines.append("EVENT\t" + json.dumps(row, ensure_ascii=False))
    return lines


def _normalized_scenario_source_type(s0: dict[str, Any] | None) -> str:
    if not s0:
        return "cell"
    t = str(s0.get("type") or "cell").strip().lower()
    if t in ("metadata", "meta", "filename"):
        return "name_extract"
    return t


def _is_name_extract_source(s0: dict[str, Any]) -> bool:
    return _normalized_scenario_source_type(s0) == "name_extract"


def _allowed_paths_after_file_filter(
    item: dict[str, Any], paths: list[str]
) -> list[str]:
    """ファイル名フィルタ後のパス一覧（ブックは開かない）。cell は全ソース OR、名前から取得は ui ブロックで同ルールを適用。"""
    from svc.svc_data_agg_extract import source_passes_file_name_filter

    sources = [s for s in (item.get("sources") or []) if isinstance(s, dict)]
    cell_sources = [s for s in sources if _normalized_scenario_source_type(s) == "cell"]
    if not cell_sources:
        return list(paths)
    allowed_paths: list[str] = []
    for fp in paths:
        if any(source_passes_file_name_filter(fp, s) for s in cell_sources):
            allowed_paths.append(fp)
    return allowed_paths


def _empty_bundle_shell() -> dict[str, Any]:
    """主キーステップ未実行などで連携・結合のみ進めた場合のプレースホルダ。"""
    return {
        "primary_values": [],
        "iteration_contexts": [],
        "link_values": {},
        "link_contexts": {},
        "join_values": {},
        "join_contexts": {},
        "path_item_values": {},
        "path_item_contexts": {},
    }


def _bundle_error_stub() -> dict[str, Any]:
    return {
        "primary_values": [None],
        "iteration_contexts": [],
        "link_values": {},
        "link_contexts": {},
        "join_values": {},
        "join_contexts": {},
        "path_item_values": {},
        "path_item_contexts": {},
    }


def fill_bundles_for_scenario_phase(
    item: dict[str, Any],
    paths: list[str],
    item_id: str,
    cache: dict[str, dict[str, Any]],
    phase_slot_index: int,
    paths_to_process: list[str] | None = None,
    progress_hook: Callable[[int, int], None] | None = None,
    phase2_primary_only: bool = True,
    max_primary_rows: int | None = None,
) -> None:
    """
    デバッグのシナリオフェーズに応じてキャッシュを更新する。
    2=主キー表示。既定は主キーのみ（phase2_primary_only）。
      連続実行では False にして主キー＋連携＋結合を一括抽出する。
    3/4=連携・結合表示（フル抽出済みなら再利用。未完了なら再抽出）。
      フェーズ2で表示上限に達して開いていないファイルは開かない。
    max_primary_rows: フェーズ2で残った主キー件数がこの件数に達したら以降のファイルを開かない。
    progress_hook: 処理ファイル進捗 (done, total)。done は 1 始まり。省略時は呼ばない。
    """
    from svc.svc_data_agg_extract import xlsx_workbook_scope

    allowed_paths = _allowed_paths_after_file_filter(item, paths)

    for old_fp in list(cache.keys()):
        if old_fp not in allowed_paths:
            cache.pop(old_fp, None)

    if paths_to_process is not None:
        proc_set = {fp for fp in paths_to_process if fp in allowed_paths}
        for fp in list(cache.keys()):
            if fp in allowed_paths and fp not in proc_set:
                cache.pop(fp, None)
        process_paths = [fp for fp in paths_to_process if fp in allowed_paths]
    else:
        process_paths = allowed_paths

    jp_hdr = str(item.get("name") or item.get("id") or "").strip()
    n_proc = len(process_paths)
    stride = 5 if n_proc >= 10 else 1
    try:
        max_pr = int(max_primary_rows) if max_primary_rows is not None else 0
    except (TypeError, ValueError):
        max_pr = 0
    if max_pr < 0:
        max_pr = 0
    kept_total = 0

    for idx, fp in enumerate(process_paths, start=1):
        if phase_slot_index >= 3:
            if fp not in cache:
                continue
        elif phase_slot_index == 2 and max_pr > 0 and kept_total >= max_pr:
            break
        remain: int | None = None
        if phase_slot_index == 2 and max_pr > 0:
            remain = max_pr - kept_total
            if remain <= 0:
                break
        with xlsx_workbook_scope():
            try:
                cache[fp] = _extract_bundles_for_matched_sheets(
                    item,
                    fp,
                    item_id=item_id,
                    jp_hdr=jp_hdr,
                    phase_slot_index=phase_slot_index,
                    existing=cache.get(fp),
                    phase2_primary_only=phase2_primary_only,
                    max_primary_rows=remain,
                )
            except Exception:
                cache[fp] = _bundle_error_stub()
        if phase_slot_index == 2:
            kept_total += len((cache.get(fp) or {}).get("primary_values") or [])

        if progress_hook is not None and n_proc > 0:
            if idx == 1 or idx % stride == 0 or idx == n_proc:
                progress_hook(idx, n_proc)


def fill_scenario_link_join_after_primary(
    item: dict[str, Any],
    paths: list[str],
    cache: dict[str, dict[str, Any]],
    item_id: str,
    *,
    cancel_check: Callable[..., None] | None = None,
    progress_hook: Callable[[int, int], None] | None = None,
) -> None:
    """主キー抽出済みキャッシュへ、連携・結合を追記する（空主キーファイルは開かない）。"""
    from svc.data_agg_cancel import DataAggCancelled
    from svc.svc_data_agg_extract import xlsx_workbook_scope

    allowed = _allowed_paths_after_file_filter(item, paths)
    jp_hdr = str(item.get("name") or item.get("id") or "").strip()
    todo = [
        fp
        for fp in allowed
        if fp in cache and not (cache.get(fp) or {}).get(_DBG_FULL_EXTRACT)
    ]
    n_todo = len(todo)
    for idx, fp in enumerate(todo, start=1):
        if cancel_check is not None:
            cancel_check()
        b = cache.get(fp)
        if not isinstance(b, dict):
            continue
        if not (b.get("primary_values") or []):
            cache[fp] = _mark_dbg_full_extract(b)
            if progress_hook is not None and n_todo > 0:
                progress_hook(idx, n_todo)
            continue
        with xlsx_workbook_scope():
            try:
                b3 = _extract_bundles_for_matched_sheets(
                    item,
                    fp,
                    item_id=item_id,
                    jp_hdr=jp_hdr,
                    phase_slot_index=3,
                    existing=b,
                )
                b4 = _extract_bundles_for_matched_sheets(
                    item,
                    fp,
                    item_id=item_id,
                    jp_hdr=jp_hdr,
                    phase_slot_index=4,
                    existing=b3,
                )
                cache[fp] = _mark_dbg_full_extract(b4)
            except DataAggCancelled:
                raise
            except Exception:
                cache[fp] = b if isinstance(b, dict) else _bundle_error_stub()
        if progress_hook is not None and n_todo > 0:
            if idx == 1 or idx % 5 == 0 or idx == n_todo:
                progress_hook(idx, n_todo)


def _name_extract_association_matches(
    item: dict[str, Any],
    hit_files: list[str],
    item_id: str,
) -> tuple[int, list[str]]:
    """
    検索通過ファイルごとに抽出し、照合キーが当該ファイルと一致し主値がある反復を数える。
    戻り値のタグは「ファイル名:反復番号（1始まり）」をソートユニーク化したリスト。
    """
    from pathlib import Path

    from svc.data_agg_path_norm import normalize_source_path, path_is_under_directory
    from svc.svc_data_agg_extract import extract_item_bundle, xlsx_workbook_scope

    sources = item.get("sources") or []
    s0 = sources[0] if sources and isinstance(sources[0], dict) else None
    pb = source_ui_block(s0) if s0 else {}
    stype = str((s0 or {}).get("source_type") or "file_name").strip().lower()
    pit_label = str((pb or {}).get("path_item") or "").strip()
    if not pit_label:
        return 0, []
    total = 0
    tags: set[str] = set()
    for fp in hit_files:
        with xlsx_workbook_scope():
            b = extract_item_bundle(
                fp,
                item,
                item_id=item_id,
                cell_positions={},
                join_path_header=None,
            )
        prim_vals = b.get("primary_values") or []
        piv = b.get("path_item_values") or {}
        pvs = piv.get(pit_label) or []
        norm_fp = normalize_source_path(fp)
        norm_dir = normalize_source_path(Path(fp).resolve().parent)
        fname = Path(fp).name.strip() or fp
        n = max(len(prim_vals), len(pvs), 1)
        for k in range(n):
            pv = prim_vals[k] if k < len(prim_vals) else None
            if pv is None or str(pv).strip() == "":
                continue
            rp = pvs[k] if k < len(pvs) else None
            if rp is None:
                continue
            rp_n = normalize_source_path(str(rp))
            ok = (
                path_is_under_directory(rp_n, norm_dir)
                if stype == "dir_name"
                else (rp_n == norm_fp)
            )
            if ok:
                total += 1
                tags.add("%s:%d" % (fname, k + 1))
    return total, sorted(tags)


def _name_extract_phase2_colvals_and_tips(
    item: dict[str, Any],
    hit_files: list[str],
    cache: dict[str, dict[str, Any]],
    scan_root: str | None,
    lbl: dict[str, str],
) -> tuple[list[str], list[str | None]]:
    """フェーズ2: ファイル×反復×PK の展開行とツールチップ（表示 NAME_EXTRACT_PHASE2_DISPLAY_MAX まで）。"""
    from svc.data_agg_path_norm import normalize_source_path, path_is_under_directory

    sources = item.get("sources") or []
    s0 = sources[0] if sources and isinstance(sources[0], dict) else {}
    pb = source_ui_block(s0) if isinstance(s0, dict) else {}
    pit_label = str((pb or {}).get("path_item") or "").strip()
    st_raw = str((s0 or {}).get("source_type") or "file_name").strip().lower()
    rows_raw: list[tuple[str, int, str, str, str, str]] = []

    for fp in hit_files:
        b = cache.get(fp) or {}
        disp = Path(fp).name
        try:
            p_res = Path(fp).resolve()
        except OSError:
            p_res = Path(fp)
        norm_fp = normalize_source_path(fp)
        norm_dir = normalize_source_path(p_res.parent)
        prim_vals = list(b.get("primary_values") or [None])
        piv = b.get("path_item_values") or {}
        pvs = (piv.get(pit_label) if pit_label else None) or []
        if not isinstance(pvs, list):
            pvs = []
        n = max(len(prim_vals), len(pvs), 1)
        for k in range(n):
            pv = prim_vals[k] if k < len(prim_vals) else None
            pk = "" if pv is None else str(pv).strip()
            if not pk:
                continue
            rp = pvs[k] if k < len(pvs) else None
            rp_s = "" if rp is None else str(rp).strip()
            cell_disp = "（内部照合あり）" if rp_s else "（空）"
            if not pit_label:
                status = lbl["COMPARE_SKIP_NO_PATH_ITEM"]
            elif rp is None or not rp_s:
                status = lbl["COMPARE_SKIP_EMPTY_CELL"]
            else:
                rp_n = normalize_source_path(str(rp))
                ok = (
                    path_is_under_directory(rp_n, norm_dir)
                    if st_raw == "dir_name"
                    else (rp_n == norm_fp)
                )
                status = lbl["COMPARE_OK"] if ok else lbl["COMPARE_NG"]
            rows_raw.append(
                (
                    disp,
                    k + 1,
                    pk,
                    cell_disp,
                    status,
                    pit_label or "—",
                )
            )

    rows_raw.sort(key=lambda r: (r[0].lower(), r[1]))
    cap = NAME_EXTRACT_PHASE2_DISPLAY_MAX
    colvals: list[str] = []
    coltips: list[str | None] = []
    for row in rows_raw[:cap]:
        disp, k1, pk, cell_disp, status, pit_l = row
        line = "[%s] #%d PK=%s 照合=%s | %s" % (disp, k1, pk, cell_disp, status)
        colvals.append(line)
        tip_pairs: list[tuple[str, str]] = [
            (lbl["TIP_FILE"], disp),
            (lbl["TIP_ITER"], str(k1)),
            (lbl["TIP_PK"], pk),
            (lbl["TIP_PATH_ITEM_READ"], cell_disp),
            (lbl["TIP_PATH_ITEM_LABEL"], pit_l),
            (lbl["TIP_COMPARE"], status),
        ]
        if status == lbl["COMPARE_NG"]:
            tip_pairs.append((lbl["TIP_EXPECTED"], _NE_PATH_USER_HIDDEN))
            tip_pairs.append((lbl["TIP_ACTUAL"], _NE_PATH_USER_HIDDEN))
        coltips.append(_lines_tip(tip_pairs))
    omitted = max(0, len(rows_raw) - cap)
    if omitted > 0:
        colvals.append(lbl["PHASE2_MORE_OMITTED_FMT"] % (omitted, cap))
        coltips.append(
            _lines_tip(
                [
                    (lbl["TIP_SCAN_ROOT"], _NE_PATH_USER_HIDDEN),
                    ("展開行計", str(len(rows_raw))),
                ]
            )
        )
    if not colvals:
        colvals = [lbl.get("PHASE2_EMPTY", "（展開行なし）")]
        coltips.append(None)
    return colvals, coltips


def _name_extract_debug_phase_result(
    item: dict[str, Any],
    paths: list[str],
    phase_slot_index: int,
    max_rows: int,
    cache: dict[str, dict[str, Any]],
    allowed: list[str],
    _nfs0: str,
    s0: dict[str, Any],
    item_id: str,
    events: list[dict[str, Any]],
    scan_root: str | None = None,
    ne_labels: dict[str, Any] | None = None,
) -> tuple[list[str], list[str], list[dict[str, Any]], list[str | None]]:
    """
    名前から取得専用。サマリ先頭4列:
    [対象ファイル数, 検索条件ヒット数, 主キー条件(ユニーク主値数), 関連付け一致数]（5列目は '-'）。
    """
    from svc.svc_data_agg_extract import name_extract_hit_files_ordered

    lbl = _ne_debug_labels(ne_labels)
    hit_files = name_extract_hit_files_ordered(allowed, s0)
    n_allowed = len(allowed)
    n_hit = len(hit_files)

    def _cap_with_tips(
        vals: list[str], tips: list[str | None]
    ) -> tuple[list[str], list[str | None]]:
        if len(vals) <= max_rows:
            return vals, tips
        return (
            vals[:max_rows] + ["…（以降省略・上限%d件）" % max_rows],
            tips[:max_rows] + [None],
        )

    if phase_slot_index == 0:
        summary = [str(n_allowed), str(n_hit), "-", "-", "-"]
        if not hit_files:
            return (
                summary,
                [lbl.get("PHASE0_EMPTY", _NE_DBG_DEFAULTS["PHASE0_EMPTY"])],
                events,
                [None],
            )
        colvals: list[str] = []
        coltips: list[str | None] = []
        for fp in hit_files:
            colvals.append(Path(fp).name)
            coltips.append(
                _lines_tip(
                    [
                        (lbl["TIP_SCAN_ROOT"], _NE_PATH_USER_HIDDEN),
                        (lbl["TIP_FILE"], _NE_PATH_USER_HIDDEN),
                    ]
                )
            )
        colvals, coltips = _cap_with_tips(colvals, coltips)
        return summary, colvals, events, coltips

    if phase_slot_index == 1:
        fill_bundles_for_scenario_phase(
            item,
            paths,
            item_id,
            cache,
            2,
            paths_to_process=hit_files,
            max_primary_rows=max_rows,
        )
        seen_prim: set[str] = set()
        ordered_unique: list[str] = []
        for fp in hit_files:
            b = cache.get(fp) or {}
            for v in b.get("primary_values") or [None]:
                s = "" if v is None else str(v).strip()
                if s and s not in seen_prim:
                    seen_prim.add(s)
                    ordered_unique.append(s)
        n_u = len(ordered_unique)
        summary = [str(n_allowed), str(n_hit), str(n_u), "-", "-"]
        if not ordered_unique:
            return (
                summary,
                [lbl.get("PHASE1_EMPTY", _NE_DBG_DEFAULTS["PHASE1_EMPTY"])],
                events,
                [None],
            )
        tips1 = [
            _lines_tip([(lbl["TIP_PK"], u), ("種別", "ユニーク主キー一覧")])
            for u in ordered_unique
        ]
        cv, ct = _cap_with_tips(list(ordered_unique), tips1)
        return summary, cv, events, ct

    if phase_slot_index == 2:
        fill_bundles_for_scenario_phase(
            item,
            paths,
            item_id,
            cache,
            2,
            paths_to_process=hit_files,
            max_primary_rows=max_rows,
        )
        seen_u: set[str] = set()
        for fp in hit_files:
            b = cache.get(fp) or {}
            for v in b.get("primary_values") or [None]:
                sx = "" if v is None else str(v).strip()
                if sx and sx not in seen_u:
                    seen_u.add(sx)
        n_u2 = len(seen_u)
        ac, _assoc_tags = _name_extract_association_matches(item, hit_files, item_id)
        summary = [str(n_allowed), str(n_hit), str(n_u2), str(ac), "-"]
        colvals2, coltips2 = _name_extract_phase2_colvals_and_tips(
            item, hit_files, cache, scan_root, lbl
        )
        return summary, colvals2, events, coltips2

    return ["-", "-", "-", "-", "-"], ["（不明なフェーズ）"], events, [None]


def scenario_debug_phase_result(
    item: dict[str, Any],
    paths: list[str],
    phase_slot_index: int,
    max_rows: int,
    cache: dict[str, dict[str, Any]],
    scan_root: str | None = None,
    name_extract_debug_labels: dict[str, Any] | None = None,
    progress_hook: Callable[[int, int], None] | None = None,
    phase2_primary_only: bool = True,
) -> tuple[list[str], list[str], list[dict[str, Any]], list[str | None]]:
    """
    シナリオフェーズ用のサマリ 5 列・値列・イベント（結合失敗等はフェーズ実行では空、抽出失敗のみ補足可）。
    phase_slot_index: 0=ファイルフィルタ … 4=結合キー（COND_KEYS 順）。
    フェーズ 0 はファイル名のみ。フェーズ 1 はシート名解決のみ（各ファイルでブックを短時間開く）。
    フェーズ 2 以降で extract_item_bundle をキャッシュに積む。
    phase2_primary_only: フェーズ2は主キーのみ（既定）。連続実行は False で一括抽出。
    戻り値第4要素は値列と同長のツールチップ（同一列インデックス）。不要時は None。
    progress_hook: セル座標系で phase_slot_index>=2 の抽出ループ進捗 (done, total)。名前取得系では未使用。
    """
    item_id = str(item.get("id") or "item")
    item_name = str(item.get("name") or item_id or "項目")
    allowed = _allowed_paths_after_file_filter(item, paths)
    for old_fp in list(cache.keys()):
        if old_fp not in allowed:
            cache.pop(old_fp, None)

    sources = item.get("sources") or []
    s0 = sources[0] if sources and isinstance(sources[0], dict) else {}
    events: list[dict[str, Any]] = []
    if not paths:
        events.append(
            {
                "reason_code": "NO_FILE",
                "scenario_id": item_id,
                "path": "",
                "detail": "検出ファイルパスが 0 件",
            }
        )

    nfs0 = str(len(allowed))
    na0 = nfs0

    if _is_name_extract_source(s0):
        return _name_extract_debug_phase_result(
            item,
            paths,
            phase_slot_index,
            max_rows,
            cache,
            allowed,
            nfs0,
            s0,
            item_id,
            events,
            scan_root=scan_root,
            ne_labels=name_extract_debug_labels,
        )

    if phase_slot_index == 0:
        names = [Path(f).name for f in allowed]
        colvals = (
            _cap_list_capped(names, max_rows)
            if names
            else ["（ファイル名フィルタ後 0 件）"]
        )
        summary = [na0, "-", "-", "-", "-"]
        return summary, colvals, events, [None] * len(colvals)

    if phase_slot_index == 1:
        colvals = _sheet_column_preview(s0, allowed, max_rows)
        summary = [nfs0, na0, "-", "-", "-"]
        return summary, colvals, events, [None] * len(colvals)

    fill_bundles_for_scenario_phase(
        item,
        paths,
        item_id,
        cache,
        phase_slot_index,
        progress_hook=progress_hook,
        phase2_primary_only=phase2_primary_only,
        max_primary_rows=max_rows,
    )

    filtered_paths = [fp for fp in paths if fp in cache]
    n_after = len(filtered_paths)
    nfs = str(n_after)
    na = str(n_after)

    prim: list[str] = []
    for fp in filtered_paths:
        if fp not in cache:
            continue
        b = cache[fp]
        for v in b.get("primary_values") or []:
            prim.append("" if v is None else str(v))
    p = source_ui_block(s0) or {}
    link_defs = p.get("link_defs") if isinstance(p.get("link_defs"), list) else []
    join_defs = p.get("join_defs") if isinstance(p.get("join_defs"), list) else []
    nl = len(link_defs)
    nj = len(join_defs)

    link_nonempty = 0
    for fp in filtered_paths:
        b = cache.get(fp) or {}
        for vals in (b.get("link_values") or {}).values():
            if vals:
                link_nonempty += 1
                break
    join_nonempty = 0
    for fp in filtered_paths:
        b = cache.get(fp) or {}
        for vals in (b.get("join_values") or {}).values():
            if vals:
                join_nonempty += 1
                break

    # link_v / join_v: 分子＝取得セル総数、分母＝キー定義数（シート登録数ではない）。
    link_count = 0
    for fp in filtered_paths:
        b = cache.get(fp) or {}
        mp = b.get("link_values") or {}
        if not isinstance(mp, dict):
            continue
        if link_defs:
            for d in link_defs:
                if not isinstance(d, dict):
                    continue
                tgt = str(d.get("item") or "").strip() or item_name
                link_count += len(mp.get(tgt) or [])
        else:
            for vals in mp.values():
                link_count += len(vals or [])
    link_v = "%d/%d" % (link_count, max(nl, 1)) if nl else "0/0"
    join_count = 0
    for fp in filtered_paths:
        b = cache.get(fp) or {}
        mp = b.get("join_values") or {}
        if not isinstance(mp, dict):
            continue
        if join_defs:
            for d in join_defs:
                if not isinstance(d, dict):
                    continue
                tgt = str(d.get("item") or "").strip() or item_name
                join_count += len(mp.get(tgt) or [])
        else:
            for vals in mp.values():
                join_count += len(vals or [])
    join_v = "%d/%d" % (join_count, max(nj, 1)) if nj else "0/0"
    npv = str(len(prim)) if prim else "0"

    events = []
    if not paths:
        events.append(
            {
                "reason_code": "NO_FILE",
                "scenario_id": item_id,
                "path": "",
                "detail": "検出ファイルパスが 0 件",
            }
        )

    if phase_slot_index == 2:
        summary = [nfs, na, npv, "-", "-"]
        colvals = _cap_list_capped(prim, max_rows) if prim else ["（主値 0 件）"]
        if prim and len(allowed) > len(filtered_paths):
            omit = "…（以降省略・上限%d件）" % max_rows
            if not colvals or colvals[-1] != omit:
                colvals = list(colvals) + [omit]
    elif phase_slot_index == 3:
        summary = [nfs, na, npv, link_v, "-"]
        colvals = _flatten_map_values_by_defs(
            cache, "link_values", link_defs, max_rows, selected_item_name=item_name
        )
    else:
        summary = [nfs, na, npv, link_v, join_v]
        colvals = _flatten_map_values_by_defs(
            cache, "join_values", join_defs, max_rows, selected_item_name=item_name
        )

    return summary, colvals, events, [None] * len(colvals)


def _patch_item_sheet_exact(
    item: dict[str, Any],
    sheet_name: str,
    *,
    workbook_sheet_names: list[str] | None = None,
) -> dict[str, Any]:
    from svc.data_agg_sheet_resolve import patch_item_sheet_exact

    return patch_item_sheet_exact(
        item, sheet_name, workbook_sheet_names=workbook_sheet_names
    )


def _merge_primary_sheet_bundles(
    parts: list[tuple[str, dict[str, Any]]],
    file_path: str,
) -> dict[str, Any]:
    """複数シートの primary バンドルを左→右順に連結する。"""
    from svc.svc_data_agg_extract import _merge_primary_sheet_bundles as _merge_ex

    return _merge_ex(parts, file_path)


def _extract_link_join_across_sheets(
    file_path: str,
    item: dict[str, Any],
    base: dict[str, Any],
    *,
    item_id: str,
    jp_hdr: str,
    scope: str,
) -> dict[str, Any]:
    """既存 primary の _sheet_parts に沿って link/join をシート順に再計算する。"""
    from svc.data_agg_sheet_resolve import list_workbook_sheet_names
    from svc.svc_data_agg_extract import (
        _align_link_join_series_to_primary,
        _append_rule_maps_with_offset,
        _mini_iter_contexts_for_sheet_part,
        _mini_spans_from_local,
        extract_item_bundle,
    )

    parts = base.get("_sheet_parts")
    out = base.copy()
    if scope == "link":
        out["link_values"] = {}
        out["link_contexts"] = {}
        values_key, contexts_key = "link_values", "link_contexts"
    else:
        out["join_values"] = {}
        out["join_contexts"] = {}
        values_key, contexts_key = "join_values", "join_contexts"

    if not isinstance(parts, list) or not parts:
        item_one = _item_with_resolved_sheet_for_debug(item, file_path)
        if item_one is None:
            return out
        return extract_item_bundle(
            file_path,
            item_one,
            item_id=item_id,
            cell_positions={},
            join_path_header=jp_hdr or None,
            debug_step_scope=scope,
            existing_bundle=base,
        )

    wb_names = list_workbook_sheet_names(file_path)
    for part in parts:
        if not isinstance(part, dict):
            continue
        sh = str(part.get("sheet_name") or "")
        g_off = int(part.get("g_off") or 0)
        n_src = int(part.get("n_src") or 0)
        if not sh or n_src < 1:
            continue
        local_spans = part.get("_cell_source_spans")
        mini_spans = _mini_spans_from_local(local_spans, n_src)
        mini_prim = list(part.get("primary_values") or [])[:n_src]
        mini = {
            "primary_values": mini_prim,
            "iteration_contexts": _mini_iter_contexts_for_sheet_part(
                file_path=str(file_path),
                sheet_name=sh,
                g_off=g_off,
                n_src=n_src,
                mini_prim=mini_prim,
                base=base,
                part=part,
            ),
            "link_values": {},
            "link_contexts": {},
            "join_values": {},
            "join_contexts": {},
            "path_item_values": {},
            "path_item_contexts": {},
            "_cell_source_spans": mini_spans,
        }
        partial = extract_item_bundle(
            file_path,
            _patch_item_sheet_exact(item, sh, workbook_sheet_names=wb_names),
            item_id=item_id,
            cell_positions={},
            join_path_header=jp_hdr or None,
            debug_step_scope=scope,
            existing_bundle=mini,
        )
        _append_rule_maps_with_offset(
            out,
            partial,
            values_key=values_key,
            contexts_key=contexts_key,
            g_off=g_off,
            file_path=file_path,
        )
    _align_link_join_series_to_primary(out)
    return out


_DBG_FULL_EXTRACT = "_dbg_full_extract"
_DBG_PRIMARY_ONLY = "_dbg_primary_only"


def _mark_dbg_full_extract(bundle: dict[str, Any]) -> dict[str, Any]:
    bundle[_DBG_FULL_EXTRACT] = True
    return bundle


def _extract_phase2_item_bundle(
    file_path: str,
    item: dict[str, Any],
    *,
    item_id: str,
    jp_hdr: str,
    primary_only: bool,
    max_primary_rows: int | None = None,
) -> dict[str, Any]:
    from svc.svc_data_agg_extract import extract_item_bundle, item_wants_skip_carry_seed

    extra: dict[str, Any] = {}
    if max_primary_rows is not None:
        extra["max_primary_rows"] = max_primary_rows
    # skip_carry_seed ON 時はスキップ行の連携値が必要なため主キーのみでは不足。一括抽出する。
    use_primary_only = bool(primary_only) and not item_wants_skip_carry_seed(item)
    if use_primary_only:
        b = extract_item_bundle(
            file_path,
            item,
            item_id=item_id,
            cell_positions={},
            join_path_header=jp_hdr or None,
            debug_step_scope="primary",
            **extra,
        )
        b[_DBG_PRIMARY_ONLY] = True
        return b
    return _mark_dbg_full_extract(
        extract_item_bundle(
            file_path,
            item,
            item_id=item_id,
            cell_positions={},
            join_path_header=jp_hdr or None,
            **extra,
        )
    )


def _extract_bundles_for_matched_sheets(
    item: dict[str, Any],
    file_path: str,
    *,
    item_id: str,
    jp_hdr: str,
    phase_slot_index: int,
    existing: dict[str, Any] | None,
    phase2_primary_only: bool = True,
    max_primary_rows: int | None = None,
) -> dict[str, Any]:
    """一致シートを左→右の順に読む。フェーズ2は主キーのみ（または一括）。3/4はキャッシュ再利用。"""
    from svc.svc_data_agg_extract import (
        extract_item_bundle,
        matching_sheets_and_names_for_item,
    )

    if (
        phase_slot_index >= 3
        and isinstance(existing, dict)
        and existing.get(_DBG_FULL_EXTRACT)
    ):
        return existing

    sources = item.get("sources") or []
    s0 = sources[0] if sources and isinstance(sources[0], dict) else None
    if not isinstance(s0, dict) or _is_name_extract_source(s0):
        if phase_slot_index == 2:
            return _extract_phase2_item_bundle(
                file_path,
                item,
                item_id=item_id,
                jp_hdr=jp_hdr,
                primary_only=phase2_primary_only,
                max_primary_rows=max_primary_rows,
            )
        base = existing if isinstance(existing, dict) else _empty_bundle_shell()
        if phase_slot_index == 3:
            return extract_item_bundle(
                file_path,
                item,
                item_id=item_id,
                cell_positions={},
                join_path_header=jp_hdr or None,
                debug_step_scope="link",
                existing_bundle=base,
            )
        return extract_item_bundle(
            file_path,
            item,
            item_id=item_id,
            cell_positions={},
            join_path_header=jp_hdr or None,
            debug_step_scope="join",
            existing_bundle=base,
        )

    sheets, wb_names = matching_sheets_and_names_for_item(file_path, item)
    if sheets is None:
        item_one = _item_with_resolved_sheet_for_debug(item, file_path)
        if item_one is None:
            return _empty_bundle_shell()
        if phase_slot_index == 2:
            return _extract_phase2_item_bundle(
                file_path,
                item_one,
                item_id=item_id,
                jp_hdr=jp_hdr,
                primary_only=phase2_primary_only,
                max_primary_rows=max_primary_rows,
            )
        base = existing if isinstance(existing, dict) else _empty_bundle_shell()
        return extract_item_bundle(
            file_path,
            item_one,
            item_id=item_id,
            cell_positions={},
            join_path_header=jp_hdr or None,
            debug_step_scope="link" if phase_slot_index == 3 else "join",
            existing_bundle=base,
        )
    if not sheets:
        return _empty_bundle_shell()

    if phase_slot_index == 2:
        return _extract_phase2_item_bundle(
            file_path,
            item,
            item_id=item_id,
            jp_hdr=jp_hdr,
            primary_only=phase2_primary_only,
            max_primary_rows=max_primary_rows,
        )

    base = existing if isinstance(existing, dict) else _empty_bundle_shell()
    scope = "link" if phase_slot_index == 3 else "join"
    return _extract_link_join_across_sheets(
        file_path,
        item,
        base,
        item_id=item_id,
        jp_hdr=jp_hdr,
        scope=scope,
    )


def _list_matching_sheet_names(
    file_path: str,
    s0: dict[str, Any],
) -> list[str] | None:
    """一致シート名一覧（左→右）。CSV などシート解決不要時は None。該当なしは []。"""
    sn = str(s0.get("sheet_name") or "").strip()
    p0 = source_ui_block(s0) or {}
    rule = str(p0.get("sheet_rule") or "")
    p = Path(file_path)
    suffix = p.suffix.lower()
    if suffix == ".csv":
        return None
    if suffix == ".xls":
        from svc.data_agg_xls_io import (
            list_xls_sheet_names,
            xls_reader_unavailable_message,
        )

        if xls_reader_unavailable_message():
            return []
        names = list_xls_sheet_names(file_path)
        if not names:
            return []
        return resolve_all_sheet_names_by_rule(names, rule, sn)
    try:
        import openpyxl  # noqa: E402
    except Exception:
        return [sn] if sn else []
    try:
        wb = openpyxl.load_workbook(p, read_only=True, data_only=True)
        names = list(wb.sheetnames or [])
        wb.close()
        return resolve_all_sheet_names_by_rule(names, rule, sn)
    except Exception:
        return [sn] if sn else []


def _item_with_resolved_sheet_for_debug(
    item: dict[str, Any],
    file_path: str,
) -> dict[str, Any] | None:
    """先頭の一致シート名へ差し替えた item（単一シート互換）。"""
    from svc.svc_data_agg_extract import matching_sheets_and_names_for_item

    sources = item.get("sources") or []
    s0 = sources[0] if sources and isinstance(sources[0], dict) else None
    if not isinstance(s0, dict) or _is_name_extract_source(s0):
        return item
    sheets, names = matching_sheets_and_names_for_item(file_path, item)
    if sheets is None:
        return item
    if not sheets:
        return None
    return _patch_item_sheet_exact(item, sheets[0], workbook_sheet_names=names)


def _sheet_column_preview(
    s0: dict[str, Any],
    file_paths: list[str],
    max_rows: int,
) -> list[str]:
    """ファイルごとに一致シートを左→右で列挙（複数は ', ' 結合）。"""
    out: list[str] = []
    for fp in file_paths:
        if len(out) >= max_rows:
            break
        suffix = Path(fp).suffix.lower()
        if suffix == ".xls":
            from svc.data_agg_xls_io import (
                list_xls_sheet_names,
                xls_reader_unavailable_message,
            )

            unavailable = xls_reader_unavailable_message()
            if unavailable:
                out.append(unavailable)
                continue
            sn = str(s0.get("sheet_name") or "").strip()
            p0 = source_ui_block(s0) or {}
            rule = str(p0.get("sheet_rule") or "")
            names = list_xls_sheet_names(fp)
            if not names:
                out.append("（.xls読取失敗）")
                continue
            matched = resolve_all_sheet_names_by_rule(names, rule, sn)
            out.append(", ".join(matched) if matched else SHEET_MISS_LABEL)
            continue
        sheets = _list_matching_sheet_names(fp, s0)
        if sheets is None:
            sn = str(s0.get("sheet_name") or "").strip()
            p0 = source_ui_block(s0) or {}
            rule = str(p0.get("sheet_rule") or "")
            one = _resolve_actual_sheet_name(fp, rule, sn)
            out.append(one if one else SHEET_MISS_LABEL)
            continue
        if not sheets:
            out.append(SHEET_MISS_LABEL)
        else:
            out.append(", ".join(sheets))
    return _cap_list_capped(out, max_rows) if out else ["（シート条件のみ）"]


def _resolve_actual_sheet_name(file_path: str, rule: str, sn: str) -> str | None:
    """先頭一致のシート名。該当なしは None。"""
    matched = _resolve_all_actual_sheet_names(file_path, rule, sn)
    if matched is None:
        if Path(file_path).suffix.lower() == ".csv":
            return "CSV"
        return sn or None
    return matched[0] if matched else None


def _resolve_all_actual_sheet_names(
    file_path: str, rule: str, sn: str
) -> list[str] | None:
    """一致シート全件。CSV は None。"""
    p = Path(file_path)
    if p.suffix.lower() == ".csv":
        return None
    if p.suffix.lower() == ".xls":
        from svc.data_agg_xls_io import (
            list_xls_sheet_names,
            xls_reader_unavailable_message,
        )

        if xls_reader_unavailable_message():
            return []
        names = list_xls_sheet_names(file_path)
        if not names:
            return []
        return resolve_all_sheet_names_by_rule(names, rule, sn)
    try:
        import openpyxl  # noqa: E402
    except Exception:
        return [sn] if sn else []
    try:
        wb = openpyxl.load_workbook(p, read_only=True, data_only=True)
        names = list(wb.sheetnames or [])
        wb.close()
        return resolve_all_sheet_names_by_rule(names, rule, sn)
    except Exception:
        return [sn] if sn else []


def _flatten_map_values(
    cache: dict[str, dict[str, Any]],
    key: str,
    max_rows: int,
) -> list[str]:
    flat: list[str] = []
    for b in cache.values():
        mp = b.get(key) or {}
        if not isinstance(mp, dict):
            continue
        for _tgt, vals in mp.items():
            for v in vals or [None]:
                if len(flat) >= max_rows:
                    return _cap_list_capped(flat, max_rows)
                flat.append("" if v is None else str(v))
    return _cap_list_capped(flat, max_rows) if flat else ["（値なし）"]


def _kept_primary_aligned_series(
    bundle: dict[str, Any],
    map_key: str,
    target: str,
) -> list[str]:
    """主キーが残った反復だけ連携／結合を返す。主キー 0 件のファイルは出さない。"""
    prim = bundle.get("primary_values") or []
    if not prim:
        return []
    mp = bundle.get(map_key)
    vals = mp.get(target) if isinstance(mp, dict) else None
    if not isinstance(vals, list):
        vals = []
    out: list[str] = []
    for i in range(len(prim)):
        if i < len(vals) and vals[i] is not None:
            out.append(str(vals[i]))
        else:
            out.append("")
    return out


def _flatten_map_values_by_defs(
    cache: dict[str, dict[str, Any]],
    key: str,
    defs: list[Any],
    max_rows: int,
    selected_item_name: str = "",
) -> list[str]:
    if not defs:
        return ["（定義なし）"]
    flat: list[str] = []
    for i, d in enumerate(defs):
        if not isinstance(d, dict):
            continue
        tgt = str(d.get("item") or "").strip() or (selected_item_name or "未指定")
        vals: list[str] = []
        for b in cache.values():
            if not isinstance(b, dict):
                continue
            vals.extend(_kept_primary_aligned_series(b, key, tgt))
        if not vals:
            vals = ["（値なし）"]
        per_key_vals = vals[:max_rows]
        for v in per_key_vals:
            flat.append("#%d[%s] %s" % (i + 1, tgt, v))
        if len(vals) > max_rows:
            flat.append("#%d[%s] …（省略・上限%d件）" % (i + 1, tgt, max_rows))
    return flat if flat else ["（値なし）"]


def format_synthetic_events_for_log(
    events: list[dict[str, Any]],
    scenario_id: str,
    default_path: str = "",
) -> list[str]:
    """UI 用に簡易イベントを §10.8 と同型の EVENT 行にする。"""
    from datetime import datetime

    ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    lines: list[str] = []
    for ev in events:
        lines.append(
            "EVENT\t%s\t%s\t%s\t%s\t%s"
            % (
                ts,
                str(ev.get("reason_code") or ""),
                scenario_id,
                str(ev.get("path") or default_path),
                json.dumps(ev, ensure_ascii=False),
            )
        )
    return lines
