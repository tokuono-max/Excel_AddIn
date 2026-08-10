# -*- coding: utf-8 -*-
"""
Python: 3.12+
Module: svc/svc_data_agg_write.py
Created: 2026-03-18
Updated: 2026-07-05
Version: 0.1.7
Purpose:
  データ集約用のマスター書き込み。行追加（append）・強制上書き（overwrite）・空き上書き（fill_in）と、
  照合キーによる行マッチを提供する。一括実行時は内部メモリで組み立てた表を終了時に一括出力、
  ステップ実行時は 1 項目分を都度反映する想定。
  svc_data_agg から呼び出され、サブモジュールとして分離する。
History (latest 3):
  - 0.1.7 (2026-07-05) freeze: SplitRow を全 Window で試行・xlwings は後回し・BaseException/PyErr_Clear で Nuitka COM 例外を抑止。
  - 0.1.6 (2026-07-05) freeze_sheet_below_header_row: SplitRow 優先・Window/FreezePanes 検証・Select 廃止・bool 戻り値。
  - 0.1.5 (2026-06-03) read_master: .xlsm を OpenXML Excel として .xlsx と同経路で読込。
  - 0.1.4 (2026-06-06) suspend_sheet_updates を restore_on_exit=False に（DONE 前の ScreenUpdating 復帰を呼び出し側に委譲）。
  - 0.1.3 (2026-04-14) write_scenario_export_table: 列幅・行の AutoFit をやめ、データ・ヘッダは折り返しなし（シナリオ定義の多列エクスポート向け）。
  - 0.1.2 (2026-04-07) データ集約レポート: 記録日時の右に「処理時間」列。一括サマリ行に wall 秒を表示。
  - 0.1.1 (2026-04-04) write_master_to_sheet に replace_full_block（Excel 上書き／指定セルのブロック置換）。
  - 0.1.0 (2026-03-18) 新規作成。read_master / write_master / apply_rows。
"""
from __future__ import annotations

import csv
import json
import re
import sys
from datetime import datetime
from pathlib import Path
from typing import Any, Optional

_path_svc = Path(__file__).resolve().parent
_root = _path_svc.parent
if str(_root) not in sys.path:
    sys.path.insert(0, str(_root))

from core.core_log import get_logger  # noqa: E402
from svc.svc_data_agg_extract import is_openxml_excel_suffix  # noqa: E402

logger = get_logger(__name__)
__version__ = "0.1.7"

# data_agg は出力行数が大きくなりやすく、全行を対象にした AutoFit が重い。
# 見た目（列幅）よりも「処理完了→Excel操作復帰」を優先し、一定行数超過時はヘッダ行のみ AutoFit に縮退する。
# 上限は config/ui_data_agg.json の EXCEL_WRITE.AUTOFIT_FULL_MAX_ROWS（省略時 3000）。
AUTOFIT_FULL_MAX_ROWS_DEFAULT = 3000

MODE_APPEND = "append"
MODE_OVERWRITE = "overwrite"
MODE_FILL_IN = "fill_in"
MODE_DUPLICATE_APPEND = "duplicate_append"
# 名前・パス系（マスタセルへの連結書き込み）
MODE_PREPEND = "prepend"
MODE_APPEND_END = "append_end"


def _is_unconditional_append_mode(mode: str) -> bool:
    """照合キーを使わず末尾（または専用分岐）へ追記するモード。"""
    m = (mode or "").strip().lower()
    return m in (MODE_APPEND, MODE_DUPLICATE_APPEND)


def excel_write_autofit_full_max_rows(cfg: dict[str, Any] | None = None) -> int:
    """
    全行 AutoFit を行う出力ブロック行数（ヘッダ含む）の上限。
    超過時はヘッダ行のみ AutoFit に縮退する。
    config/ui_data_agg.json の EXCEL_WRITE.AUTOFIT_FULL_MAX_ROWS を参照する。
    """
    raw_cfg = cfg
    if raw_cfg is None:
        try:
            from core import core_cst as cst  # noqa: WPS433

            raw_cfg = cst.get_ui_config_from_file_required("data_agg")
        except Exception:
            return AUTOFIT_FULL_MAX_ROWS_DEFAULT
    ew = raw_cfg.get("EXCEL_WRITE") if isinstance(raw_cfg, dict) else {}
    if not isinstance(ew, dict):
        ew = {}
    try:
        n = int(ew.get("AUTOFIT_FULL_MAX_ROWS", AUTOFIT_FULL_MAX_ROWS_DEFAULT))
    except (TypeError, ValueError):
        n = AUTOFIT_FULL_MAX_ROWS_DEFAULT
    return max(1, n)


def _autofit_max_row_for_block(
    top_row: int,
    block_bottom_row: int,
    *,
    cfg: dict[str, Any] | None = None,
) -> int:
    """AutoFit 対象の最終行。ブロック行数が上限超ならヘッダ行のみ。"""
    tr = max(1, int(top_row))
    bottom = max(tr, int(block_bottom_row))
    limit = excel_write_autofit_full_max_rows(cfg)
    if (bottom - tr + 1) > limit:
        return tr
    return bottom


def _trim_cell_text(v: Any) -> str:
    if v is None:
        return ""
    return str(v).strip()


def merge_cell_for_write_mode(old: Any, new: Any, mode: str) -> Any:
    """
    照合キー一致時の 1 セル更新。prepend / append_end は既存を trim して空なら new のみ（§2 項6）。
    """
    m = (mode or "").strip().lower()
    if m == MODE_OVERWRITE:
        return new
    if m == MODE_FILL_IN:
        if old is None or old == "":
            return new
        return old
    if m == MODE_PREPEND:
        if _trim_cell_text(old) == "":
            return new
        return "%s%s" % (new, old)
    if m == MODE_APPEND_END:
        if _trim_cell_text(old) == "":
            return new
        return "%s%s" % (old, new)
    if m in (MODE_APPEND, MODE_DUPLICATE_APPEND):
        return new
    if old is None or old == "":
        return new
    return old

EVENT_LOG_SHEET = "データ集約レポート"
EVENT_LOG_SHEET_LEGACY = "DataAgg_EventLog"
EVENT_LOG_HEADERS = [
    "記録日時",
    "処理時間",
    "区分",
    "書込み方式",
    "出力シート名",
    "シナリオID",
    "対象パス",
    "詳細",
]


def format_elapsed_ms_ja(ms: int) -> str:
    """ユーザー表示・レポート列用の経過時間ラベル（ミリ秒から）。"""
    if ms < 0:
        ms = 0
    if ms < 1000:
        return "%d ms" % ms
    sec_f = ms / 1000.0
    if sec_f < 60:
        return "%.2f 秒" % sec_f
    total_s = int(round(ms / 1000))
    mi, se = divmod(total_s, 60)
    if mi < 60:
        return "%d 分 %d 秒" % (mi, se)
    ho, mi2 = divmod(mi, 60)
    return "%d 時間 %d 分 %d 秒" % (ho, mi2, se)


def _event_log_reason_ja(code: str) -> str:
    """ログ列「区分」用。未知コードはそのまま。"""
    c = str(code or "").strip()
    m = {
        "PATH_TRACE_PRE_NAME": "パス追跡（名前取得・実行前）",
        "PATH_TRACE_POST_NAME": "パス追跡（名前取得・実行後）",
        "BATCH_OK": "一括実行・完了",
        "BATCH_FAIL": "一括実行・失敗",
        "BATCH_CANCEL": "一括実行・中止",
    }
    return m.get(c, c)


def _locate_event_log_sheet(book: Any) -> Any:
    """データ集約レポートシートを探す。旧名のみなら新名へリネームを試みる。"""
    legacy = None
    for sh in book.sheets:
        try:
            nm = str(sh.name)
        except Exception:
            continue
        if nm == EVENT_LOG_SHEET:
            return sh
        if nm == EVENT_LOG_SHEET_LEGACY:
            legacy = sh
    if legacy is not None:
        try:
            legacy.name = EVENT_LOG_SHEET
        except Exception:
            pass
        return legacy
    return None


def format_path_trace_for_event_log(
    scenario_id: str,
    file_path: str | Path,
    reason_code: str,
    path_col: str,
    headers: list[str],
    snapshot_rows: list[dict[str, Any]],
) -> list[list[Any]]:
    """
    名前取得の照合デバッ用: マージ行の path 関連フィールドを JSON で 1 行にまとめて記録する。
    snapshot_rows は JSON 直列化可能な dict のリスト（行ごと）。
    """
    if not snapshot_rows:
        return []
    ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    sid = str(scenario_id or "")
    fp = str(file_path)
    detail = json.dumps(
        {
            "種別": "パス追跡",
            "パス列": path_col,
            "列見出し": headers,
            "行スナップショット": snapshot_rows,
        },
        ensure_ascii=False,
    )
    return [[ts, "", _event_log_reason_ja(str(reason_code)), "", "", sid, fp, detail]]


def format_join_events_for_event_log(
    scenario_id: str,
    file_path: str | Path,
    events: list[dict[str, Any]],
) -> list[list[Any]]:
    """§10.8 用: 結合イベントをログシート 1 行ずつ（詳細は JSON）に整形する。"""
    if not events:
        return []
    ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    sid = str(scenario_id or "")
    fp = str(file_path)
    return [
        [
            ts,
            "",
            _event_log_reason_ja(str(ev.get("reason_code") or "")),
            "",
            "",
            sid,
            fp,
            json.dumps(ev, ensure_ascii=False),
        ]
        for ev in events
    ]


def format_batch_run_summary_row(
    scenario_id: str,
    scenario_path: str,
    *,
    ok: bool,
    files: int = 0,
    output_rows: int = 0,
    append: int = 0,
    update: int = 0,
    join_events: int = 0,
    compute_ms: int | None = None,
    write_ms: int | None = None,
    total_ms: int | None = None,
    error: str | None = None,
    excel_write_summary: str = "",
    output_sheet_name: str = "",
) -> list[Any]:
    """一括実行の成否サマリをイベントログ 1 行分にする（データ集約レポート）。"""
    ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    err_s = str(error or "").strip().lower()
    is_cancel = err_s == "cancelled"
    if ok:
        code_en = "BATCH_OK"
        result_ja = "成功"
    elif is_cancel:
        code_en = "BATCH_CANCEL"
        result_ja = "中止"
    else:
        code_en = "BATCH_FAIL"
        result_ja = "失敗"
    detail: dict[str, Any] = {
        "種別": "一括実行サマリ",
        "結果": result_ja,
        "ファイル数": files,
        "出力行数": output_rows,
        "追加行数": append,
        "更新行数": update,
        "結合イベント数": join_events,
    }
    if compute_ms is not None:
        detail["集計処理ミリ秒"] = compute_ms
    if write_ms is not None:
        detail["書込ミリ秒"] = write_ms
    if total_ms is not None:
        detail["合計ミリ秒"] = total_ms
    if error:
        detail["エラー"] = error
    elapsed_cell = format_elapsed_ms_ja(int(total_ms)) if total_ms is not None else ""
    return [
        ts,
        elapsed_cell,
        _event_log_reason_ja(code_en),
        str(excel_write_summary or ""),
        str(output_sheet_name or ""),
        str(scenario_id or ""),
        str(scenario_path or ""),
        json.dumps(detail, ensure_ascii=False),
    ]


def append_event_log_rows(book: Any, rows: list[list[Any]]) -> None:
    """アクティブブックにイベントログシートを用意し、行を追記する（§10.8 / Q1）。"""
    if not rows:
        return
    try:
        from core import core_xlc  # noqa: E402
    except ImportError:
        logger.warning("[DATA_AGG_WRITE] core_xlc が無くイベントログをスキップします")
        return
    try:
        ws = _locate_event_log_sheet(book)
        if ws is None:
            ws = book.sheets.add()
            try:
                ws.name = EVENT_LOG_SHEET
            except Exception:
                for i in range(1, 99):
                    alt = f"{EVENT_LOG_SHEET}_{i}"
                    try:
                        ws.name = alt
                        break
                    except Exception:
                        continue
            core_xlc.write_chunk(ws, 1, 1, [EVENT_LOG_HEADERS], text_mode=True)
        else:
            # 旧レイアウト（列数不足）のログシートへ追記する前にヘッダ行を新形式に揃える
            try:
                ur0 = getattr(ws, "used_range", None)
                lc0 = getattr(ur0, "last_cell", None) if ur0 is not None else None
                nc0 = int(lc0.column) if lc0 is not None else 0
                if 0 < nc0 < len(EVENT_LOG_HEADERS):
                    core_xlc.write_chunk(ws, 1, 1, [EVENT_LOG_HEADERS], text_mode=True)
            except Exception:
                pass
        ur = getattr(ws, "used_range", None)
        if ur is None:
            start_row = 2
        else:
            lc = getattr(ur, "last_cell", None)
            last_r = int(lc.row) if lc is not None else 1
            start_row = last_r + 1
        core_xlc.write_chunk(ws, start_row, 1, rows, text_mode=True)
        end_row = start_row + len(rows) - 1
        n_col = len(EVENT_LOG_HEADERS)
        try:
            core_xlc.autofit_sheet_columns(
                ws,
                min_row=1,
                min_col=1,
                max_row=max(1, end_row),
                max_col=max(1, n_col),
                sheet_name_for_visible_fallback=str(
                    getattr(ws, "name", "") or EVENT_LOG_SHEET
                ),
            )
        except Exception:
            pass
        logger.info("[DATA_AGG_WRITE] イベントログ追記 件数=%s", len(rows))
    except Exception as e:
        logger.warning("[DATA_AGG_WRITE] イベントログ追記失敗: %s", e)


def _overlay_row_values(dest: list[Any], src_row: list[Any]) -> None:
    """dest を src_row の値で上書きし、長い方に合わせて延長する（overwrite と同等の列対応）。"""
    for j in range(min(len(dest), len(src_row))):
        dest[j] = src_row[j]
    if len(src_row) > len(dest):
        dest.extend(src_row[len(dest) :])


def read_master(path: str | Path) -> tuple[list[str], list[list[Any]]]:
    """
    マスターファイルを読み、ヘッダ行とデータ行のリストを返す。

    【概要】
      Excel (.xlsx/.xlsm) は OpenPyXL、CSV は csv モジュールで読む。先頭行をヘッダ、2 行目以降をデータとする。
      ファイルが存在しない場合は空のヘッダ・空の行リストを返す。

    【引数】
      path: マスターファイルのパス。

    【戻り値】
      (headers, rows)。headers は列名のリスト、rows は行ごとの値のリストのリスト。
    """
    p = Path(path).resolve()
    if not p.is_file():
        return ([], [])
    suf = p.suffix.lower()
    if suf == ".csv":
        return _read_csv_master(p)
    if is_openxml_excel_suffix(suf):
        return _read_excel_master(p)
    return ([], [])


def _read_csv_master(path: Path) -> tuple[list[str], list[list[Any]]]:
    """CSV マスターを読む。"""
    try:
        with path.open("r", encoding="utf-8-sig", newline="") as f:
            reader = list(csv.reader(f))
        if not reader:
            return ([], [])
        headers = list(reader[0])
        rows = [list(r) for r in reader[1:]]
        return (headers, rows)
    except Exception as e:
        logger.warning("[DATA_AGG_WRITE] CSV 読込エラー %s: %s", path, e)
        return ([], [])


def _read_excel_master(path: Path) -> tuple[list[str], list[list[Any]]]:
    """Excel マスターを読む。先頭シートの先頭行をヘッダとする。"""
    try:
        import openpyxl  # noqa: E402
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        if ws is None:
            return ([], [])
        rows_iter = list(ws.iter_rows(values_only=True))
        wb.close()
        if not rows_iter:
            return ([], [])
        headers = [str(c) if c is not None else "" for c in rows_iter[0]]
        rows = [[c for c in r] for r in rows_iter[1:]]
        return (headers, rows)
    except Exception as e:
        logger.warning("[DATA_AGG_WRITE] Excel 読込エラー %s: %s", path, e)
        return ([], [])


def _write_csv_master(path: Path, headers: list[str], rows: list[list[Any]]) -> None:
    """CSV マスターにヘッダと行を書き込む。"""
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8-sig", newline="") as f:
        w = csv.writer(f)
        w.writerow(headers)
        for r in rows:
            w.writerow(r)


def _write_excel_master(path: Path, headers: list[str], rows: list[list[Any]]) -> None:
    """Excel マスターにヘッダと行を書き込む。"""
    try:
        import openpyxl  # noqa: E402
    except ImportError:
        logger.warning("[DATA_AGG_WRITE] openpyxl が利用できません")
        return
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.Workbook()
    ws = wb.active
    if ws is None:
        return
    ws.append(headers)
    for r in rows:
        ws.append(r)
    wb.save(path)


def _row_key(row: list[Any], key_indices: list[int]) -> tuple:
    """行から照合キーとなる列の値をタプルで返す。"""
    from core.core_join_compare import join_compare_display_key  # noqa: WPS433

    return tuple(
        join_compare_display_key(row[i] if i < len(row) else None) for i in key_indices
    )


def parse_a1_to_row_col_1based(cell_ref: str) -> Optional[tuple[int, int]]:
    """
    A1 形式を (行, 列) の 1 始まりインデックスに変換。失敗時は None。
    """
    from svc.svc_data_agg_extract import _parse_cell_ref  # noqa: WPS433 — 既存の列行パーサを再利用

    s = (cell_ref or "").strip()
    if not s:
        return None
    col0, row0 = _parse_cell_ref(s)
    if col0 is None or row0 is None:
        return None
    return (row0 + 1, col0 + 1)


def _natural_sort_key_for_cell(val: Any) -> list[Any]:
    """セル値を自然順比較用キーにする（文字列化して数字部分を数値化）。"""
    from svc.svc_data_agg_scan import _natural_text_key  # noqa: WPS433

    return _natural_text_key("" if val is None else str(val))


def sort_table_rows_for_excel_options(
    headers: list[str],
    table_rows: list[list[Any]],
    excel_options: dict[str, Any],
) -> list[list[Any]]:
    """
    excel_options["sort_keys"] に従い table_rows を安定ソートする。
    空の item キーは無視。該当列が無いキーも無視。
    """
    sk_raw = excel_options.get("sort_keys")
    if not isinstance(sk_raw, list) or not sk_raw:
        return table_rows
    keys_use: list[tuple[int, bool, bool]] = []
    for ent in sk_raw:
        if not isinstance(ent, dict):
            continue
        name = str(ent.get("item") or "").strip()
        if not name or name not in headers:
            continue
        idx = headers.index(name)
        desc = str(ent.get("order") or "asc").strip().lower() == "desc"
        natural = bool(ent.get("natural"))
        keys_use.append((idx, natural, desc))
    if not keys_use:
        return table_rows
    out = [list(r) for r in table_rows]

    def _plain_key(v: Any) -> str:
        return "" if v is None else str(v)

    for col_idx, natural, desc in reversed(keys_use):

        def _mk_key(row: list[Any], ci: int = col_idx, nat: bool = natural) -> Any:
            if ci >= len(row):
                v = None
            else:
                v = row[ci]
            if nat:
                return _natural_sort_key_for_cell(v)
            return _plain_key(v)

        out.sort(key=_mk_key, reverse=desc)
    return out


def _cell_effectively_blank(v: Any) -> bool:
    return v is None or (isinstance(v, str) and str(v).strip() == "")


def _matrix_all_blank(matrix: list[list[Any]]) -> bool:
    for row in matrix:
        for v in row:
            if not _cell_effectively_blank(v):
                return False
    return True


def _strip_leading_row_if_matches_header(
    body_rows: list[list[Any]], header_row: list[str]
) -> list[list[Any]]:
    """
    body の先頭行が header_row と列見出しとして同一なら除く（ヘッダは常に先頭 1 行にのみ出力するため）。
    """
    if not body_rows or not header_row:
        return body_rows
    hdr = [str(h) if h is not None else "" for h in header_row]
    n_hdr = len(hdr)
    r0 = list(body_rows[0])
    if len(r0) < n_hdr:
        return body_rows
    r0cmp = [str(r0[i]) if r0[i] is not None else "" for i in range(n_hdr)]
    if r0cmp == hdr:
        return body_rows[1:]
    return body_rows


def append_start_row_after_region_read(
    top_row: int,
    headers: list[str],
    rows: list[list[Any]],
) -> int:
    """
    _read_sheet_as_master_region の結果から、データ追記を開始する行（1 始まり）を返す。
    領域が空または内容がすべて空なら top_row。それ以外は既存ブロック（先頭行を含む）の次の行。
    """
    if not headers and not rows:
        return top_row
    mat: list[list[Any]] = [list(headers)] + [list(r) for r in rows]
    if _matrix_all_blank(mat):
        return top_row
    return top_row + len(mat)


def _append_data_start_row_1based(sheet: Any, top_row: int, top_col: int) -> int:
    ph, pr = _read_sheet_as_master_region(sheet, top_row, top_col)
    return append_start_row_after_region_read(top_row, ph, pr)


def _read_sheet_as_master_region(
    sheet: Any, top_row: int, top_col: int
) -> tuple[list[str], list[list[Any]]]:
    """
    (top_row, top_col) を左上とし、シートの UsedRange 右下角までをマスタ形式で読む。
    """
    try:
        from core import core_xlc  # noqa: E402

        ur = getattr(sheet, "used_range", None)
        if ur is None:
            return ([], [])
        api = getattr(ur, "api", None)
        if api is None:
            return ([], [])
        row0 = core_xlc.com_excel_scalar_int(getattr(api, "Row", None), 0)
        col0 = core_xlc.com_excel_scalar_int(getattr(api, "Column", None), 0)
        rows_o = getattr(api, "Rows", None)
        cols_o = getattr(api, "Columns", None)
        nrows = (
            core_xlc.com_excel_scalar_int(getattr(rows_o, "Count", None), 0)
            if rows_o is not None
            else 0
        )
        ncols_ur = (
            core_xlc.com_excel_scalar_int(getattr(cols_o, "Count", None), 0)
            if cols_o is not None
            else 0
        )
        last_row = row0 + nrows - 1
        last_col = col0 + ncols_ur - 1
        if last_row < top_row or last_col < top_col:
            return ([], [])
        raw = sheet.range((top_row, top_col), (last_row, last_col)).value
        if raw is None:
            return ([], [])
        if isinstance(raw, (list, tuple)):
            if len(raw) > 0 and not isinstance(raw[0], (list, tuple)):
                list_2d = [list(raw)]
            else:
                list_2d = [
                    list(row) if isinstance(row, (list, tuple)) else [row]
                    for row in raw
                ]
        else:
            list_2d = [[raw]]
        if not list_2d:
            return ([], [])
        headers = [str(c) if c is not None else "" for c in list_2d[0]]
        rows = [[c for c in r] for r in list_2d[1:]]
        return (headers, rows)
    except Exception as e:
        logger.warning("[DATA_AGG_WRITE] シート領域読込エラー: %s", e)
        return ([], [])


def _read_sheet_as_master(sheet: Any) -> tuple[list[str], list[list[Any]]]:
    """
    xlwings シートからヘッダ行とデータ行を読み、マスター形式で返す（左上 A1）。
    """
    return _read_sheet_as_master_region(sheet, 1, 1)


def _unique_sheet_name_in_book(book: Any, base: str) -> str:
    names = {str(s.name) for s in book.sheets}
    cand = (base or "集約")[:31]
    if cand not in names:
        return cand
    for i in range(1, 9999):
        suf = "_%d" % i
        cand = ((base or "集約")[: 31 - len(suf)] + suf)[:31]
        if cand not in names:
            return cand
    return "集約"[:31]


def _sanitize_sheet_name_prefix(scenario_id: str) -> str:
    """シナリオ ID をシート名の接頭辞に使えるよう整形（Excel 禁止文字除去・長さ制限）。"""
    raw = (scenario_id or "").strip() or "集約"
    s = re.sub(r"[\[\]:*?/\\]", "", raw)
    s = s.strip() or "集約"
    return s[:25]


def sanitize_excel_tab_name(s: str) -> str:
    """操作者入力のシート名を Excel 制約に合わせて整形（禁止文字除去・最大 31 文字）。"""
    t = re.sub(r"[\[\]:*?/\\]", "", (s or "").strip())
    return t[:31]


def _unique_sheet_name_from_existing(names: set[str], base: str) -> str:
    """base をそのまま使えれば使い、重複時は base_2 … で 31 文字以内に収める。"""
    b = (base or "集約").strip() or "集約"
    b = b[:31]
    if b not in names:
        return b
    for i in range(2, 10000):
        suf = "_%d" % i
        cand = (b[: max(1, 31 - len(suf))] + suf)[:31]
        if cand not in names:
            return cand
    return b[:28] + "_99"


def add_data_agg_output_sheet(
    book: Any,
    name_rule: str,
    scenario_id: str,
    *,
    custom_sheet_name: str = "",
) -> Any:
    """データ集約の「新しいシート」を追加して返す。

    name_rule が custom_sheet_name のときは custom_sheet_name をシート名のベースとする（重複時は _連番）。
    それ以外はシナリオ名_連番（scenario_datetime / scenario_seq は scenario_name_seq と同じ扱い）。
    """
    act = book.sheets.active
    names = {str(s.name) for s in book.sheets}
    rule = (name_rule or "").strip().lower()
    if rule == "custom_sheet_name":
        base = sanitize_excel_tab_name(custom_sheet_name)
        if not base:
            base = "集約"
        name_final = _unique_sheet_name_from_existing(names, base)
        return book.sheets.add(name=name_final, after=act)
    _ = name_rule
    prefix = _sanitize_sheet_name_prefix(scenario_id)
    max_n = 0
    for nm in names:
        if nm.startswith(prefix + "_"):
            tail = nm[len(prefix) + 1 :]
            if tail.isdigit():
                try:
                    max_n = max(max_n, int(tail))
                except ValueError:
                    pass
    name_final = ""
    for n in range(max_n + 1, max_n + 10000):
        cand = ("%s_%d" % (prefix, n))[:31]
        if cand not in names:
            name_final = cand
            break
    if not name_final:
        name_final = _unique_sheet_name_in_book(book, (prefix + "_1")[:31])
    return book.sheets.add(name=name_final, after=act)


def sanitize_excel_defined_name_base(s: str) -> str:
    """定義名に使えるようベース文字列を整形（空ならプレースホルダ）。"""
    t = re.sub(r"[^\w\u3040-\u30ff\u4e00-\u9fff]", "_", (s or "").strip())
    t = re.sub(r"_+", "_", t).strip("_")
    if not t:
        return "DataAggJump"
    if t[0].isdigit():
        t = "N_" + t
    return t[:200]


def _abs_a1_cell_1based(row_1: int, col_1: int) -> str:
    from svc.svc_data_agg_extract import _col_row_to_cell_ref  # noqa: WPS433

    core = _col_row_to_cell_ref(col_1 - 1, row_1 - 1)
    m = re.match(r"^([A-Za-z]+)(\d+)$", core, re.I)
    if not m:
        return "$A$1"
    return "$%s$%s" % (m.group(1).upper(), m.group(2))


def try_register_data_agg_jump_range(
    book: Any,
    sheet: Any,
    name_base: str,
    top_row: int,
    top_col: int,
    n_rows: int,
    n_cols: int,
) -> None:
    """書き込み矩形にジャンプ用の定義名を付与する（ベストエフォート）。

    定義名は「シート名_連番」形式（ブック内で重複しない連番）。name_base は互換のため無視する。
    """
    _ = name_base
    if n_rows <= 0 or n_cols <= 0 or top_row < 1 or top_col < 1:
        return
    br = top_row + n_rows - 1
    bc = top_col + n_cols - 1
    shnm = str(getattr(sheet, "name", None) or "Sheet1")
    stem = sanitize_excel_defined_name_base(shnm)
    try:
        api = book.api
        existing: list[str] = []
        try:
            for i in range(1, int(api.Names.Count) + 1):
                try:
                    existing.append(str(api.Names(i).Name))
                except Exception:
                    pass
        except Exception:
            pass
        name_use = ""
        for seq in range(1, 10000):
            cand = ("%s_%d" % (stem[:180], seq))[:255]
            if cand not in existing:
                name_use = cand
                break
        if not name_use:
            name_use = (stem[:200] + "_1")[:255]
        tl = _abs_a1_cell_1based(top_row, top_col)
        br_a = _abs_a1_cell_1based(br, bc)
        if n_rows == 1 and n_cols == 1:
            addr = tl
        else:
            addr = "%s:%s" % (tl, br_a)
        sh_esc = "'" + shnm.replace("'", "''") + "'"
        ref = "=%s!%s" % (sh_esc, addr)
        try:
            api.Names(name_use).Delete()
        except Exception:
            pass
        api.Names.Add(Name=name_use, RefersTo=ref)
    except Exception as e:
        logger.debug("[DATA_AGG_WRITE] ジャンプ定義名の登録をスキップ: %s", e)


def _merge_keyed_row(
    old_row: list[Any],
    new_row: list[Any],
    mode: str,
    column_modes: Optional[list[str]],
) -> None:
    """照合一致時に行をセル単位でマージ（項目ごとに異なる write_mode を column_modes で渡せる）。"""
    while len(old_row) < len(new_row):
        old_row.append(None)
    for j in range(len(new_row)):
        cm = column_modes[j] if column_modes and j < len(column_modes) else mode
        old_row[j] = merge_cell_for_write_mode(old_row[j], new_row[j], cm)


def write_master(
    path: str | Path,
    headers: list[str],
    rows: list[list[Any]],
    mode: str = MODE_APPEND,
    match_key_indices: Optional[list[int]] = None,
    existing_headers: Optional[list[str]] = None,
    existing_rows: Optional[list[list[Any]]] = None,
    column_modes: Optional[list[str]] = None,
) -> tuple[int, int]:
    """
    マスターファイルにデータを反映する。照合キーに応じて追加・上書き・空き上書きを行う。

    【概要】
      existing_* が渡された場合は既存データとして使い、渡されない場合は path から読む。
      mode が append のときは既存の末尾に rows を追加。
      mode が overwrite / fill_in のときは match_key_indices で行を照合し、一致すれば上書き（fill_in は空セルのみ）。

    【引数】
      path: マスターファイルの保存先パス。
      headers: 列名のリスト（新規作成時または既存と同一順）。
      rows: 書き込むデータ行のリスト。
      mode: "append" | "overwrite" | "fill_in"。
      match_key_indices: 照合キーとなる列のインデックスリスト。None のときは照合せず append 扱い。
      existing_headers: 既存ヘッダ。None のときは path から読む。
      existing_rows: 既存データ行。None のときは path から読む。

    【戻り値】
      (追加行数, 更新行数) のタプル。
    """
    p = Path(path).resolve()
    if existing_headers is not None and existing_rows is not None:
        prev_headers = existing_headers
        prev_rows = [list(r) for r in existing_rows]
    else:
        prev_headers, prev_rows = read_master(p)
    key_indices = match_key_indices or []
    append_count = 0
    update_count = 0
    if _is_unconditional_append_mode(mode) or not key_indices:
        prev_rows.extend(rows)
        append_count = len(rows)
    else:
        # キーでマッチして更新 or 追加（overwrite / fill_in）。不一致時は常に行追加。
        existing_keys = {_row_key(r, key_indices): i for i, r in enumerate(prev_rows)}
        for new_row in rows:
            k = _row_key(new_row, key_indices)
            if k in existing_keys:
                idx = existing_keys[k]
                old_row = prev_rows[idx]
                _merge_keyed_row(old_row, new_row, mode, column_modes)
                update_count += 1
            else:
                prev_rows.append(new_row)
                append_count += 1
    if prev_headers and not headers:
        pass  # 既存ヘッダを維持
    else:
        prev_headers = headers
    if p.suffix.lower() == ".csv":
        _write_csv_master(p, prev_headers, prev_rows)
    else:
        _write_excel_master(p, prev_headers, prev_rows)
    logger.info(
        "[DATA_AGG_WRITE] 書き込み マスター=%s モード=%s 追加=%s 更新=%s",
        p,
        mode,
        append_count,
        update_count,
    )
    return (append_count, update_count)


def write_master_to_sheet(
    sheet: Any,
    headers: list[str],
    rows: list[list[Any]],
    mode: str = MODE_APPEND,
    match_key_indices: Optional[list[int]] = None,
    existing_headers: Optional[list[str]] = None,
    existing_rows: Optional[list[list[Any]]] = None,
    column_modes: Optional[list[str]] = None,
    *,
    top_left_row: int = 1,
    top_left_col: int = 1,
    jump_register: bool = False,
    jump_name_base: str = "",
    book_for_jump: Any = None,
    append_chunk_no_header: bool = False,
    replace_full_block: bool = False,
) -> tuple[int, int]:
    """
    アクティブシート（xlwings）にデータを反映する。照合キーに応じて追加・上書き・空き上書きを行う。

    【概要】
      existing_* が渡された場合は既存データとして使い、渡されない場合は sheet から読む。
      mode が append のときは既存の末尾に rows を追加。
      mode が overwrite / fill_in のときは match_key_indices で行を照合し、一致すれば上書き（fill_in は空セルのみ）。

    【引数】
      sheet: xlwings の Sheet オブジェクト。
      headers: 列名のリスト（新規作成時または既存と同一順）。
      rows: 書き込むデータ行のリスト。
      mode: "append" | "overwrite" | "fill_in"。
      match_key_indices: 照合キーとなる列のインデックスリスト。None のときは照合せず append 扱い。
      existing_headers: 既存ヘッダ。None のときは sheet から読む。
      existing_rows: 既存データ行。None のときは sheet から読む。
      top_left_row / top_left_col: 書き込み・読込の左上セル（1 始まり）。既定は A1。
      jump_register: True のとき、書き込み後に定義名でジャンプ先を登録する。
      jump_name_base: 互換用（無視）。定義名は「シート名_連番」で付与する。
      book_for_jump: 定義名を付けるブック。None のときは sheet.book を使用。
      append_chunk_no_header: True のとき Excel「追加」モード。ブロック先頭行（空のとき）なら
      1 行目にヘッダ＋続けてデータ、既存表の下ならデータのみ write_chunk する。
      replace_full_block: True のときシート上の既存ブロックを読まず、(tr,tc) から
      ヘッダ 1 行＋今回の rows のみを write_chunk し、余白をクリアする（Excel 上書き／指定セル用）。

    【戻り値】
      (追加行数, 更新行数) のタプル。
    """
    tr = max(1, int(top_left_row))
    tc = max(1, int(top_left_col))
    if replace_full_block:
        # キー照合・既存行 extend せず、結果ブロックだけを張り替える。
        if not headers and not rows:
            logger.info("[DATA_AGG_WRITE] excel全置換スキップ: ヘッダもデータ行も無し")
            return (0, 0)
        hdr_line = [str(h) if h is not None else "" for h in (headers or [])]
        ncols = len(hdr_line)
        if rows:
            ncols = max(ncols, max(len(r) for r in rows))
        if ncols < 1:
            ncols = 1
        while len(hdr_line) < ncols:
            hdr_line.append("")
        hdr_line = hdr_line[:ncols]

        def _pad_row(r: list[Any]) -> list[Any]:
            rr = list(r)
            while len(rr) < ncols:
                rr.append(None)
            return rr[:ncols]

        chunk_2d = [hdr_line] + [_pad_row(list(r)) for r in rows]
        try:
            from core import core_xlc  # noqa: E402

            bk = book_for_jump if book_for_jump is not None else getattr(sheet, "book", None)
            with core_xlc.suspend_sheet_updates(sheet, restore_on_exit=False):
                core_xlc.write_chunk(sheet, tr, tc, chunk_2d, text_mode=True)
                full_bottom = tr + len(chunk_2d) - 1
                n_rows_rect = len(chunk_2d)
                if tr == 1 and tc == 1:
                    core_xlc.clear_used_range_overflow(sheet, full_bottom, ncols)
                else:
                    core_xlc.clear_used_range_overflow_at(
                        sheet, tr, tc, n_rows_rect, ncols
                    )
            try:
                # 行数が大きい場合はヘッダ行のみ AutoFit（Excel完了待ち短縮）
                max_row_af = _autofit_max_row_for_block(tr, full_bottom)
                core_xlc.autofit_sheet_columns(
                    sheet,
                    min_row=tr,
                    min_col=tc,
                    max_row=max_row_af,
                    max_col=max(tc, tc + ncols - 1),
                )
            except Exception:
                pass
            if jump_register and bk is not None:
                try_register_data_agg_jump_range(
                    bk,
                    sheet,
                    str(jump_name_base or "DataAgg"),
                    tr,
                    tc,
                    n_rows_rect,
                    ncols,
                )
        except Exception as e:
            logger.warning(
                "[DATA_AGG_WRITE] シート全置換書込エラー: %s", e, exc_info=True
            )
            raise
        logger.info(
            "[DATA_AGG_WRITE] excel全置換 top_left=(%s,%s) 行=%s 列=%s",
            tr,
            tc,
            len(chunk_2d),
            ncols,
        )
        return (0, len(rows))
    if append_chunk_no_header:
        if not rows:
            logger.info("[DATA_AGG_WRITE] excel追記スキップ: データ行 0")
            return (0, 0)
        ncols = len(headers) if headers else (len(rows[0]) if rows else 1)
        if ncols < 1:
            ncols = 1
        ph: list[str] = []
        pr: list[list[Any]] = []
        if existing_headers is not None and existing_rows is not None:
            if len(existing_headers) == 0 and len(existing_rows) == 0:
                start_row = tr
                logger.debug(
                    "[DATA_AGG_WRITE] excel追記: 既存明示空のため start_row=top_left (%s)", tr
                )
            else:
                ph, pr = _read_sheet_as_master_region(sheet, tr, tc)
                start_row = append_start_row_after_region_read(tr, ph, pr)
        else:
            ph, pr = _read_sheet_as_master_region(sheet, tr, tc)
            start_row = append_start_row_after_region_read(tr, ph, pr)
        include_header = start_row == tr
        hdr_line = [str(h) if h is not None else "" for h in (headers or [])]
        while len(hdr_line) < ncols:
            hdr_line.append("")
        hdr_line = hdr_line[:ncols]
        try:
            from core import core_xlc  # noqa: E402

            bk = book_for_jump if book_for_jump is not None else getattr(sheet, "book", None)
            with core_xlc.suspend_sheet_updates(sheet, restore_on_exit=False):
                if include_header:
                    chunk_2d = [hdr_line] + [list(r) for r in rows]
                    core_xlc.write_chunk(sheet, tr, tc, chunk_2d, text_mode=True)
                    full_bottom = tr + len(chunk_2d) - 1
                else:
                    core_xlc.write_chunk(sheet, start_row, tc, rows, text_mode=True)
                    full_bottom = start_row + len(rows) - 1
                n_rows_rect = full_bottom - tr + 1
                if tr == 1 and tc == 1:
                    core_xlc.clear_used_range_overflow(sheet, full_bottom, ncols)
                else:
                    core_xlc.clear_used_range_overflow_at(
                        sheet, tr, tc, n_rows_rect, ncols
                    )
            try:
                max_row_af = _autofit_max_row_for_block(tr, full_bottom)
                core_xlc.autofit_sheet_columns(
                    sheet,
                    min_row=tr,
                    min_col=tc,
                    max_row=max_row_af,
                    max_col=max(tc, tc + ncols - 1),
                )
            except Exception:
                pass
            if jump_register and bk is not None:
                try_register_data_agg_jump_range(
                    bk,
                    sheet,
                    str(jump_name_base or "DataAgg"),
                    tr,
                    tc,
                    n_rows_rect,
                    ncols,
                )
        except Exception as e:
            logger.warning(
                "[DATA_AGG_WRITE] シート追記書込エラー: %s", e, exc_info=True
            )
        logger.info(
            "[DATA_AGG_WRITE] excel追記 top_left=(%s,%s) start_row=%s include_header=%s "
            "ncols=%s 今回データ行=%s 読取ヘッダ列=%s 読取データ行=%s",
            tr,
            tc,
            start_row,
            include_header,
            ncols,
            len(rows),
            len(ph),
            len(pr),
        )
        return (len(rows), 0)
    if existing_headers is not None and existing_rows is not None:
        prev_headers = existing_headers
        prev_rows = [list(r) for r in existing_rows]
    else:
        prev_headers, prev_rows = _read_sheet_as_master_region(sheet, tr, tc)
    key_indices = match_key_indices or []
    append_count = 0
    update_count = 0
    if _is_unconditional_append_mode(mode) or not key_indices:
        prev_rows.extend(rows)
        append_count = len(rows)
    else:
        existing_keys = {_row_key(r, key_indices): i for i, r in enumerate(prev_rows)}
        for new_row in rows:
            k = _row_key(new_row, key_indices)
            if k in existing_keys:
                idx = existing_keys[k]
                old_row = prev_rows[idx]
                _merge_keyed_row(old_row, new_row, mode, column_modes)
                update_count += 1
            else:
                prev_rows.append(new_row)
                append_count += 1
    if prev_headers and not headers:
        pass
    else:
        prev_headers = headers
    prev_rows = _strip_leading_row_if_matches_header(prev_rows, prev_headers)
    # シートへ書き込み（先頭 1 行のみヘッダ）
    try:
        from core import core_xlc  # noqa: E402

        data_2d = [prev_headers] + prev_rows
        if data_2d:
            bk = book_for_jump if book_for_jump is not None else getattr(sheet, "book", None)
            logger.info(
                "[DATA_AGG_WRITE] シート全量書込 top_left=(%s,%s) 出力行数=%s (先頭1行=ヘッダ) "
                "mode=%s merge_append=%s merge_update=%s",
                tr,
                tc,
                len(data_2d),
                mode,
                append_count,
                update_count,
            )
            with core_xlc.suspend_sheet_updates(sheet, restore_on_exit=False):
                core_xlc.write_chunk(sheet, tr, tc, data_2d, text_mode=True)
                if tr == 1 and tc == 1:
                    core_xlc.clear_used_range_overflow(
                        sheet, len(data_2d), len(prev_headers)
                    )
                else:
                    core_xlc.clear_used_range_overflow_at(
                        sheet, tr, tc, len(data_2d), len(prev_headers)
                    )
            try:
                max_row_af = _autofit_max_row_for_block(tr, max(tr, tr + len(data_2d) - 1))
                core_xlc.autofit_sheet_columns(
                    sheet,
                    min_row=tr,
                    min_col=tc,
                    max_row=max_row_af,
                    max_col=max(tc, tc + len(prev_headers) - 1),
                )
            except Exception:
                pass
            if jump_register and bk is not None:
                try_register_data_agg_jump_range(
                    bk,
                    sheet,
                    str(jump_name_base or "DataAgg"),
                    tr,
                    tc,
                    len(data_2d),
                    len(prev_headers),
                )
    except Exception as e:
        logger.warning("[DATA_AGG_WRITE] シート書込エラー: %s", e)
    logger.info(
        "[DATA_AGG_WRITE] シート書込 モード=%s 追加=%s 更新=%s",
        mode,
        append_count,
        update_count,
    )
    return (append_count, update_count)


def _col_1based_to_letters(col: int) -> str:
    n = max(1, int(col))
    letters = ""
    while n > 0:
        n, rem = divmod(n - 1, 26)
        letters = chr(65 + rem) + letters
    return letters


def _a1_cell_1based(row: int, col: int) -> str:
    return "%s%d" % (_col_1based_to_letters(col), max(1, int(row)))


def _clear_native_exception_state() -> None:
    """pywin32 com_error 捕捉後に C レベルで残る例外状態をクリアする（Nuitka 等）。"""
    try:
        import ctypes

        ctypes.pythonapi.PyErr_Clear()
    except Exception:
        pass


def _activate_sheet_best_effort(sheet: Any) -> None:
    try:
        sheet.activate()
    except Exception:
        _clear_native_exception_state()


def _sheet_api_name(sheet_api: Any) -> str:
    try:
        return str(getattr(sheet_api, "Name", "") or "")
    except Exception:
        _clear_native_exception_state()
    return ""


def _window_displays_sheet(aw: Any, sheet_api: Any) -> bool:
    """Window が対象 Worksheet を表示しているか。"""
    target = _sheet_api_name(sheet_api)
    if not target:
        return False
    try:
        active = getattr(aw, "ActiveSheet", None)
        if active is None:
            return False
        return str(getattr(active, "Name", "") or "") == target
    except Exception:
        _clear_native_exception_state()
    return False


def _workbook_app_api(sheet: Any) -> Any | None:
    try:
        book = getattr(sheet, "book", None)
        app = getattr(book, "app", None) if book is not None else None
        return getattr(app, "api", None) if app is not None else None
    except Exception:
        _clear_native_exception_state()
    return None


def _iter_workbook_windows(sheet_api: Any):
    """ブックに属する Window を列挙する（best-effort）。"""
    try:
        wins = sheet_api.Parent.Windows
        count = int(getattr(wins, "Count", 0) or 0)
        for idx in range(1, count + 1):
            yield wins(idx)
    except Exception:
        _clear_native_exception_state()


def _get_window_for_sheet(sheet: Any) -> Any | None:
    """対象シートを表示している Window を取得する（誤 Window への固定を避ける）。"""
    api = getattr(sheet, "api", None)
    if api is None:
        return None

    _activate_sheet_best_effort(sheet)

    app_api = _workbook_app_api(sheet)
    if app_api is not None:
        try:
            aw = getattr(app_api, "ActiveWindow", None)
            if aw is not None and _window_displays_sheet(aw, api):
                return aw
        except Exception:
            _clear_native_exception_state()

    try:
        wins = api.Parent.Windows
        count = int(getattr(wins, "Count", 0) or 0)
        for idx in range(1, count + 1):
            win = wins(idx)
            if _window_displays_sheet(win, api):
                return win
    except Exception:
        _clear_native_exception_state()

    _activate_sheet_best_effort(sheet)
    if app_api is not None:
        try:
            aw = getattr(app_api, "ActiveWindow", None)
            if aw is not None:
                return aw
        except Exception:
            _clear_native_exception_state()
    return None


def _verify_freeze_applied(aw: Any, *, header_row: int, left_col: int) -> bool:
    """FreezePanes と Split 位置が期待どおりか。"""
    hr = max(1, int(header_row))
    lc = max(1, int(left_col))
    try:
        if not bool(getattr(aw, "FreezePanes", False)):
            return False
        split_row = int(getattr(aw, "SplitRow", 0) or 0)
        split_col = int(getattr(aw, "SplitColumn", 0) or 0)
        return split_row == hr and split_col == max(0, lc - 1)
    except Exception:
        _clear_native_exception_state()
    return False


def _prepare_window_view_for_freeze(aw: Any, *, header_row: int, left_col: int) -> None:
    """Select を使わず表示位置だけ整える（固定直後の見え方を安定させる）。"""
    hr = max(1, int(header_row))
    lc = max(1, int(left_col))
    try:
        aw.ScrollRow = max(1, hr)
        aw.ScrollColumn = max(1, lc)
    except Exception:
        _clear_native_exception_state()


def _freeze_window_via_split(
    aw: Any,
    *,
    header_row: int,
    left_col: int,
) -> bool:
    """Select を使わず SplitRow/SplitColumn でウィンドウ枠を固定する。"""
    hr = max(1, int(header_row))
    lc = max(1, int(left_col))
    try:
        aw.FreezePanes = False
        aw.SplitRow = hr
        aw.SplitColumn = max(0, lc - 1)
        aw.FreezePanes = True
        if _verify_freeze_applied(aw, header_row=hr, left_col=lc):
            _prepare_window_view_for_freeze(aw, header_row=hr, left_col=lc)
            return True
    except Exception:
        _clear_native_exception_state()
    return False


def _try_freeze_via_xlwings(
    sheet: Any,
    freeze_cell: str,
    *,
    header_row: int,
    left_col: int,
) -> bool:
    """xlwings freeze_at を試し、FreezePanes を検証してから True を返す。"""
    hr = max(1, int(header_row))
    lc = max(1, int(left_col))
    try:
        fp = getattr(sheet, "freeze_panes", None)
        if fp is None:
            return False
        freeze_at = getattr(fp, "freeze_at", None)
        if not callable(freeze_at):
            return False
        unfreeze = getattr(fp, "unfreeze", None)
        if callable(unfreeze):
            unfreeze()
        freeze_at(freeze_cell)
        _clear_native_exception_state()
        aw = _get_window_for_sheet(sheet)
        if aw is not None and _verify_freeze_applied(aw, header_row=hr, left_col=lc):
            _prepare_window_view_for_freeze(aw, header_row=hr, left_col=lc)
            return True
    except Exception:
        _clear_native_exception_state()
    return False


def freeze_sheet_below_header_row(
    sheet: Any,
    header_row: int,
    *,
    left_col: int = 1,
) -> bool:
    """ヘッダ行の直下でウィンドウ枠を固定する（header_row=1 で1行目固定）。

    Returns:
        FreezePanes が期待位置で True になったとき True。
    """
    hr = max(1, int(header_row))
    lc = max(1, int(left_col))
    freeze_cell = _a1_cell_1based(hr + 1, lc)

    _clear_native_exception_state()
    try:
        _activate_sheet_best_effort(sheet)

        api = getattr(sheet, "api", None)
        if api is not None:
            for aw in _iter_workbook_windows(api):
                _activate_sheet_best_effort(sheet)
                if _freeze_window_via_split(aw, header_row=hr, left_col=lc):
                    return True

            aw = _get_window_for_sheet(sheet)
            if aw is not None and _freeze_window_via_split(
                aw, header_row=hr, left_col=lc
            ):
                return True

        if _try_freeze_via_xlwings(
            sheet, freeze_cell, header_row=hr, left_col=lc
        ):
            return True
        return False
    except BaseException:
        _clear_native_exception_state()
        return False
    finally:
        _clear_native_exception_state()


def _clear_worksheet_autofilter(ws: Any) -> None:
    """既存オートフィルタを解除する（best-effort）。"""
    try:
        if bool(ws.AutoFilterMode):
            ws.AutoFilter.ShowAllData()
            return
    except Exception:
        pass
    try:
        ws.AutoFilterMode = False
    except Exception:
        pass


def apply_autofilter_to_block(
    sheet: Any,
    *,
    top_row: int,
    left_col: int,
    n_rows: int,
    n_cols: int,
) -> bool:
    """矩形範囲の先頭行をヘッダとしてオートフィルタを付ける。成功時 True。"""
    nr = max(1, int(n_rows))
    nc = max(1, int(n_cols))
    tr = max(1, int(top_row))
    tc = max(1, int(left_col))
    br = tr + nr - 1
    bc = tc + nc - 1
    addr = "%s:%s" % (_a1_cell_1based(tr, tc), _a1_cell_1based(br, bc))
    try:
        try:
            sheet.activate()
        except Exception:
            pass
        ws = getattr(sheet, "api", None)
        if ws is None:
            return False
        _clear_worksheet_autofilter(ws)
        rng = ws.Range(addr)
        applied = False
        try:
            import pythoncom

            rng.AutoFilter(
                pythoncom.Missing,
                pythoncom.Missing,
                pythoncom.Missing,
                pythoncom.Missing,
                pythoncom.Missing,
            )
            applied = bool(ws.AutoFilterMode)
        except Exception:
            pass
        if not applied:
            try:
                rng.AutoFilter()
                applied = bool(ws.AutoFilterMode)
            except Exception:
                pass
        if not applied:
            try:
                sheet.range(addr).api.AutoFilter()
                applied = bool(ws.AutoFilterMode)
            except Exception:
                pass
        if not applied:
            logger.warning(
                "[DATA_AGG_WRITE] オートフィルタ未適用 addr=%s AutoFilterMode=%s",
                addr,
                getattr(ws, "AutoFilterMode", None),
            )
        return applied
    except Exception as e:
        logger.warning("[DATA_AGG_WRITE] オートフィルタ例外 addr=%s: %s", addr, e)
        return False


def apply_new_sheet_view_options(
    sheet: Any,
    *,
    top_left_row: int = 1,
    top_left_col: int = 1,
    n_rows_including_header: int,
    n_cols: int,
    freeze_header_row: bool = False,
    autofilter: bool = False,
) -> None:
    """新規シート出力直後: ヘッダ行固定とオートフィルタ（best-effort）。"""
    if not freeze_header_row and not autofilter:
        return
    tr = max(1, int(top_left_row))
    tc = max(1, int(top_left_col))
    nr = max(1, int(n_rows_including_header))
    nc = max(1, int(n_cols))
    try:
        sheet.activate()
    except Exception:
        pass
    if autofilter:
        try:
            ok_af = apply_autofilter_to_block(
                sheet,
                top_row=tr,
                left_col=tc,
                n_rows=nr,
                n_cols=nc,
            )
            if ok_af:
                logger.info(
                    "[DATA_AGG_WRITE] 新規シート: オートフィルタ rows=%s cols=%s top=(%s,%s)",
                    nr,
                    nc,
                    tr,
                    tc,
                )
        except Exception as e:
            logger.warning("[DATA_AGG_WRITE] オートフィルタエラー: %s", e)
    if freeze_header_row:
        _clear_native_exception_state()
        ok_fr = freeze_sheet_below_header_row(sheet, tr, left_col=tc)
        _clear_native_exception_state()
        if ok_fr:
            logger.info(
                "[DATA_AGG_WRITE] 新規シート: ヘッダ行固定 header_row=%s left_col=%s",
                tr,
                tc,
            )
        else:
            logger.warning(
                "[DATA_AGG_WRITE] ヘッダ行固定未適用 header_row=%s left_col=%s",
                tr,
                tc,
            )


def _scenario_export_apply_layout(
    sheet_pointer: Any,
    *,
    n_col: int,
    hdr_row: int,
    data_start: int,
    end_row: int,
    has_title: bool,
) -> None:
    """タイトル・ヘッダ・データを左寄せ。シナリオ定義表は折り返しなし（長文は隣列へはみ出し可）。"""
    xl_left = -4131
    xl_top = -4160
    try:
        if has_title:
            try:
                rng_t = sheet_pointer.range(1, 1).resize(1, n_col)
                rng_t.api.HorizontalAlignment = xl_left
                rng_t.api.VerticalAlignment = xl_top
            except Exception:
                pass
        try:
            rng_h = sheet_pointer.range((hdr_row, 1), (hdr_row, n_col))
            rng_h.api.HorizontalAlignment = xl_left
            rng_h.api.VerticalAlignment = xl_top
            rng_h.api.WrapText = False
        except Exception:
            pass
        if end_row >= data_start:
            try:
                rng_d = sheet_pointer.range((data_start, 1), (end_row, n_col))
                rng_d.api.HorizontalAlignment = xl_left
                rng_d.api.VerticalAlignment = xl_top
                rng_d.api.WrapText = False
            except Exception:
                pass
    except Exception:
        pass


def write_scenario_export_table(
    sheet_pointer: Any,
    headers: list[str],
    rows: list[list[Any]],
    *,
    sheet_title: str = "",
) -> None:
    """シナリオ定義エクスポート: 1 行目タイトル（結合・左寄せ）・2 行目ヘッダ・3 行目以降データ。列幅・行高の自動調整は行わない。"""
    if not headers:
        return
    try:
        from core import core_xlc  # noqa: E402
    except ImportError:
        logger.warning("[DATA_AGG_WRITE] core_xlc が無くシナリオ出力をスキップします")
        return
    n_col = max(1, len(headers))
    title = (sheet_title or "").strip()
    has_title = bool(title)
    try:
        if title:
            sheet_pointer.range(1, 1).value = title
            try:
                rng = sheet_pointer.range(1, 1).resize(1, n_col)
                rng.merge()
            except Exception:
                pass
            hdr_row = 2
            data_start = 3
        else:
            hdr_row = 1
            data_start = 2
        core_xlc.write_chunk(sheet_pointer, hdr_row, 1, [headers], text_mode=True)
        if rows:
            core_xlc.write_chunk(sheet_pointer, data_start, 1, rows, text_mode=True)
        end_row = (data_start + len(rows) - 1) if rows else hdr_row
        _scenario_export_apply_layout(
            sheet_pointer,
            n_col=n_col,
            hdr_row=hdr_row,
            data_start=data_start,
            end_row=end_row,
            has_title=has_title,
        )
    except Exception:
        logger.warning(
            "[DATA_AGG_WRITE] write_scenario_export_table failed", exc_info=True
        )
