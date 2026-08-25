# -*- coding: utf-8 -*-
"""
Python: 3.12+
Module: svc/svc_data_agg_extract.py
Created: 2026-03-18
Updated: 2026-08-24
Version: 0.1.19
Purpose:
  データ集約用の抽出エンジン。座標（絶対セル）・メタデータ（パス・フォルダ名・ファイル名）・
  ファイル名からの文字列抽出（範囲・デリミタ・正規表現）を提供する。OpenPyXL / csv で Excel/CSV を直接読む。
  svc_data_agg から呼び出され、サブモジュールとして分離する。
History (latest 3):
  - 0.1.19 (2026-08-24) 空スキップ後の非連番 rule_iter でも連携/結合を列一括読取して拾う。
  - 0.1.18 (2026-08-24) 連携/結合列が主キーと同じ長さ・同じ iter なら揃えコピーを省略。
  - 0.1.17 (2026-08-24) 一部ソースのみの連携/結合列を主キー行に揃える。短い列の iter 振り直しで他ソース行へ載る不具合を修正。
"""
from __future__ import annotations

import csv
import importlib
import os
import re
import sys
import threading
import time
from collections.abc import Iterator
from contextlib import contextmanager
from pathlib import Path
from typing import Any, Callable, Optional

_path_svc = Path(__file__).resolve().parent
_root = _path_svc.parent
if str(_root) not in sys.path:
    sys.path.insert(0, str(_root))

from core.core_log import get_logger  # noqa: E402
from svc.data_agg_extract_limit import (  # noqa: E402
    record_extract_truncation_if_needed,
    skip_extract_truncation_peek,
    resolve_extract_repeat_limit,
)
from svc.data_agg_primary_end import (  # noqa: E402
    END_MODE_N_COUNT,
    END_MODE_UNTIL_EMPTY,
    END_MODE_UNTIL_LAST,
    apply_until_last_trim,
    effective_skip_primary_tokens,
    is_blank_primary_value,
    primary_value_matches_skip_tokens,
    source_end_mode,
    source_keep_empty_primary_slots,
    source_wants_skip_primary,
)
from svc.data_agg_source_ui import source_ui_block  # noqa: E402
from svc.data_agg_sheet_resolve import (  # noqa: E402
    list_workbook_sheet_names,
    parse_comma_separated_patterns,
    patch_item_sheet_exact,
    resolve_all_sheet_names_by_rule,
    source_skips_sheet_extract,
)
from svc.data_agg_excel_read import (  # noqa: E402
    extract_read_openpyxl_cell,
    extract_read_openpyxl_row,
)
from svc.data_agg_value_post import (  # noqa: E402
    postprocess_cell_primary,
    postprocess_cell_primary_batch,
    postprocess_link_rule_value,
    postprocess_link_rule_value_batch,
    postprocess_metadata_like_primary,
    postprocess_name_extract_primary,
)

logger = get_logger(__name__)
__version__ = "0.1.16"


class DataAggCsvReadError(Exception):
    """バッチスコープ内で CSV 行列化に失敗した場合（ファイル再読込にはフォールバックしない）。"""

_OPENXML_EXCEL_SUFFIXES = frozenset({".xlsx", ".xlsm"})


def is_openxml_excel_suffix(suffix: str) -> bool:
    """OpenPyXL で read_only 読取可能な Excel 拡張子（.xlsx / .xlsm）。"""
    return (suffix or "").lower() in _OPENXML_EXCEL_SUFFIXES


def _poll_cancel_check(
    cancel_check: Optional[Callable[..., None]], *, force: bool = False
) -> None:
    if cancel_check is not None:
        cancel_check(force=force)

# compute_batch 等で同一 .xlsx を複数項目・複数セル参照するときの重複 load_workbook を避ける。
_tls_wb_scope = threading.local()
# DATA_AGG_PER_FILE_TIMING=1: load_workbook の秒をパスキー別に累積（consume で取り出し）
_tls_wb_open_sec = threading.local()
_SHEET_MATERIALIZE_THRESHOLD = 8
# 縦/横反復がこの件数以上のとき、ファイル処理開始時にシートを先に materialize する。
_SHEET_PRECACHE_REPEAT_MIN = 64


def _per_file_workbook_timing_enabled() -> bool:
    from core import core_env

    return core_env.data_agg_file_timing_enabled()


def _add_workbook_open_seconds(path_key: str, seconds: float) -> None:
    if seconds <= 0 or not _per_file_workbook_timing_enabled():
        return
    d = getattr(_tls_wb_open_sec, "by_path", None)
    if d is None:
        d = {}
        _tls_wb_open_sec.by_path = d
    d[path_key] = float(d.get(path_key, 0.0)) + float(seconds)


def consume_workbook_open_ms_for_path(file_path: str) -> int:
    """
    当該ファイルパスについてスコープ内で計測した openpyxl.load_workbook 合計を ms で返し、内部を消す。
    OpenXML Excel（.xlsx/.xlsm）以外や未ロード時は 0。
    """
    if not _per_file_workbook_timing_enabled():
        return 0
    d = getattr(_tls_wb_open_sec, "by_path", None)
    if not d:
        return 0
    try:
        key = str(Path(file_path).resolve())
    except Exception:
        key = str(file_path)
    sec = float(d.pop(key, 0.0))
    return int(sec * 1000.0 + 0.5)


def new_workbook_cache_frame() -> dict[str, Any]:
    """
    workbook / シート行列 / CSV キャッシュ用の空フレーム。
    マスタ項目単位など、呼び出し側が寿命管理する共有キャッシュに使う。
    """
    return {
        "wbs": {},
        "sheet_mats": {},
        "sheet_hits": {},
        "csv_mats": {},
        "csv_dfs": {},
        "xls_books": {},
        "xls_sheet_mats": {},
    }


def close_workbook_cache_frame(frame: Optional[dict[str, Any]]) -> None:
    """フレーム内の Workbook / xlrd Book を閉じ、辞書を空にする。"""
    if not frame:
        return
    for wb in list((frame.get("wbs") or {}).values()):
        try:
            wb.close()
        except Exception:
            pass
    try:
        from svc.data_agg_xls_io import close_xls_workbook
    except Exception:
        close_xls_workbook = None  # type: ignore[assignment]
    for book in list((frame.get("xls_books") or {}).values()):
        if close_xls_workbook is not None:
            try:
                close_xls_workbook(book)
            except Exception:
                pass
        else:
            try:
                release = getattr(book, "release_resources", None)
                if callable(release):
                    release()
            except Exception:
                pass
    for k in (
        "wbs",
        "sheet_mats",
        "sheet_hits",
        "csv_mats",
        "csv_dfs",
        "xls_books",
        "xls_sheet_mats",
    ):
        d = frame.get(k)
        if isinstance(d, dict):
            d.clear()


def _xlsx_scope_stack() -> list[dict[str, Any]]:
    stack: list[dict[str, Any]] = getattr(_tls_wb_scope, "stack", None)
    if stack is None:
        stack = []
        _tls_wb_scope.stack = stack
    return stack


@contextmanager
def xlsx_workbook_scope(
    *, shared_frame: Optional[dict[str, Any]] = None
) -> Iterator[None]:
    """
    入れ子可。スレッドローカル。同一スコープ内では resolve 済みパス文字列キーで read_only Workbook を1つだけ保持し、
    終了時にまとめて close する。スコープ外では従来どおり extract_cell ごとに開閉する。
    各フレームは wbs（パス→Workbook）と sheet_mats（(パスキー, シート名)→行行列）を持つ。

    shared_frame を渡した場合は既存フレームを TLS に載せるだけ（exit で close しない）。
    マスタ項目単位キャッシュなど、別スレッドから同じ dict を bind し直す用途。
    """
    stack = _xlsx_scope_stack()
    owned = shared_frame is None
    frame = new_workbook_cache_frame() if owned else shared_frame
    assert frame is not None
    stack.append(frame)
    try:
        yield
    finally:
        stack.pop()
        if owned:
            close_workbook_cache_frame(frame)


@contextmanager
def bind_workbook_cache_frame(frame: dict[str, Any]) -> Iterator[None]:
    """既存フレームを現在スレッドの TLS に載せる（exit で close しない）。"""
    with xlsx_workbook_scope(shared_frame=frame):
        yield


def _xlsx_workbook_cache_top() -> Optional[dict[str, Any]]:
    stack = getattr(_tls_wb_scope, "stack", None)
    if not stack:
        return None
    return stack[-1]


def xlsx_workbook_scope_active() -> bool:
    """スレッド内に workbook キャッシュスコープがあるとき True。"""
    return _xlsx_workbook_cache_top() is not None


def _xlsx_cache_path_key(file_path: str | Path) -> str:
    try:
        return str(Path(file_path).resolve())
    except Exception:
        return str(file_path)


def xlsx_workbook_path_cached(file_path: str | Path) -> bool:
    """
    現在の TLS スコープ先頭フレームに、当該パスのブック／CSV 実体が既にあるとき True。
    （シート行列だけのヒットは「未オープン」扱い → 進捗は [F]）
    """
    frame = _xlsx_workbook_cache_top()
    if frame is None:
        return False
    key = _xlsx_cache_path_key(file_path)
    if key in (frame.get("wbs") or {}):
        return True
    if key in (frame.get("xls_books") or {}):
        return True
    if key in (frame.get("csv_mats") or {}):
        return True
    if key in (frame.get("csv_dfs") or {}):
        return True
    return False


def xlsx_progress_cache_mark(file_path: str | Path) -> str:
    """進捗文言用: キャッシュ再利用なら '[C] '、新規読込なら '[F] '。"""
    return "[C] " if xlsx_workbook_path_cached(file_path) else "[F] "


def _load_workbook_readonly(path: Path) -> Any:
    """read_only + data_only。リンク展開を省略して読込を軽くする。"""
    import openpyxl  # noqa: E402

    return openpyxl.load_workbook(
        path,
        read_only=True,
        data_only=True,
        keep_links=False,
    )


def _xlsx_workbook_from_cache(path: Path) -> Optional[Any]:
    """スコープ内ならキャッシュから取得または load して登録。スコープ外は None。"""
    frame = _xlsx_workbook_cache_top()
    if frame is None:
        return None
    if not is_openxml_excel_suffix(path.suffix):
        return None
    wbs: dict[str, Any] = frame.setdefault("wbs", {})
    key = str(path.resolve())
    wb = wbs.get(key)
    if wb is not None:
        return wb
    try:
        from svc.data_agg_cancel import poll_active_cancel  # noqa: WPS433

        poll_active_cancel(force=True)
    except Exception:
        pass
    try:
        import openpyxl  # noqa: E402
    except ImportError:
        return None
    try:
        if _per_file_workbook_timing_enabled():
            t_ld = time.perf_counter()
            wb = _load_workbook_readonly(path)
            _add_workbook_open_seconds(key, time.perf_counter() - t_ld)
        else:
            wb = _load_workbook_readonly(path)
    except Exception as e:
        logger.debug("[DATA_AGG_EXTRACT] Excel 読込エラー %s: %s", path, e)
        return None
    wbs[key] = wb
    return wb

_POLARS_MODULE: Any | None = None
_POLARS_CHECKED = False


def _get_polars() -> Any | None:
    global _POLARS_MODULE, _POLARS_CHECKED
    if _POLARS_CHECKED:
        return _POLARS_MODULE
    try:
        _POLARS_MODULE = importlib.import_module("polars")
    except Exception:
        _POLARS_MODULE = None
    _POLARS_CHECKED = True
    return _POLARS_MODULE


def _has_polars() -> bool:
    return _get_polars() is not None


def source_passes_file_name_filter(file_path: str | Path, src: dict[str, Any]) -> bool:
    """
    セル系ソースの UI 保存ブロック（`ui_scenario_source_v1`、レガシーキーは source_ui_block で吸収）内の file_pattern / file_name_rule を評価する。
    file_pattern が空（トークンなし）ならフィルタなし（True）。大小文字は区別しない（§10.3.3）。

    複数パターンはカンマ区切り（``parse_comma_separated_patterns``。シート名条件と同じ）。
    - 完全一致／含む: いずれかのトークンに該当（OR）
    - 含まない: いずれのトークンも含まない（AND of exclusions）
    """
    stype = (src.get("type") or "cell").strip().lower()
    if stype != "cell":
        return True
    block = source_ui_block(src)
    if not isinstance(block, dict):
        return True
    p = Path(file_path)
    stem_l = p.stem.lower()
    ext_l = p.suffix.lower()
    ext_checked = block.get("ext_checked")
    if isinstance(ext_checked, list) and ext_checked:
        norm_exts = {
            (str(e).strip().lower() if str(e).strip().startswith(".") else "." + str(e).strip().lower())
            for e in ext_checked
            if str(e).strip()
        }
        if norm_exts and ext_l not in norm_exts:
            return False
    tokens = parse_comma_separated_patterns(block.get("file_pattern"))
    if not tokens:
        return True
    rule = str(block.get("file_name_rule") or "含む").strip()
    toks_l = [t.lower() for t in tokens]
    if "完全一致" in rule or rule.lower() in ("exact", "equals"):
        return stem_l in set(toks_l)
    if "含まない" in rule or rule.lower() in ("exclude", "not_contains"):
        return all(t not in stem_l for t in toks_l)
    return any(t in stem_l for t in toks_l)


def matching_sheets_for_item(
    file_path: str | Path,
    item_config: dict[str, Any],
) -> list[str] | None:
    """
    項目内の cell ソース（ファイル名条件に合うもの）のシート条件を和集合し、
    ブック左→右順で返す。名前取得のみ・CSV は None。該当なしは []。
    """
    sheets, _names = matching_sheets_and_names_for_item(file_path, item_config)
    return sheets


def matching_sheets_and_names_for_item(
    file_path: str | Path,
    item_config: dict[str, Any],
) -> tuple[list[str] | None, list[str] | None]:
    """matching_sheets_for_item に加え、ブックのシート名一覧を返す（CSV 等は両方 None）。"""
    sources = item_config.get("sources") or []
    has_cell = False
    for src in sources:
        if not isinstance(src, dict) or source_skips_sheet_extract(src):
            continue
        stype = str(src.get("type") or "cell").strip().lower()
        if stype in ("name_extract", "metadata", "meta", "filename"):
            continue
        has_cell = True
        break
    if not has_cell:
        return None, None
    names = list_workbook_sheet_names(file_path)
    if names is None:
        return None, None
    wanted: set[str] = set()
    for src in sources:
        if not isinstance(src, dict) or source_skips_sheet_extract(src):
            continue
        stype = str(src.get("type") or "cell").strip().lower()
        if stype in ("name_extract", "metadata", "meta", "filename"):
            continue
        if not source_passes_file_name_filter(file_path, src):
            continue
        sn = str(src.get("sheet_name") or "").strip()
        pb = source_ui_block(src) or {}
        rule = str(pb.get("sheet_rule") or "")
        wanted.update(resolve_all_sheet_names_by_rule(names, rule, sn))
    return [n for n in names if n in wanted], names


def _mini_spans_from_local(
    local_spans: Any,
    n_src: int,
) -> dict[int, tuple[int, int]]:
    """シート部分の _cell_source_spans を (開始, 行数) のままミニバンドル用にコピーする。"""
    mini: dict[int, tuple[int, int]] = {}
    if isinstance(local_spans, dict):
        for si, sp in local_spans.items():
            if not isinstance(sp, (tuple, list)) or len(sp) != 2:
                continue
            try:
                start = int(sp[0])
                ln = int(sp[1])
            except (TypeError, ValueError):
                continue
            if ln > 0:
                mini[int(si)] = (start, ln)
    if not mini:
        return {0: (0, n_src)} if n_src else {}
    return mini


def _merged_cell_source_spans(
    parts: list[tuple[str, dict[str, Any]]],
) -> dict[int, tuple[int, int]]:
    """シート連結後もソース番号を保ち、連続区間なら (開始, 行数) にまとめる。"""
    by_si: dict[int, list[tuple[int, int]]] = {}
    g_off = 0
    for _sh, b in parts:
        if not isinstance(b, dict):
            continue
        n = len(b.get("primary_values") or [])
        local_spans = b.get("_cell_source_spans")
        if not isinstance(local_spans, dict):
            local_spans = {0: (0, n)} if n else {}
        for si, sp in local_spans.items():
            if not isinstance(sp, (tuple, list)) or len(sp) != 2:
                continue
            try:
                local_off = int(sp[0])
                ln = int(sp[1])
            except (TypeError, ValueError):
                continue
            if ln < 1:
                continue
            by_si.setdefault(int(si), []).append((g_off + local_off, ln))
        g_off += n
    out: dict[int, tuple[int, int]] = {}
    for si, segs in by_si.items():
        segs.sort(key=lambda x: x[0])
        start = segs[0][0]
        total = 0
        expected = start
        contiguous = True
        for st, ln in segs:
            if st != expected:
                contiguous = False
                break
            total += ln
            expected = st + ln
        if contiguous and total > 0:
            out[si] = (start, total)
    return out


def file_paths_for_source_extract(
    file_paths: Sequence[str | Path],
    src: dict[str, Any],
) -> list[str]:
    """
    当該ソースの file_pattern / name_extract 規則に合うパスのみ返す（入力順・重複除去）。

    extract_item_bundle と同じ source_passes_file_name_filter / name_extract 順序を
    マスタ列抽出の走査前に適用し、無関係なファイルを開かない。
    file_pattern が空の cell ソースはフィルタなし（全候補）。
    """
    ordered: list[str] = []
    seen: set[str] = set()
    for fp in file_paths:
        fps = str(fp).strip()
        if not fps or fps in seen:
            continue
        seen.add(fps)
        ordered.append(fps)
    if not ordered:
        return []
    typ = str(src.get("type") or "cell").strip().lower()
    if typ in ("metadata", "meta", "filename", "name_extract"):
        return list(name_extract_hit_files_ordered(ordered, src))
    return [fp for fp in ordered if source_passes_file_name_filter(fp, src)]


# メタデータのソース種別（後方互換用）
SOURCE_FULL_PATH = "full_path"
SOURCE_DIR_NAME = "dir_name"
SOURCE_FILE_NAME = "file_name"

# 名前から抽出の開始モード
START_MODE_HEAD = "head"
START_MODE_DELIMITER = "delimiter"
START_MODE_POSITION = "position"

# 名前から抽出の長さモード
LENGTH_MODE_CHAR = "char"
LENGTH_MODE_COUNT = "count"
LENGTH_MODE_END = "end"


def extract_metadata(
    file_path: str | Path,
    source_type: str = SOURCE_FILE_NAME,
) -> str:
    """
    ファイルパスからメタデータ（フルパス・フォルダ名・ファイル名）を抽出する。

    【概要】
      指定した source_type に応じて、フルパス・ディレクトリ名・ファイル名のいずれかを返す。

    【引数】
      file_path: 対象ファイルのパス。
      source_type: "full_path" | "dir_name" | "file_name"。

    【戻り値】
      抽出した文字列。取得できない場合は空文字。
    """
    from svc.data_agg_path_norm import normalize_source_path  # noqa: E402

    p = Path(file_path).resolve()
    if source_type == SOURCE_FULL_PATH:
        return normalize_source_path(p)
    if source_type == SOURCE_DIR_NAME:
        return p.parent.name if p.parent else ""
    if source_type == SOURCE_FILE_NAME:
        return p.name
    return ""


def _extract_from_string(
    text: str,
    start: int = 0,
    length: Optional[int] = None,
    delimiter: Optional[str] = None,
    part_index: Optional[int] = None,
    pattern: Optional[str] = None,
) -> str:
    """
    文字列に範囲・デリミタ・正規表現の抽出ルールを適用する。
    extract_from_filename およびメタデータの抽出ルールで共通利用。
    """
    if pattern:
        m = re.search(pattern, text)
        if m:
            return m.group(1) if m.lastindex and m.lastindex >= 1 else m.group(0)
        return ""
    if delimiter is not None and part_index is not None:
        parts = text.split(delimiter)
        if 0 <= part_index < len(parts):
            return parts[part_index]
        return ""
    if length is not None:
        return text[start : start + length]
    return text[start:] if start < len(text) else ""


def extract_from_filename(
    file_path: str | Path,
    start: int = 0,
    length: Optional[int] = None,
    delimiter: Optional[str] = None,
    part_index: Optional[int] = None,
    pattern: Optional[str] = None,
) -> str:
    """
    ファイル名（またはパス）から、範囲指定・デリミタ分割・正規表現で文字列を抽出する。

    【概要】
      length 指定時は name[start:start+length]。delimiter と part_index 指定時は分割して第 n 要素。
      pattern 指定時は正規表現の最初のグループを返す（グループが無い場合はマッチ全体）。

    【引数】
      file_path: 対象ファイルのパス。
      start: 開始文字位置（0 始まり）。
      length: 取得文字数。None のときは末尾まで。
      delimiter: 区切り文字。指定時は name を分割する。
      part_index: 分割後のインデックス（0 始まり）。delimiter と併用。
      pattern: 正規表現パターン。マッチの第1グループまたはマッチ全体を返す。

    【戻り値】
      抽出した文字列。
    """
    p = Path(file_path).resolve()
    return _extract_from_string(
        p.name, start, length, delimiter, part_index, pattern
    )


def extract_from_name(
    file_path: str | Path,
    source_type: str = "file_name",
    search_condition: Optional[str] = None,
    search_text: Optional[str] = None,
    search_keywords: Optional[list[str]] = None,  # deprecated
    keyword_logic: Optional[str] = None,  # deprecated
    start_mode: str = "head",
    start_value: Any = 0,
    length_mode: str = "end",
    length_value: Any = None,
    delimiter: Optional[str] = None,
    part_index: Optional[int] = None,
    pattern: Optional[str] = None,
    replacement: Optional[str] = None,
) -> str:
    """
    ファイル名またはフォルダ名から、開始位置・抽出長さ・正規表現加工で文字列を抽出する。

    【概要】
      source_type で file_name / dir_name を選択。
      search_condition と search_text で抽出前のフィルタ（含む/含まない）を適用。
      start_mode: head（先頭）→ length_mode で抽出長さを指定。
      start_mode: delimiter → 区切記号＋part_index で分割し、該当部分を取得（抽出長さは不要）。
      start_mode: position → start_value を開始位置として length_mode で抽出。
      抽出後に pattern と replacement で re.sub による加工を適用。

    【引数】
      file_path: 対象ファイルのパス。
      source_type: "file_name" | "dir_name"
      search_condition: "include"（含む）| "exclude"（含まない）| None（フィルタなし）
      search_text: 検索文字列。空の場合はフィルタなし。
      search_keywords: 互換用（未使用）。
      keyword_logic: 互換用（未使用）。
      start_mode: "head" | "delimiter" | "position"
      start_value: position 時の開始位置（UI は 1 始まり＝先頭文字を 1。内部で 0 始まりに変換）。
      length_mode: "char" | "count" | "end"
        - char: length_value を終端文字として、その文字を含めて抽出。
        - count: length_value 文字数だけ抽出。
        - end: 末尾まで。
      length_value: char/count 時の値。
      delimiter: 区切記号（delimiter モード時）。
      part_index: 分割後のインデックス（UI は 1 始まり。内部で 0 始まりに変換）。
      pattern: 抽出後の正規表現パターン（re.sub）。
      replacement: 置換文字列。

    【戻り値】
      抽出・加工した文字列。
    """
    p = Path(file_path).resolve()
    st = str(source_type or "file_name").strip().lower()
    text = _name_extract_text_for_match_and_extract(p, st)

    # 検索条件フィルタ（単一文字列）: 条件を満たさない場合は空文字を返す（name_extract_search_matches と整合）
    kw = (search_text or "").strip()
    if kw and search_condition:
        t_l = text.lower()
        k_l = kw.lower()
        sc = str(search_condition).strip().lower()
        if sc in ("exclude", "含まない"):
            if k_l in t_l:
                return ""
        elif sc in ("exact", "equals", "完全一致"):
            if k_l != t_l:
                return ""
        else:
            if k_l not in t_l:
                return ""

    def _apply_regex(t: str) -> str:
        if pattern:
            try:
                # 置換欄はUI非表示。replacement 未指定時は空文字置換で整形する。
                return re.sub(pattern, replacement if replacement is not None else "", t)
            except re.error:
                pass
        return t

    if start_mode == START_MODE_DELIMITER:
        if delimiter is not None and str(delimiter) != "" and part_index is not None:
            parts = text.split(str(delimiter))
            try:
                idx0 = int(part_index)
            except (TypeError, ValueError):
                idx0 = 1
            idx0 = max(0, idx0 - 1)
            extracted = parts[idx0] if 0 <= idx0 < len(parts) else ""
        else:
            extracted = text
        return _apply_regex(extracted)

    if start_mode == START_MODE_HEAD:
        start = 0
    elif start_mode == START_MODE_POSITION:
        try:
            sv = int(start_value) if start_value is not None else 1
        except (TypeError, ValueError):
            sv = 1
        start = max(0, sv - 1)
    else:
        start = 0

    if start >= len(text):
        return _apply_regex("")

    if length_mode == LENGTH_MODE_CHAR:
        end_char = str(length_value) if length_value is not None else ""
        if end_char:
            pos = text.find(end_char, start)
            if pos >= 0:
                hi = min(len(text), pos + len(end_char))
                extracted = text[start:hi]
            else:
                extracted = text[start:]
        else:
            extracted = text[start:]
    elif length_mode == LENGTH_MODE_COUNT:
        cnt = int(length_value) if length_value is not None else None
        extracted = (
            text[start : start + cnt]
            if cnt is not None and cnt > 0
            else text[start:]
        )
    else:
        extracted = text[start:]

    return _apply_regex(extracted)


def _name_extract_text_for_match_and_extract(path: Path, source_type: str) -> str:
    """
    名前取得の検索フィルタ・抽出で共有する対象文字列。
    file_name: 拡張子なし（stem）。dir_name: 親フォルダ名。
    """
    st = str(source_type or "file_name").strip().lower()
    if st == "dir_name":
        return (path.parent.name if path.parent else "") or ""
    return path.stem or ""


def name_extract_search_matches(file_path: str | Path, src: dict[str, Any]) -> bool:
    """
    extract_from_name と同じ対象文字列（dir: 親フォルダ名 / file: stem）で
    search_text・search_condition を評価する。キーワード空または条件未設定はフィルタ無し。
    """
    kw = (src.get("search_text") or "").strip()
    if not kw or not src.get("search_condition"):
        return True
    p = Path(file_path).resolve()
    st = str(src.get("source_type") or "file_name").strip().lower()
    text = _name_extract_text_for_match_and_extract(p, st).lower()
    kw_l = kw.lower()
    hit_sub = kw_l in text
    hit_eq = kw_l == text
    sc = str(src.get("search_condition") or "include").strip().lower()
    if sc in ("exclude", "含まない"):
        return not hit_sub
    if sc in ("exact", "equals", "完全一致"):
        return hit_eq
    return hit_sub


def name_extract_join_comparison_key(file_path: str | Path, src: dict[str, Any]) -> str:
    """
    名前取得の結合パス照合キー（正規化済み）。file_name: ファイルの正規化フルパス。
    dir_name: 親ディレクトリの正規化パス。svc_data_agg._apply_name_extract_path_assignment と一致させる。
    """
    from svc.data_agg_path_norm import normalize_source_path

    p = Path(file_path).resolve()
    st = str(src.get("source_type") or "file_name").strip().lower()
    if st == "dir_name":
        return normalize_source_path(p.parent)
    return normalize_source_path(p)


def name_extract_unique_join_keys_ordered(allowed_paths: list[str], src: dict[str, Any]) -> list[str]:
    """検索一致ファイルについて、照合キーを allowed の出現順で重複なく列挙。"""
    seen: set[str] = set()
    out: list[str] = []
    for fp in allowed_paths:
        if not name_extract_search_matches(fp, src):
            continue
        key = name_extract_join_comparison_key(fp, src)
        if key not in seen:
            seen.add(key)
            out.append(key)
    return out


def name_extract_hit_files_ordered(allowed_paths: list[str], src: dict[str, Any]) -> list[str]:
    """検索一致したファイルパスを allowed の順で列挙（同一フォルダ内の複数ファイルは別行のまま）。"""
    return [fp for fp in allowed_paths if name_extract_search_matches(fp, src)]


def extract_cells_repeat(
    file_path: str | Path,
    sheet_name: Optional[str] = None,
    cell_ref: str = "A1",
    repeat_direction: str = "vertical",
    repeat_until_empty: bool = True,
    repeat_max: Optional[int] = None,
    *,
    cancel_check: Optional[Callable[..., None]] = None,
) -> list[Any]:
    """
    開始セルから縦または横に繰り返しセルを取得する。

    【引数】
      repeat_direction: "vertical"（下方向）| "horizontal"（右方向）
      repeat_until_empty: True で空セルで終了
      repeat_max: 最大取得数。None で制限なし。

    【戻り値】
      取得した値のリスト。
    """
    p = Path(file_path).resolve()
    if not p.is_file():
        return []
    cap = resolve_extract_repeat_limit(
        repeat_max=repeat_max,
        repeat_until_empty=repeat_until_empty,
    )
    if p.suffix.lower() == ".csv":
        col, row = _parse_cell_ref(cell_ref)
        if col is None or row is None:
            return []
        if _xlsx_workbook_cache_top() is not None:
            series = _read_repeated_series_from_csv_cached(
                p,
                base_col=col,
                base_row=row,
                row_step=1 if repeat_direction == "vertical" else 0,
                col_step=1 if repeat_direction == "horizontal" else 0,
                limit=cap,
                repeat_until_empty=repeat_until_empty,
                cancel_check=cancel_check,
            )
            if series is not None:
                return series
        mat = _get_csv_matrix(p, create=False)
        if mat is None and _xlsx_workbook_cache_top() is not None:
            mat = _get_csv_matrix(p, create=True)
        if mat is not None:
            dc = 1 if repeat_direction == "horizontal" else 0
            dr = 1 if repeat_direction == "vertical" else 0
            out_mat: list[Any] = []
            c, r = col, row
            for _ in range(cap):
                if r < 0 or c < 0:
                    break
                v = _matrix_cell_value(mat, c, r)
                if _stop_repeat_on_empty(v, repeat_until_empty=repeat_until_empty):
                    break
                out_mat.append(v)
                c += dc
                r += dr
                if len(out_mat) >= cap:
                    break
            return out_mat
        pl_mod = _get_polars()
        if pl_mod is not None:
            try:
                df = pl_mod.read_csv(str(p), has_header=False, encoding="utf8-lossy")
                col, row = _parse_cell_ref(cell_ref)
                if col is None or row is None:
                    return []
                out: list[Any] = []
                dc = 1 if repeat_direction == "horizontal" else 0
                dr = 1 if repeat_direction == "vertical" else 0
                for _ in range(cap):
                    if row >= df.height or col >= df.width or row < 0 or col < 0:
                        break
                    v = df.row(row)[col]
                    if _stop_repeat_on_empty(v, repeat_until_empty=repeat_until_empty):
                        break
                    out.append(v)
                    col += dc
                    row += dr
                    if len(out) >= cap:
                        break
                return out
            except Exception:
                pass
    col, row = _parse_cell_ref(cell_ref)
    if col is None or row is None:
        return []
    results: list[Any] = []
    delta_col = 1 if repeat_direction == "horizontal" else 0
    delta_row = 1 if repeat_direction == "vertical" else 0
    for _ in range(cap):
        _poll_cancel_check(cancel_check)
        cr = _col_row_to_cell_ref(col, row)
        v = extract_cell(p, sheet_name, cr)
        if _stop_repeat_on_empty(v, repeat_until_empty=repeat_until_empty):
            break
        results.append(v)
        col += delta_col
        row += delta_row
        if len(results) >= cap:
            break
    return results


def extract_cell(
    file_path: str | Path,
    sheet_name: Optional[str] = None,
    cell_ref: str = "A1",
) -> Any:
    """
    指定ファイルのセル値を取得する。Excel (.xlsx/.xlsm) は OpenPyXL、CSV は csv モジュールで読む。

    【概要】
      絶対位置のセル参照（例: "B5"）で値を返す。CSV の場合は sheet_name を無視し、先頭行を 1 として行・列で解釈する
      （cell_ref は "A1" 形式のみ対応。列 A=0, 行 1=0 のインデックスでアクセス）。

    【引数】
      file_path: 対象ファイルのパス。
      sheet_name: シート名。None のときは先頭シート（Excel）または唯一のシート（CSV）。
      cell_ref: セル番地（例: "A1", "B5"）。CSV では A=列0, 1=行0。

    【戻り値】
      セルの値。取得できない場合は None。
    """
    p = Path(file_path).resolve()
    if not p.is_file():
        return None
    suf = p.suffix.lower()
    if suf == ".csv":
        return _get_csv_cell(p, cell_ref)
    if is_openxml_excel_suffix(suf) or suf == ".xls":
        return _get_excel_cell(p, sheet_name, cell_ref)
    return None


def _csv_path_key(path: Path) -> str:
    """スコープ内では resolve 結果を frame にキャッシュ（大量セル参照時の resolve コスト削減）。"""
    frame = _xlsx_workbook_cache_top()
    if frame is None:
        return str(path.resolve())
    cache: dict[str, str] = frame.setdefault("csv_path_key_cache", {})
    raw = str(path)
    hit = cache.get(raw)
    if hit is not None:
        return hit
    resolved = str(path.resolve())
    cache[raw] = resolved
    return resolved


def _load_csv_polars_df(path: Path) -> Any | None:
    """Polars で CSV を1回読む（legacy 互換: utf8-lossy, infer_schema_length=0）。"""
    pl_mod = _get_polars()
    if pl_mod is None:
        return None
    try:
        return pl_mod.read_csv(
            str(path),
            has_header=False,
            encoding="utf8-lossy",
            infer_schema_length=0,
            try_parse_dates=False,
        )
    except Exception as e:
        logger.debug("[DATA_AGG_EXTRACT] CSV Polars 読込エラー %s: %s", path, e)
        return None


def _load_csv_reader_matrix(path: Path) -> list[list[Any]]:
    """Polars 不可時: csv.reader（utf-8-sig）。"""
    try:
        with path.open("r", encoding="utf-8-sig", newline="") as f:
            return [list(row) for row in csv.reader(f)]
    except Exception as e:
        logger.debug("[DATA_AGG_EXTRACT] CSV 読込エラー %s: %s", path, e)
        return []


def _materialize_csv_matrix(path: Path) -> list[list[Any]]:
    """
    CSV を1回だけ読み、行×列の行列にする（パリティ検証・legacy 参照用）。
    バッチ hot path では Polars DF キャッシュを使い list 化しない。
    """
    df = _load_csv_polars_df(path)
    if df is not None:
        return [list(row) for row in df.rows()]
    return _load_csv_reader_matrix(path)


def _df_cell_value(df: Any, col: int, row: int) -> Any:
    if col < 0 or row < 0 or row >= df.height or col >= df.width:
        return None
    return df.row(row)[col]


def _df_column_slice(df: Any, col: int, start_row: int, length: int) -> list[Any]:
    """列 col の start_row から最大 length 件（Polars slice）。"""
    if col < 0 or col >= df.width or start_row < 0 or length <= 0:
        return []
    if start_row >= df.height:
        return []
    n = min(length, df.height - start_row)
    if n <= 0:
        return []
    return df[df.columns[col]].slice(start_row, n).to_list()


def _ensure_csv_cache(path: Path) -> None:
    """xlsx_workbook_scope 内: CSV を DF（優先）または行列（Polars 不可時）に1回だけ載せる。"""
    frame = _xlsx_workbook_cache_top()
    if frame is None:
        return
    if path.suffix.lower() != ".csv":
        return
    path_key = _csv_path_key(path)
    dfs: dict[str, Any] = frame.setdefault("csv_dfs", {})
    mats: dict[str, list[list[Any]]] = frame.setdefault("csv_mats", {})
    if path_key in dfs or path_key in mats:
        return
    try:
        from svc.data_agg_cancel import poll_active_cancel  # noqa: WPS433

        poll_active_cancel(force=True)
    except Exception:
        pass
    df = _load_csv_polars_df(path)
    if df is not None:
        dfs[path_key] = df
        return
    mat = _load_csv_reader_matrix(path)
    if mat:
        mats[path_key] = mat


def _get_csv_df(path: Path, *, create: bool = True) -> Any | None:
    """スコープ内 Polars DF。create=True で未キャッシュなら _ensure_csv_cache。"""
    frame = _xlsx_workbook_cache_top()
    if frame is None or path.suffix.lower() != ".csv":
        return None
    path_key = _csv_path_key(path)
    dfs: dict[str, Any] = frame.setdefault("csv_dfs", {})
    hit = dfs.get(path_key)
    if hit is not None or not create:
        return hit
    _ensure_csv_cache(path)
    return dfs.get(path_key)


def _get_csv_matrix(path: Path, *, create: bool = True) -> Optional[list[list[Any]]]:
    """スコープ内 csv.reader 行列（Polars 不可時のみ）。hot path では DF を優先。"""
    frame = _xlsx_workbook_cache_top()
    if frame is None or path.suffix.lower() != ".csv":
        return None
    path_key = _csv_path_key(path)
    mats: dict[str, list[list[Any]]] = frame.setdefault("csv_mats", {})
    mat = mats.get(path_key)
    if mat is not None or not create:
        return mat if mat else None
    _ensure_csv_cache(path)
    return mats.get(path_key)


def _csv_cell_value_rc(path: Path, col: int, row: int) -> Any:
    """キャッシュ済み CSV から (col,row) 0始まりで1セル。"""
    df = _get_csv_df(path, create=False)
    if df is not None:
        return _df_cell_value(df, col, row)
    mat = _get_csv_matrix(path, create=False)
    if mat is not None:
        return _matrix_cell_value(mat, col, row)
    return None


def _read_repeated_series_from_df(
    df: Any,
    *,
    base_col: int,
    base_row: int,
    row_step: int,
    col_step: int,
    limit: int,
    repeat_until_empty: bool,
    cancel_check: Optional[Callable[..., None]] = None,
) -> list[Any]:
    """Polars DF から反復系列（縦 step=1 は列 slice 高速経路）。"""
    if limit <= 0:
        return []
    if row_step > 0 and col_step == 0 and row_step == 1:
        raw = _df_column_slice(df, base_col, base_row, limit)
        if not repeat_until_empty:
            return raw[:limit]
        out: list[Any] = []
        for v in raw:
            if v is None or v == "":
                break
            out.append(v)
            if len(out) >= limit:
                break
        return out
    vals: list[Any] = []
    if row_step == 0 and col_step == 0:
        v0 = _df_cell_value(df, base_col, base_row)
        if _stop_repeat_on_empty(v0, repeat_until_empty=repeat_until_empty):
            return []
        return [v0] * limit
    n = 0
    while len(vals) < limit:
        if n % 64 == 0:
            _poll_cancel_check(cancel_check)
        if row_step > 0 and col_step == 0:
            v = _df_cell_value(df, base_col, base_row + row_step * n)
        elif row_step == 0 and col_step > 0:
            v = _df_cell_value(df, base_col + col_step * n, base_row)
        else:
            break
        if _stop_repeat_on_empty(v, repeat_until_empty=repeat_until_empty):
            break
        vals.append(v)
        n += 1
    return vals


def _read_repeated_series_from_csv_cached(
    path: Path,
    *,
    base_col: int,
    base_row: int,
    row_step: int,
    col_step: int,
    limit: int,
    repeat_until_empty: bool,
    cancel_check: Optional[Callable[..., None]] = None,
) -> Optional[list[Any]]:
    """スコープ内 CSV キャッシュから反復系列。"""
    df = _get_csv_df(path, create=True)
    if df is not None:
        return _read_repeated_series_from_df(
            df,
            base_col=base_col,
            base_row=base_row,
            row_step=row_step,
            col_step=col_step,
            limit=limit,
            repeat_until_empty=repeat_until_empty,
            cancel_check=cancel_check,
        )
    mat = _get_csv_matrix(path, create=True)
    if mat is not None:
        return _read_repeated_series_from_matrix(
            mat,
            base_col=base_col,
            base_row=base_row,
            row_step=row_step,
            col_step=col_step,
            limit=limit,
            repeat_until_empty=repeat_until_empty,
            cancel_check=cancel_check,
        )
    return None


def _matrix_cell_value_at_ref(mat: list[list[Any]], cell_ref: str) -> Any:
    col, row = _parse_cell_ref(cell_ref)
    if col is None or row is None:
        return None
    return _matrix_cell_value(mat, col, row)


def _peek_repeat_cell_from_csv(
    path: Path,
    *,
    base_col: int,
    base_row: int,
    row_off: int,
    col_off: int,
    index: int,
) -> Any:
    col_n = base_col + (col_off * index)
    row_n = base_row + (row_off * index)
    if _xlsx_workbook_cache_top() is not None:
        return _csv_cell_value_rc(path, col_n, row_n)
    return extract_cell(path, cell_ref=_col_row_to_cell_ref(col_n, row_n))


def _peek_repeat_cell_from_matrix(
    mat: list[list[Any]],
    *,
    base_col: int,
    base_row: int,
    row_off: int,
    col_off: int,
    index: int,
) -> Any:
    """反復系列 index 番目の peek（打ち切り検知用）。"""
    return _matrix_cell_value(mat, base_col + (col_off * index), base_row + (row_off * index))


def _finish_repeated_cell_vals(
    *,
    vals: list[Any],
    results: list[Any],
    cell_start: int,
    cell_source_spans_out: Optional[dict[int, tuple[int, int]]],
    si: int,
    limit: int,
    repeat_until_empty: bool,
    file_path: str | Path,
    item_label: str,
    item_id: Optional[str],
    positions: dict[str, tuple[int, int]],
    cell_ref: str,
    row_off: int,
    col_off: int,
    ui_blk: dict[str, Any] | None,
    src: dict[str, Any],
    max_primary_rows: Optional[int],
    peek_v: Any,
    skip_trunc_peek: bool = False,
) -> None:
    """縦/横反復の主値抽出後処理（xlsx / CSV 共通）。"""
    if source_end_mode(src) == END_MODE_UNTIL_LAST:
        vals = apply_until_last_trim(vals, until_last=True)
    if vals:
        n_last = len(vals) - 1
        cell_ref_last = _resolve_cell_with_offset(cell_ref, row_off * n_last, col_off * n_last)
        c0, r0 = _parse_cell_ref(cell_ref_last)
        if c0 is not None and r0 is not None and item_id:
            positions[item_id] = (c0, r0)
        if not skip_trunc_peek:
            record_extract_truncation_if_needed(
                vals,
                limit=limit,
                peek_next=peek_v,
                file_path=file_path,
                item_label=item_label,
                source_index=si,
            )
    _append_postprocessed_cell_vals(
        results,
        vals,
        ui_blk,
        src,
        max_primary_rows=max_primary_rows,
    )
    if cell_source_spans_out is not None:
        cell_source_spans_out[si] = (cell_start, len(results) - cell_start)


def precache_csv_matrix_for_file(
    file_path: str | Path,
    *,
    progress_hook: Optional[Callable[[str], None]] = None,
) -> None:
    """xlsx_workbook_scope 内で CSV を先にキャッシュ（Polars DF 優先）。"""
    if _xlsx_workbook_cache_top() is None:
        return
    p_abs = Path(file_path).resolve()
    if p_abs.suffix.lower() != ".csv":
        return
    if progress_hook is not None:
        try:
            progress_hook("CSV読込中: %s" % p_abs.name)
        except Exception:
            pass
    _ensure_csv_cache(p_abs)
    path_key = _csv_path_key(p_abs)
    frame = _xlsx_workbook_cache_top()
    if frame is None:
        return
    ok = path_key in frame.get("csv_dfs", {}) or path_key in frame.get("csv_mats", {})
    if not ok and p_abs.is_file():
        raise DataAggCsvReadError("CSV キャッシュに失敗しました: %s" % p_abs)


def _csv_cache_loaded(path: Path) -> bool:
    frame = _xlsx_workbook_cache_top()
    if frame is None:
        return False
    path_key = _csv_path_key(path)
    return path_key in frame.get("csv_dfs", {}) or path_key in frame.get("csv_mats", {})


def _get_csv_cell_legacy_file_read(path: Path, cell_ref: str) -> Any:
    """スコープ外の従来経路: セル参照のたびにファイルを読む（互換・テスト用）。"""
    if _xlsx_workbook_cache_top() is not None:
        raise DataAggCsvReadError(
            "バッチスコープ内で CSV ファイル再読込（legacy）は禁止: %s" % path
        )
    col, row = _parse_cell_ref(cell_ref)
    if col is None or row is None:
        return None
    try:
        df = _load_csv_polars_df(path)
        if df is not None:
            if 0 <= row < df.height and 0 <= col < df.width:
                return df.row(row)[col]
        with path.open("r", encoding="utf-8-sig", newline="") as f:
            reader = list(csv.reader(f))
        if row < len(reader) and col < len(reader[row]):
            return reader[row][col]
    except Exception as e:
        logger.debug("[DATA_AGG_EXTRACT] CSV 読込エラー %s: %s", path, e)
    return None


def _get_csv_cell(path: Path, cell_ref: str) -> Any:
    """CSV ファイルから A1 形式のセル参照で値を取得する。"""
    if _xlsx_workbook_cache_top() is not None:
        _ensure_csv_cache(path)
        if not _csv_cache_loaded(path) and path.is_file():
            raise DataAggCsvReadError(
                "CSV キャッシュ未取得（バッチ処理中）: %s" % path
            )
        col, row = _parse_cell_ref(cell_ref)
        if col is None or row is None:
            return None
        return _csv_cell_value_rc(path, col, row)
    return _get_csv_cell_legacy_file_read(path, cell_ref)


def _col_row_to_cell_ref(col: int, row: int) -> str:
    """0始まりの (列, 行) を A1 形式に変換。"""
    if col < 0 or row < 0:
        return "A1"
    col_letters = ""
    c = col + 1
    while c > 0:
        c, r = divmod(c - 1, 26)
        col_letters = chr(ord("A") + r) + col_letters
    return col_letters + str(row + 1)


def _parse_cell_ref(cell_ref: str) -> tuple[Optional[int], Optional[int]]:
    """A1 形式を (列インデックス, 行インデックス) に変換。行・列は 0 始まり。"""
    s = (cell_ref or "").strip().upper()
    if not s:
        return (None, None)
    col_str = ""
    i = 0
    while i < len(s) and s[i].isalpha():
        col_str += s[i]
        i += 1
    row_str = s[i:] if i < len(s) else ""
    if not col_str or not row_str:
        return (None, None)
    try:
        row = int(row_str) - 1  # 1-based to 0-based
    except ValueError:
        return (None, None)
    col = 0
    for c in col_str:
        col = col * 26 + (ord(c) - ord("A") + 1)
    col -= 1  # 1-based to 0-based
    return (col, row)


def _get_excel_cell(
    path: Path,
    sheet_name: Optional[str],
    cell_ref: str,
) -> Any:
    """Excel ファイルからセル値を取得。.xlsx/.xlsm は OpenPyXL、.xls は xlrd。"""
    if path.suffix.lower() == ".xls":
        try:
            from svc.data_agg_xls_io import read_xls_cell
        except Exception:
            logger.debug("[DATA_AGG_EXTRACT] .xls 読取モジュール不可: %s", path)
            return None
        return read_xls_cell(path, sheet_name, cell_ref)
    try:
        import openpyxl  # noqa: E402
    except ImportError:
        logger.warning("[DATA_AGG_EXTRACT] openpyxl が利用できません")
        return None
    wb_cached = _xlsx_workbook_from_cache(path)
    if wb_cached is not None:
        try:
            return _xlsx_cell_value_open_workbook(
                wb_cached,
                sheet_name,
                cell_ref,
                path=path,
            )
        except Exception:
            return None
    try:
        wb = _load_workbook_readonly(path)
        if sheet_name:
            if sheet_name not in wb.sheetnames:
                ws = wb.active
            else:
                ws = wb[sheet_name]
        else:
            ws = wb.active
        if ws is None:
            return None
        val = extract_read_openpyxl_cell(ws[cell_ref])
        wb.close()
        return val
    except Exception as e:
        logger.debug("[DATA_AGG_EXTRACT] Excel 読込エラー %s: %s", path, e)
    return None


def _resolve_readonly_worksheet(wb: Any, sheet_name: Optional[str]) -> tuple[Any, str]:
    """read_only Workbook から対象シートと実効シート名を返す。"""
    names = getattr(wb, "sheetnames", None) or []
    if sheet_name and sheet_name in names:
        return wb[sheet_name], sheet_name
    ws = wb.active
    title = getattr(ws, "title", "") if ws is not None else ""
    return ws, str(title)


def _materialize_readonly_sheet_matrix(ws: Any) -> list[list[Any]]:
    """read_only シートを一度だけ走査し抽出用スカラーの行リストにする（以降 ws[ref] は使わない）。"""
    if ws is None:
        return []
    rows: list[list[Any]] = []
    for tup in ws.iter_rows(values_only=False):
        rows.append(extract_read_openpyxl_row(tup))
    return rows


def _matrix_cell_value(matrix: list[list[Any]], col: int, row: int) -> Any:
    if row < 0 or col < 0:
        return None
    if row >= len(matrix):
        return None
    r = matrix[row]
    if col >= len(r):
        return None
    return r[col]


def _get_readonly_sheet_matrix(
    wb: Any,
    sheet_name: Optional[str],
    path: Path,
    *,
    create: bool = True,
) -> Optional[list[list[Any]]]:
    """xlsx_workbook_scope 内の (path, シート) 行列。create=False なら未構築時は None。"""
    frame = _xlsx_workbook_cache_top()
    if frame is None:
        return None
    ws, resolved_name = _resolve_readonly_worksheet(wb, sheet_name)
    if ws is None:
        return None
    path_key = str(path.resolve())
    mats_store: dict[Any, list[list[Any]]] = frame.setdefault("sheet_mats", {})
    store_key = (path_key, resolved_name)
    mat = mats_store.get(store_key)
    if mat is not None or not create:
        return mat
    try:
        mat = _materialize_readonly_sheet_matrix(ws)
    except Exception:
        return None
    mats_store[store_key] = mat
    return mat


def _collect_xlsx_sheets_for_precache(items: list[dict[str, Any]]) -> set[Optional[str]]:
    """
    同一ファイル内で materialize するシート名（None は既定シート）。
    - 縦/横反復が _SHEET_PRECACHE_REPEAT_MIN 行以上
    - またはセル座標ソースが参照するシート（複数項目・品名などの都度読取を1回化）
    """
    sheets: set[Optional[str]] = set()
    for it in items:
        if not isinstance(it, dict):
            continue
        for src in it.get("sources") or []:
            if not isinstance(src, dict):
                continue
            if (src.get("type") or "cell").strip().lower() != "cell":
                continue
            sn = src.get("sheet_name")
            sheet_key = str(sn).strip() if sn is not None and str(sn).strip() else None
            sheets.add(sheet_key)
            rd = (src.get("repeat_direction") or "").strip().lower()
            if rd not in ("vertical", "horizontal"):
                continue
            try:
                until_empty, until_last, repeat_max = _resolve_cell_source_repeat(src)
            except Exception:
                until_empty, until_last, repeat_max = True, False, None
            limit = resolve_extract_repeat_limit(
                repeat_max=repeat_max,
                repeat_until_empty=until_empty,
                repeat_until_last=until_last,
            )
            if limit >= _SHEET_PRECACHE_REPEAT_MIN:
                sheets.add(sheet_key)
    return sheets


def precache_xlsx_workbook_sheets_for_items(
    file_path: str | Path,
    items: list[dict[str, Any]],
) -> None:
    """
    xlsx_workbook_scope 内で、大量反復が見込まれるシートを先に materialize する（Phase B）。
    スコープ外では何もしない。
    """
    if _xlsx_workbook_cache_top() is None:
        return
    p_abs = Path(file_path).resolve()
    if not is_openxml_excel_suffix(p_abs.suffix):
        return
    sheets = _collect_xlsx_sheets_for_precache(items)
    if not sheets:
        return
    wb = _xlsx_workbook_from_cache(p_abs)
    if wb is None:
        return
    for sheet_name in sheets:
        _get_readonly_sheet_matrix(wb, sheet_name, p_abs, create=True)


def _read_repeated_series_from_matrix(
    mat: list[list[Any]],
    *,
    base_col: int,
    base_row: int,
    row_step: int,
    col_step: int,
    limit: int,
    repeat_until_empty: bool,
    cancel_check: Optional[Callable[..., None]] = None,
) -> list[Any]:
    """materialize 済み行列から反復系列を読む（縦/横/固定セル反復）。"""
    vals: list[Any] = []
    if row_step == 0 and col_step == 0:
        v0 = _matrix_cell_value(mat, base_col, base_row)
        if _stop_repeat_on_empty(v0, repeat_until_empty=repeat_until_empty):
            return []
        return [v0] * limit
    n = 0
    while len(vals) < limit:
        if n % 64 == 0:
            _poll_cancel_check(cancel_check)
        if row_step > 0 and col_step == 0:
            v = _matrix_cell_value(mat, base_col, base_row + row_step * n)
        elif row_step == 0 and col_step > 0:
            v = _matrix_cell_value(mat, base_col + col_step * n, base_row)
        else:
            break
        if _stop_repeat_on_empty(v, repeat_until_empty=repeat_until_empty):
            break
        vals.append(v)
        n += 1
    return vals


def _xlsx_cell_value_open_workbook(
    wb: Any,
    sheet_name: Optional[str],
    cell_ref: str,
    *,
    path: Optional[Path] = None,
    ephemeral_sheet_cache: bool = False,
) -> Any:
    """
    既に開いた read_only Workbook から 1 セル読み。
    xlsx_workbook_scope 内、または ephemeral_sheet_cache=True（同一 wb での繰り返し）では
    シートを iter_rows で一度具体化し、メモリ参照に切り替える。
    """
    col, row = _parse_cell_ref(cell_ref)
    if col is None or row is None:
        return None
    return _xlsx_cell_value_open_workbook_rc(
        wb,
        sheet_name,
        col,
        row,
        path=path,
        ephemeral_sheet_cache=ephemeral_sheet_cache,
    )


def _xlsx_cell_value_open_workbook_rc(
    wb: Any,
    sheet_name: Optional[str],
    col: int,
    row: int,
    *,
    path: Optional[Path] = None,
    ephemeral_sheet_cache: bool = False,
) -> Any:
    """
    既に開いた read_only Workbook から列/行インデックス（0 始まり）で 1 セル読み。
    A1 文字列生成・再パースを避け、大量反復時のオーバーヘッドを抑える。
    """
    if col < 0 or row < 0:
        return None

    frame = _xlsx_workbook_cache_top()
    use_mat = frame is not None or ephemeral_sheet_cache
    if not use_mat:
        try:
            ws, _ = _resolve_readonly_worksheet(wb, sheet_name)
            if ws is None:
                return None
            cell_ref = _col_row_to_cell_ref(col, row)
            return extract_read_openpyxl_cell(ws[cell_ref])
        except Exception:
            return None

    ws, resolved_name = _resolve_readonly_worksheet(wb, sheet_name)
    if ws is None:
        return None

    path_key = str(path.resolve()) if path is not None else f"id:{id(wb)}"
    mat: Optional[list[list[Any]]] = None
    mats_store: Optional[dict[Any, list[list[Any]]]] = None
    store_key: Any = None

    if frame is not None:
        mats_store = frame.setdefault("sheet_mats", {})
        store_key = (path_key, resolved_name)
        mat = mats_store.get(store_key)
    elif ephemeral_sheet_cache:
        mats_store = getattr(wb, "_da_sheet_mats", None)
        if mats_store is None:
            mats_store = {}
            setattr(wb, "_da_sheet_mats", mats_store)
        store_key = resolved_name
        mat = mats_store.get(store_key)

    # xlsx_workbook_scope 内は同一シートへ多数アクセスされる前提のため、
    # 早い段階で materialize した方が総時間が安定して短い。
    # （read_only の ws[cell_ref] は都度走査コストが高く、ここでは逆効果になりやすい）
    if mat is None and frame is None and ephemeral_sheet_cache:
        hits_store = getattr(wb, "_da_sheet_hits", None)
        if hits_store is None:
            hits_store = {}
            setattr(wb, "_da_sheet_hits", hits_store)
        hits = int(hits_store.get(store_key, 0)) + 1
        hits_store[store_key] = hits
        if hits < _SHEET_MATERIALIZE_THRESHOLD:
            try:
                cell_ref = _col_row_to_cell_ref(col, row)
                return extract_read_openpyxl_cell(ws[cell_ref])
            except Exception:
                return None

    if mat is None:
        try:
            mat = _materialize_readonly_sheet_matrix(ws)
        except Exception:
            return None
        if mats_store is not None and store_key is not None:
            mats_store[store_key] = mat

    return _matrix_cell_value(mat, col, row)


def _xlsx_read_repeated_series_open_workbook(
    wb: Any,
    sheet_name: Optional[str],
    *,
    base_col: int,
    base_row: int,
    row_step: int,
    col_step: int,
    limit: int,
    repeat_until_empty: bool,
    path: Optional[Path] = None,
    cancel_check: Optional[Callable[..., None]] = None,
) -> Optional[list[Any]]:
    """
    反復セル抽出の高速経路（OpenXML Excel / read_only）。
    xlsx_workbook_scope 内では materialize 済み行列から読む。
    それ以外は対象1列/1行の iter_rows で直接読み出す。
    非対応パターンは None を返し、呼び出し側で従来経路へフォールバックする。
    """
    if limit <= 0 or base_col < 0 or base_row < 0:
        return []
    # 互換性優先: 一軸反復（縦/横）を高速化。その他は従来経路へ戻す。
    # 例: row=1,col=0 / row=0,col=7 / row=3,col=0 / row=0,col=0(固定セル反復)
    vertical = row_step > 0 and col_step == 0
    horizontal = row_step == 0 and col_step > 0
    if row_step == 0 and col_step == 0:
        vertical = horizontal = False
    elif not vertical and not horizontal:
        return None

    if path is not None and _xlsx_workbook_cache_top() is not None:
        mat = _get_readonly_sheet_matrix(wb, sheet_name, path, create=True)
        if mat is not None:
            try:
                return _read_repeated_series_from_matrix(
                    mat,
                    base_col=base_col,
                    base_row=base_row,
                    row_step=row_step,
                    col_step=col_step,
                    limit=limit,
                    repeat_until_empty=repeat_until_empty,
                    cancel_check=cancel_check,
                )
            except Exception:
                return None

    if row_step == 0 and col_step == 0:
        try:
            c_ref = _col_row_to_cell_ref(base_col, base_row)
            v0 = _xlsx_cell_value_open_workbook(wb, sheet_name, c_ref, path=path)
        except Exception:
            return None
        if _stop_repeat_on_empty(v0, repeat_until_empty=repeat_until_empty):
            return []
        return [v0] * limit

    if not vertical and not horizontal:
        return None
    try:
        ws, _ = _resolve_readonly_worksheet(wb, sheet_name)
    except Exception:
        return None
    if ws is None:
        return []
    vals: list[Any] = []
    try:
        if vertical:
            min_col = base_col + 1
            min_row = base_row + 1
            step = row_step
            max_row = min_row + max(0, step * (limit - 1))
            picked = 0
            for i, tup in enumerate(
                ws.iter_rows(
                    min_row=min_row,
                    max_row=max_row,
                    min_col=min_col,
                    max_col=min_col,
                    values_only=False,
                )
            ):
                if i % 64 == 0:
                    _poll_cancel_check(cancel_check)
                if step > 1 and (i % step) != 0:
                    continue
                v = extract_read_openpyxl_cell(tup[0]) if tup else None
                if _stop_repeat_on_empty(v, repeat_until_empty=repeat_until_empty):
                    break
                vals.append(v)
                picked += 1
                if picked >= limit:
                    break
        else:
            min_row = base_row + 1
            min_col = base_col + 1
            step = col_step
            max_col = min_col + max(0, step * (limit - 1))
            for i, tup in enumerate(
                ws.iter_rows(
                    min_row=min_row,
                    max_row=min_row,
                    min_col=min_col,
                    max_col=max_col,
                    values_only=False,
                )
            ):
                # 1行のみだが要素数は limit 相当
                picked = 0
                for j, cell in enumerate(tup):
                    if j % 64 == 0:
                        _poll_cancel_check(cancel_check)
                    if step > 1 and (j % step) != 0:
                        continue
                    v = extract_read_openpyxl_cell(cell)
                    if _stop_repeat_on_empty(v, repeat_until_empty=repeat_until_empty):
                        return vals
                    vals.append(v)
                    picked += 1
                    if picked >= limit:
                        return vals
                break
    except Exception:
        return None
    return vals


def _read_cell_at_repeat_index(
    *,
    wb_ctx: Any,
    wb_owned: bool,
    file_path: str | Path,
    sheet_name: Optional[str],
    base_col: int,
    base_row: int,
    row_off: int,
    col_off: int,
    index: int,
) -> Any:
    """反復系列の index 番目（0 始まり）のセル値を読む（打ち切り peek 用）。"""
    col_n = base_col + (col_off * index)
    row_n = base_row + (row_off * index)
    cell_ref_n = _col_row_to_cell_ref(col_n, row_n)
    p_abs = Path(file_path).resolve()
    if wb_ctx is None and p_abs.suffix.lower() == ".csv":
        if _xlsx_workbook_cache_top() is not None:
            return _csv_cell_value_rc(p_abs, col_n, row_n)
    if wb_ctx is not None:
        return _xlsx_cell_value_open_workbook(
            wb_ctx,
            sheet_name,
            cell_ref_n,
            path=p_abs if not wb_owned else None,
            ephemeral_sheet_cache=wb_owned,
        )
    return extract_cell(file_path, sheet_name=sheet_name, cell_ref=cell_ref_n)


def _append_postprocessed_cell_vals(
    results: list[Any],
    vals: list[Any],
    ui_block: dict[str, Any] | None,
    src: dict[str, Any],
    *,
    max_primary_rows: Optional[int] = None,
) -> None:
    """反復セル読取結果をフィルタし、主値後処理して results に追記する。"""
    allow_empty = bool(src.get("allow_empty"))
    # 終端／N件+スキップ: 空スロットを残し、後段でスキップフィルタする。
    keep_empty_slots = source_keep_empty_primary_slots(src)
    raw: list[Any] = []
    for v in vals:
        if keep_empty_slots:
            raw.append(v)
        elif v is not None and (v != "" or allow_empty):
            raw.append(v)
        if (
            max_primary_rows is not None
            and max_primary_rows > 0
            and len(results) + len(raw) >= max_primary_rows
        ):
            break
    if max_primary_rows is not None and max_primary_rows > 0:
        room = max_primary_rows - len(results)
        if room <= 0:
            return
        raw = raw[:room]
    if not raw:
        return
    if len(raw) == 1:
        results.append(postprocess_cell_primary(raw[0], ui_block))
        return
    results.extend(postprocess_cell_primary_batch(raw, ui_block))


def _is_blank_primary_value(v: Any) -> bool:
    """主キーが空欄／空（空白のみ含む）か。"""
    return is_blank_primary_value(v)


def _source_wants_skip_empty_primary(src: dict[str, Any]) -> bool:
    """スキップ有効かつ実効トークンがあるとき True（空白までは空欄トークンを除外）。"""
    return source_wants_skip_primary(src)


def _iter_contexts_rule_indices_contiguous(iter_contexts: list[dict[str, Any]]) -> bool:
    """rule_iter_index が 0..n-1 連番なら高速系列読取が使える。"""
    for i, ic in enumerate(iter_contexts):
        try:
            ri = int(ic.get("rule_iter_index", i))
        except (TypeError, ValueError):
            return False
        if ri != i:
            return False
    return True


def _bundle_file_path(bundle: dict[str, Any]) -> str:
    """iteration_contexts 先頭から file_path を取る。"""
    ictx = bundle.get("iteration_contexts") or []
    if ictx and isinstance(ictx[0], dict):
        return str(ictx[0].get("file_path") or "")
    return ""


def _ctx_iter_index(ctx: Any, fallback: int) -> int:
    """コンテキストの iter_index。不正時は fallback。"""
    if not isinstance(ctx, dict):
        return int(fallback)
    try:
        return int(ctx.get("iter_index", fallback))
    except (TypeError, ValueError):
        return int(fallback)


def _series_already_aligned_to_primary(
    values: list[Any],
    contexts: list[Any],
    n_prim: int,
) -> bool:
    """値・コンテキストが主キー行数と同じで、iter_index が 0..n-1 のとき True。"""
    if n_prim < 1 or len(values) != n_prim:
        return False
    if not isinstance(contexts, list) or len(contexts) != n_prim:
        return False
    for i, ctx in enumerate(contexts):
        if _ctx_iter_index(ctx, i) != i:
            return False
    return True


def _align_one_rule_series(
    values: list[Any],
    contexts: list[Any],
    n_prim: int,
    *,
    file_path: str,
) -> tuple[list[Any], list[Any]]:
    """
    値列を主キー反復の長さに揃える。iter_index がある値はその行へ置く（先勝ち）。
    カード専用列など短い系列がリスト先頭＝ユニット行になるのを防ぐ。
    すでに同じ長さ・同じ iter ならコピーしない。
    """
    if n_prim < 1:
        return [], []
    if _series_already_aligned_to_primary(values, contexts, n_prim):
        return values, contexts
    placed_v: list[Any] = [None] * n_prim
    placed_c: list[Any] = [None] * n_prim
    seen: set[int] = set()
    n_ctx = len(contexts) if isinstance(contexts, list) else 0
    n_pair = max(len(values), n_ctx)
    for i in range(n_pair):
        v = values[i] if i < len(values) else None
        ctx = contexts[i] if i < n_ctx else None
        ix = _ctx_iter_index(ctx, i)
        if ix < 0 or ix >= n_prim or ix in seen:
            continue
        seen.add(ix)
        placed_v[ix] = v
        if isinstance(ctx, dict):
            c = dict(ctx)
            c["iter_index"] = int(ix)
            placed_c[ix] = c
        else:
            placed_c[ix] = {"file_path": str(file_path), "iter_index": int(ix)}
    for i in range(n_prim):
        if placed_c[i] is None:
            placed_c[i] = {"file_path": str(file_path), "iter_index": int(i)}
    return placed_v, placed_c


def _align_link_join_series_to_primary(bundle: dict[str, Any]) -> None:
    """link/join/path_item の値列を primary_values の行数・iter_index に揃える。"""
    n_prim = len(bundle.get("primary_values") or [])
    if n_prim < 1:
        return
    fp = _bundle_file_path(bundle)
    for vkey, ckey in (
        ("link_values", "link_contexts"),
        ("join_values", "join_contexts"),
        ("path_item_values", "path_item_contexts"),
    ):
        vd = bundle.get(vkey)
        if not isinstance(vd, dict) or not vd:
            continue
        cd = bundle.get(ckey)
        if not isinstance(cd, dict):
            cd = {}
            bundle[ckey] = cd
        for t in list(vd.keys()):
            vals = vd.get(t)
            if not isinstance(vals, list):
                continue
            ctxs = cd.get(t)
            if not isinstance(ctxs, list):
                ctxs = []
            if _series_already_aligned_to_primary(vals, ctxs, n_prim):
                continue
            nv, nc = _align_one_rule_series(vals, ctxs, n_prim, file_path=fp)
            vd[t] = nv
            cd[t] = nc


def _compress_series_by_kept_iters(
    values: list[Any],
    contexts: list[Any],
    old_to_new: dict[int, int],
    n_kept: int,
    file_path: str,
) -> tuple[list[Any], list[Any]]:
    """スキップ後、元 iter_index が残った行だけを新しい連番へ移す。"""
    new_v: list[Any] = [None] * n_kept
    new_c: list[Any] = [None] * n_kept
    n_ctx = len(contexts) if isinstance(contexts, list) else 0
    n_pair = max(len(values), n_ctx)
    for i in range(n_pair):
        v = values[i] if i < len(values) else None
        ctx = contexts[i] if i < n_ctx else None
        old_ix = _ctx_iter_index(ctx, i)
        new_i = old_to_new.get(old_ix)
        if new_i is None or new_c[new_i] is not None:
            continue
        new_v[new_i] = v
        if isinstance(ctx, dict):
            c = dict(ctx)
            c["iter_index"] = int(new_i)
            new_c[new_i] = c
        else:
            new_c[new_i] = {"file_path": str(file_path), "iter_index": int(new_i)}
    for i in range(n_kept):
        if new_c[i] is None:
            new_c[i] = {"file_path": str(file_path), "iter_index": int(i)}
    return new_v, new_c


def _apply_skip_empty_primary_filter(bundle: dict[str, Any], sources: list[Any]) -> None:
    """
    skip_empty_primary 対象ソースの空主キー反復を、連携・結合・path_item も含めて落とす。
    rule_iter_index（シート上の元オフセット）は残す。
    """
    prim = list(bundle.get("primary_values") or [])
    if not prim:
        return
    span_map_raw = bundle.get("_cell_source_spans")
    span_map: dict[Any, Any] = span_map_raw if isinstance(span_map_raw, dict) else {}
    drop = [False] * len(prim)
    any_skip = False
    for si, src in enumerate(sources):
        if not isinstance(src, dict):
            continue
        if (src.get("type") or "cell").strip().lower() != "cell":
            continue
        tokens = effective_skip_primary_tokens(src)
        if not tokens:
            continue
        any_skip = True
        if source_skips_sheet_extract(src):
            continue
        sp = span_map.get(si)
        if not sp or not isinstance(sp, (tuple, list)) or len(sp) != 2:
            if isinstance(span_map_raw, dict) and span_map:
                continue
            g_off, n_src = 0, len(prim)
        else:
            try:
                g_off = int(sp[0])
                n_src = int(sp[1])
            except (TypeError, ValueError):
                continue
        for local_i in range(max(0, n_src)):
            gi = g_off + local_i
            if 0 <= gi < len(prim) and primary_value_matches_skip_tokens(prim[gi], tokens):
                drop[gi] = True
    if not any_skip or not any(drop):
        return
    keep_idx = [i for i, d in enumerate(drop) if not d]
    old_to_new = {old_i: new_i for new_i, old_i in enumerate(keep_idx)}
    n_kept = len(keep_idx)
    fp = _bundle_file_path(bundle)

    def _take(lst: Any) -> list[Any]:
        if not isinstance(lst, list):
            return []
        return [lst[i] for i in keep_idx if i < len(lst)]

    bundle["primary_values"] = _take(prim)
    old_ictx = bundle.get("iteration_contexts") or []
    new_ictx: list[Any] = []
    new_prim = bundle["primary_values"]
    for new_i, old_i in enumerate(keep_idx):
        if old_i < len(old_ictx) and isinstance(old_ictx[old_i], dict):
            ctx = dict(old_ictx[old_i])
        else:
            ctx = {}
        ctx["iter_index"] = int(new_i)
        if "rule_iter_index" not in ctx:
            # スパン内ローカル位置を推定
            ri = old_i
            for si, src in enumerate(sources):
                if not isinstance(src, dict):
                    continue
                sp = span_map.get(si)
                if not sp or not isinstance(sp, (tuple, list)) or len(sp) != 2:
                    continue
                try:
                    g0 = int(sp[0])
                    n0 = int(sp[1])
                except (TypeError, ValueError):
                    continue
                if g0 <= old_i < g0 + n0:
                    ri = old_i - g0
                    break
            ctx["rule_iter_index"] = int(ri)
        ctx["primary_value"] = new_prim[new_i] if new_i < len(new_prim) else None
        new_ictx.append(ctx)
    bundle["iteration_contexts"] = new_ictx

    # 値列はリスト位置ではなく元 iter_index で残す（短いカード専用列がユニット行に載らないようにする）
    for vkey, ckey in (
        ("link_values", "link_contexts"),
        ("join_values", "join_contexts"),
        ("path_item_values", "path_item_contexts"),
    ):
        vd = bundle.get(vkey)
        if not isinstance(vd, dict):
            continue
        cd = bundle.get(ckey)
        if not isinstance(cd, dict):
            cd = {}
            bundle[ckey] = cd
        for t in list(vd.keys()):
            vals = vd.get(t)
            if not isinstance(vals, list):
                continue
            ctxs = cd.get(t)
            if not isinstance(ctxs, list):
                ctxs = []
            nv, nc = _compress_series_by_kept_iters(vals, ctxs, old_to_new, n_kept, fp)
            vd[t] = nv
            cd[t] = nc

    # スパンを圧縮後の長さに更新
    new_spans: dict[int, tuple[int, int]] = {}
    cursor = 0
    for si, src in enumerate(sources):
        if not isinstance(src, dict):
            continue
        if (src.get("type") or "cell").strip().lower() != "cell":
            continue
        sp = span_map.get(si)
        if not sp or not isinstance(sp, (tuple, list)) or len(sp) != 2:
            continue
        try:
            g_off = int(sp[0])
            n_src = int(sp[1])
        except (TypeError, ValueError):
            continue
        kept = sum(1 for i in range(g_off, g_off + n_src) if i < len(drop) and not drop[i])
        if kept > 0:
            new_spans[si] = (cursor, kept)
            cursor += kept
        else:
            new_spans[si] = (cursor, 0)
    bundle["_cell_source_spans"] = new_spans


def link_def_wants_carry_empty(ld: dict[str, Any] | None) -> bool:
    """連携キー定義の空欄前置保持が有効か。"""
    if not isinstance(ld, dict):
        return False
    v = ld.get("carry_empty")
    if v is True or v == 1:
        return True
    if isinstance(v, str) and v.strip().lower() in ("1", "true", "yes", "on"):
        return True
    return False


def _is_blank_link_carry_value(v: Any) -> bool:
    """前置保持用の空欄判定（空白のみ・Excel テキストの先頭 ' のみも空）。"""
    if v is None:
        return True
    s = str(v).strip()
    if s.startswith("'"):
        s = s[1:].strip()
    return s == ""


def _apply_carry_empty_link_values(bundle: dict[str, Any], sources: list[Any]) -> None:
    """
    carry_empty の連携キー列について、ソース区間内で空欄を直前の非空値で埋める。
    シート／ファイルをまたがない（呼び出し側がシート単位の bundle であること）。
    """
    vals_map = bundle.get("link_values")
    if not isinstance(vals_map, dict) or not vals_map:
        return
    span_map_raw = bundle.get("_cell_source_spans")
    span_map: dict[Any, Any] = span_map_raw if isinstance(span_map_raw, dict) else {}
    n_all = len(bundle.get("primary_values") or [])
    for si, src in enumerate(sources):
        if not isinstance(src, dict):
            continue
        if (src.get("type") or "cell").strip().lower() != "cell":
            continue
        if source_skips_sheet_extract(src):
            continue
        ui_blk = source_ui_block(src)
        if not isinstance(ui_blk, dict):
            continue
        sp = span_map.get(si) if span_map else None
        if sp and isinstance(sp, (tuple, list)) and len(sp) == 2:
            try:
                g_off = int(sp[0])
                n_src = int(sp[1])
            except (TypeError, ValueError):
                continue
        else:
            if span_map:
                continue
            g_off, n_src = 0, n_all
        if n_src < 1:
            continue
        for ld in ui_blk.get("link_defs") or []:
            if not isinstance(ld, dict):
                continue
            if not link_def_wants_carry_empty(ld):
                continue
            mode = str(ld.get("mode") or "セル座標").strip()
            if "固定" in mode or mode.lower() in ("fixed", "literal"):
                continue
            target = str(ld.get("item") or "").strip()
            if not target:
                continue
            series = vals_map.get(target)
            if not isinstance(series, list) or not series:
                continue
            last: Any = None
            for local_i in range(n_src):
                gi = g_off + local_i
                if gi < 0 or gi >= len(series):
                    break
                cur = series[gi]
                if _is_blank_link_carry_value(cur):
                    if last is not None:
                        series[gi] = last
                else:
                    last = cur


def _resolve_cell_source_repeat(src: dict[str, Any]) -> tuple[bool, bool, Optional[int]]:
    """(repeat_until_empty, repeat_until_last, repeat_max_for_n_count)。"""
    mode = source_end_mode(src)
    until_empty = mode == END_MODE_UNTIL_EMPTY
    until_last = mode == END_MODE_UNTIL_LAST
    repeat_max: Optional[int] = None
    if mode == END_MODE_N_COUNT:
        try:
            repeat_max = int(src["repeat_max"]) if src.get("repeat_max") is not None else None
        except (TypeError, ValueError):
            repeat_max = None
    return until_empty, until_last, repeat_max


def _stop_repeat_on_empty(v: Any, *, repeat_until_empty: bool) -> bool:
    return bool(repeat_until_empty) and is_blank_primary_value(v)


def extract_item_values(
    file_path: str | Path,
    item_config: dict[str, Any],
    item_id: Optional[str] = None,
    cell_positions: Optional[dict[str, tuple[int, int]]] = None,
    max_primary_rows: Optional[int] = None,
    *,
    cell_source_spans_out: Optional[dict[int, tuple[int, int]]] = None,
    cancel_check: Optional[Callable[..., None]] = None,
) -> list[Any]:
    """
    1 項目分の設定に基づき、指定ファイルから抽出した値のリストを返す。

    【概要】
      item_config の sources を走査し、座標・メタデータ・ファイル名の各ソースから値を集める。
      sources が空のときは [""] のみを返す（ファイル名での主値フォールバックは行わない）。
      複数ソースがある場合は順に取得し、空でない値をリストに追加する。
      単一セル・単一メタデータの場合は要素 1 のリストを返す。
      セル座標の相対指定（anchor + row_offset, col_offset）に対応。

    【引数】
      file_path: 対象ファイルのパス。
      item_config: 項目設定（sources: [ { type, sheet_name?, cell_ref?, anchor?, row_offset?, col_offset? } ] 等）。
      item_id: 項目ID。セル位置の記録に使用。
      cell_positions: 項目ID→(col, row) の辞書。相対指定の基準と、抽出したセル位置の記録に使用。

    【戻り値】
      抽出した値のリスト。1 ステップで 1 項目の「列」に書き込む値の並び。

    cell_source_spans_out:
      指定時、セル系ソースごとに (結合主値リスト上の開始インデックス, 行数) を格納する。
    """
    results: list[Any] = []
    sources = item_config.get("sources") or []
    positions = cell_positions if cell_positions is not None else {}
    item_label = str(item_config.get("name") or item_config.get("id") or item_id or "-")
    if not sources:
        # ソース未設定列は主値を載せない（ファイル名フォールバックは一括で誤出力になるため）
        return [""]
    for si, src in enumerate(sources):
        _poll_cancel_check(cancel_check)
        if not isinstance(src, dict):
            continue
        stype = (src.get("type") or "cell").strip().lower()
        if stype == "name_extract":
            if not name_extract_search_matches(file_path, src):
                continue
            ui_blk = source_ui_block(src)
            ex_mode = str((ui_blk or {}).get("extract_mode") or "extract").strip().lower()
            if ex_mode == "fixed":
                raw_f = src.get("length_value")
                v_fix = "" if raw_f is None else str(raw_f).strip()
                if v_fix:
                    results.append(
                        postprocess_name_extract_primary(
                            v_fix, ui_blk if isinstance(ui_blk, dict) else None
                        )
                    )
                continue
            v = extract_from_name(
                file_path,
                source_type=src.get("source_type") or "file_name",
                search_condition=src.get("search_condition"),
                search_text=src.get("search_text"),
                start_mode=src.get("start_mode") or START_MODE_HEAD,
                start_value=src.get("start_value"),
                length_mode=src.get("length_mode") or LENGTH_MODE_END,
                length_value=src.get("length_value"),
                delimiter=src.get("delimiter"),
                part_index=int(x) if (x := src.get("part_index")) is not None else None,
                pattern=src.get("pattern"),
                replacement=src.get("replacement"),
            )
            if v is not None and str(v) != "":
                results.append(
                    postprocess_name_extract_primary(v, ui_blk if isinstance(ui_blk, dict) else None)
                )
        elif stype == "metadata" or stype == "meta":
            meta_type = src.get("source_type") or SOURCE_FILE_NAME
            raw = extract_metadata(file_path, meta_type)
            start = src.get("start")
            length = src.get("length")
            delimiter = src.get("delimiter")
            part_index = src.get("part_index")
            pattern = src.get("pattern")
            if (
                start is not None
                or length is not None
                or delimiter is not None
                or part_index is not None
                or pattern
            ):
                ln_val = int(length) if length is not None else None
                if ln_val is not None and ln_val <= 0:
                    ln_val = None
                v = _extract_from_string(
                    raw,
                    start=int(start) if start is not None else 0,
                    length=ln_val,
                    delimiter=delimiter,
                    part_index=int(part_index) if part_index is not None else None,
                    pattern=pattern,
                )
            else:
                v = raw
            if v is not None:
                ui_blk = source_ui_block(src)
                results.append(
                    postprocess_metadata_like_primary(v, ui_blk if isinstance(ui_blk, dict) else None)
                )
        elif stype == "filename":
            v = extract_from_filename(
                file_path,
                start=src.get("start", 0),
                length=src.get("length"),
                delimiter=src.get("delimiter"),
                part_index=src.get("part_index"),
                pattern=src.get("pattern"),
            )
            ui_blk = source_ui_block(src)
            results.append(
                postprocess_metadata_like_primary(v, ui_blk if isinstance(ui_blk, dict) else None)
            )
        else:
            # cell / coordinate（絶対 or 相対）
            if not source_passes_file_name_filter(file_path, src):
                continue
            if source_skips_sheet_extract(src):
                continue
            cell_start = len(results)
            sheet_name = src.get("sheet_name")
            cell_ref = src.get("cell_ref") or "A1"
            anchor = src.get("anchor")
            row_off = int(src.get("row_offset") or 0)
            col_off = int(src.get("col_offset") or 0)
            if anchor and anchor in positions:
                base_col, base_row = positions[anchor]
                col = base_col + col_off
                row = base_row + row_off
                cell_ref = _col_row_to_cell_ref(col, row)
            ui_blk = source_ui_block(src)
            _blk = ui_blk if isinstance(ui_blk, dict) else None
            repeat_dir = (src.get("repeat_direction") or "").strip().lower()
            if repeat_dir in ("vertical", "horizontal"):
                repeat_until_empty, repeat_until_last, repeat_max = _resolve_cell_source_repeat(src)
                limit = resolve_extract_repeat_limit(
                    repeat_max=repeat_max,
                    repeat_until_empty=repeat_until_empty,
                    repeat_until_last=repeat_until_last,
                    max_primary_rows=max_primary_rows,
                )
                skip_trunc_peek = skip_extract_truncation_peek(
                    repeat_max=repeat_max,
                    repeat_until_empty=repeat_until_empty,
                    repeat_until_last=repeat_until_last,
                )
                vals: list[Any] = []
                base_col, base_row = _parse_cell_ref(cell_ref)
                if base_col is None or base_row is None:
                    base_col, base_row = _parse_cell_ref("A1")
                if base_col is None or base_row is None:
                    base_col, base_row = 0, 0
                p_abs = Path(file_path).resolve()
                wb_ctx: Any = None
                wb_owned = False
                rd_row_step = row_off if repeat_dir == "vertical" else 0
                rd_col_step = col_off if repeat_dir == "horizontal" else 0
                if p_abs.suffix.lower() == ".csv":
                    if _xlsx_workbook_cache_top() is not None:
                        vals_opt = _read_repeated_series_from_csv_cached(
                            p_abs,
                            base_col=base_col,
                            base_row=base_row,
                            row_step=rd_row_step,
                            col_step=rd_col_step,
                            limit=limit,
                            repeat_until_empty=repeat_until_empty,
                            cancel_check=cancel_check,
                        )
                        if vals_opt is None:
                            raise DataAggCsvReadError(
                                "CSV 一括抽出に失敗しました: %s" % p_abs
                            )
                        vals = vals_opt
                        peek_v = None
                        if not skip_trunc_peek:
                            peek_v = _peek_repeat_cell_from_csv(
                                p_abs,
                                base_col=base_col,
                                base_row=base_row,
                                row_off=row_off,
                                col_off=col_off,
                                index=len(vals),
                            )
                        _finish_repeated_cell_vals(
                            vals=vals,
                            results=results,
                            cell_start=cell_start,
                            cell_source_spans_out=cell_source_spans_out,
                            si=si,
                            limit=limit,
                            repeat_until_empty=repeat_until_empty,
                            file_path=file_path,
                            item_label=item_label,
                            item_id=item_id,
                            positions=positions,
                            cell_ref=cell_ref,
                            row_off=row_off,
                            col_off=col_off,
                            ui_blk=_blk,
                            src=src,
                            max_primary_rows=max_primary_rows,
                            peek_v=peek_v,
                            skip_trunc_peek=skip_trunc_peek,
                        )
                        continue
                    csv_mat = _get_csv_matrix(p_abs, create=False)
                    if csv_mat is not None:
                        vals = _read_repeated_series_from_matrix(
                            csv_mat,
                            base_col=base_col,
                            base_row=base_row,
                            row_step=rd_row_step,
                            col_step=rd_col_step,
                            limit=limit,
                            repeat_until_empty=repeat_until_empty,
                            cancel_check=cancel_check,
                        )
                        peek_v = None
                        if not skip_trunc_peek:
                            peek_v = _peek_repeat_cell_from_matrix(
                                csv_mat,
                                base_col=base_col,
                                base_row=base_row,
                                row_off=row_off,
                                col_off=col_off,
                                index=len(vals),
                            )
                        _finish_repeated_cell_vals(
                            vals=vals,
                            results=results,
                            cell_start=cell_start,
                            cell_source_spans_out=cell_source_spans_out,
                            si=si,
                            limit=limit,
                            repeat_until_empty=repeat_until_empty,
                            file_path=file_path,
                            item_label=item_label,
                            item_id=item_id,
                            positions=positions,
                            cell_ref=cell_ref,
                            row_off=row_off,
                            col_off=col_off,
                            ui_blk=_blk,
                            src=src,
                            max_primary_rows=max_primary_rows,
                            peek_v=peek_v,
                            skip_trunc_peek=skip_trunc_peek,
                        )
                        continue
                try:
                    if is_openxml_excel_suffix(p_abs.suffix):
                        wb_ctx = _xlsx_workbook_from_cache(p_abs)
                        if wb_ctx is None:
                            try:
                                import openpyxl  # noqa: E402

                                wb_ctx = _load_workbook_readonly(p_abs)
                                wb_owned = True
                            except Exception:
                                wb_ctx = None
                    elif p_abs.suffix.lower() == ".xls":
                        from svc.data_agg_xls_io import read_xls_cell, read_xls_repeated_series

                        vals_xls = read_xls_repeated_series(
                            p_abs,
                            sheet_name,
                            base_col=base_col,
                            base_row=base_row,
                            row_step=rd_row_step,
                            col_step=rd_col_step,
                            limit=limit,
                            repeat_until_empty=repeat_until_empty,
                        )
                        peek_v = None
                        if not skip_trunc_peek:
                            peek_ref = _col_row_to_cell_ref(
                                base_col + (col_off * len(vals_xls)),
                                base_row + (row_off * len(vals_xls)),
                            )
                            peek_v = read_xls_cell(p_abs, sheet_name, peek_ref)
                        _finish_repeated_cell_vals(
                            vals=vals_xls,
                            results=results,
                            cell_start=cell_start,
                            cell_source_spans_out=cell_source_spans_out,
                            si=si,
                            limit=limit,
                            repeat_until_empty=repeat_until_empty,
                            file_path=file_path,
                            item_label=item_label,
                            item_id=item_id,
                            positions=positions,
                            cell_ref=cell_ref,
                            row_off=row_off,
                            col_off=col_off,
                            ui_blk=_blk,
                            src=src,
                            max_primary_rows=max_primary_rows,
                            peek_v=peek_v,
                            skip_trunc_peek=skip_trunc_peek,
                        )
                        continue
                    if wb_ctx is not None:
                        vals_fast = _xlsx_read_repeated_series_open_workbook(
                            wb_ctx,
                            sheet_name,
                            base_col=base_col,
                            base_row=base_row,
                            row_step=row_off,
                            col_step=col_off,
                            limit=limit,
                            repeat_until_empty=repeat_until_empty,
                            path=p_abs if not wb_owned else None,
                            cancel_check=cancel_check,
                        )
                        if isinstance(vals_fast, list):
                            vals = vals_fast
                            peek_v = None
                            if not skip_trunc_peek:
                                peek_v = _read_cell_at_repeat_index(
                                    wb_ctx=wb_ctx,
                                    wb_owned=wb_owned,
                                    file_path=file_path,
                                    sheet_name=sheet_name,
                                    base_col=base_col,
                                    base_row=base_row,
                                    row_off=row_off,
                                    col_off=col_off,
                                    index=len(vals),
                                )
                            _finish_repeated_cell_vals(
                                vals=vals,
                                results=results,
                                cell_start=cell_start,
                                cell_source_spans_out=cell_source_spans_out,
                                si=si,
                                limit=limit,
                                repeat_until_empty=repeat_until_empty,
                                file_path=file_path,
                                item_label=item_label,
                                item_id=item_id,
                                positions=positions,
                                cell_ref=cell_ref,
                                row_off=row_off,
                                col_off=col_off,
                                ui_blk=_blk,
                                src=src,
                                max_primary_rows=max_primary_rows,
                                peek_v=peek_v,
                                skip_trunc_peek=skip_trunc_peek,
                            )
                            continue
                    if p_abs.suffix.lower() == ".csv" and _xlsx_workbook_cache_top() is not None:
                        raise DataAggCsvReadError(
                            "CSV 一括抽出に失敗（バッチ内逐次ループ禁止）: %s" % p_abs
                        )
                    # 取得座標 = 基準セル + (行/列オフセット * N)（N は 0 始まり）
                    for n in range(limit):
                        _poll_cancel_check(cancel_check)
                        col_n = base_col + (col_off * n)
                        row_n = base_row + (row_off * n)
                        cell_ref_n = _col_row_to_cell_ref(col_n, row_n)
                        if wb_ctx is not None:
                            v = _xlsx_cell_value_open_workbook(
                                wb_ctx,
                                sheet_name,
                                cell_ref_n,
                                path=p_abs,
                                ephemeral_sheet_cache=wb_owned,
                            )
                        else:
                            v = extract_cell(
                                file_path,
                                sheet_name=sheet_name,
                                cell_ref=cell_ref_n,
                            )
                        if _stop_repeat_on_empty(v, repeat_until_empty=repeat_until_empty):
                            break
                        vals.append(v)
                    if vals:
                        if repeat_until_last:
                            vals = apply_until_last_trim(vals, until_last=True)
                        if not skip_trunc_peek:
                            peek_v = _read_cell_at_repeat_index(
                                wb_ctx=wb_ctx,
                                wb_owned=wb_owned,
                                file_path=file_path,
                                sheet_name=sheet_name,
                                base_col=base_col,
                                base_row=base_row,
                                row_off=row_off,
                                col_off=col_off,
                                index=len(vals),
                            )
                            record_extract_truncation_if_needed(
                                vals,
                                limit=limit,
                                peek_next=peek_v,
                                file_path=file_path,
                                item_label=item_label,
                                source_index=si,
                            )
                finally:
                    if wb_owned and wb_ctx is not None:
                        try:
                            wb_ctx.close()
                        except Exception:
                            pass
                _append_postprocessed_cell_vals(
                    results,
                    vals,
                    _blk,
                    src,
                    max_primary_rows=max_primary_rows,
                )
                if vals and item_id:
                    n_last = len(vals) - 1
                    cell_ref_last = _resolve_cell_with_offset(
                        cell_ref,
                        row_off * n_last,
                        col_off * n_last,
                    )
                    c0, r0 = _parse_cell_ref(cell_ref_last)
                    if c0 is not None and r0 is not None:
                        positions[item_id] = (c0, r0)
            else:
                v = extract_cell(
                    file_path,
                    sheet_name=sheet_name,
                    cell_ref=cell_ref,
                )
                if v is not None and (v != "" or src.get("allow_empty")):
                    results.append(postprocess_cell_primary(v, _blk))
                    col, row = _parse_cell_ref(cell_ref)
                    if col is not None and row is not None and item_id:
                        positions[item_id] = (col, row)
            if cell_source_spans_out is not None:
                cell_source_spans_out[si] = (cell_start, len(results) - cell_start)
        if max_primary_rows is not None and max_primary_rows > 0 and len(results) >= max_primary_rows:
            break
    if max_primary_rows is not None and max_primary_rows > 0:
        results = results[:max_primary_rows]
        if cell_source_spans_out is not None:
            for k in list(cell_source_spans_out.keys()):
                g0, ln = cell_source_spans_out[k]
                if g0 + ln > len(results):
                    cell_source_spans_out[k] = (g0, max(0, len(results) - g0))
    return results if results else [None]


def _resolve_cell_with_offset(base_cell: str, row_off: Any = 0, col_off: Any = 0) -> str:
    """A1 形式セルへ行/列オフセットを適用して A1 形式を返す。"""
    c0, r0 = _parse_cell_ref(base_cell or "A1")
    if c0 is None or r0 is None:
        c0, r0 = 0, 0
    try:
        ro = int(row_off or 0)
    except (TypeError, ValueError):
        ro = 0
    try:
        co = int(col_off or 0)
    except (TypeError, ValueError):
        co = 0
    return _col_row_to_cell_ref(c0 + co, r0 + ro)


def _extract_from_cell_rule(
    file_path: str | Path,
    src: dict[str, Any],
    rule: dict[str, Any],
) -> Any:
    """link_defs / join_defs の 1 ルールから値を抽出する。"""
    mode = str(rule.get("mode") or "セル座標").strip()
    rdict = rule if isinstance(rule, dict) else {}
    if "固定" in mode or mode.lower() in ("fixed", "literal"):
        return postprocess_link_rule_value(rule.get("cell"), rdict)
    sheet_name = src.get("sheet_name")
    base_cell = str(rule.get("cell") or src.get("cell_ref") or "A1")
    # 連携キー/結合キーとも同一の座標解決規約（基準セル + 独自 row/col offset）を使う。
    cell_ref = _resolve_cell_with_offset(base_cell, rule.get("row"), rule.get("col"))
    v = extract_cell(file_path, sheet_name=sheet_name, cell_ref=cell_ref)
    return postprocess_link_rule_value(v, rdict)


def _cell_ref_with_iteration(base_cell: str, repeat_direction: str, iter_index: int) -> str:
    """主キーの反復方向に合わせて基準セルを iter_index 分だけ進める。"""
    c0, r0 = _parse_cell_ref(base_cell or "A1")
    if c0 is None or r0 is None:
        c0, r0 = 0, 0
    rd = (repeat_direction or "").strip().lower()
    if rd == "vertical":
        r0 += max(0, int(iter_index))
    elif rd == "horizontal":
        c0 += max(0, int(iter_index))
    return _col_row_to_cell_ref(c0, r0)


def _extract_from_cell_rule_with_context(
    file_path: str | Path,
    src: dict[str, Any],
    rule: dict[str, Any],
    iter_ctx: dict[str, Any],
) -> Any:
    """
    反復コンテキスト（file_path + base_cell）に基づいて link/join 値を取得する。
    rule の row/col は link/join 側の独自オフセットとして常に適用する。
    シート上の段進みには rule_iter_index（ソース内 0 始まり）を使い、未指定時は iter_index
    （結合主値リスト上の行。後方互換）にフォールバックする。
    """
    mode = str(rule.get("mode") or "セル座標").strip()
    rdict = rule if isinstance(rule, dict) else {}
    if "固定" in mode or mode.lower() in ("fixed", "literal"):
        return postprocess_link_rule_value(rule.get("cell"), rdict)
    sheet_name = src.get("sheet_name")
    try:
        rule_iter = int(iter_ctx.get("rule_iter_index", iter_ctx.get("iter_index", 0)))
    except (TypeError, ValueError):
        rule_iter = 0
    base_cell = str(rule.get("cell") or iter_ctx.get("base_cell") or src.get("cell_ref") or "A1")

    def _to_int(v: Any, default: int = 0) -> int:
        try:
            return int(v)
        except (TypeError, ValueError):
            return default

    # 移動オフセットは 0 開始。rule_iter=0 は 0、1 は offset …（ソース内の何行目か）
    row_step = _to_int(rule.get("row"), 0)
    col_step = _to_int(rule.get("col"), 0)
    row_off = row_step * max(0, rule_iter)
    col_off = col_step * max(0, rule_iter)
    # 取得座標 = 設定座標 + (オフセット * N)（N は 0 開始）
    # ここでは基準セルを二重に進めず、offset*N のみを適用する。
    cell_ref = _resolve_cell_with_offset(base_cell, row_off, col_off)
    v = extract_cell(file_path, sheet_name=sheet_name, cell_ref=cell_ref)
    return postprocess_link_rule_value(v, rdict)


def _extract_cell_rule_series_fast(
    file_path: str | Path,
    src: dict[str, Any],
    rule: dict[str, Any],
    *,
    n_src: int,
    cancel_check: Optional[Callable[..., None]] = None,
) -> Optional[list[Any]]:
    """
    link/join ルールの反復値を高速取得する（OpenXML Excel / CSV の典型セル座標パターン）。
    非対応時は None を返し、呼び出し側で行列キャッシュ経由の 1 セルずつ取得へフォールバックする。
    """
    if n_src < 1:
        return []
    mode = str(rule.get("mode") or "セル座標").strip()
    if "固定" in mode or mode.lower() in ("fixed", "literal"):
        return [postprocess_link_rule_value(rule.get("cell"), rule)] * n_src
    p_abs = Path(file_path).resolve()
    is_csv = p_abs.suffix.lower() == ".csv"
    is_xls = p_abs.suffix.lower() == ".xls"
    if not is_csv and not is_openxml_excel_suffix(p_abs.suffix) and not is_xls:
        return None
    base_cell = str(rule.get("cell") or src.get("cell_ref") or "A1").strip()
    c0, r0 = _parse_cell_ref(base_cell)
    if c0 is None or r0 is None:
        return None
    try:
        row_step = int(rule.get("row") or 0)
    except (TypeError, ValueError):
        row_step = 0
    try:
        col_step = int(rule.get("col") or 0)
    except (TypeError, ValueError):
        col_step = 0
    if is_xls:
        from svc.data_agg_xls_io import read_xls_repeated_series

        raw = read_xls_repeated_series(
            p_abs,
            src.get("sheet_name"),
            base_col=c0,
            base_row=r0,
            row_step=row_step,
            col_step=col_step,
            limit=n_src,
            repeat_until_empty=False,
        )
        out = postprocess_link_rule_value_batch(raw, rule)
        if len(out) < n_src:
            out.extend([None] * (n_src - len(out)))
        return out[:n_src]
    if is_csv:
        df = _get_csv_df(p_abs, create=True)
        if df is not None:
            raw = _read_repeated_series_from_df(
                df,
                base_col=c0,
                base_row=r0,
                row_step=row_step,
                col_step=col_step,
                limit=n_src,
                repeat_until_empty=False,
                cancel_check=cancel_check,
            )
        else:
            mat = _get_csv_matrix(p_abs, create=True)
            if mat is None:
                return None
            raw = _read_repeated_series_from_matrix(
                mat,
                base_col=c0,
                base_row=r0,
                row_step=row_step,
                col_step=col_step,
                limit=n_src,
                repeat_until_empty=False,
                cancel_check=cancel_check,
            )
        out = postprocess_link_rule_value_batch(raw, rule)
        if len(out) < n_src:
            out.extend([None] * (n_src - len(out)))
        return out[:n_src]
    wb_ctx = _xlsx_workbook_from_cache(p_abs)
    wb_owned = False
    if wb_ctx is None:
        try:
            import openpyxl  # noqa: E402

            wb_ctx = _load_workbook_readonly(p_abs)
            wb_owned = True
        except Exception:
            wb_ctx = None
    if wb_ctx is None:
        return None
    try:
        raw = _xlsx_read_repeated_series_open_workbook(
            wb_ctx,
            src.get("sheet_name"),
            base_col=c0,
            base_row=r0,
            row_step=row_step,
            col_step=col_step,
            limit=n_src,
            repeat_until_empty=False,
            path=p_abs if not wb_owned else None,
            cancel_check=cancel_check,
        )
        if not isinstance(raw, list):
            return None
        out = [postprocess_link_rule_value(v, rule) for v in raw]
        if len(out) < n_src:
            out.extend([None] * (n_src - len(out)))
        return out[:n_src]
    finally:
        if wb_owned:
            try:
                wb_ctx.close()
            except Exception:
                pass


def _extract_cell_rules_series_fast_map(
    file_path: str | Path,
    src: dict[str, Any],
    rules: list[dict[str, Any]],
    *,
    n_src: int,
    cancel_check: Optional[Callable[..., None]] = None,
) -> Optional[dict[int, list[Any]]]:
    """
    複数 link/join ルールを 1 回の列範囲走査でまとめて取得する高速経路。
    対応条件（厳しめ）:
      - OpenXML Excel（.xlsx / .xlsm）または CSV
      - 非固定値ルールは同一基準行・同一 row_step（col=0 の縦反復）
    非対応時は None（呼び出し側で行列キャッシュ経由の 1 セルずつ取得へフォールバック）。
    部分適用: 開始行や歩幅が違うルールは個別高速経路へ回す。
    """
    if n_src < 1:
        return {}
    # 逆効果回避:
    # - 反復が少ないケースは準備コストが勝ちやすい
    # - ルール数が少ないケースは個別高速経路の方が軽い
    if n_src < 8:
        return None
    p_abs = Path(file_path).resolve()
    is_csv = p_abs.suffix.lower() == ".csv"
    if not is_csv and not is_openxml_excel_suffix(p_abs.suffix):
        return None
    if not isinstance(rules, list) or not rules:
        return {}
    if len(rules) < 3:
        return None
    non_fixed: list[tuple[int, int, int, dict[str, Any]]] = []
    fixed: list[tuple[int, Any, dict[str, Any]]] = []
    base_row_ref: Optional[int] = None
    row_step_ref: Optional[int] = None
    for i, r in enumerate(rules):
        if not isinstance(r, dict):
            return None
        mode = str(r.get("mode") or "セル座標").strip()
        if "固定" in mode or mode.lower() in ("fixed", "literal"):
            fixed.append((i, r.get("cell"), r))
            continue
        base_cell = str(r.get("cell") or src.get("cell_ref") or "A1").strip()
        c0, r0 = _parse_cell_ref(base_cell)
        if c0 is None or r0 is None:
            return None
        try:
            row_step = int(r.get("row") or 0)
            col_step = int(r.get("col") or 0)
        except (TypeError, ValueError):
            return None
        # ここで全体を落とすと、混在ルール時に高速化が効かない。
        # 対応可能な一軸反復（縦/横、固定セル含む）だけ部分適用する。
        if not (
            (row_step > 0 and col_step == 0)
            or (row_step == 0 and col_step > 0)
            or (row_step == 0 and col_step == 0)
        ):
            continue
        # 現実装のバッチ走査は同一基準行・同一歩幅の縦反復のみ対象。
        # それ以外（横反復/固定セル/開始行違い/歩幅違い）は個別高速経路へ回す。
        if not (row_step > 0 and col_step == 0):
            continue
        if base_row_ref is None:
            base_row_ref = r0
            row_step_ref = row_step
        elif base_row_ref != r0 or row_step != row_step_ref:
            continue
        non_fixed.append((i, c0, r0, r))
    out: dict[int, list[Any]] = {}
    for i, fv, r in fixed:
        out[i] = [postprocess_link_rule_value(fv, r)] * n_src
    if not non_fixed:
        return out
    step = int(row_step_ref or 1)
    if step < 1:
        step = 1
    if is_csv:
        df = _get_csv_df(p_abs, create=True)
        row0_csv = base_row_ref or 0
        if df is not None:
            for i, c, _r, rule in non_fixed:
                raw = _read_repeated_series_from_df(
                    df,
                    base_col=c,
                    base_row=row0_csv,
                    row_step=step,
                    col_step=0,
                    limit=n_src,
                    repeat_until_empty=False,
                    cancel_check=cancel_check,
                )
                if len(raw) < n_src:
                    raw = raw + [None] * (n_src - len(raw))
                out[i] = postprocess_link_rule_value_batch(raw[:n_src], rule)
            return out
        mat = _get_csv_matrix(p_abs, create=True)
        if mat is None:
            return None
        for i, c, _r, rule in non_fixed:
            raw = [
                _matrix_cell_value(mat, c, row0_csv + ri * step) for ri in range(n_src)
            ]
            out[i] = postprocess_link_rule_value_batch(raw, rule)
        return out
    wb_ctx = _xlsx_workbook_from_cache(p_abs)
    wb_owned = False
    if wb_ctx is None:
        try:
            import openpyxl  # noqa: E402

            wb_ctx = _load_workbook_readonly(p_abs)
            wb_owned = True
        except Exception:
            wb_ctx = None
    if wb_ctx is None:
        return None
    try:
        cols = [x[1] for x in non_fixed]
        if cols and (max(cols) - min(cols)) > 128:
            return None
        bucket: dict[int, list[Any]] = {i: [] for i, _c, _r, _rule in non_fixed}
        row0 = base_row_ref or 0
        in_scope = _xlsx_workbook_cache_top() is not None and not wb_owned
        mat = (
            _get_readonly_sheet_matrix(wb_ctx, src.get("sheet_name"), p_abs, create=True)
            if in_scope
            else None
        )
        if mat is not None:
            for ri in range(n_src):
                if ri % 64 == 0:
                    _poll_cancel_check(cancel_check)
                src_row = row0 + ri * step
                for i, c, _r, rule in non_fixed:
                    v = _matrix_cell_value(mat, c, src_row)
                    bucket[i].append(postprocess_link_rule_value(v, rule))
        else:
            ws, _ = _resolve_readonly_worksheet(wb_ctx, src.get("sheet_name"))
            if ws is None:
                return None
            min_col = min(cols) + 1
            max_col = max(cols) + 1
            min_row = row0 + 1
            max_row = min_row + max(0, (n_src - 1) * step)
            col_pos = {c: (c + 1 - min_col) for c in cols}
            take_i = 0
            for ri, tup in enumerate(
                ws.iter_rows(
                    min_row=min_row,
                    max_row=max_row,
                    min_col=min_col,
                    max_col=max_col,
                    values_only=False,
                )
            ):
                if take_i >= n_src:
                    break
                if ri % step != 0:
                    continue
                if take_i % 64 == 0:
                    _poll_cancel_check(cancel_check)
                for i, c, _r, rule in non_fixed:
                    pos = col_pos.get(c, -1)
                    cell = tup[pos] if (0 <= pos < len(tup)) else None
                    v = extract_read_openpyxl_cell(cell)
                    bucket[i].append(postprocess_link_rule_value(v, rule))
                take_i += 1
        for i in bucket:
            if len(bucket[i]) < n_src:
                bucket[i].extend([None] * (n_src - len(bucket[i])))
            out[i] = bucket[i][:n_src]
        return out
    finally:
        if wb_owned:
            try:
                wb_ctx.close()
            except Exception:
                pass



def _empty_item_bundle() -> dict[str, Any]:
    return {
        "primary_values": [],
        "iteration_contexts": [],
        "link_values": {},
        "link_contexts": {},
        "join_values": {},
        "join_contexts": {},
        "path_item_values": {},
        "path_item_contexts": {},
    }


def _rule_iter_indices_for_sheet_slice(
    base: dict[str, Any],
    *,
    g_off: int,
    n_src: int,
    part: dict[str, Any] | None = None,
) -> list[int]:
    """
    シート部分の各行について、シート上の元オフセット (rule_iter_index) を返す。
    空スキップ後は連番でないことがある。
    """
    if isinstance(part, dict):
        raw = part.get("rule_iter_indices")
        if isinstance(raw, list) and len(raw) >= n_src:
            out: list[int] = []
            for i in range(n_src):
                try:
                    out.append(int(raw[i]))
                except (TypeError, ValueError):
                    out.append(i)
            return out
    base_ictx = base.get("iteration_contexts") or []
    out2: list[int] = []
    for i in range(n_src):
        ri = i
        gi = g_off + i
        if gi < len(base_ictx) and isinstance(base_ictx[gi], dict):
            raw_ri = base_ictx[gi].get("rule_iter_index")
            if raw_ri is not None:
                try:
                    ri = int(raw_ri)
                except (TypeError, ValueError):
                    ri = i
        out2.append(ri)
    return out2


def _mini_iter_contexts_for_sheet_part(
    *,
    file_path: str,
    sheet_name: str,
    g_off: int,
    n_src: int,
    mini_prim: list[Any],
    base: dict[str, Any],
    part: dict[str, Any] | None = None,
) -> list[dict[str, Any]]:
    """link/join 再計算用のミニ iteration_contexts（rule_iter_index 維持）。"""
    ris = _rule_iter_indices_for_sheet_slice(base, g_off=g_off, n_src=n_src, part=part)
    out: list[dict[str, Any]] = []
    for i in range(n_src):
        out.append(
            {
                "file_path": str(file_path),
                "iter_index": int(i),
                "rule_iter_index": int(ris[i] if i < len(ris) else i),
                "sheet_name": sheet_name,
                "base_cell": None,
                "primary_value": mini_prim[i] if i < len(mini_prim) else None,
            }
        )
    return out


def _merge_primary_sheet_bundles(
    parts: list[tuple[str, dict[str, Any]]],
    file_path: str,
) -> dict[str, Any]:
    merged = _empty_item_bundle()
    sheet_parts: list[dict[str, Any]] = []
    g_off = 0
    prim_all: list[Any] = []
    ictx_all: list[dict[str, Any]] = []
    for sh, b in parts:
        if not isinstance(b, dict):
            continue
        prim = list(b.get("primary_values") or [])
        n = len(prim)
        local_spans = b.get("_cell_source_spans")
        if not isinstance(local_spans, dict):
            local_spans = {0: (0, n)} if n else {}
        rule_iter_indices: list[int] = []
        for i, v in enumerate(prim):
            prim_all.append(v)
            b_ictx = b.get("iteration_contexts") or []
            if i < len(b_ictx) and isinstance(b_ictx[i], dict):
                src_ctx = dict(b_ictx[i])
            else:
                src_ctx = {
                    "file_path": str(file_path),
                    "base_cell": None,
                    "base_row": None,
                    "base_col": None,
                    "filter_snapshot": {"file_path": str(file_path), "passed": True},
                    "primary_value": v,
                }
            src_ctx["file_path"] = str(file_path)
            src_ctx["iter_index"] = int(g_off + i)
            src_ctx["sheet_name"] = sh
            src_ctx["primary_value"] = v
            if src_ctx.get("rule_iter_index") is None:
                src_ctx["rule_iter_index"] = int(i)
            try:
                rule_iter_indices.append(int(src_ctx.get("rule_iter_index", i)))
            except (TypeError, ValueError):
                rule_iter_indices.append(i)
            ictx_all.append(src_ctx)
        # 連携/結合もシート連結オフセットを付与（シートごとに 0.. だと後段で上書き・欠落する）
        for key in (
            "link_values",
            "link_contexts",
            "join_values",
            "join_contexts",
            "path_item_values",
            "path_item_contexts",
        ):
            mp = b.get(key)
            if not isinstance(mp, dict):
                continue
            dest = merged.setdefault(key, {})
            for tgt, vals in mp.items():
                if not isinstance(vals, list):
                    continue
                if key.endswith("_contexts"):
                    out_list = dest.setdefault(tgt, [])
                    for i, ctx in enumerate(vals):
                        c = dict(ctx) if isinstance(ctx, dict) else {}
                        c["file_path"] = str(c.get("file_path") or file_path)
                        # シート内の iter_index（ソース区間）を維持し、シート連結分だけずらす。
                        # リスト位置 i で振り直すと、カード専用列がユニット行（iter 0）に載る。
                        local_ix = _ctx_iter_index(c, i)
                        c["iter_index"] = int(g_off + local_ix)
                        c["sheet_name"] = sh
                        out_list.append(c)
                else:
                    dest.setdefault(tgt, []).extend(vals)
        sheet_parts.append(
            {
                "sheet_name": sh,
                "g_off": g_off,
                "n_src": n,
                "primary_values": prim,
                "rule_iter_indices": rule_iter_indices,
                "_cell_source_spans": dict(local_spans),
            }
        )
        g_off += n
    merged["primary_values"] = prim_all
    merged["iteration_contexts"] = ictx_all
    merged["_sheet_parts"] = sheet_parts
    merged["_cell_source_spans"] = _merged_cell_source_spans(parts)
    _align_link_join_series_to_primary(merged)
    return merged


def _append_rule_maps_with_offset(
    dest: dict[str, Any],
    src_bundle: dict[str, Any],
    *,
    values_key: str,
    contexts_key: str,
    g_off: int,
    file_path: str,
) -> None:
    svm = src_bundle.get(values_key) or {}
    scm = src_bundle.get(contexts_key) or {}
    if not isinstance(svm, dict):
        return
    for target, vals in svm.items():
        if not isinstance(vals, list):
            continue
        dest.setdefault(values_key, {}).setdefault(target, []).extend(vals)
        ctxs = scm.get(target) if isinstance(scm, dict) else None
        out_ctx = dest.setdefault(contexts_key, {}).setdefault(target, [])
        for i, _v in enumerate(vals):
            if isinstance(ctxs, list) and i < len(ctxs) and isinstance(ctxs[i], dict):
                c = dict(ctxs[i])
            else:
                c = {"file_path": str(file_path), "base_cell": None}
            c["file_path"] = str(c.get("file_path") or file_path)
            local_ix = _ctx_iter_index(c, i)
            c["iter_index"] = int(g_off + local_ix)
            out_ctx.append(c)


def _extract_link_join_across_sheet_parts(
    file_path: str | Path,
    item_config: dict[str, Any],
    base: dict[str, Any],
    *,
    item_id: Optional[str],
    join_path_header: Optional[str],
    scope: str,
    max_primary_rows: Optional[int],
    cancel_check: Optional[Callable[..., None]],
) -> dict[str, Any]:
    parts = base.get("_sheet_parts")
    out = base.copy()
    if scope == "link":
        out["link_values"] = {}
        out["link_contexts"] = {}
        values_key, contexts_key = "link_values", "link_contexts"
    else:
        out["join_values"] = {}
        out["join_contexts"] = {}
        values_key, contexts_key = "join_values", "join_contexts"
    if not isinstance(parts, list) or not parts:
        return _extract_item_bundle_impl(
            file_path,
            item_config,
            item_id=item_id,
            cell_positions={},
            join_path_header=join_path_header,
            debug_step_scope=scope,
            existing_bundle=base,
            max_primary_rows=max_primary_rows,
            cancel_check=cancel_check,
        )
    fp = str(file_path)
    wb_names = list_workbook_sheet_names(file_path)
    for part in parts:
        if not isinstance(part, dict):
            continue
        sh = str(part.get("sheet_name") or "")
        g_off = int(part.get("g_off") or 0)
        n_src = int(part.get("n_src") or 0)
        if not sh or n_src < 1:
            continue
        local_spans = part.get("_cell_source_spans")
        mini_spans = _mini_spans_from_local(local_spans, n_src)
        mini_prim = list(part.get("primary_values") or [])[:n_src]
        mini = {
            "primary_values": mini_prim,
            "iteration_contexts": _mini_iter_contexts_for_sheet_part(
                file_path=fp,
                sheet_name=sh,
                g_off=g_off,
                n_src=n_src,
                mini_prim=mini_prim,
                base=base,
                part=part,
            ),
            "link_values": {},
            "link_contexts": {},
            "join_values": {},
            "join_contexts": {},
            "path_item_values": {},
            "path_item_contexts": {},
            "_cell_source_spans": mini_spans,
        }
        partial = _extract_item_bundle_impl(
            file_path,
            patch_item_sheet_exact(item_config, sh, workbook_sheet_names=wb_names),
            item_id=item_id,
            cell_positions={},
            join_path_header=join_path_header,
            debug_step_scope=scope,
            existing_bundle=mini,
            max_primary_rows=max_primary_rows,
            cancel_check=cancel_check,
        )
        _append_rule_maps_with_offset(
            out,
            partial,
            values_key=values_key,
            contexts_key=contexts_key,
            g_off=g_off,
            file_path=fp,
        )
    _align_link_join_series_to_primary(out)
    return out


def extract_item_bundle(
    file_path: str | Path,
    item_config: dict[str, Any],
    item_id: Optional[str] = None,
    cell_positions: Optional[dict[str, tuple[int, int]]] = None,
    join_path_header: Optional[str] = None,
    *,
    debug_step_scope: Optional[str] = None,
    existing_bundle: Optional[dict[str, Any]] = None,
    max_primary_rows: Optional[int] = None,
    cancel_check: Optional[Callable[..., None]] = None,
) -> dict[str, Any]:
    """
    1 項目分の抽出結果を主値・連携値・結合キー値に分けて返す。

    シート名条件（左端／完全一致／含む／含まない）を解決し、
    複数一致時はブック左端から右へ順に読み取って連結する。
    マスタデバッグ／シナリオデバッグ／本番一括で共通。
    """
    if debug_step_scope in ("link", "join"):
        if (
            isinstance(existing_bundle, dict)
            and isinstance(existing_bundle.get("_sheet_parts"), list)
            and existing_bundle.get("_sheet_parts")
        ):
            return _extract_link_join_across_sheet_parts(
                file_path,
                item_config,
                existing_bundle,
                item_id=item_id,
                join_path_header=join_path_header,
                scope=str(debug_step_scope),
                max_primary_rows=max_primary_rows,
                cancel_check=cancel_check,
            )
        return _extract_item_bundle_impl(
            file_path,
            item_config,
            item_id=item_id,
            cell_positions=cell_positions,
            join_path_header=join_path_header,
            debug_step_scope=debug_step_scope,
            existing_bundle=existing_bundle,
            max_primary_rows=max_primary_rows,
            cancel_check=cancel_check,
        )

    sheets, wb_names = matching_sheets_and_names_for_item(file_path, item_config)
    if sheets is None:
        return _extract_item_bundle_impl(
            file_path,
            item_config,
            item_id=item_id,
            cell_positions=cell_positions,
            join_path_header=join_path_header,
            debug_step_scope=debug_step_scope,
            existing_bundle=existing_bundle,
            max_primary_rows=max_primary_rows,
            cancel_check=cancel_check,
        )
    if not sheets:
        return _empty_item_bundle()
    if len(sheets) == 1:
        return _extract_item_bundle_impl(
            file_path,
            patch_item_sheet_exact(
                item_config, sheets[0], workbook_sheet_names=wb_names
            ),
            item_id=item_id,
            cell_positions=cell_positions,
            join_path_header=join_path_header,
            debug_step_scope=debug_step_scope,
            existing_bundle=existing_bundle,
            max_primary_rows=max_primary_rows,
            cancel_check=cancel_check,
        )
    parts: list[tuple[str, dict[str, Any]]] = []
    remain = max_primary_rows
    for sh in sheets:
        if remain is not None and remain <= 0:
            break
        b = _extract_item_bundle_impl(
            file_path,
            patch_item_sheet_exact(item_config, sh, workbook_sheet_names=wb_names),
            item_id=item_id,
            cell_positions=cell_positions,
            join_path_header=join_path_header,
            debug_step_scope=debug_step_scope,
            existing_bundle=None,
            max_primary_rows=remain,
            cancel_check=cancel_check,
        )
        parts.append((sh, b))
        if remain is not None:
            remain = max(0, int(remain) - len(b.get("primary_values") or []))
    return _merge_primary_sheet_bundles(parts, str(file_path))


def _build_source_iter_contexts(
    *,
    file_path: str | Path,
    src_base: str,
    row_step: int,
    col_step: int,
    g_off: int,
    n_src: int,
    bundle: dict[str, Any],
) -> list[dict[str, Any]]:
    """
    セルソースの反復コンテキストを組む。
    既存 iteration_contexts に rule_iter_index があれば（空スキップ後）それを優先する。
    """
    iter_contexts: list[dict[str, Any]] = []
    b_ictx = bundle.get("iteration_contexts") or []
    for local_i in range(n_src):
        gi = g_off + local_i
        ri = local_i
        if gi < len(b_ictx) and isinstance(b_ictx[gi], dict):
            raw_ri = b_ictx[gi].get("rule_iter_index")
            if raw_ri is not None:
                try:
                    ri = int(raw_ri)
                except (TypeError, ValueError):
                    ri = local_i
        base_cell_i = _resolve_cell_with_offset(
            src_base,
            row_step * ri,
            col_step * ri,
        )
        iter_contexts.append(
            {
                "file_path": str(file_path),
                "iter_index": int(g_off + local_i),
                "rule_iter_index": int(ri),
                "base_cell": base_cell_i,
            }
        )
    for local_i, ic in enumerate(iter_contexts):
        gi = g_off + local_i
        if gi < len(b_ictx) and isinstance(b_ictx[gi], dict):
            b_ictx[gi]["base_cell"] = ic.get("base_cell")
            b_ictx[gi]["rule_iter_index"] = ic.get("rule_iter_index")
    return iter_contexts


def _rule_iter_indices_list(iter_contexts: list[dict[str, Any]]) -> list[int]:
    """各コンテキストの rule_iter_index（不正時はリスト位置）。"""
    out: list[int] = []
    for i, ic in enumerate(iter_contexts):
        if not isinstance(ic, dict):
            out.append(i)
            continue
        try:
            out.append(int(ic.get("rule_iter_index", i)))
        except (TypeError, ValueError):
            out.append(i)
    return out


def _append_rule_series_to_bundle(
    *,
    bundle: dict[str, Any],
    values_key: str,
    contexts_key: str,
    target: str,
    file_path: str | Path,
    src: dict[str, Any],
    rule: dict[str, Any],
    iter_contexts: list[dict[str, Any]],
    n_src: int,
    cancel_check: Optional[Callable[..., None]] = None,
) -> None:
    """link/join の系列値を bundle に追記。非連番 rule_iter も列一括読取して拾う。"""
    ris = _rule_iter_indices_list(iter_contexts)
    n_read = int(n_src)
    if ris:
        n_read = max(n_read, max(ris) + 1)
    vals_fast = (
        _extract_cell_rule_series_fast(
            file_path, src, rule, n_src=n_read, cancel_check=cancel_check
        )
        if n_read > 0
        else []
    )
    if isinstance(vals_fast, list):
        for local_i, iter_ctx in enumerate(iter_contexts):
            ri = ris[local_i] if local_i < len(ris) else local_i
            v = vals_fast[ri] if 0 <= ri < len(vals_fast) else None
            bundle[values_key].setdefault(target, []).append(v)
            bundle[contexts_key].setdefault(target, []).append(
                {
                    "file_path": str(iter_ctx.get("file_path") or file_path),
                    "iter_index": int(iter_ctx.get("iter_index", 0)),
                    "base_cell": iter_ctx.get("base_cell"),
                }
            )
        return
    for iter_ctx in iter_contexts:
        v = _extract_from_cell_rule_with_context(file_path, src, rule, iter_ctx)
        bundle[values_key].setdefault(target, []).append(v)
        bundle[contexts_key].setdefault(target, []).append(
            {
                "file_path": str(iter_ctx.get("file_path") or file_path),
                "iter_index": int(iter_ctx.get("iter_index", 0)),
                "base_cell": iter_ctx.get("base_cell"),
            }
        )


def _extract_item_bundle_impl(
    file_path: str | Path,
    item_config: dict[str, Any],
    item_id: Optional[str] = None,
    cell_positions: Optional[dict[str, tuple[int, int]]] = None,
    join_path_header: Optional[str] = None,
    *,
    debug_step_scope: Optional[str] = None,
    existing_bundle: Optional[dict[str, Any]] = None,
    max_primary_rows: Optional[int] = None,
    cancel_check: Optional[Callable[..., None]] = None,
) -> dict[str, Any]:
    """
    1 項目分の抽出結果を主値・連携値・結合キー値に分けて返す。

    debug_step_scope:
      None … 本番相当（主キー＋連携＋結合を一括）。
      "primary" … 主キー（と path_item／正規化フルパス）のみ。デバッグの主キーステップ用。
      "link" … existing_bundle 必須。連携キーだけ再計算してマージ。
      "join" … existing_bundle 必須。結合キーだけ再計算してマージ。

    戻り値:
      {
        "primary_values": list[Any],
        "iteration_contexts": list[dict[str, Any]],
        "link_values": dict[item_name, list[Any]],
        "link_contexts": dict[item_name, list[dict[str, Any]]],
        "join_values": dict[item_name, list[Any]],
        "join_contexts": dict[item_name, list[dict[str, Any]]],
        "path_item_values": dict[item_name, list[Any]],  # 名前系のフルパス関連付け用
        "path_item_contexts": dict[item_name, list[dict[str, Any]]],
      }
    """
    sources = item_config.get("sources") or []
    _poll_cancel_check(cancel_check, force=True)

    if debug_step_scope in ("link", "join"):
        if not existing_bundle or not isinstance(existing_bundle, dict):
            return {
                "primary_values": [],
                "iteration_contexts": [],
                "link_values": {},
                "link_contexts": {},
                "join_values": {},
                "join_contexts": {},
                "path_item_values": {},
                "path_item_contexts": {},
            }
        # 浅いコピー: 主値・連携・path_item 等は共有し、再計算する link/join だけ新規 dict（デバッグ段階実行のコスト削減）。
        bundle = existing_bundle.copy()
        if debug_step_scope == "link":
            bundle["link_values"] = {}
            bundle["link_contexts"] = {}
        else:
            bundle["join_values"] = {}
            bundle["join_contexts"] = {}
        span_map = bundle.get("_cell_source_spans")
        legacy_link_iter = not isinstance(span_map, dict)
        n_all = max(1, len(bundle.get("primary_values") or []))
        for si, src in enumerate(sources):
            if not isinstance(src, dict):
                continue
            stype = (src.get("type") or "cell").strip().lower()
            ui_blk = source_ui_block(src)
            if not isinstance(ui_blk, dict):
                continue
            if stype != "cell":
                continue
            if not source_passes_file_name_filter(file_path, src):
                continue
            if source_skips_sheet_extract(src):
                continue
            if legacy_link_iter:
                g_off, n_src = 0, n_all
            else:
                sp = span_map.get(si) if isinstance(span_map, dict) else None
                if not sp or not isinstance(sp, (tuple, list)) or len(sp) != 2:
                    continue
                try:
                    g_off = int(sp[0])
                    n_src = int(sp[1])
                except (TypeError, ValueError):
                    continue
                if n_src < 1:
                    continue
            src_base = str(src.get("cell_ref") or "A1")
            try:
                row_step = int(src.get("row_offset") or 0)
            except (TypeError, ValueError):
                row_step = 0
            try:
                col_step = int(src.get("col_offset") or 0)
            except (TypeError, ValueError):
                col_step = 0
            iter_contexts = _build_source_iter_contexts(
                file_path=file_path,
                src_base=src_base,
                row_step=row_step,
                col_step=col_step,
                g_off=g_off,
                n_src=n_src,
                bundle=bundle,
            )
            if debug_step_scope == "link":
                for ld in ui_blk.get("link_defs") or []:
                    if not isinstance(ld, dict):
                        continue
                    target = str(ld.get("item") or "").strip()
                    if not target:
                        continue
                    _append_rule_series_to_bundle(
                        bundle=bundle,
                        values_key="link_values",
                        contexts_key="link_contexts",
                        target=target,
                        file_path=file_path,
                        src=src,
                        rule=ld,
                        iter_contexts=iter_contexts,
                        n_src=n_src,
                        cancel_check=cancel_check,
                    )
            else:
                for jd in ui_blk.get("join_defs") or []:
                    if not isinstance(jd, dict):
                        continue
                    target = str(jd.get("item") or "").strip()
                    if not target:
                        continue
                    _append_rule_series_to_bundle(
                        bundle=bundle,
                        values_key="join_values",
                        contexts_key="join_contexts",
                        target=target,
                        file_path=file_path,
                        src=src,
                        rule=jd,
                        iter_contexts=iter_contexts,
                        n_src=n_src,
                        cancel_check=cancel_check,
                    )
        if debug_step_scope == "link":
            _align_link_join_series_to_primary(bundle)
            _apply_carry_empty_link_values(bundle, sources)
        elif debug_step_scope == "join":
            _align_link_join_series_to_primary(bundle)
        return bundle

    cell_spans: dict[int, tuple[int, int]] = {}
    prim_vals = extract_item_values(
        file_path,
        item_config,
        item_id=item_id,
        cell_positions=cell_positions,
        max_primary_rows=max_primary_rows,
        cell_source_spans_out=cell_spans,
        cancel_check=cancel_check,
    )
    bundle: dict[str, Any] = {
        "primary_values": prim_vals,
        "iteration_contexts": [
            {
                "file_path": str(file_path),
                "iter_index": int(i),
                "base_cell": None,
                "base_row": None,
                "base_col": None,
                "filter_snapshot": {
                    "file_path": str(file_path),
                    "passed": True,
                },
                "primary_value": prim_vals[i] if i < len(prim_vals) else None,
            }
            for i in range(len(prim_vals or []))
        ],
        "link_values": {},
        "link_contexts": {},
        "join_values": {},
        "join_contexts": {},
        "path_item_values": {},
        "path_item_contexts": {},
        "_cell_source_spans": dict(cell_spans),
    }
    run_link_join = debug_step_scope is None
    for si, src in enumerate(sources):
        _poll_cancel_check(cancel_check)
        if not isinstance(src, dict):
            continue
        stype = (src.get("type") or "cell").strip().lower()
        ui_blk = source_ui_block(src)
        if not isinstance(ui_blk, dict):
            continue
        if stype == "cell":
            if not source_passes_file_name_filter(file_path, src):
                continue
            if source_skips_sheet_extract(src):
                continue
            sp = cell_spans.get(si)
            if not sp or not isinstance(sp, (tuple, list)) or len(sp) != 2:
                continue
            try:
                g_off = int(sp[0])
                n_src = int(sp[1])
            except (TypeError, ValueError):
                continue
            if n_src < 1:
                continue
            src_base = str(src.get("cell_ref") or "A1")
            try:
                row_step = int(src.get("row_offset") or 0)
            except (TypeError, ValueError):
                row_step = 0
            try:
                col_step = int(src.get("col_offset") or 0)
            except (TypeError, ValueError):
                col_step = 0
            iter_contexts = _build_source_iter_contexts(
                file_path=file_path,
                src_base=src_base,
                row_step=row_step,
                col_step=col_step,
                g_off=g_off,
                n_src=n_src,
                bundle=bundle,
            )
            if run_link_join:
                contiguous = _iter_contexts_rule_indices_contiguous(iter_contexts)
                link_defs = [x for x in (ui_blk.get("link_defs") or []) if isinstance(x, dict)]
                link_fast = (
                    _extract_cell_rules_series_fast_map(
                        file_path, src, link_defs, n_src=n_src, cancel_check=cancel_check
                    )
                    if contiguous
                    else None
                )
                for ldi, ld in enumerate(link_defs):
                    if not isinstance(ld, dict):
                        continue
                    target = str(ld.get("item") or "").strip()
                    if not target:
                        continue
                    vals_fast = link_fast.get(ldi) if isinstance(link_fast, dict) else None
                    if isinstance(vals_fast, list):
                        for local_i, iter_ctx in enumerate(iter_contexts):
                            v = vals_fast[local_i] if local_i < len(vals_fast) else None
                            bundle["link_values"].setdefault(target, []).append(v)
                            bundle["link_contexts"].setdefault(target, []).append(
                                {
                                    "file_path": str(iter_ctx.get("file_path") or file_path),
                                    "iter_index": int(iter_ctx.get("iter_index", 0)),
                                    "base_cell": iter_ctx.get("base_cell"),
                                }
                            )
                    else:
                        _append_rule_series_to_bundle(
                            bundle=bundle,
                            values_key="link_values",
                            contexts_key="link_contexts",
                            target=target,
                            file_path=file_path,
                            src=src,
                            rule=ld,
                            iter_contexts=iter_contexts,
                            n_src=n_src,
                            cancel_check=cancel_check,
                        )
                join_defs = [x for x in (ui_blk.get("join_defs") or []) if isinstance(x, dict)]
                join_fast = (
                    _extract_cell_rules_series_fast_map(
                        file_path, src, join_defs, n_src=n_src, cancel_check=cancel_check
                    )
                    if contiguous
                    else None
                )
                for jdi, jd in enumerate(join_defs):
                    if not isinstance(jd, dict):
                        continue
                    target = str(jd.get("item") or "").strip()
                    if not target:
                        continue
                    vals_fast = join_fast.get(jdi) if isinstance(join_fast, dict) else None
                    if isinstance(vals_fast, list):
                        for local_i, iter_ctx in enumerate(iter_contexts):
                            v = vals_fast[local_i] if local_i < len(vals_fast) else None
                            bundle["join_values"].setdefault(target, []).append(v)
                            bundle["join_contexts"].setdefault(target, []).append(
                                {
                                    "file_path": str(iter_ctx.get("file_path") or file_path),
                                    "iter_index": int(iter_ctx.get("iter_index", 0)),
                                    "base_cell": iter_ctx.get("base_cell"),
                                }
                            )
                    else:
                        _append_rule_series_to_bundle(
                            bundle=bundle,
                            values_key="join_values",
                            contexts_key="join_contexts",
                            target=target,
                            file_path=file_path,
                            src=src,
                            rule=jd,
                            iter_contexts=iter_contexts,
                            n_src=n_src,
                            cancel_check=cancel_check,
                        )
        elif stype == "name_extract":
            if not name_extract_search_matches(file_path, src):
                continue
            target = str(ui_blk.get("path_item") or "").strip()
            if target:
                # 名前系の連携キー（フルパスで結合）は、まずフルパス列として保持する。
                pv = extract_metadata(file_path, SOURCE_FULL_PATH)
                bundle["path_item_values"].setdefault(target, []).append(pv)
                bundle["path_item_contexts"].setdefault(target, []).append(
                    {
                        "file_path": str(file_path),
                        "iter_index": int(len(bundle["path_item_values"][target]) - 1),
                    }
                )
    # セル座標系: 照合列（名前取得の path_item で決まる列、レガシーは join_path_item_id）へ正規化フルパスを付与
    from svc.svc_data_agg_scenario import LINEAGE_CELL, infer_item_lineage  # noqa: E402

    jh = (join_path_header or "").strip()
    if jh and infer_item_lineage(sources) == LINEAGE_CELL:
        from svc.data_agg_path_norm import normalize_source_path  # noqa: E402

        norm = normalize_source_path(file_path)
        prim = bundle.get("primary_values") or []
        n = max(1, len(prim))
        bundle["path_item_values"][jh] = [norm] * n
        bundle["path_item_contexts"][jh] = [
            {"file_path": str(file_path), "iter_index": int(i)} for i in range(n)
        ]
    _align_link_join_series_to_primary(bundle)
    _apply_skip_empty_primary_filter(bundle, sources if isinstance(sources, list) else [])
    _apply_carry_empty_link_values(bundle, sources if isinstance(sources, list) else [])
    return bundle
