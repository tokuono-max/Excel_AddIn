# -*- coding: utf-8 -*-
"""シナリオのシート名条件（左端／完全一致／含む／含まない）の解決。"""

from __future__ import annotations

from collections.abc import Sequence
from pathlib import Path
from typing import Any, Literal

SheetRuleKind = Literal["left", "exact", "contains", "not_contains"]

SHEET_MISS_LABEL = "（該当なし）"


def parse_comma_separated_patterns(raw: str | None) -> list[str]:
    """
    カンマ区切りのパターン入力をトークン化する（シート名・ファイル名で共通）。

    - 先頭・末尾・連続カンマによる空要素は捨てる（例: ``,R_,実装,`` → ``["R_", "実装"]``）
    - 各要素は strip（前後空白除去）。空白のみの要素も捨てる
    - 全体が空／空白のみ → []
    """
    s = "" if raw is None else str(raw)
    if s.strip() == "":
        return []
    out: list[str] = []
    for part in s.split(","):
        tok = str(part).strip()
        if tok:
            out.append(tok)
    return out


def parse_sheet_name_patterns(raw: str | None) -> list[str]:
    """シート名入力をカンマ区切りトークン化する。``parse_comma_separated_patterns`` と同じ。"""
    return parse_comma_separated_patterns(raw)


def classify_sheet_rule(rule: str | None) -> SheetRuleKind:
    """UI 文言／英語エイリアスを正規化する。空・不明は left（左端）扱い。"""
    r = str(rule or "").strip()
    if not r:
        return "left"
    rl = r.lower()
    if "左端" in r or rl in ("left", "leftmost"):
        return "left"
    # 「含まない」を先に判定（「含む」より長い／優先）
    if "含まない" in r or rl in ("exclude", "not_contains", "not-contains"):
        return "not_contains"
    if "含む" in r or rl in ("contains", "include"):
        return "contains"
    if "完全一致" in r or rl in ("exact", "equals"):
        return "exact"
    return "left"


def resolve_all_sheet_names_by_rule(
    sheetnames: Sequence[str],
    rule: str | None,
    pattern: str | None,
    *,
    case_sensitive: bool = True,
) -> list[str]:
    """
    ブックのシート名一覧から、条件に合うシート名を左端から順にすべて返す。

    - left: 先頭シートのみ（1 件）。pattern は無視
    - exact: パターンのいずれかと完全一致（OR）。0 件以上
    - contains: パターンのいずれかを含む（OR）
    - not_contains: パターンのいずれも含まない（＝1つでも含めば除外）
    - pattern が空（トークンなし）の exact/contains/not_contains: 空リスト
    - 複数パターンはカンマ区切り（``parse_comma_separated_patterns``）
    """
    names = [str(x) for x in sheetnames if str(x).strip() != ""]
    if not names:
        return []
    kind = classify_sheet_rule(rule)
    if kind == "left":
        return [names[0]]
    tokens = parse_comma_separated_patterns(pattern)
    if not tokens:
        return []

    def _key(s: str) -> str:
        return s if case_sensitive else s.casefold()

    toks_k = [_key(t) for t in tokens]
    if kind == "exact":
        tok_set = set(toks_k)
        return [x for x in names if _key(x) in tok_set]
    if kind == "contains":
        return [x for x in names if any(t in _key(x) for t in toks_k)]
    if kind == "not_contains":
        return [x for x in names if all(t not in _key(x) for t in toks_k)]
    return [names[0]]


def resolve_sheet_name_by_rule(
    sheetnames: Sequence[str],
    rule: str | None,
    pattern: str | None,
    *,
    case_sensitive: bool = True,
) -> str | None:
    """条件に合う最初のシート名。該当なしは None。"""
    matched = resolve_all_sheet_names_by_rule(
        sheetnames, rule, pattern, case_sensitive=case_sensitive
    )
    return matched[0] if matched else None


def list_workbook_sheet_names(file_path: str | Path) -> list[str] | None:
    """
    ブックのシート名一覧。
    CSV などシート概念なしは None。読取失敗は []。
    """
    p = Path(file_path)
    suffix = p.suffix.lower()
    if suffix == ".csv":
        return None
    if suffix == ".xls":
        from svc.data_agg_xls_io import list_xls_sheet_names, xls_reader_unavailable_message

        if xls_reader_unavailable_message():
            return []
        return list_xls_sheet_names(p)
    if suffix in (".xlsx", ".xlsm"):
        try:
            import openpyxl  # noqa: E402
        except Exception:
            return []
        try:
            wb = openpyxl.load_workbook(p, read_only=True, data_only=True)
            names = list(wb.sheetnames or [])
            wb.close()
            return [str(x) for x in names if str(x).strip() != ""]
        except Exception:
            return []
    return []


SKIP_SHEET_EXTRACT_KEY = "_data_agg_skip_sheet"


def source_skips_sheet_extract(src: dict[str, Any] | None) -> bool:
    """当該シート向けに抑制した cell ソースか。"""
    return isinstance(src, dict) and bool(src.get(SKIP_SHEET_EXTRACT_KEY))


def _is_cell_extract_source(src: dict[str, Any] | None) -> bool:
    if not isinstance(src, dict):
        return False
    typ = str(src.get("type") or "cell").strip().lower()
    return typ not in ("name_extract", "metadata", "meta", "filename")


def matching_sheets_for_cell_source(
    file_path: str | Path,
    src: dict[str, Any] | None,
) -> list[str] | None:
    """
    セル系ソースのシート名条件に合うシート一覧（左→右）。
    名前取得系・CSV などシート解決不要時は None。
    該当なし・読取失敗は []。
    """
    if not isinstance(src, dict):
        return None
    typ = str(src.get("type") or "cell").strip().lower()
    if typ in ("name_extract", "metadata", "meta", "filename"):
        return None
    from svc.data_agg_source_ui import source_ui_block

    sn = str(src.get("sheet_name") or "").strip()
    pb = source_ui_block(src) or {}
    rule = str(pb.get("sheet_rule") or "")
    names = list_workbook_sheet_names(file_path)
    if names is None:
        return None
    if not names:
        return []
    return resolve_all_sheet_names_by_rule(names, rule, sn)


def patch_item_sheet_exact(
    item: dict[str, Any],
    sheet_name: str,
    *,
    workbook_sheet_names: Sequence[str] | None = None,
) -> dict[str, Any]:
    """
    cell ソースの sheet_name を実名にし、sheet_rule を完全一致にする。
    workbook_sheet_names があるとき、元のシート条件に合わないソースは抽出抑制する
    （インデックスを保ったまま当該シートでは読まない）。
    """
    import copy

    from svc.data_agg_source_ui import ensure_source_ui_block, source_ui_block

    out = copy.deepcopy(item)
    names = [str(x) for x in (workbook_sheet_names or []) if str(x).strip() != ""]
    for src in out.get("sources") or []:
        if not _is_cell_extract_source(src):
            continue
        if names:
            orig_sn = str(src.get("sheet_name") or "").strip()
            orig_rule = str((source_ui_block(src) or {}).get("sheet_rule") or "")
            matched = resolve_all_sheet_names_by_rule(names, orig_rule, orig_sn)
            if sheet_name not in matched:
                src[SKIP_SHEET_EXTRACT_KEY] = True
                continue
        src.pop(SKIP_SHEET_EXTRACT_KEY, None)
        src["sheet_name"] = sheet_name
        ensure_source_ui_block(src)["sheet_rule"] = "完全一致"
    return out
